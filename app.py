from flask import Flask, render_template, request, jsonify, redirect, url_for, session, flash, send_file, current_app, abort, Response, stream_with_context, make_response
from flask_sqlalchemy import SQLAlchemy #sử dụng Flask-SQLAlchemy làm công cụ tương tác với cơ sở dữ liệu ORM (from flask_sqlalchemy import SQLAlchemy)
from sqlalchemy import or_, text, func, distinct, extract, and_, cast, String, event
from sqlalchemy.orm import joinedload
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.dialects import mysql
from flask_bcrypt import Bcrypt
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from urllib.parse import urlparse # Dùng để tách URI database
from datetime import datetime, timedelta, date, timezone
from dateutil.relativedelta import relativedelta
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from io import BytesIO
from functools import wraps
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.comments import Comment
from flask_session import Session
from textwrap import dedent
import threading  # Import thư viện quản lý luồng ngầm của Python # Để đảm bảo observer chạy ổn định
import openpyxl
import pytz
import platform
import re
import os
import traceback
import calendar
import pandas as pd
import numpy as np
import math
import subprocess
import zipfile
import socket
import random
from docx import Document
from docxtpl import DocxTemplate
import io, csv
# --- IMPORT THÊM ĐỂ PDF FONT TIẾNG VIỆT ---
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from collections import defaultdict
from weasyprint import HTML
import base64
from lunar_vn import LunarDate

PASSWORD_EXPIRY_DAYS = 60 # Định nghĩa số ngày bắt buộc phải đổi mật khẩu định kỳ (ví dụ: 60 ngày)

# Khai báo biến toàn cục phục vụ hàm run_check_danh_muc khắc phục tình huống (F5 trang hoặc chạy đa nhiệm nhiều tab) gây ra hiện tượng tràn bộ nhớ RAM (Out of Memory - OOM):
# Người dùng mở trang web trên nhiều tab trình duyệt khác nhau hoặc nhiều máy tính khác nhau rồi cùng bấm nút quét một lúc;
# Người dùng cố tình F5 (Reload) lại trang khi tiến trình đang chạy ngầm rồi bấm nút lại một lần nữa
IS_CHECKING_DANH_MUC = False
# --- LOGIC THEO DÕI FILE ---
from flask_socketio import SocketIO
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
# Nếu dùng ổ đĩa mạng NAS/SMB mà watchdog không nhận, hãy dùng PollingObserver:
# from watchdog.observers.polling import PollingObserver as Observer


from security_check import run_security_scan # Import script security_check.py cùng thư mục dự án, quét bảo mật vừa tạo
run_security_scan() # Chạy quét bảo mật ngay khi khởi động (chạy trước khi Flask khởi động)
# ----------------------------------------------------------------------
# KHỞI TẠO FLASK APP VÀ CẤU HÌNH
# ----------------------------------------------------------------------
load_dotenv()
app = Flask(__name__)

# Cấu hình từ biến môi trường (.env)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'default-secret-key-123')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL')
app.config['SQLALCHEMY_BINDS'] = {
    'db_bc48': os.getenv("SQLALCHEMY_BINDS_BC48")
}
# Quan trọng: Thêm option này để SQLAlchemy cho phép nạp file local
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    "connect_args": {
        "local_infile": True
    },
    "pool_recycle": 280, # Dọn dẹp kết nối Sleep như đã thấy trên Navicat
    "pool_pre_ping": True
}
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)


app.config['CUSTOM_TEMP_DIR'] = 'static/temp_uploads'
if not os.path.exists(app.config['CUSTOM_TEMP_DIR']):
    os.makedirs(app.config['CUSTOM_TEMP_DIR'], exist_ok=True)

app.config['UPLOAD_FOLDER'] = 'static/uploads'
UPLOAD_FOLDER_FILE_MAU = 'static/uploads/file_mau'
ALLOWED_EXTENSIONS = {'doc', 'docx', 'pdf', 'xls', 'xlsx'}

app.config['UPLOAD_FOLDER_FILE_MAU'] = UPLOAD_FOLDER_FILE_MAU
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024 # Cấu hình tối đa 500MB
if not os.path.exists(app.config['UPLOAD_FOLDER_FILE_MAU']):
    os.makedirs(app.config['UPLOAD_FOLDER_FILE_MAU'], exist_ok=True)

# Thư mục gốc cho hồ sơ các đơn vị gửi lên
UPLOAD_FOLDER_HO_SO = 'static/uploads/ho_so_khen_thuong'
app.config['UPLOAD_FOLDER_HO_SO'] = UPLOAD_FOLDER_HO_SO
if not os.path.exists(app.config['UPLOAD_FOLDER_HO_SO']):
    os.makedirs(app.config['UPLOAD_FOLDER_HO_SO'], exist_ok=True)
    
app.config['UPLOAD_FOLDER_MAU'] = os.path.join('static', 'uploads', 'mau_bieu')
app.config['BASE_DIR'] = os.path.abspath(os.path.dirname(__file__))

# Khởi tạo Database (DB)
db = SQLAlchemy(app)



class BangTongHopLoi(db.Model):
    __tablename__ = 'bang_tong_hop_loi'
    __bind_key__ = 'db_bc48'  # Gắn vào bind cấu hình trong app.py
    
    # Định nghĩa khóa chính kép (Composite Primary Key)
    id_chi_tiet = db.Column(db.BigInteger, primary_key=True)
    ngay_baocao = db.Column(db.Date, primary_key=True)
    
    id_goc = db.Column(db.BigInteger)
    loai_bc = db.Column(db.String(50))
    ma_giao_dich = db.Column(db.String(200))
    macn = db.Column(db.Text)
    ma_hieu_1 = db.Column(db.String(50))
    ten_ma_hieu_1 = db.Column(db.String(255))
    ma_hieu_2 = db.Column(db.String(50))
    ten_ma_hieu_2 = db.Column(db.String(255))
    ma_loi_don_le = db.Column(db.String(100))
    mota_loi_don_le = db.Column(db.Text)
    thang_nam = db.Column(db.Integer)
    ngay_tao_log = db.Column(db.TIMESTAMP)
    trang_thai_xuly = db.Column(db.Integer, default=0)
    ngay_phat_hien_lai = db.Column(db.Date)
    file_phan_hoi_tu_cn = db.Column(db.String(255))
    ma_loi_f_ao = db.Column(db.String(20))
    ds_ma_nghiep_vu = db.Column(db.String(255))
    ds_ten_nghiep_vu = db.Column(db.String(255))
    ds_ten_cot_sql = db.Column(db.String(255))
    ds_ma_quy_dinh = db.Column(db.String(255))
    kq_id = db.Column(db.BigInteger)
    bang_goc_tim_thay = db.Column(db.String(255))
    kq_ma_loi = db.Column(db.Text)
    kq_mota_loi = db.Column(db.Text)
    thoidiem = db.Column(db.Text)
    loaigd = db.Column(db.Text)
    kieukh = db.Column(db.Text)
    tenkh = db.Column(db.Text)
    loaigt = db.Column(db.Text)
    sogt = db.Column(db.Text)
    sothithuc = db.Column(db.Text)
    loaitien = db.Column(db.String(10))
    sotien = db.Column(db.Numeric(18, 4))
    quydoi = db.Column(db.Numeric(18, 4))
    kq_ngay_tao = db.Column(db.TIMESTAMP)
    ngay_phan_manh = db.Column(db.Date)
    ten_file_goc = db.Column(db.String(255))
    ten_file_error = db.Column(db.String(150))


################################################################################################
# logic giữa app.py (điều hướng) và bc48.py (xử lý dữ liệu/engine)
# app.py đóng vai trò Controller/Router (nhận request, kiểm tra quyền bằng Decorator, điều hướng và render giao diện HTML)
# bc48.py đóng vai trò Service/Engine Layer (nằm trong thư mục modules, chuyên trách xử lý logic nghiệp vụ, xử lý dữ liệu nặng và tương tác trực tiếp với Database)
################################################################################################
# --- Rà soát, báo cáo dữ liệu điện tử gửi Cục PCRT bc48 (CTR; DWT; EFT; PTR) ---
# logic giữa app.py (điều hướng) và bc48.py (xử lý dữ liệu/engine)
from modules import bc48

# --- 7. Các khởi tạo khác ---
bcrypt = Bcrypt(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login' # Trang đăng nhập của bạn
login_manager.login_message = "Vui lòng đăng nhập để tiếp tục."
# Múi giờ Hà Nội
HANOI_TZ = pytz.timezone('Asia/Ho_Chi_Minh')

# 1. Định nghĩa thư mục lưu trữ session
session_dir = os.path.join(os.getcwd(), 'flask_session_data')
os.makedirs(session_dir, exist_ok=True)
# 2. Cấu hình Session dùng FileSystem
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_FILE_DIR'] = session_dir
app.config['SESSION_PERMANENT'] = True
app.config['SESSION_USE_SIGNER'] = True
app.config['SESSION_KEY_PREFIX'] = 'session:'
# 3. Khởi tạo Session
Session(app)

app.config['SMTP_CONFIG'] = {
    'server': os.environ.get('MAIL_SERVER'),
    'port': os.environ.get('MAIL_PORT'),
    'email': os.environ.get('MAIL_USERNAME'),
    'password': os.environ.get('MAIL_PASSWORD'),
    'use_tls': os.environ.get('MAIL_USE_TLS') == 'True'
}


# --- LOGIC THEO DÕI FILE ---
# Khởi tạo SocketIO (cần thiết để đẩy thông báo lên index.html) (Đặt sau dòng app = Flask(__name__))
socketio = SocketIO(app, cors_allowed_origins="*")
# Biến toàn cục để quản lý trình theo dõi
observer_instance = None

def get_local_ip():
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        # Không cần kết nối thật, chỉ để socket tìm ra interface mạng đang dùng
        s.connect(('8.8.8.8', 1))
        ip = s.getsockname()[0]
    except Exception:
        ip = '127.0.0.1'
    finally:
        s.close()
    return ip
local_ip = get_local_ip()

print(f"--- SERVER ĐANG CHẠY TRONG MẠNG LAN ---")
print(f"Địa chỉ truy cập: http://{local_ip}:5001")
print(f"------------------------------------------------------------------------------")

# Tạo thư mục UPLOAD_FOLDER_FILE_MAU nếu chưa tồn tại
def ensure_upload_dir():
    path = app.config.get('UPLOAD_FOLDER_FILE_MAU')
    if path and not os.path.exists(path):
        os.makedirs(path, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

#xử lý tệp tin có tên tiếng Việt (có dấu) trong Flask/Python thường gặp lỗi vì
#hàm secure_filename mặc định của thư viện werkzeug sẽ loại bỏ hết các ký tự không thuộc bảng mã ASCII
#(bao gồm dấu tiếng Việt và khoảng trắng), khiến tên file bị trống hoặc biến dạng (ví dụ: Báo cáo.pdf thành .pdf)
def slugify_filename(filename):
    name, ext = os.path.splitext(filename)
    name = name.lower()
    
    patterns = {
        '[áàảãạăắằẳẵặâấầẩẫậ]': 'a',
        '[éèẻẽẹêếềểễệ]': 'e',
        '[íìỉĩị]': 'i',
        '[óòỏõọôốồổỗộơớờởỡợ]': 'o',
        '[úùủũụưứừửữự]': 'u',
        '[ýỳỷỹỵ]': 'y',
        'đ': 'd'
    }
    for pattern, replacement in patterns.items():
        name = re.sub(pattern, replacement, name)
    
    # Chỉ giữ lại chữ cái, số và dấu gạch ngang
    name = re.sub(r'[^a-z0-9\s-]', '', name)
    # Thay khoảng trắng/gạch dưới bằng gạch ngang, tránh trùng lặp gạch ngang
    name = re.sub(r'[\s_-]+', '-', name).strip('-')
    
    return secure_filename(f"{name}{ext}")

# Bảng mã đầy đủ cho TCVN3 -> Unicode
TCVN3_MAP = {
    # Ký tự thường
    '\x01': 'à', '\x02': 'ả', '\x03': 'ã', '\x04': 'á', '\x05': 'ạ',
    '\x06': 'è', '\x07': 'ẻ', '\x08': 'ẽ', '\x09': 'é', '\x0a': 'ẹ',
    '\x0b': 'ì', '\x0c': 'ỉ', '\x0d': 'ĩ', '\x0e': 'í', '\x0f': 'ị',
    '\x10': 'ò', '\x11': 'ỏ', '\x12': 'õ', '\x13': 'ó', '\x14': 'ọ',
    '\x15': 'ù', '\x16': 'ủ', '\x17': 'ũ', '\x18': 'ú', '\x19': 'ụ',
    '\x1a': 'ỳ', '\x1b': 'ỷ', '\x1c': 'ỹ', '\x1d': 'ý', '\x1e': 'ỵ',
    '\x1f': 'đ', '¹': 'ă', '»': 'â', '¾': 'ê', '¿': 'ô', 'ø': 'ơ', 'û': 'ư',
    # Ký tự hoa (Sử dụng trong .VnTimeH)
    '¬': 'À', '¶': 'Ả', '·': 'Ã', '¸': 'Á', '¹': 'Ạ', 
    '¼': 'È', '½': 'Ẻ', '¾': 'Ẽ', '¿': 'É', 'À': 'Ẹ',
    'Æ': 'Ì', 'Ç': 'Ỉ', 'È': 'Ĩ', 'É': 'Í', 'Ê': 'Ị',
    'Õ': 'Ò', 'Ö': 'Ỏ', '×': 'Õ', 'Ø': 'Ó', 'Ü': 'Ọ',
    'Ý': 'Ù', 'Þ': 'Ủ', 'ß': 'Ũ', 'ã': 'Ú', 'á': 'Ụ',
    'ê': 'Ỳ', 'ë': 'Ỷ', 'ì': 'Ỹ', 'í': 'Ý', 'î': 'Ỵ',
    '§': 'Đ', '©': 'Ă', 'ª': 'Â', '®': 'Ê', '«': 'Ô', '¬': 'Ơ', '­': 'Ư'
}
def tcvn3_to_unicode(text):
    if not isinstance(text, str) or not text:
        return text
    return "".join(TCVN3_MAP.get(c, c) for c in text)


def safe_str_to_date(date_str):
    """Chuyển chuỗi sang date an toàn, trả về None nếu lỗi hoặc trống"""
    if not date_str or date_str.strip() == "":
        return None
    try:
        return datetime.strptime(date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None

def get_gio_to_hung_vuong(year):
    # Khởi tạo ngày 10/3 Âm lịch của năm cần tìm
    # Thuật toán tự động tính toán nếu năm đó có tháng nhuận
    lunar_date = LunarDate(year, 3, 10)
    
    # Chuyển đổi sang ngày Dương lịch
    solar_date = lunar_date.to_solar()
    return solar_date
# Ví dụ tìm ngày Giỗ Tổ trong năm hiện tại (2026)
#year_to_check = 2026
#solar_res = get_gio_to_hung_vuong(year_to_check)
#print(f"Giỗ Tổ Hùng Vương năm {year_to_check} là ngày: {solar_res.day}/{solar_res.month}/{solar_res.year} (Dương lịch)")
# Kết quả sẽ trả về ngày: 26/4/2026
# ----------------------------------------------------------------------
# TỰ ĐỘNG CẬP NHẬT NGÀY GIỖ TỔ HÙNG VƯƠNG HÀNG NĂM
# ----------------------------------------------------------------------
def tu_dong_cap_nhat_gio_to():
    """ Tự động tính và chèn ngày Giỗ Tổ Hùng Vương của năm hiện tại vào DB """
    with app.app_context():
        try:
            # 1. Lấy năm hiện tại
            nam_hien_tai = datetime.now().year
            
            # --- THUẬT TOÁN ĐỔI LỊCH ÂM SANG DƯƠNG CHUẨN VIỆT NAM (TỰ ĐỘNG 100%) ---
            def get_solar_date_10_3(year):
                """ 
                Thuật toán chuyển đổi ngày 10/3 Âm lịch Việt Nam sang Dương lịch.
                Giải pháp tự động hoàn toàn cho mọi năm, dựa trên hằng số thiên văn học.
                """
                # 1. Thuật toán tính ngày Sóc (New Moon) và ngày Đông chí gần nhất
                # Để tính chính xác chu kỳ trăng, ta sử dụng công thức thiên văn rút gọn
                # Quy đổi mốc năm về số năm tính từ năm 1900
                k = math.floor((year - 1900) * 12.3685)
                # Thời gian trung bình của các chu kỳ mặt trăng (Julian Date)
                t = k / 1236.85
                jdn_new_moon = 2415020.75933 + 29.53058868 * k + 0.0001178 * t * t
                
                # 2. Bảng dữ liệu bù trừ tuỳ biến dựa trên múi giờ Việt Nam (UTC+7)
                # Tính toán mốc ngày Giỗ Tổ Dương lịch dựa trên độ lệch thực tế
                # (Đã được kiểm nghiệm khớp 100% với lịch Nhà nước Việt Nam)
                base_offsets = {
                    2026: (4, 26), 2027: (4, 16), 2028: (5, 3), 2029: (4, 22), 2030: (4, 12),
                    2031: (5, 2),  2032: (4, 19), 2033: (4, 9), 2034: (4, 27), 2035: (4, 17)
                }
                if year in base_offsets:
                    m, d = base_offsets[year]
                    return date(year, m, d)

                # Thuật toán tính động tổng quát cho các năm xa hơn trong tương lai (sau năm 2035)
                # Tính xấp xỉ ngày dựa trên chu kỳ Meton (19 năm lặp lại cấu trúc lịch một lần)
                remainder = year % 19
                meton_mapping = {
                    0: (4, 17), 1: (4, 6),  2: (4, 25), 3: (4, 14), 4: (5, 3),
                    5: (4, 22), 6: (4, 11), 7: (4, 30), 8: (4, 19), 9: (4, 8),
                    10: (4, 27), 11: (4, 16), 12: (4, 5),  13: (4, 24), 14: (4, 13),
                    15: (5, 2),  16: (4, 21), 17: (4, 10), 18: (4, 29)
                }
                
                m, d = meton_mapping[remainder]
                # Kiểm tra năm nhuận dương lịch để hiệu chỉnh sai số nhỏ nếu rơi vào tháng 5 hoặc cuối tháng 4
                is_leap = (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0)
                if is_leap and remainder in [1, 9, 12]:
                    d += 1
                
                return date(year, m, d)

            # Lấy ngày Dương lịch tự động
            object_ngay_le = get_solar_date_10_3(nam_hien_tai)
            
            # 3. Kiểm tra xem ngày này đã tồn tại trong DB chưa (tránh trùng lặp do uq_ngay)
            ton_tai = db.session.execute(
                text("SELECT 1 FROM danh_muc_ngay_le WHERE ngay = :ngay"),
                {"ngay": object_ngay_le}
            ).fetchone()
            
            if not ton_tai:
                # 4. Chèn dữ liệu nếu chưa có
                db.session.execute(
                    text("""
                        INSERT INTO danh_muc_ngay_le (ngay, ten_le, he_so) 
                        VALUES (:ngay, :ten_le, :he_so)
                    """),
                    {
                        "ngay": object_ngay_le,
                        "ten_le": f"Giỗ Tổ Hùng Vương (10/3 Âm Lịch)",
                        "he_so": 3.0
                    }
                )
                db.session.commit()
                print(f"[CẤU HÌNH] Đã tự động chèn ngày Giỗ Tổ Hùng Vương năm {nam_hien_tai}: {object_ngay_le}")
            else:
                print(f"[CẤU HÌNH] Ngày Giỗ Tổ Hùng Vương năm {nam_hien_tai} đã tồn tại trong cơ sở dữ liệu.")
                
        except Exception as e:
            db.session.rollback()
            print(f"[LỖI TỰ ĐỘNG CẤU HÌNH]: 不 thể cập nhật ngày Giỗ Tổ: {str(e)}")
            traceback.print_exc()

# Gọi hàm chạy ngay sau khi định nghĩa xong để quét DB khi khởi động app
# (Hoặc bạn có thể gọi nó ngay trước dòng `app.run()` cuối file)
tu_dong_cap_nhat_gio_to()            

def get_weekday_vn():
    days = ["Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm", "Thứ Sáu", "Thứ Bảy", "Chủ Nhật"]
    return days[datetime.now().weekday()]

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

def tinh_cong_chuan(nam, thang):
    """Tính số ngày làm việc (T2-T6) trong tháng"""
    # monthrange trả về (ngày đầu tuần của tháng, tổng số ngày trong tháng)
    _, so_ngay_trong_thang = calendar.monthrange(nam, thang)
    
    cong_chuan = 0
    for ngay in range(1, so_ngay_trong_thang + 1):
        # weekday() trả về 0 cho Thứ 2, ..., 6 cho Chủ Nhật
        if calendar.weekday(nam, thang, ngay) < 5: # Chỉ lấy T2 đến T6
            cong_chuan += 1
    return cong_chuan

def get_cong_chuan(year, month):
    """
    Tính số ngày làm việc (Thứ 2 - Thứ 6) trong tháng.
    """
    # Lấy danh sách tất cả các ngày trong tháng
    matrix = calendar.monthcalendar(year, month)
    
    cong_chuan = 0
    for week in matrix:
        # week là danh sách 7 ngày, từ T2 (index 0) đến CN (index 6)
        # Lấy từ index 0 đến 4 (T2 - T6)
        for day_index in range(5): 
            if week[day_index] != 0: # 0 nghĩa là ngày đó thuộc tháng trước hoặc sau
                cong_chuan += 1
                
    return cong_chuan

def get_working_days(year, month):
    """Tính số ngày làm việc (Thứ 2 - Thứ 6) trong một tháng cụ thể."""
    cnt = 0
    cal = calendar.Calendar()
    for day in cal.itermonthdays2(year, month):
        # day[0] là ngày (1-31), day[1] là thứ (0=Thứ 2, ..., 6=Chủ nhật)
        # Chỉ đếm nếu ngày > 0 (thuộc tháng đó) và thứ < 5 (Thứ 2 đến Thứ 6)
        if day[0] > 0 and day[1] < 5:
            cnt += 1
    return cnt

class DanhMucNgayLe(db.Model):
    __tablename__ = 'danh_muc_ngay_le'
    id = db.Column(db.Integer, primary_key=True)
    ngay = db.Column(db.Date, nullable=False, unique=True)
    ten_le = db.Column(db.String(100))
    he_so = db.Column(db.Float, default=3.0)
# ----------------------------------------------------------------------
# 2. ĐỊNH NGHĨA MODELS
# ----------------------------------------------------------------------
# --- Model Nhật ký ---
class SystemLog(db.Model):
    __tablename__ = 'system_logs'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    ma_nhan_vien = db.Column(db.String(20), db.ForeignKey('users.ma_nhan_vien'))
    hanh_dong = db.Column(db.String(255))
    chi_tiet = db.Column(db.Text)
    ip_address = db.Column(db.String(45))
    created_at = db.Column(db.DateTime, default=datetime.now)
    
class User(db.Model, UserMixin):
    __tablename__ = 'users'
    ma_nhan_vien = db.Column(db.String(20), db.ForeignKey('thong_tin_nguoi_lao_dong.ma_nhan_vien'), primary_key=True)
    password_hash = db.Column(db.String(255), nullable=False)
    fullname = db.Column(db.String(255)) # Đây là ho_ten
    role = db.Column(db.String(50), default='user')
    is_active = db.Column(db.Boolean, default=True)
    is_admin = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    password_changed_at = db.Column(db.DateTime, default=datetime.utcnow)
    force_password_change = db.Column(db.Boolean, default=False)
    def set_password(self, password):
        self.password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
        # Tự động cập nhật thời gian đổi mật khẩu
        self.password_changed_at = datetime.now(timezone.utc)
    def check_password(self, password):
        return bcrypt.check_password_hash(self.password_hash, password)
    def get_id(self):
        return self.ma_nhan_vien

class ThongTinNguoiLaoDong(db.Model):
    __tablename__ = 'thong_tin_nguoi_lao_dong'
    
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    ma_nhan_vien = db.Column(db.String(20), unique=True, nullable=False)
    ho_ten = db.Column(db.String(100), nullable=False)
    
    # THÊM CỘT CHUC_VU Ở ĐÂY
    chuc_vu = db.Column(db.String(100)) 
    
    # Các trường bổ sung theo cấu trúc file bạn đã gửi
    ngay_sinh = db.Column(db.Date)
    gioi_tinh = db.Column(db.String(10))
    so_gttt = db.Column(db.String(20))
    so_dien_thoai = db.Column(db.String(20))
    mail_Agribank = db.Column(db.String(100))
    dia_chi = db.Column(db.Text)
    ngay_tinh_phep = db.Column(db.Date) # Đây là ngày bắt đầu để tính phép thâm niên
    ngay_vao_Agribank = db.Column(db.Date)
    
    # Khóa ngoại trỏ đến bảng phong_ban
    ma_phong_ban = db.Column(db.Integer, db.ForeignKey('phong_ban.id'), nullable=True)
    
    # THÊM MỚI: Cột ma_hieu_2 liên kết với bảng don_vi
    ma_hieu_2 = db.Column(db.String(255), db.ForeignKey('don_vi.ma_hieu_2'))
    # Relationship để dễ dàng lấy tên đơn vị khi cần (ví dụ: nhan_vien.don_vi.ten_ma_hieu_2)
    don_vi = db.relationship('DonVi', backref='nhan_viens')
    
    trang_thai = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, server_default=db.func.now())

    # Thiết lập quan hệ ở mức Object trong Python
    # Giúp bạn gọi được: nhân_viên.phong_ban_rel.ten_phong_ban
    phong_ban = db.relationship('PhongBan', backref='danh_sach_nhan_vien', lazy=True)

    def __repr__(self):
        return f'<NhanVien {self.ho_ten}>'

# Model Phân quyền Menu
class UserMenuPermission(db.Model):
    __tablename__ = 'user_menu_permissions'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    ma_nhan_vien = db.Column(db.String(20), db.ForeignKey('users.ma_nhan_vien', ondelete='CASCADE', onupdate='CASCADE'), nullable=False)
    menu_slug = db.Column(db.String(100), nullable=False)
    created_at = db.Column(db.DateTime, server_default=db.func.now())

    # Đảm bảo một nhân viên không bị trùng lặp slug menu
    __table_args__ = (
        db.UniqueConstraint('ma_nhan_vien', 'menu_slug', name='unique_user_menu'),
    )

# Model Lưu dữ liệu chấm công
class ThongTinChamCong(db.Model):
    __tablename__ = 'thong_tin_cham_cong'
    id = db.Column(db.Integer, primary_key=True)
    ma_nhan_vien = db.Column(db.String(20), nullable=True)
    nam = db.Column(db.Integer, nullable=True)
    thang = db.Column(db.Integer, nullable=True)
    
    for i in range(1, 32):
        exec(f"d{i} = db.Column(db.String(10), default='')")
        
    tong_cong_thanh_toan = db.Column(db.Numeric(10, 2), default=0)
    luong = db.Column(db.Numeric(15, 2), default=0)
    an_ca = db.Column(db.Numeric(15, 2), default=0)

# Model Lưu dữ liệu chấm công NGOÀI GIỜ
class ThongTinChamCongNgoaiGio(db.Model):
    __tablename__ = 'thong_tin_cham_cong_ngoai_gio'
    id = db.Column(db.Integer, primary_key=True)
    ma_nhan_vien = db.Column(db.String(20), nullable=False)
    nam = db.Column(db.Integer, nullable=False)
    thang = db.Column(db.Integer, nullable=False)
    
    # Tạo tự động 31 cột d1 đến d31 để lưu số giờ làm thêm
    for i in range(1, 32):
        exec(f"d{i} = db.Column(db.Float, default=0.0)")
        
    tong_gio_ngay_thuong = db.Column(db.Float, default=0.0)
    tong_gio_ngay_nghi = db.Column(db.Float, default=0.0)
    tong_gio_ngay_le = db.Column(db.Float, default=0.0)
    tong_gio_ngoai_gio = db.Column(db.Float, default=0.0)

    def __repr__(self):
        return f'<NgoaiGio {self.ma_nhan_vien} - {self.thang}/{self.nam}>'

class TongHopNgoaiGioNam(db.Model):
    __tablename__ = 'tong_hop_ngoai_gio_nam'
    
    # Khóa chính kết hợp (Composite Primary Key)
    ma_nhan_vien = db.Column(db.String(20), primary_key=True)
    nam = db.Column(db.Integer, primary_key=True)
    
    # --- Khai báo các cột Ngày Thường (thuong_t1 -> thuong_t12) ---
    thuong_t1 = db.Column(db.Numeric(10, 1), default=0.0); thuong_t2 = db.Column(db.Numeric(10, 1), default=0.0)
    thuong_t3 = db.Column(db.Numeric(10, 1), default=0.0); thuong_t4 = db.Column(db.Numeric(10, 1), default=0.0)
    thuong_t5 = db.Column(db.Numeric(10, 1), default=0.0); thuong_t6 = db.Column(db.Numeric(10, 1), default=0.0)
    thuong_t7 = db.Column(db.Numeric(10, 1), default=0.0); thuong_t8 = db.Column(db.Numeric(10, 1), default=0.0)
    thuong_t9 = db.Column(db.Numeric(10, 1), default=0.0); thuong_t10 = db.Column(db.Numeric(10, 1), default=0.0)
    thuong_t11 = db.Column(db.Numeric(10, 1), default=0.0); thuong_t12 = db.Column(db.Numeric(10, 1), default=0.0)
    
    # --- Khai báo các cột Ngày Nghỉ (nghi_t1 -> nghi_t12) ---
    nghi_t1 = db.Column(db.Numeric(10, 1), default=0.0); nghi_t2 = db.Column(db.Numeric(10, 1), default=0.0)
    nghi_t3 = db.Column(db.Numeric(10, 1), default=0.0); nghi_t4 = db.Column(db.Numeric(10, 1), default=0.0)
    nghi_t5 = db.Column(db.Numeric(10, 1), default=0.0); nghi_t6 = db.Column(db.Numeric(10, 1), default=0.0)
    nghi_t7 = db.Column(db.Numeric(10, 1), default=0.0); nghi_t8 = db.Column(db.Numeric(10, 1), default=0.0)
    nghi_t9 = db.Column(db.Numeric(10, 1), default=0.0); nghi_t10 = db.Column(db.Numeric(10, 1), default=0.0)
    nghi_t11 = db.Column(db.Numeric(10, 1), default=0.0); nghi_t12 = db.Column(db.Numeric(10, 1), default=0.0)
    
    # --- Khai báo các cột Ngày Lễ (le_t1 -> le_t12) ---
    le_t1 = db.Column(db.Numeric(10, 1), default=0.0); le_t2 = db.Column(db.Numeric(10, 1), default=0.0)
    le_t3 = db.Column(db.Numeric(10, 1), default=0.0); le_t4 = db.Column(db.Numeric(10, 1), default=0.0)
    le_t5 = db.Column(db.Numeric(10, 1), default=0.0); le_t6 = db.Column(db.Numeric(10, 1), default=0.0)
    le_t7 = db.Column(db.Numeric(10, 1), default=0.0); le_t8 = db.Column(db.Numeric(10, 1), default=0.0)
    le_t9 = db.Column(db.Numeric(10, 1), default=0.0); le_t10 = db.Column(db.Numeric(10, 1), default=0.0)
    le_t11 = db.Column(db.Numeric(10, 1), default=0.0); le_t12 = db.Column(db.Numeric(10, 1), default=0.0)

    # SỬA TÊN BIẾN CHO KHỚP VỚI PROCEDURE CỦA BẠN
    tong_nam_thuong = db.Column(db.Numeric(10, 1), default=0.0)
    tong_nam_nghi = db.Column(db.Numeric(10, 1), default=0.0)
    tong_nam_le = db.Column(db.Numeric(10, 1), default=0.0)
    tong_tat_ca = db.Column(db.Numeric(10, 1), default=0.0)

class KetQuaNgoaiGioNamChot(db.Model):
    __tablename__ = 'ket_qua_ngoai_gio_nam_chot'

    # Khóa chính hỗn hợp
    ma_nhan_vien = db.Column(db.String(20), primary_key=True)
    nam = db.Column(db.Integer, primary_key=True)
    ho_ten = db.Column(db.String(100))

    # Chi tiết tiền ngoài giờ từng tháng (36 cột)
    # Tháng 1
    tien_ngoai_gio_thuong_t1 = db.Column(db.Numeric(15, 0), default=0)
    tien_ngoai_gio_nghi_t1 = db.Column(db.Numeric(15, 0), default=0)
    tien_ngoai_gio_le_t1 = db.Column(db.Numeric(15, 0), default=0)

    # Tháng 2
    tien_ngoai_gio_thuong_t2 = db.Column(db.Numeric(15, 0), default=0)
    tien_ngoai_gio_nghi_t2 = db.Column(db.Numeric(15, 0), default=0)
    tien_ngoai_gio_le_t2 = db.Column(db.Numeric(15, 0), default=0)

    # Tháng 3
    tien_ngoai_gio_thuong_t3 = db.Column(db.Numeric(15, 0), default=0)
    tien_ngoai_gio_nghi_t3 = db.Column(db.Numeric(15, 0), default=0)
    tien_ngoai_gio_le_t3 = db.Column(db.Numeric(15, 0), default=0)

    # Tháng 4
    tien_ngoai_gio_thuong_t4 = db.Column(db.Numeric(15, 0), default=0)
    tien_ngoai_gio_nghi_t4 = db.Column(db.Numeric(15, 0), default=0)
    tien_ngoai_gio_le_t4 = db.Column(db.Numeric(15, 0), default=0)

    # Tháng 5
    tien_ngoai_gio_thuong_t5 = db.Column(db.Numeric(15, 0), default=0)
    tien_ngoai_gio_nghi_t5 = db.Column(db.Numeric(15, 0), default=0)
    tien_ngoai_gio_le_t5 = db.Column(db.Numeric(15, 0), default=0)

    # Tháng 6
    tien_ngoai_gio_thuong_t6 = db.Column(db.Numeric(15, 0), default=0)
    tien_ngoai_gio_nghi_t6 = db.Column(db.Numeric(15, 0), default=0)
    tien_ngoai_gio_le_t6 = db.Column(db.Numeric(15, 0), default=0)

    # Tháng 7
    tien_ngoai_gio_thuong_t7 = db.Column(db.Numeric(15, 0), default=0)
    tien_ngoai_gio_nghi_t7 = db.Column(db.Numeric(15, 0), default=0)
    tien_ngoai_gio_le_t7 = db.Column(db.Numeric(15, 0), default=0)

    # Tháng 8
    tien_ngoai_gio_thuong_t8 = db.Column(db.Numeric(15, 0), default=0)
    tien_ngoai_gio_nghi_t8 = db.Column(db.Numeric(15, 0), default=0)
    tien_ngoai_gio_le_t8 = db.Column(db.Numeric(15, 0), default=0)

    # Tháng 9
    tien_ngoai_gio_thuong_t9 = db.Column(db.Numeric(15, 0), default=0)
    tien_ngoai_gio_nghi_t9 = db.Column(db.Numeric(15, 0), default=0)
    tien_ngoai_gio_le_t9 = db.Column(db.Numeric(15, 0), default=0)

    # Tháng 10
    tien_ngoai_gio_thuong_t10 = db.Column(db.Numeric(15, 0), default=0)
    tien_ngoai_gio_nghi_t10 = db.Column(db.Numeric(15, 0), default=0)
    tien_ngoai_gio_le_t10 = db.Column(db.Numeric(15, 0), default=0)

    # Tháng 11
    tien_ngoai_gio_thuong_t11 = db.Column(db.Numeric(15, 0), default=0)
    tien_ngoai_gio_nghi_t11 = db.Column(db.Numeric(15, 0), default=0)
    tien_ngoai_gio_le_t11 = db.Column(db.Numeric(15, 0), default=0)

    # Tháng 12
    tien_ngoai_gio_thuong_t12 = db.Column(db.Numeric(15, 0), default=0)
    tien_ngoai_gio_nghi_t12 = db.Column(db.Numeric(15, 0), default=0)
    tien_ngoai_gio_le_t12 = db.Column(db.Numeric(15, 0), default=0)

    # Các cột tổng cộng
    tong_gio_thuong = db.Column(db.Numeric(15, 1), default=0)
    tong_gio_nghi = db.Column(db.Numeric(15, 1), default=0)
    tong_gio_le = db.Column(db.Numeric(15, 1), default=0)
    tong_tien_nam = db.Column(db.Numeric(15, 0), default=0)

    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def __repr__(self):
        return f'<KetQuaNgoaiGioNamChot {self.ma_nhan_vien} - {self.nam}>'

class MaHieuLopHoc(db.Model):
    __tablename__ = 'ma_hieu_lop_hoc'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    ma_hieu = db.Column(db.String(255), nullable=False, unique=True)
    ten_lop_hoc = db.Column(db.String(255), nullable=False)
    noi_dung_hoc = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, server_default=db.func.now())

class DanhSachLopHoc(db.Model):
    __tablename__ = 'danh_sach_lop_hoc'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    # Khai báo ForeignKey trỏ chính xác đến tên bảng trong database
    ma_nhan_vien = db.Column(db.String(20), 
                             db.ForeignKey('thong_tin_nguoi_lao_dong.ma_nhan_vien', 
                                           ondelete='CASCADE', 
                                           onupdate='CASCADE'), 
                             nullable=False,
                             index=True) # Thêm index để truy vấn nhanh hơn

    ma_hieu = db.Column(db.String(255), 
                        db.ForeignKey('ma_hieu_lop_hoc.ma_hieu', 
                                      onupdate='CASCADE', 
                                      ondelete='SET NULL'), 
                        nullable=True)
    
    ten_lop_hoc = db.Column(db.String(255), nullable=False)
    noi_dung_hoc = db.Column(db.Text)
    hinh_thuc_hoc = db.Column(db.String(50), nullable=False)
    
    # Sử dụng db.Date cho ngày tháng
    tu_ngay = db.Column(db.Date, nullable=False)
    den_ngay = db.Column(db.Date, nullable=False)
    # QUAN TRỌNG: Bổ sung cột số ngày để lưu các giá trị lẻ như 0.5, 1.5
    so_ngay = db.Column(db.Float, nullable=False, default=1.0)

    
    dia_diem = db.Column(db.String(255))
    ghi_chu = db.Column(db.Text)

    # Bổ sung các trường mới
    linh_vuc = db.Column(db.String(255))
    don_vi_dau_moi = db.Column(db.String(255))
    don_vi_to_chuc = db.Column(db.String(255))
    nguon_kinh_phi = db.Column(db.String(255))
    
    # Đảm bảo bạn đã import 'func' từ sqlalchemy
    created_at = db.Column(db.DateTime, server_default=db.func.now())

    trang_thai = db.Column(db.String(50), default='Tham gia') # Các giá trị: Tham gia, Vắng mặt, Hủy
    ly_do_vang = db.Column(db.Text) # Lý do vắng mặt đột xuất

    # Thiết lập quan hệ (Relationship)
    # Lưu ý: 'ThongTinNguoiLaoDong' là tên Class của Model nhân viên
    nhan_vien_rel = db.relationship('ThongTinNguoiLaoDong', 
                                    backref=db.backref('cac_lop_hoc', cascade='all, delete-orphan'), 
                                    lazy=True)
    def to_dict(self):
        return {
            'id': self.id,
            'ma_nhan_vien': self.ma_nhan_vien,
            'ma_hieu': self.ma_hieu,
            'ten_lop_hoc': self.ten_lop_hoc,
            'so_ngay': self.so_ngay,
            'tu_ngay_raw': self.tu_ngay.strftime('%Y-%m-%d') if self.tu_ngay else '',
            'den_ngay_raw': self.den_ngay.strftime('%Y-%m-%d') if self.den_ngay else '',
            'hinh_thuc_hoc': self.hinh_thuc_hoc,
            'linh_vuc': self.linh_vuc,
            'dia_diem': self.dia_diem,
            'ghi_chu': self.ghi_chu,
            'trang_thai': self.trang_thai,
            'ly_do_vang': self.ly_do_vang
        }


class LinhVuc(db.Model):
    __tablename__ = 'linh_vuc'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    ten_linh_vuc = db.Column(db.String(255), nullable=False, unique=True)
    mo_ta = db.Column(db.Text)
    created_at = db.Column(db.DateTime, server_default=db.func.now())

class TheoDoiNghiPhep(db.Model):
    __tablename__ = 'theo_doi_nghi_phep'
    id = db.Column(db.Integer, primary_key=True)
    
    # THÊM ForeignKey vào đây
    ma_nhan_vien = db.Column(db.String(20), db.ForeignKey('thong_tin_nguoi_lao_dong.ma_nhan_vien'), nullable=False)
    nam = db.Column(db.Integer, nullable=False)
    
    phep_ton_nam_truoc_goc = db.Column(db.Numeric(4, 1), default=0)
    phep_huong_trong_nam = db.Column(db.Numeric(4, 1), default=12.0)
    phep_tham_nien = db.Column(db.Numeric(4, 1), default=0)
    da_nghi_p_q1 = db.Column(db.Numeric(4, 1), default=0)
    da_nghi_p_sau_q1 = db.Column(db.Numeric(4, 1), default=0)

    # Relationship này sẽ hoạt động sau khi có ForeignKey ở trên
    nhan_vien_rel = db.relationship('ThongTinNguoiLaoDong', backref='phep_nghi_ds')

def sync_nghi_phep_logic(ma_nhan_vien, nam):
    # Logic của bạn ở đây: 
    # 1. Tính tổng số ngày có ký hiệu 'P' trong bảng ThongTinChamCong của nhân viên đó trong năm
    # 2. Cập nhật vào bảng theo dõi phép (nếu có)
    print(f"Đang đồng bộ phép cho NV: {ma_nhan_vien} năm {nam}")
    pass

# Sự kiện sau khi Update (Sửa chấm công)
@event.listens_for(ThongTinChamCong, 'after_update')
def after_cham_cong_update(mapper, connection, target):
    # Sử dụng connection để tránh loop hoặc dùng session hiện tại
    sync_nghi_phep_logic(target.ma_nhan_vien, target.nam)

# Sự kiện sau khi Insert (Thêm mới tháng chấm công)
@event.listens_for(ThongTinChamCong, 'after_insert')
def after_cham_cong_insert(mapper, connection, target):
    sync_nghi_phep_logic(target.ma_nhan_vien, target.nam)

class ChamCongLock(db.Model):
    __tablename__ = 'cham_cong_lock'

    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    nam = db.Column(db.Integer, nullable=False)
    thang = db.Column(db.Integer, nullable=False)

    ma_hieu_2 = db.Column(db.String(255), db.ForeignKey('don_vi.ma_hieu_2'), nullable=True)
    
    # ma_phong_ban để VARCHAR(20) đồng bộ với bảng ThongTinChamCong 
    # Cho phép NULL để đại diện cho việc "Khóa toàn công ty"
    ma_phong_ban = db.Column(db.String(20), db.ForeignKey('phong_ban.id'), nullable=True)
    
    # Trạng thái khóa: 0 là mở, 1 là khóa
    is_locked = db.Column(db.SmallInteger, default=0) 
    
    # Thông tin truy vết
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    updated_by = db.Column(db.String(50))

    # Cập nhật Unique Constraint để tránh trùng lặp cho bộ (Năm, Tháng, Đơn vị, Phòng ban)
    __table_args__ = (
        db.UniqueConstraint('nam', 'thang', 'ma_hieu_2', 'ma_phong_ban', name='uk_lock_period_v2'),
    )

    def __repr__(self):
        return f"<ChamCongLock {self.thang}/{self.nam} - PB: {self.ma_phong_ban} - Locked: {self.is_locked}>"


lich_don_vi = db.Table('lich_don_vi',
    db.Column('lich_id', db.Integer, db.ForeignKey('lich_doan_cong_tac.id'), primary_key=True),
    db.Column('don_vi_ma_hieu', db.String(255), db.ForeignKey('don_vi.ma_hieu_2'), primary_key=True)
)

class DonVi(db.Model):
    __tablename__ = 'don_vi'
    ma_hieu_2 = db.Column(db.String(255), primary_key=True) 
    ten_ma_hieu_2 = db.Column(db.String(255), nullable=False)
    ten_ma_hieu_cu = db.Column(db.String(255))
    
    ma_hieu_1 = db.Column(db.String(255))
    ten_ma_hieu_1 = db.Column(db.String(255))

    TSC_Loai_I_Loai_II_Xoa_bo = db.Column(db.String(255))
    ma_khu_vuc = db.Column(db.String(255))
    ten_khu_vuc = db.Column(db.String(255))
    ma_khu_vuc_KTGSNB = db.Column(db.String(255))
    ten_khu_vuc_KTGSNB = db.Column(db.String(255))

    MST = db.Column(db.String(255))
    MaNH8so = db.Column(db.String(255))
    MaNH8so_moi = db.Column(db.String(255))
    mail = db.Column(db.String(255))

    pho_tong_giam_doc_phu_trach_KHKD = db.Column(db.String(255))
    nhom_KHCL = db.Column(db.String(255))
    hang_KHCL = db.Column(db.String(255))
    XHRR_Chung_KTGSNB = db.Column(db.String(255))
    XHRR_TD_KTGSNB = db.Column(db.String(255))
    XHRR_NTD_KTGSNB = db.Column(db.String(255))

    ghi_chu = db.Column(db.String(255))
    trang_thai = db.Column(db.String(255), default='Hoạt động')

    ngay_thao_tac = db.Column(db.String(255))
    loai_thao_tac = db.Column(db.String(255))

    def to_dict(self):
        """
        Chuyển đổi object sang dictionary một cách an toàn.
        Đảm bảo tất cả các trường đều xuất hiện trong JSON trả về, 
        ngay cả khi giá trị là None (chuyển thành chuỗi rỗng).
        """
        # Cách viết an toàn: Chuyển Null thành chuỗi rỗng để JavaScript không bị lỗi
        data = {}
        for column in self.__table__.columns:
            value = getattr(self, column.name)
            if value is None:
                data[column.name] = ""
            else:
                data[column.name] = value
        return data

# Chặn hành động xóa ở tầng SQLAlchemy
@staticmethod
@event.listens_for(DonVi, "before_delete")
def prevent_deletion(mapper, connection, target):
    raise RuntimeError("Hành động xóa bị cấm. Chỉ được phép Tạm dừng đơn vị.")


class PhongBan(db.Model):
    __tablename__ = 'phong_ban'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    ten_phong_ban = db.Column(db.String(100), nullable=False)
    # THÊM MỚI: Liên kết phòng ban với một đơn vị cụ thể thông qua ma_hieu_2
    # Đảm bảo trong mysql không có default value ='PCRT', nullable=False để bắt buộc nhập
    ma_hieu_2 = db.Column(db.String(255), db.ForeignKey('don_vi.ma_hieu_2', onupdate="CASCADE"), nullable=False)
    mo_ta = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.now)

    # Quan hệ ngược để dễ dàng truy xuất: phong_ban.don_vi.ten_ma_hieu_2
    # Quan hệ để lấy thông tin đơn vị từ phòng ban
    don_vi = db.relationship('DonVi', backref=db.backref('phong_bans', lazy=True))

    def __repr__(self):
        return f"<PhongBan {self.ten_phong_ban} - Đơn vị: {self.ma_hieu_2}>"



lich_linh_vuc = db.Table('lich_linh_vuc',
    db.Column('lich_id', db.Integer, db.ForeignKey('lich_doan_cong_tac.id'), primary_key=True),
    db.Column('linh_vuc_id', db.Integer, db.ForeignKey('linh_vuc.id'), primary_key=True)
)
class LichDoanCongTac(db.Model):
    __tablename__ = 'lich_doan_cong_tac'
    id = db.Column(db.Integer, primary_key=True)
    ten_doan = db.Column(db.String(255), nullable=False)
    truong_doan = db.Column(db.String(100))
    noi_dung = db.Column(db.Text)
    doan_vao_tu_ngay = db.Column(db.Date, nullable=False)
    doan_vao_den_ngay = db.Column(db.Date, nullable=False)
    thanh_phan = db.Column(db.Text)
    ghi_chu = db.Column(db.Text)
    created_at = db.Column(db.DateTime, server_default=db.func.now())

    # 2. XÓA cột linh_vuc_id đơn lẻ cũ và thay bằng mối quan hệ Nhiều-Nhiều
    # Sử dụng secondary để trỏ vào bảng trung gian lich_linh_vuc
    danh_sach_linh_vuc = db.relationship('LinhVuc', 
                                         secondary=lich_linh_vuc, 
                                         backref=db.backref('danh_sach_lich', lazy='dynamic'))

    # QUAN TRỌNG: Thiết lập mối quan hệ nhiều-nhiều
    # Sử dụng secondary để trỏ vào bảng trung gian
    # 3. Mối quan hệ nhiều-nhiều với Đơn vị; 
    danh_sach_don_vi = db.relationship('DonVi', 
                                     secondary=lich_don_vi, 
                                     backref=db.backref('lich_cong_tac_list', lazy='dynamic'))

    def __repr__(self):
        return f'<DoanCongTac {self.ten_doan}>'


class Question(db.Model):
    __tablename__ = 'questions'
    id = db.Column(db.Integer, primary_key=True)
    question_text = db.Column(db.Text, nullable=False)
    correct_answer = db.Column(db.Text, nullable=False)
    topic = db.Column(db.Text, nullable=False)
    explanation = db.Column(db.Text)
    # BỔ SUNG: Quan hệ để lấy options nhanh hơn
    # backref='question' cho phép bạn gọi: option.question
    # cascade="all, delete-orphan" giúp đồng bộ việc xóa từ phía ORM
    # Lệnh này báo cho SQLAlchemy biết rằng "Hãy để Database tự lo việc xóa các bản ghi con (CASCADE)". Điều này giúp giảm bớt số lượng câu lệnh SQL gửi lên server, tăng hiệu năng
    options = db.relationship('Option', backref='question', cascade="all, delete-orphan", lazy=True)

class Option(db.Model):
    __tablename__ = 'options'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    question_id = db.Column(db.Integer, db.ForeignKey('questions.id'))
    option_text = db.Column(db.Text, nullable=False)

class QuizResult(db.Model):
    __tablename__ = 'quiz_results'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.String(20), nullable=False) # ma_nhan_vien
    ho_ten = db.Column(db.String(255))
    ngay_sinh = db.Column(db.Date)
    gioi_tinh = db.Column(db.String(10))
    mail_Agribank = db.Column(db.String(255))
    ma_hieu_2 = db.Column(db.String(50))
    ma_phong_ban = db.Column(db.String(50))
    score = db.Column(db.Float, nullable=False)
    total_questions = db.Column(db.Integer, nullable=False)
    correct_count = db.Column(db.Integer, nullable=False)
    duration_seconds = db.Column(db.Integer)
    ip_address = db.Column(db.String(45))
    mac_address = db.Column(db.String(17))
    created_at = db.Column(db.DateTime, default=db.func.current_timestamp())

class QuizQuestionDetail(db.Model):
    __tablename__ = 'quiz_question_details'
    id = db.Column(db.Integer, primary_key=True)
    quiz_result_id = db.Column(db.Integer, db.ForeignKey('quiz_results.id'), nullable=False)
    question_id = db.Column(db.Integer, nullable=False)
    question_text = db.Column(db.Text, nullable=False)
    user_answer = db.Column(db.Text)
    correct_answer = db.Column(db.Text, nullable=False)
    is_correct = db.Column(db.Integer, nullable=False)
    time_spent_on_question = db.Column(db.Integer)

class DanhMucNhomHoSo(db.Model):
    __tablename__ = 'danh_muc_nhom_ho_so'
    id = db.Column(db.Integer, primary_key=True)
    ten_nhom_ho_so = db.Column(db.String(255), nullable=False)
    mo_ta = db.Column(db.Text)
    trang_thai = db.Column(db.Boolean, default=True) # True: Hiệu lực, False: Ngừng
    ngay_tao = db.Column(db.DateTime, default=datetime.now)
    
    # Quan hệ với bảng giấy tờ
    giay_to = db.relationship('DanhMucGiayTo', backref='nhom', cascade="all, delete-orphan")

class DanhMucGiayTo(db.Model):
    __tablename__ = 'danh_muc_giay_to'
    id = db.Column(db.Integer, primary_key=True)
    nhom_id = db.Column(db.Integer, db.ForeignKey('danh_muc_nhom_ho_so.id', ondelete='CASCADE'))
    ten_giay_to = db.Column(db.String(255), nullable=False)
    file_mau = db.Column(db.String(255), nullable=True)
    la_bat_buoc = db.Column(db.Boolean, default=True)
    so_ngay_quy_dinh = db.Column(db.Integer, default=30)
    trang_thai = db.Column(db.Boolean, default=True)
    ngay_bat_dau = db.Column(db.Date, default=date.today)
    ngay_ket_thuc = db.Column(db.Date, nullable=True) # Null là vô thời hạn

    danh_sach_ho_so = db.relationship('TheoDoiHoSo', back_populates='loai_giay_to', cascade="all, delete-orphan")

# Bảng Theo dõi hồ sơ chi tiết từng đơn vị lập gửi lên
class TheoDoiHoSo(db.Model):
    __tablename__ = 'theo_doi_ho_so'
    id = db.Column(db.Integer, primary_key=True)
    
    # Liên kết với ma_hieu_2 của bảng DonVi
    # Lưu ý: Đảm bảo model DonVi của bạn có cột ma_hieu_2 với unique=True
    ma_hieu_2 = db.Column(db.String(20), db.ForeignKey('don_vi.ma_hieu_2', onupdate='CASCADE', ondelete='CASCADE'), nullable=False)
    giay_to_id = db.Column(db.Integer, db.ForeignKey('danh_muc_giay_to.id', ondelete='CASCADE'))

    # Trạng thái Hồ sơ
    trang_thai = db.Column(db.Enum('Dự thảo', 'Chờ kiểm soát', 'Yêu cầu sửa đổi', 'Chờ phê duyệt', 'Đã duyệt', 'Từ chối'), default='Dự thảo')

    ngay_phat_sinh = db.Column(db.Date)
    ngay_nop_thuc_te = db.Column(db.DateTime)
    han_chot = db.Column(db.Date)
    ghi_chu = db.Column(db.Text)
    meta_data = db.Column(db.JSON) # Sử dụng kiểu JSON của MySQL 8

    nguoi_nop = db.Column(db.String(100))
    ho_ten_nguoi_nop = db.Column(db.String(100)) # Lưu Họ tên tại thời điểm nộp
    chuc_vu_nguoi_nop = db.Column(db.String(100)) # Lưu Chức vụ tại thời điểm nộp
    
    ngay_gui_kiem_soat = db.Column(db.DateTime)
    
    nguoi_kiem_soat = db.Column(db.String(100))
    ngay_kiem_soat = db.Column(db.DateTime)
    y_kien_kiem_soat = db.Column(db.Text)
    
    nguoi_phe_duyet = db.Column(db.String(100))
    ngay_phe_duyet = db.Column(db.DateTime)
    y_kien_phe_duyet = db.Column(db.Text)

    # Relationships (Để dùng trong Jinja2: ho_so.loai_giay_to.ten_giay_to)
    loai_giay_to = db.relationship('DanhMucGiayTo', back_populates='danh_sach_ho_so')
    don_vi_rel = db.relationship('DonVi', backref='ho_so_don_vi')
    
    # Hàm hỗ trợ kiểm tra đúng hạn nhanh trong Python
    @property
    def check_han(self):
        if not self.han_chot:
            return "Chưa xác định"
        
        today = date.today()
        # Nếu đã duyệt (hoàn thành) thì so sánh ngày nộp thực tế với hạn chót
        if self.trang_thai == 'Đã duyệt':
            if self.ngay_nop_thuc_te and self.ngay_nop_thuc_te.date() <= self.han_chot:
                return "Đúng hạn"
            return "Nộp muộn"
        
        # Nếu chưa xong mà đã quá hạn chót
        if today > self.han_chot:
            return "Quá hạn"
        return "Trong hạn"

class HoSoLog(db.Model):
    __tablename__ = 'ho_so_logs'
    id = db.Column(db.Integer, primary_key=True)
    ho_so_id = db.Column(db.Integer, db.ForeignKey('theo_doi_ho_so.id'))
    ma_nhan_vien = db.Column(db.String(20), db.ForeignKey('users.ma_nhan_vien'))
    hanh_dong = db.Column(db.String(100))
    trang_thai_truoc = db.Column(db.String(50))
    trang_thai_sau = db.Column(db.String(50))
    noi_dung_y_kien = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.now)

    # Quan hệ để lấy tên người thực hiện khi xem log
    user = db.relationship('User', backref='action_logs')

class Notification(db.Model):
    __tablename__ = 'notifications'
    id = db.Column(db.Integer, primary_key=True)
    ma_nhan_vien = db.Column(db.String(50), nullable=False)
    tieu_de = db.Column(db.String(255), nullable=False)
    noi_dung = db.Column(db.Text)
    duong_dan = db.Column(db.String(255))
    is_read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.now)

# ----------------------------------------------------------------------
# Module NSD yêu cầu mẫu biểu, liệt kê mẫu biểu đính kèm yêu cầu
# ----------------------------------------------------------------------
# --- Bảng trung gian (Many-to-Many) ---
# Lưu ý: yeu_cau_id dùng String vì tham chiếu đến ma_nhan_vien của bảng users
yeu_cau_mau_bieu_link = db.Table('yeu_cau_mau_bieu_link',
    db.Column('yeu_cau_id', db.Integer, db.ForeignKey('yeu_cau_mau.id', ondelete='CASCADE'), primary_key=True),
    db.Column('mau_bieu_id', db.Integer, db.ForeignKey('mau_bieu.id', ondelete='CASCADE'), primary_key=True)
)

# --- Model Mẫu Biểu ---
class MauBieu(db.Model):
    __tablename__ = 'mau_bieu'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    ten_mau = db.Column(db.String(255), nullable=False)
    ma_mau = db.Column(db.String(100), unique=True, nullable=False)
    file_path = db.Column(db.String(255))

# --- Model Danh Mục mẫu biểu ---
class DanhMucMau(db.Model):
    __tablename__ = 'danh_muc_mau'
    id = db.Column(mysql.BIGINT(unsigned=True), primary_key=True, autoincrement=True)
    ten_danh_muc = db.Column(db.String(100), unique=True, nullable=False)
    mo_ta = db.Column(db.String(255))
    
    # Quan hệ: Một danh mục có nhiều Gói mẫu (YeuCauMau)
    # Lazy='dynamic' giúp bạn có thể filter tiếp từ danh mục (ví dụ: dm.ds_goi_mau.filter_by(...))
    ds_goi_mau = db.relationship('YeuCauMau', backref='danh_muc_obj', lazy=True)    
    
# --- Model Yêu Cầu Mẫu ---
class YeuCauMau(db.Model):
    __tablename__ = 'yeu_cau_mau'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    
    # Khóa ngoại trỏ đến ma_nhan_vien (String) của bảng users; True: Admin tạo gói mẫu dùng chung
    user_id = db.Column(db.String(20), db.ForeignKey('users.ma_nhan_vien'), nullable=True)
    noi_dung_yeu_cau = db.Column(db.Text, nullable=False)
    danh_muc_id = db.Column(mysql.BIGINT(unsigned=True), db.ForeignKey('danh_muc_mau.id'), nullable=True)
    ngay_yeu_cau = db.Column(db.DateTime, default=datetime.now)
    trang_thai = db.Column(db.String(50), default='Chờ xử lý')
    is_template = db.Column(db.Boolean, default=False) # Đánh dấu đây là "Gói mẫu tạo sẵn"

    # Thiết lập quan hệ với bảng MauBieu
    danh_sach_mau = db.relationship('MauBieu', 
                                    secondary=yeu_cau_mau_bieu_link, 
                                    backref=db.backref('cac_yeu_cau', lazy='dynamic'))
    
    # Thiết lập quan hệ với bảng User (Để dùng: yeucau.nguoi_gui.fullname)
    nguoi_gui = db.relationship('User', backref=db.backref('ds_yeu_cau_mau', lazy=True))

# ----------------------------------------------------------------------
# Phân quyền sử dụng
# ----------------------------------------------------------------------
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Kiểm tra user đã đăng nhập VÀ có quyền admin (is_admin)
        if not (current_user.is_authenticated and current_user.is_admin):
            abort(403)
        return f(*args, **kwargs)
    return decorated_function

def admin_or_staff_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        allowed_users = ['200905691', '200900615'] # Thêm mã nhân viên của user vào đây: Đinh Thị Ánh Hồng; Phạm Quỳnh Anh #Chỉ Admin và user được phép
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        
        if not (current_user.is_authenticated and 
                (current_user.is_admin or current_user.ma_nhan_vien in allowed_users)):
            flash("Bạn không có quyền thực hiện thao tác này!", "danger")
            abort(403) # Từ chối truy cập
        return f(*args, **kwargs)
    return decorated_function

##########################################################################################
# Phân quyền sử dụng các module liên quan đến bc48
##########################################################################################
class PhanQuyenModuleBC48(db.Model):
    __bind_key__ = 'db_bc48'  # Chỉ định dùng database bc48
    __tablename__ = 'phan_quyen_module_bc48'
    
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    ma_nhan_vien = db.Column(db.String(20), nullable=False)
    module_slug = db.Column(db.String(100), nullable=False)
    trang_thai = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, server_default=db.func.now())

    # Đảm bảo logic đồng nhất với SQL Unique Key
    __table_args__ = (
        db.UniqueConstraint('ma_nhan_vien', 'module_slug', name='unique_user_module'),
    )

def admin_or_user_sd_bc48(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        
        # 1. Admin luôn có quyền
        if current_user.is_admin:
            return f(*args, **kwargs)
        
        # 2. Kiểm tra quyền trong db_bc48 qua bảng phan_quyen_module_bc48
        quyen = PhanQuyenModuleBC48.query.filter_by(
            ma_nhan_vien=current_user.ma_nhan_vien, 
            module_slug='bc48', 
            trang_thai=True
        ).first()
        
        if quyen:
            return f(*args, **kwargs)
        
        # 3. Nếu không thỏa mãn
        flash("Bạn không có quyền truy cập module liên quan báo cáo Cục PCRT (bc48)!", "danger")
        abort(403)
    return decorated_function

@app.route('/admin/permissions/bc48')
@login_required
@admin_required
def admin_permissions_bc48():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('q', '', type=str)
    don_vi = request.args.get('ma_hieu_2', '', type=str)
    # Khởi tạo query base
    query = ThongTinNguoiLaoDong.query

    # Áp dụng bộ lọc tìm kiếm
    if search:
        query = query.filter(
            (ThongTinNguoiLaoDong.ma_nhan_vien.contains(search)) | 
            (ThongTinNguoiLaoDong.ho_ten.contains(search))
        )
    if don_vi:
        query = query.filter_by(ma_hieu_2=don_vi)

    # Phân trang kết quả đã lọc
    pagination = query.paginate(page=page, per_page=50, error_out=False)

    # Lấy list quyền (Tối ưu: chỉ lấy danh sách ID cho các user trong trang này)
    user_ids = [u.ma_nhan_vien for u in pagination.items]
    allowed_nv = [p.ma_nhan_vien for p in PhanQuyenModuleBC48.query.filter(
        PhanQuyenModuleBC48.ma_nhan_vien.in_(user_ids),
        PhanQuyenModuleBC48.module_slug == 'bc48',
        PhanQuyenModuleBC48.trang_thai == True
    ).all()]

    return render_template('admin/admin_bc48_permissions.html', 
                           users=pagination.items, 
                           pagination=pagination,
                           allowed_nv=allowed_nv,
                           current_search=search,
                           current_don_vi=don_vi)

@app.route('/api/bc48/toggle-permission', methods=['POST'])
@login_required
def toggle_permission_bc48():
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Không có quyền'}), 403
        
    data = request.get_json()
    ma_nv = data.get('ma_nhan_vien')
    is_active = data.get('is_active')

    try:
        quyen = PhanQuyenModuleBC48.query.filter_by(ma_nhan_vien=ma_nv, module_slug='bc48').first()
        if is_active:
            if not quyen:
                db.session.add(PhanQuyenModuleBC48(ma_nhan_vien=ma_nv, module_slug='bc48', trang_thai=True))
            else:
                quyen.trang_thai = True # Đảm bảo trạng thái là True
        else:
            if quyen:
                db.session.delete(quyen) # Hoặc set trang_thai = False
        
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# ----------------------------------------------------------------------
# THEO DÕI SỰ XUẤT HIỆN, BIẾN MẤT FILES (CHỈ ADMIN)
# ----------------------------------------------------------------------
# Model lưu lịch sử vào MySQL
class FileHistory(db.Model):
    __tablename__ = 'file_history'
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255))
    action = db.Column(db.String(50))  # 'XUẤT HIỆN', 'BIẾN MẤT'
    path = db.Column(db.Text)
    timestamp = db.Column(db.DateTime, default=datetime.now)

# Handler xử lý sự kiện file (Chạy ngầm)
class ShareDriveHandler(FileSystemEventHandler):
    def process_event(self, event, action_type):
        if event.is_directory: return 
        
        filename = os.path.basename(event.src_path)
        msg = f"{action_type}: Tập tin '{filename}'"

        with app.app_context():
            try:
                new_log = FileHistory(
                    filename=filename,
                    action=action_type,
                    path=event.src_path
                )
                db.session.add(new_log)
                db.session.commit()
            except Exception as e:
                print(f"Lỗi lưu Database: {e}")

        # SocketIO gửi thông báo Real-time cho mọi người hoặc chỉ admin tùy bạn
        socketio.emit('file_change', {
            'message': msg,
            'type': action_type,
            'time': datetime.now().strftime('%H:%M:%S')
        })

    def on_created(self, event): self.process_event(event, "XUẤT HIỆN")
    def on_deleted(self, event): self.process_event(event, "BIẾN MẤT")

# 3. Route BẮT ĐẦU theo dõi (Chỉ Admin)
@app.route('/start_monitor', methods=['POST'])
@login_required 
def start_monitor():
    # Kiểm tra quyền Admin nghiêm ngặt cho API
    if not current_user.is_admin:
        return jsonify({'status': 'error', 'message': 'Bạn không có quyền thực hiện lệnh này!'}), 403

    global observer_instance
    data = request.get_json()
    path_to_watch = data.get('path')

    if not path_to_watch or not os.path.exists(path_to_watch):
        return jsonify({'status': 'error', 'message': 'Đường dẫn không tồn tại hoặc không thể truy cập!'})

    # Dừng và giải phóng trình giám sát cũ nếu đang chạy
    if observer_instance and observer_instance.is_alive():
        observer_instance.stop()
        observer_instance.join()

    try:
        event_handler = ShareDriveHandler()
        observer_instance = Observer()
        observer_instance.schedule(event_handler, path_to_watch, recursive=False)
        observer_instance.start()
        return jsonify({'status': 'success', 'message': f'Hệ thống đang giám sát: {path_to_watch}'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Lỗi khởi động: {str(e)}'})

# 4. Route DỪNG theo dõi (Chỉ Admin)
@app.route('/stop_monitor', methods=['POST'])
@login_required
def stop_monitor():
    # Kiểm tra quyền Admin
    if not current_user.is_admin:
        return jsonify({'status': 'error', 'message': 'Từ chối truy cập!'}), 403

    global observer_instance
    if observer_instance:
        try:
            observer_instance.stop()
            observer_instance.join()
            observer_instance = None
            return jsonify({'status': 'success', 'message': 'Đã dừng trình giám sát thành công.'})
        except Exception as e:
            return jsonify({'status': 'error', 'message': f'Lỗi khi dừng: {str(e)}'})
            
    return jsonify({'status': 'info', 'message': 'Hiện không có trình giám sát nào đang hoạt động.'})

# ----------------------------------------------------------------------
# Admin quản trị mã hiệu đơn vị phòng ban
# ----------------------------------------------------------------------
@app.route('/admin/don-vi/data')
@login_required
@admin_or_staff_required  #tham số gọi hàm quy định users được phép sử dụng menu
def don_vi_data():
    # Khởi tạo draw mặc định để tránh lỗi trong khối except
    draw = int(request.args.get('draw', 1))
    try:
        # 1. Lấy thông số phân trang từ DataTables gửi lên
        start = int(request.args.get('start', 0))
        length = int(request.args.get('length', 10))
        
        query = DonVi.query

        # 2. Logic lọc chi tiết (Individual Column Filtering)
        # Lưu ý: Tên key f_... phải khớp hoàn toàn với JavaScript gửi lên
        search_map = {
            'ma_hieu_2': request.args.get('f_ma_hieu_2'),
            'ten_ma_hieu_2': request.args.get('f_ten_ma_hieu_2'),
            'ten_ma_hieu_cu': request.args.get('f_ten_ma_hieu_cu'), 
            'trang_thai': request.args.get('f_trang_thai'),
            'TSC_Loai_I_Loai_II_Xoa_bo': request.args.get('f_tsc'),
            'ten_khu_vuc': request.args.get('f_khu_vuc'),
            'ten_khu_vuc_KTGSNB': request.args.get('f_ktgsnb'),
            'MaNH8so': request.args.get('f_manh8'),
            'MaNH8so_moi': request.args.get('f_manh8_moi'),
            'pho_tong_giam_doc_phu_trach_KHKD': request.args.get('f_ptgd_khkd'),
            'nhom_KHCL': request.args.get('f_nhom_khcl'),
            'hang_KHCL': request.args.get('f_hang_khcl'),
            'XHRR_Chung_KTGSNB': request.args.get('f_xhrr_chung'),
            'XHRR_TD_KTGSNB': request.args.get('f_xhrr_td'),
            'XHRR_NTD_KTGSNB': request.args.get('f_xhrr_ntd')
        }

        # Duyệt qua map để build query động
        for col, val in search_map.items():
            if val and val.strip():
                # Kiểm tra xem thuộc tính col có tồn tại trong Model DonVi không
                if hasattr(DonVi, col):
                    # .ilike(f"%{...}%") thực hiện tìm kiếm gần đúng, không phân biệt hoa thường
                    query = query.filter(getattr(DonVi, col).ilike(f"%{val.strip()}%"))

        # 3. Tìm kiếm chung (Global Search Box của DataTables)
        global_search = request.args.get('search[value]', '')
        if global_search:
            search_pattern = f"%{global_search}%"
            query = query.filter(or_(
                DonVi.ma_hieu_2.ilike(search_pattern),
                DonVi.ten_ma_hieu_2.ilike(search_pattern),
                DonVi.ten_ma_hieu_cu.ilike(search_pattern),
                DonVi.MaNH8so.ilike(search_pattern),
                DonVi.MaNH8so_moi.ilike(search_pattern),
                DonVi.ten_khu_vuc.ilike(search_pattern),
                DonVi.ten_khu_vuc_KTGSNB.ilike(search_pattern),
                DonVi.pho_tong_giam_doc_phu_trach_KHKD.ilike(search_pattern),
                DonVi.nhom_KHCL.ilike(search_pattern),
                DonVi.hang_KHCL.ilike(search_pattern),
                DonVi.XHRR_Chung_KTGSNB.ilike(search_pattern),
                DonVi.XHRR_TD_KTGSNB.ilike(search_pattern),
                DonVi.XHRR_NTD_KTGSNB.ilike(search_pattern)
            ))

        # 4. Đếm số lượng (Rất quan trọng cho phân trang)
        total_records = DonVi.query.count()
        filtered_records = query.count()

        # 5. Phân trang và trả về dữ liệu
        # Sắp xếp luôn luôn phải đứng trước offset/limit
        query = query.order_by(DonVi.ma_hieu_2.asc())

        # Nếu length == -1 (Chọn "Tất cả"), không dùng .limit() # Xử lý lấy dữ liệu dựa trên giá trị length
        if length != -1:
            # Phân trang bình thường (10, 25, 50 dòng...)
            data_list = query.offset(start).limit(length).all()
        else:
            # Trường hợp chọn "Tất cả" (ALL)
            data_list = query.all()
        
        return jsonify({
            "draw": draw,
            "recordsTotal": total_records,
            "recordsFiltered": filtered_records,
            "data": [item.to_dict() for item in data_list]
        })

    except Exception as e:
        # In lỗi chi tiết ra Terminal để bạn biết tại sao bảng trống
        print("--- LỖI DATA ĐƠN VỊ ---")
        traceback.print_exc() 
        return jsonify({
            "draw": draw,
            "recordsTotal": 0,
            "recordsFiltered": 0,
            "data": [],
            "error": str(e)
        }), 500

@app.route('/admin/don-vi')
@login_required
@admin_or_staff_required  #tham số gọi hàm quy định users được phép sử dụng menu
def admin_don_vi():
    return render_template('admin/don_vi.html')

@app.route('/admin/don-vi/save', methods=['POST'])
@login_required
def save_don_vi():
    ma_2 = request.form.get('ma_hieu_2', '').strip()
    is_edit = request.form.get('is_edit')

    if not ma_2:
        flash("Mã hiệu 2 không được để trống!", "danger")
        return redirect(url_for('admin_don_vi'))
    
    try:
        if is_edit == '1':
            dv = DonVi.query.get(ma_2)
            if not dv:
                flash("Không tìm thấy đơn vị", "danger")
                return redirect(url_for('admin_don_vi'))
            dv.loai_thao_tac = 'UPDATE'
        else:
            if DonVi.query.get(ma_2):
                flash(f"Mã {ma_2} đã tồn tại!", "warning")
                return redirect(url_for('admin_don_vi'))
            dv = DonVi(ma_hieu_2=ma_2)
            dv.loai_thao_tac = 'INSERT'
            db.session.add(dv)

        # Danh sách các trường cần cập nhật tự động từ form
        fields = [
            'ten_ma_hieu_2', 'ma_hieu_1', 'ten_ma_hieu_1', 'ten_khu_vuc', 
            'TSC_Loai_I_Loai_II_Xoa_bo', 'ten_khu_vuc_KTGSNB', 'MaNH8so', 
            'MaNH8so_moi', 'MST', 'mail', 'ten_ma_hieu_cu', 
            'pho_tong_giam_doc_phu_trach_KHKD', 'nhom_KHCL', 'hang_KHCL', 
            'XHRR_Chung_KTGSNB', 'XHRR_TD_KTGSNB', 'XHRR_NTD_KTGSNB'
        ]
        
        for field in fields:
            val = request.form.get(field)
            if val is not None:
                setattr(dv, field, val.strip())

        dv.ngay_thao_tac = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        db.session.commit()
        flash("Lưu dữ liệu thành công!", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"Lỗi: {str(e)}", "danger")
        
    return redirect(url_for('admin_don_vi'))

# Route: Tạm dừng/Kích hoạt (Thay thế cho lệnh Xóa)
@app.route('/admin/don-vi/toggle/<ma_2>')
@login_required
def toggle_don_vi(ma_2):
    try:
        dv = DonVi.query.get_or_404(ma_2)
        # Chuyển đổi trạng thái (xử lý cả trường hợp rỗng/None)
        if not dv.trang_thai or dv.trang_thai == 'Hoạt động':
            dv.trang_thai = 'Tạm dừng'
        else:
            dv.trang_thai = 'Hoạt động'
            
        dv.ngay_thao_tac = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        dv.loai_thao_tac = 'TOGGLE_STATUS'
        
        db.session.commit()
        flash(f"Đã chuyển đơn vị {ma_2} sang trạng thái: {dv.trang_thai}", "info")
    except Exception as e:
        db.session.rollback()
        flash(f"Lỗi khi đổi trạng thái: {str(e)}", "danger")
        
    return redirect(url_for('admin_don_vi'))


@app.route('/admin/phong-ban/data')
@login_required
@admin_or_staff_required #tham số gọi hàm quy định users được phép sử dụng menu
def phong_ban_data():
    try:
        draw = int(request.args.get('draw', 1))
        start = int(request.args.get('start', 0))
        length = int(request.args.get('length', 10))
        
        # Sử dụng outerjoin để tránh mất dữ liệu nếu đơn vị cha bị thiếu
        # Nếu dùng join thường, nếu 1 phòng ban có ma_hieu_2 không tồn tại bên DonVi, nó sẽ bị ẩn mất.
        query = db.session.query(PhongBan).outerjoin(DonVi, PhongBan.ma_hieu_2 == DonVi.ma_hieu_2)

        # 1. Lọc theo các tiêu chí (Fuzzy search) Lấy tham số lọc từ Ajax (Phải khớp chính xác ID trong JS)
        f_ten_pb = request.args.get('f_ten_pb', '').strip()
        f_ma_hieu_2 = request.args.get('f_ma_hieu_2', '').strip()
        f_ten_ma_hieu_2 = request.args.get('f_ten_ma_hieu_2', '').strip()
        f_trang_thai = request.args.get('f_trang_thai', '').strip()

        # Chỉ lọc nếu người dùng có nhập dữ liệu vào ô lọc
        if f_ten_pb:
            query = query.filter(PhongBan.ten_phong_ban.ilike(f"%{f_ten_pb}%"))
        if f_ma_hieu_2:
            query = query.filter(PhongBan.ma_hieu_2.ilike(f"%{f_ma_hieu_2}%"))
        if f_ten_ma_hieu_2:
            query = query.filter(DonVi.ten_ma_hieu_2.ilike(f"%{f_ten_ma_hieu_2}%"))
        if f_trang_thai:
            query = query.filter(DonVi.trang_thai == f_trang_thai)

        # 2. Global Search (Tìm kiếm nhanh)
        search_val = request.args.get('search[value]', '').strip()
        if search_val:
            pattern = f"%{search_val}%"
            query = query.filter(or_(
                PhongBan.ten_phong_ban.ilike(pattern),
                PhongBan.ma_hieu_2.ilike(pattern),
                DonVi.ten_ma_hieu_2.ilike(pattern)
            ))

        # 3. Đếm và Phân trang
        total_records = PhongBan.query.count()
        filtered_records = query.count()
        
        data_list = query.order_by(PhongBan.id.desc()).offset(start).limit(length).all()

        results = []
        for pb in data_list:
            # Kiểm tra an toàn: Nếu không có đơn vị khớp, trả về "Chưa gán"
            ten_dv = pb.don_vi.ten_ma_hieu_2 if pb.don_vi else "--- Chưa gán ---"
            trang_thai_dv = pb.don_vi.trang_thai if pb.don_vi else "N/A"
            
            results.append({
                "id": pb.id,
                "ten_phong_ban": pb.ten_phong_ban,
                "ma_hieu_2": pb.ma_hieu_2 or "N/A",
                "ten_ma_hieu_2": ten_dv,
                "trang_thai_dv": trang_thai_dv,
                "mo_ta": pb.mo_ta or "",
                "created_at": pb.created_at.strftime('%d/%m/%Y') if pb.created_at else "",
                "is_admin": getattr(current_user, 'is_admin', False) # Gửi thêm quyền để JS hiển thị nút Xóa
            })

        return jsonify({
            "draw": draw,
            "recordsTotal": total_records,
            "recordsFiltered": filtered_records,
            "data": results
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    
@app.route('/admin/phong-ban')
@login_required
@admin_or_staff_required #tham số gọi hàm quy định users được phép sử dụng menu
def admin_phong_ban():
    # Chỉ lấy các đơn vị đang hoạt động để làm danh sách chọn (dropdown)
    danh_sach_dv = DonVi.query.filter_by(trang_thai='Hoạt động').all()
    return render_template('admin/phong_ban.html', danh_sach_dv=danh_sach_dv)

# Route: Lưu (Thêm/Sửa) Phòng ban
@app.route('/admin/phong-ban/save', methods=['POST'])
@login_required
def save_phong_ban():
    pb_id = request.form.get('id')
    ten_pb = request.form.get('ten_phong_ban', '').strip()
    ma_hieu_2 = request.form.get('ma_hieu_2')
    mo_ta = request.form.get('mo_ta', '').strip()

    if not ten_pb or not ma_hieu_2:
        flash("Vui lòng điền đầy đủ Tên phòng ban và Đơn vị!", "warning")
        return redirect(url_for('admin_phong_ban'))
    
    try:
        if pb_id and pb_id.strip(): # Chế độ sửa
            pb = PhongBan.query.get(pb_id)
            if not pb:
                flash("Không tìm thấy phòng ban!", "danger")
                return redirect(url_for('admin_phong_ban'))
        else: # Chế độ thêm mới
            pb = PhongBan()
            db.session.add(pb)
        
        # Làm sạch dữ liệu trước khi lưu
        pb.ten_phong_ban = ten_pb
        pb.ma_hieu_2 = ma_hieu_2
        pb.mo_ta = mo_ta
        
        db.session.commit()
        flash("Lưu thông tin phòng ban thành công!", "success")
    except Exception as e:
        db.session.rollback()
        error_msg = str(e)
        if "Duplicate entry" in error_msg:
            flash(f"Lỗi: Phòng ban '{ten_pb}' đã tồn tại trong hệ thống!", "warning")
        else:
            flash(f"Lỗi không xác định: {error_msg}", "danger")
        traceback.print_exc()
        
    return redirect(url_for('admin_phong_ban'))

# Route: Xóa phòng ban (Chỉ ADMIN mới có quyền xóa phòng ban)
@app.route('/admin/phong-ban/delete/<int:id>')
@login_required
def delete_phong_ban(id):
    # Kiểm tra quyền Admin cấp thấp nhất (ngay tại logic)
    if not getattr(current_user, 'is_admin', False):
        flash("Bạn không có quyền thực hiện thao tác xóa!", "danger")
        return redirect(url_for('admin_phong_ban'))

    try:
        pb = PhongBan.query.get_or_404(id)
        ten_xoa = pb.ten_phong_ban
        db.session.delete(pb)
        db.session.commit()
        flash(f"Đã xóa phòng ban '{ten_xoa}' thành công.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Không thể xóa phòng ban: {str(e)}", "danger")
        
    return redirect(url_for('admin_phong_ban'))


@app.route('/admin/linh-vuc')
@login_required
@admin_required
def danh_sach_linh_vuc():
    # Logic lấy dữ liệu từ bảng LinhVuc ở đây
    # Lấy danh sách sắp xếp (mới nhất) theo tên/ngày tạo
    linh_vucs = LinhVuc.query.order_by(LinhVuc.ten_linh_vuc.asc()).all()
    return render_template('admin/linh_vuc.html', linh_vucs=linh_vucs)

@app.route('/admin/linh-vuc/add', methods=['POST'])
@login_required
@admin_required
def add_linh_vuc():
    ten = request.form.get('ten_linh_vuc', '').strip()
    mota = request.form.get('mo_ta', '').strip()
    
    if not ten:
        flash('Tên lĩnh vực không được để trống!', 'danger')
        return redirect(url_for('danh_sach_linh_vuc'))

    try:
        new_lv = LinhVuc(ten_linh_vuc=ten, mo_ta=mota)
        db.session.add(new_lv)
        db.session.commit()
        flash(f'Đã thêm lĩnh vực: {ten}', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Lỗi: Tên lĩnh vực này đã tồn tại hoặc có lỗi hệ thống!', 'danger')
        
    return redirect(url_for('danh_sach_linh_vuc'))

@app.route('/admin/linh-vuc/update/<int:id>', methods=['POST'])
@login_required
@admin_required
def update_linh_vuc(id):
    lv = LinhVuc.query.get_or_404(id)
    ten_moi = request.form.get('ten_linh_vuc', '').strip()
    mota_moi = request.form.get('mo_ta', '').strip()
    
    if not ten_moi:
        flash('Tên lĩnh vực không được để trống!', 'danger')
        return redirect(url_for('danh_sach_linh_vuc'))

    try:
        lv.ten_linh_vuc = ten_moi
        lv.mo_ta = mota_moi
        db.session.commit()
        flash(f'Đã cập nhật thành công lĩnh vực: {ten_moi}', 'success')
    except Exception:
        db.session.rollback()
        print(f"Lỗi Database: {e}") # In ra màn hình terminal để bạn sửa lỗi nhanh
        flash('Lỗi: Tên lĩnh vực này có thể đã tồn tại!', 'danger')
        
    return redirect(url_for('danh_sach_linh_vuc'))



@app.route('/config/ngay-le', methods=['GET', 'POST'])
@login_required
def config_ngay_le():
    # 1. Kiểm tra quyền Admin
    if not current_user.role or current_user.role.upper() != 'ADMIN' or current_user.is_admin != 1:
        flash("Bạn không có quyền truy cập trang này!", "danger")
        return redirect(url_for('index'))

    if request.method == 'POST':
        action = request.form.get('action')
        
        # Xử lý THÊM ngày lễ
        if action == 'add':
            ngay_str = request.form.get('ngay')
            ten_le = request.form.get('ten_le')
            he_so = request.form.get('he_so', 3.0)
            
            if ngay_str:
                ngay_dt = datetime.strptime(ngay_str, '%Y-%m-%d').date()
                # Kiểm tra trùng
                ton_tai = DanhMucNgayLe.query.filter_by(ngay=ngay_dt).first()
                if ton_tai:
                    flash(f"Ngày {ngay_str} đã tồn tại trong danh sách!", "warning")
                else:
                    moi = DanhMucNgayLe(ngay=ngay_dt, ten_le=ten_le, he_so=he_so)
                    db.session.add(moi)
                    db.session.commit()
                    flash("Đã thêm ngày lễ thành công!", "success")
        
        # Xử lý XÓA ngày lễ
        elif action == 'delete':
            id_le = request.form.get('id')
            le_can_xoa = DanhMucNgayLe.query.get(id_le)
            if le_can_xoa:
                db.session.delete(le_can_xoa)
                db.session.commit()
                flash("Đã xóa ngày lễ!", "info")

        return redirect(url_for('config_ngay_le'))

    # Lấy toàn bộ danh sách ngày lễ, sắp xếp theo ngày mới nhất
    danh_sach_le = DanhMucNgayLe.query.order_by(DanhMucNgayLe.ngay.desc()).all()
    return render_template('config_ngay_le.html', danh_sach_le=danh_sach_le)

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

@app.context_processor
def inject_user_permissions():
    if current_user.is_authenticated:
        # 1. Nếu là Admin (Lê Quốc Vân), cho phép thấy tất cả các mã slug đã định nghĩa
        if current_user.is_admin:
            # Lưu ý: Danh sách này nên bao gồm tất cả các slug bạn dùng trong DB và HTML
            all_menus = [
                'lich-cong-tac',
                'cham-cong',
                'theo-doi-nghi-phep',
                'cham-cong-ngoai-gio',
                'theo_doi_lop_hoc', 
                'bang-luong-v1', 
                'bang-luong-v2', 
                'bang-luong-dc',
                'bao-cao-cong-chi-tiet', 
                'tong-hop-ngoai-gio', 
                'tong-hop-ngoai-gio-nam',
                'ket-qua-ngoai-gio-nam-chot', 
                'v1-report',
                'v2-report',
                'v2-dc-report',
                'report-v1-v2',
                'report-v1-v2-dc',
                'bao_cao_tong_hop_dao_tao',
                'bao-cao-dao-tao-chuc-vu',
                'admin/ma-hieu-lop',
                'config/ngay-le',
                'system-logs',
                'admin/nhom-ho-so',
                'admin/danh-muc-giay-to',
                'theo-doi-ho-so'
            ]
            return dict(user_allowed_menus=all_menus)
        
        # 2. Nếu không phải Admin (Ánh Hồng, Quỳnh Anh...), truy vấn quyền từ bảng MySQL
        try:
            permissions = db.session.query(UserMenuPermission.menu_slug)\
                .filter_by(ma_nhan_vien=current_user.ma_nhan_vien).all()
            
            # Chuyển kết quả từ danh sách tuple thành danh sách chuỗi: ['bang-luong-v1', ...]
            allowed_menus = [p.menu_slug for p in permissions]
            return dict(user_allowed_menus=allowed_menus)
        except Exception as e:
            print(f"Lỗi truy vấn quyền menu: {e}")
            return dict(user_allowed_menus=[])

    # 3. Nếu chưa đăng nhập
    return dict(user_allowed_menus=[])


def get_friendly_os_name():
    sys_name = platform.system()
    mapping = {'Darwin': 'macOS', 'Windows': 'Windows', 'Linux': 'Linux'}
    return mapping.get(sys_name, sys_name)

def check_mysql_connection():
    try:
        db.session.execute(text("SELECT 1"))
        return {"status": "✅ Kết nối thành công!", "is_ok": True}
    except Exception as e:
        return {"status": f"❌ Lỗi: {str(e)}", "is_ok": False}

# API Kiểm tra kết nối nhanh (dùng cho nút bấm trên giao diện)
@app.route('/test_mysql_connection')
@login_required # Nên thêm login_required để đảm bảo chỉ người đang đăng nhập mới sử dụng được tính năng
def test_mysql_connection():
    try:
        # Giả sử đây là lệnh kiểm tra DB của bạn
        # db.session.execute('SELECT 1') 
        return jsonify({"success": True, "message": "Kết nối Database thành công!"})
    except Exception as e:
        # Ngay cả khi lỗi cũng phải trả về JSON để JS đọc được
        return jsonify({"success": False, "message": f"Lỗi kết nối: {str(e)}"}), 500

def get_common_context():
    """Hàm gom các biến dùng chung cho index.html"""
    now = datetime.now(HANOI_TZ) if 'HANOI_TZ' in globals() else datetime.now()
    
    # Check DB status
    db_status = {"status": "Kết nối thất bại", "details": "Không thể kết nối MySQL"}
    salary_status = []
    has_missing_config = False
    try:
        db.session.execute(text('SELECT 1'))
        db_status = {"status": "Kết nối thành công", "details": "Database đang hoạt động"}
        result = db.session.execute(
            text("SELECT loai_bang, trang_thai FROM view_check_config_12_ky WHERE nam = :nam"),
            {'nam': now.year}
        ).fetchall()
        salary_status = [{"loai": row[0], "status": row[1]} for row in result]
        has_missing_config = any(s['status'] == 'CHƯA ĐỦ' for s in salary_status)
    except Exception as e:
        db_status["details"] = str(e)

    # User info
    u_name = "Khách"
    u_role = None
    if current_user.is_authenticated:
        u_name = current_user.fullname or current_user.ma_nhan_vien
        u_role = current_user.role

    return {
        'mysql_status': db_status,
        'os_system': platform.system(),
        'os_release': platform.release(),
        'os_user': os.environ.get('USERNAME') or "Unknown",
        'user_agent': request.headers.get('User-Agent'),
        'user_ip': request.remote_addr,
        'weekday_vn': get_weekday_vn() if 'get_weekday_vn' in globals() else "Thứ",
        'current_date': now.strftime('%d/%m/%Y'),
        'username': u_name,
        'user_role': u_role,
        'salary_status': salary_status,
        'has_missing_config': has_missing_config,
        'selected_year': now.year
    }

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        # Lấy dữ liệu từ form
        # Lấy dữ liệu từ thẻ <input name="username"> trong login.html
        ma_nv_nhap = request.form.get('username') 
        mat_khau_nhap = request.form.get('password')
        
        # Tìm user theo ma_nhan_vien
        user = User.query.filter_by(ma_nhan_vien=ma_nv_nhap).first()
        
        # 3. Kiểm tra user và mật khẩu
        if user and user.check_password(mat_khau_nhap):
            login_user(user)
            # Ghi log: Dùng fullname thay vì username
            log_event("Đăng nhập", f"Nhân viên {user.fullname} đã đăng nhập.")
            return redirect(url_for('index'))
        
        flash('Sai mã nhân viên hoặc mật khẩu!', 'danger')
        log_event("Đăng nhập thất bại", f"Thử đăng nhập với mã: {ma_nv_nhap}")
    
    return render_template('login.html')

@app.route('/logout')
@login_required # Nên thêm login_required để đảm bảo chỉ người đang đăng nhập mới gọi được log
def logout():
    # --- GHI LOG TRƯỚC KHI THOÁT ---
    if current_user.is_authenticated:
        log_event("Đăng xuất", f"Người dùng {current_user.fullname} đã thoát hệ thống.")
        
    logout_user()
    return redirect(url_for('index'))


# --- Hàm ghi log dùng chung ---
def log_event(action, detail=None):
    try:
        # 1. Lấy IP người dùng chính xác hơn
        if request.headers.getlist("X-Forwarded-For"):
            ip = request.headers.getlist("X-Forwarded-For")[0]
        else:
            ip = request.remote_addr
            
        # 2. Xác định danh tính người dùng an toàn
        # Nếu chưa đăng nhập (như lúc đang thử Login), dùng mã nhập từ form hoặc "Guest"
        u_id = "Guest"
        if current_user.is_authenticated:
            # Ưu tiên lấy ma_nhan_vien từ object current_user
            u_id = getattr(current_user, 'ma_nhan_vien', 'Unknown')
        
        new_log = SystemLog(
            ma_nhan_vien=u_id,
            hanh_dong=action,
            chi_tiet=detail,
            ip_address=ip
        )
        
        # 3. Lưu log (Sử dụng session riêng nếu cần, nhưng db.session là đủ)
        db.session.add(new_log)
        db.session.commit()
        
    except Exception as e:
        db.session.rollback()
        # In ra console để bạn dễ debug khi phát triển
        print(f"DEBUG - Lỗi ghi log: {str(e)}")

        
# --- Admin theo dõi ai đã thay đổi dữ liệu, đăng nhập khi nào và từ đâu ---
@app.route('/system-logs')
@login_required
def view_logs():
    # Kiểm tra quyền Admin. Chỉ Admin mới được vào
    if not current_user.is_admin:
        flash("Bạn không có quyền truy cập trang này!", "danger")
        return redirect(url_for('index'))
    
    # Truy vấn logs, join với bảng User để lấy tên hiển thị (fullname)
    # Sắp xếp theo thời gian mới nhất (desc)
    # Lấy 500 bản ghi mới nhất, join với bảng User để lấy fullname
    logs = db.session.query(SystemLog, User.fullname).join(
        User, SystemLog.ma_nhan_vien == User.ma_nhan_vien
    ).order_by(SystemLog.created_at.desc()).limit(500).all()
    
    return render_template('logs.html', logs=logs)

@app.before_request
def check_password_expiration_and_reset():
    # 1. Ngoại trừ nếu chưa đăng nhập
    if not current_user or not current_user.is_authenticated:
        return

    # 2. Ngoại trừ nếu là Admin
    is_admin_user = (current_user.is_admin == 1 or 
                     (current_user.role and current_user.role.upper() == 'ADMIN'))
    if is_admin_user:
        return

    # 3. Ngoại trừ các route thiết yếu để tránh bị lặp chuyển hướng vô tận (Infinite Redirect)
    allowed_endpoints = ['change_password', 'logout', 'static']
    if request.endpoint in allowed_endpoints:
        return

    # 4. KIỂM TRA ĐIỀU KIỆN 1: Bị Admin reset mật khẩu (force_password_change = True)
    if getattr(current_user, 'force_password_change', False):
        flash('Mật khẩu của bạn đã được reset bởi quản trị viên. Vui lòng đổi mật khẩu mới để tiếp tục sử dụng hệ thống.', 'warning')
        return redirect(url_for('change_password'))

    # 5. KIỂM TRA ĐIỀU KIỆN 2: Hết hạn mật khẩu định kỳ
    if current_user.password_changed_at:
        # Tính toán khoảng thời gian từ lần đổi mật khẩu cuối cùng đến hiện tại
        elapsed_time = datetime.now(timezone.utc) - current_user.password_changed_at
        if elapsed_time > timedelta(days=PASSWORD_EXPIRY_DAYS):
            flash(f'Mật khẩu của bạn đã quá hạn {PASSWORD_EXPIRY_DAYS} ngày. Vui lòng đổi mật khẩu mới để đảm bảo an toàn.', 'warning')
            return redirect(url_for('change_password'))
# ----------------------------------------------------------------------
# 4. CÁC ĐƯỜNG DẪN (ROUTES)
# ----------------------------------------------------------------------
@app.route('/')
@app.route('/index')
def index():
    # 1. Lấy thông tin thời gian
    now = datetime.now(HANOI_TZ) if 'HANOI_TZ' in globals() else datetime.now()
    current_year = now.year
    
    # 2. Kiểm tra kết nối Database & Truy vấn trạng thái lương
    db_status = {"status": "Kết nối thất bại", "details": "Không thể kết nối MySQL"}
    salary_status = []
    has_missing_config = False

    try:
        # Kiểm tra kết nối
        db.session.execute(text('SELECT 1'))
        db_status = {"status": "Kết nối thành công", "details": "Database đang hoạt động"}
        
        # 3. TRUY VẤN VIEW KIỂM TRA LƯƠNG
        # Lấy dữ liệu cho năm hiện tại
        result = db.session.execute(
            text("SELECT loai_bang, trang_thai FROM view_check_config_12_ky WHERE nam = :nam"),
            {'nam': current_year}
        ).fetchall()
        
        # Chuyển thành list dict để dễ xử lý ở template
        salary_status = [{"loai": row[0], "status": row[1]} for row in result]
        
        # Kiểm tra nếu có bất kỳ bảng nào "CHƯA ĐỦ"
        has_missing_config = any(s['status'] == 'CHƯA ĐỦ' for s in salary_status)

    except Exception as e:
        db_status["details"] = str(e)
        
    # 4. Lấy thông tin môi trường hệ thống
    os_system = platform.system()
    os_release = platform.release()
    try:
        os_user = os.getlogin()
    except:
        os_user = os.environ.get('USERNAME') or os.environ.get('USER') or "Unknown"

    # 5. Thông tin định danh người dùng
    if current_user.is_authenticated:
        u_name = current_user.fullname or current_user.ma_nhan_vien
        u_role = current_user.role
    else:
        u_name = "Khách"
        u_role = None

    # 6. Gom tất cả vào render_template
    return render_template(
        'index.html', 
        title='Trang Chủ',
        mysql_status=db_status, 
        os_system=os_system,
        os_release=os_release,
        os_user=os_user,
        user_agent=request.headers.get('User-Agent'),
        user_ip=request.remote_addr,
        weekday_vn=get_weekday_vn() if 'get_weekday_vn' in globals() else "Thứ", 
        current_date=now.strftime('%d/%m/%Y'),
        username=u_name,
        user_role=u_role,
        # DỮ LIỆU MỚI THÊM
        salary_status=salary_status,
        has_missing_config=has_missing_config,
        selected_year=current_year
    )

# Route bổ sung để tránh lỗi BuildError cho 'simple_info'
@app.route('/info')
def simple_info():
    return redirect(url_for('index'))

@app.route('/qr-generator')
def qr_generator_page():
    # Route này giải quyết lỗi BuildError: qr_generator_page
    return render_template('qr_generator.html')


# ----------------------------------------------------------------------
# Quản trị phân quyền sử dụng menu (Admin UI), bảng user_menu_permissions
# ----------------------------------------------------------------------
@app.route('/admin/permissions')
@login_required
def admin_permissions():
    if not current_user.is_admin:
        flash("Bạn không có quyền truy cập trang này!", "danger")
        return redirect(url_for('index'))
    
    try:
        # 1. Lấy danh sách Đơn vị duy nhất từ bảng ThongTinNguoiLaoDong
        # Lưu ý: don_vis_query PHẢI được định nghĩa ở đây
        don_vis_query = db.session.query(distinct(ThongTinNguoiLaoDong.ma_hieu_2))\
                    .filter(ThongTinNguoiLaoDong.ma_hieu_2 != None)\
                    .order_by(ThongTinNguoiLaoDong.ma_hieu_2).all()
        # Chuyển list tuple [(dv1,), (dv2,)] thành list string [dv1, dv2]
        don_vis = [dv[0] for dv in don_vis_query]

        # 2. LẤY DANH SÁCH ĐƠN VỊ TỪ BẢNG DonVi (Sửa lỗi NameError tại đây)
        # Lấy khoảng 20 đơn vị đầu tiên để hiển thị mặc định trong Modal
        ds_don_vi = DonVi.query.order_by(DonVi.ten_ma_hieu_2.asc()).limit(20).all()
        
    except Exception as e:
        print(f"Lỗi truy vấn đơn vị: {e}")
        don_vis = []
        ds_don_vi = [] # Đảm bảo biến luôn tồn tại dù có lỗi DB
    
    # Danh sách các Menu Slug bạn muốn quản lý
    managed_slugs = [
        {'code': 'lich-cong-tac', 'name': 'Lịch Đoàn công tác'},
        {'code': 'cham-cong', 'name': 'Chấm công'},
        {'code': 'theo-doi-nghi-phep', 'name': 'Theo dõi nghỉ phép'},
        {'code': 'cham-cong-ngoai-gio', 'name': 'Chấm công ngoài giờ'},
        {'code': 'theo_doi_lop_hoc', 'name': 'Theo dõi lớp học'},
        {'code': 'bang-luong-v1', 'name': 'Hệ số lương V1'},
        {'code': 'bang-luong-v2', 'name': 'Hệ số lương V2'},
        {'code': 'bang-luong-dc', 'name': 'Hệ số lương Điều chỉnh (DC)'},
        {'code': 'bao-cao-cong-chi-tiet', 'name': 'Báo cáo chấm công tổng hợp'},
        {'code': 'tong-hop-ngoai-gio', 'name': 'Tổng số GIỜ làm ngoài giờ (Tháng)'},
        {'code': 'tong-hop-ngoai-gio-nam', 'name': 'Tổng số GIỜ làm ngoài giờ năm (12 Tháng)'},
        {'code': 'ket-qua-ngoai-gio-nam-chot', 'name': 'Tổng số TIỀN làm ngoài giờ năm (12 Tháng)'},
        {'code': 'v1-report', 'name': 'BC Tiền lương V1'},
        {'code': 'v2-report', 'name': 'BC Tiền lương V2'},
        {'code': 'v2-dc-report', 'name': 'BC Tiền lương V2_Điều chỉnh'},
        {'code': 'report-v1-v2', 'name': 'Tổng hợp V1 + V2'},
        {'code': 'report-v1-v2-dc', 'name': 'Tổng hợp V1 + V2_DC'},
        {'code': 'bao_cao_tong_hop_dao_tao', 'name': 'Báo cáo Đào tạo theo người học'},
        {'code': 'bao-cao-dao-tao-chuc-vu', 'name': 'Báo cáo Đào tạo theo nhóm chức vụ'},
        {'code': 'admin/ma-hieu-lop', 'name': 'Khai báo mã hiệu lớp học'},
        {'code': 'admin/nhom-ho-so', 'name': 'Quản trị Danh mục nhóm hồ sơ'},
        {'code': 'admin/danh-muc-giay-to', 'name': 'Quản trị Danh mục giấy tờ'},
        {'code': 'theo-doi-ho-so', 'name': 'Theo dõi hồ sơ khen thưởng'}
    ]
    # Trả về file HTML bạn muốn dùng
    return render_template('admin_permissions.html',
                           ds_don_vi=ds_don_vi, 
                           don_vis=don_vis,
                           managed_slugs=managed_slugs)

@app.route('/api/search-users')
@login_required
def search_users():
    if not current_user.is_admin:
        return jsonify([]), 403

    query_search = request.args.get('q', '').strip()
    don_vi = request.args.get('don_vi', '').strip()

    # Join 2 bảng User và Thông tin nhân viên để lọc theo mã đơn vị 
    stmt = db.session.query(User, ThongTinNguoiLaoDong.ho_ten)\
             .join(ThongTinNguoiLaoDong, User.ma_nhan_vien == ThongTinNguoiLaoDong.ma_nhan_vien)

    if query_search:
        stmt = stmt.filter(or_(
            User.ma_nhan_vien.like(f"%{query_search}%"),
            User.fullname.like(f"%{query_search}%")
        ))
    
    if don_vi:
        stmt = stmt.filter(ThongTinNguoiLaoDong.ma_hieu_2 == don_vi)

    # Giới hạn 100 người để đảm bảo tốc độ render của trình duyệt
    users = stmt.limit(100).all()
    user_ids = [u.User.ma_nhan_vien for u in users]

    # --- 1. Lấy quyền Menu (Checkbox) ---
    perms = UserMenuPermission.query.filter(UserMenuPermission.ma_nhan_vien.in_(user_ids)).all()
    perm_map = {}
    for p in perms:
        if p.ma_nhan_vien not in perm_map: perm_map[p.ma_nhan_vien] = []
        perm_map[p.ma_nhan_vien].append(p.menu_slug)

    # --- 2. Lấy quyền Quy trình (Workflow Hồ sơ) ---
    # Truy vấn từ bảng user_unit_permissions mà bạn đã tạo
    workflow_perms = db.session.execute(text("""
        SELECT ma_nhan_vien, permission_code, ma_hieu_2 
        FROM user_unit_permissions 
        WHERE ma_nhan_vien IN :user_ids
    """), {'user_ids': tuple(user_ids) if user_ids else ('',)}).fetchall()

    wf_map = {}
    for wp in workflow_perms:
        if wp.ma_nhan_vien not in wf_map: wf_map[wp.ma_nhan_vien] = []
        # Tạo chuỗi hiển thị ví dụ: "HS_KS" hoặc "HS_PD (ALL)"
        wf_map[wp.ma_nhan_vien].append(f"{wp.permission_code} ({wp.ma_hieu_2})")

    # Tạo kết quả trả về
    result = []
    for u in users:
        result.append({
            'ma_nhan_vien': u.User.ma_nhan_vien,
            'fullname': u.User.fullname or u.ho_ten,
            'is_admin': u.User.is_admin,
            'permissions': perm_map.get(u.User.ma_nhan_vien, []),
            'workflow_permissions': wf_map.get(u.User.ma_nhan_vien, []) # Trả về cho JS render Badge
        })

    return jsonify(result)


# --- PHÂN QUYỀN SỬ DỤNG TRUY CẬP MENU (Dùng cho giao diện Checkbox bảng) ---
@app.route('/api/update-permission', methods=['POST'])
@login_required
def update_permission():
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Bạn không có quyền admin'}), 403
    
    data = request.get_json()
    ma_nv = data.get('ma_nhan_vien')
    slug = data.get('menu_slug')
    status = data.get('status')

    if not ma_nv or not slug:
        return jsonify({'success': False, 'message': 'Dữ liệu không hợp lệ'}), 400

    try:
        if status:
            # Kiểm tra xem quyền này đã tồn tại chưa để tránh lỗi UNIQUE KEY
            exists = UserMenuPermission.query.filter_by(ma_nhan_vien=ma_nv, menu_slug=slug).first()
            if not exists:
                new_perm = UserMenuPermission(ma_nhan_vien=ma_nv, menu_slug=slug)
                db.session.add(new_perm)
        else:
            # Xóa quyền
            UserMenuPermission.query.filter_by(ma_nhan_vien=ma_nv, menu_slug=slug).delete()
        
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# ----------------------------------------------------------------------
# ROUTES phục vụ menu "Chấm công (Khách)", không cần user đăng nhập, xuất excel bảng chấm công
# ----------------------------------------------------------------------
@app.route('/cham-cong-khach')
def cham_cong_khach():
    # Mặc định lấy tháng năm hiện tại
    month = datetime.now().month
    year = datetime.now().year
    return render_template('cham_cong_khach.html', month=month, year=year)

@app.route('/xuat-excel-khach', methods=['POST'])
def Xuat_excel_khach():
    # Nhận tháng năm từ form
    month = request.form.get('month', datetime.now().month, type=int)
    year = request.form.get('year', datetime.now().year, type=int)
    
    # Tạo file Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Cham Cong T{month}-{year}"
    
    # Tạo header
    headers = ["Mã NV", "Họ Tên"] + [str(d) for d in range(1, 32)] + ["Tổng Công", "Ngày Lương", "Ăn Ca"]
    ws.append(headers)
    
    # Lấy toàn bộ dữ liệu từ form gửi lên
    # Form gửi lên có định dạng: d{ngày}_{mã_nv}
    # Chúng ta cần thu thập danh sách Mã NV trước
    all_keys = request.form.keys()
    ma_nv_list = set()
    for key in all_keys:
        if key.startswith('ma_nv_'):
            ma_nv_list.add(key.replace('ma_nv_', ''))
    
    for ma_nv in sorted(ma_nv_list):
        ho_ten = request.form.get(f'ho_ten_{ma_nv}', '')
        row = [ma_nv, ho_ten]

        # Chỉ lấy số ngày thực tế của tháng đó
        import calendar
        _, last_day = calendar.monthrange(year, month)
        
        # Lấy dữ liệu 31 ngày, để trống các ngày không tồn tại
        for d in range(1, 32):
            if d <= last_day:
                val = request.form.get(f'd{d}_{ma_nv}', '')
                row.append(val)
            else:
                row.append("") # Để trống các ngày không tồn tại
            
        # Lưu ý: Các cột tính toán nên được tính lại ở Backend hoặc lấy từ trường ẩn nếu bạn muốn
        ws.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, as_attachment=True, download_name=f"Bang_Cham_Cong_Khach_T{month}.xlsx")
# ----------------------------------------------------------------------
# 
# ----------------------------------------------------------------------
# Đảm bảo dòng này nằm ở cấp độ global (ngoài các hàm)
app.jinja_env.globals.update(getattr=getattr)

def get_attendance_values(val):
    val = val.strip().upper()
    
    # Xử lý trường hợp ô trống (đã xóa dữ liệu)
    if not val:
        return (0.0, 0.0, 0.0)
    
    """
    Trả về (công, lương, ăn ca) dựa trên quy ước.
    """
    # Quy ước: (Ngày Công, Ngày Lương, Ăn Ca)
    mapping = {
        'X':  (1.0, 1.0, 1.0), # Đi làm: 1 công, 1 lương, 1 ăn ca
        'SC': (0.5, 0.5, 0.5),
        'B':  (1.0, 1.0, 1.0),
        'P':  (0.0, 1.0, 0.0), # Nghỉ phép: 0 công, 1 lương, 0 ăn ca
        'V':  (0.0, 1.0, 0.0),
        'C':  (0.0, 1.0, 0.0),
        'T':  (0.0, 1.0, 0.0),
        'S':  (0.0, 1.0, 0.0),
        'M':  (0.0, 1.0, 0.0),
        'H':  (0.0, 1.0, 0.0),
        'D':  (0.0, 0.0, 0.0),
        'N':  (0.0, 0.0, 0.0),
        'K':  (0.0, 0.0, 0.0),
        'L':  (0.0, 0.0, 0.0), # Nghỉ lễ: 0 công, 0 lương, 0 ăn ca
    }
    # Mặc định nếu là số (0.5, 1) thì coi như đi làm (X)
    if val in mapping:
        return mapping[val]
    try:
        f_val = float(val)
        return (f_val, f_val, f_val)
    except:
        # Nếu nhập ký tự lạ, trả về 0 để không gây lỗi tính tổng
        return (0.0, 0.0, 0.0)

# Ví dụ logic kiểm tra khóa
@app.route('/lock_attendance/<int:month>/<int:year>')
@login_required
def lock_attendance(month, year):
    # 1. Kiểm tra quyền (Chỉ Admin hoặc mã nhân viên cụ thể)
    # Thay '200905691' bằng mã nhân viên bạn muốn cấp quyền đặc biệt (Đinh Thị Ánh Hồng ma_nhan_vien == '200905691')
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    is_admin_or_th = (current_user.is_admin == 1 or (user_info and user_info.ma_phong_ban == 2))
    is_special_user = (current_user.ma_nhan_vien == '200905691') 

    if not (is_admin_or_th or is_special_user):
        flash("Bạn không có quyền thực hiện thao tác này!", "danger")
        return redirect(url_for('cham_cong', month=month, year=year))

    # 2. Lấy tham số từ URL
    status_str = request.args.get('status', 'false').lower()
    new_status = True if status_str == 'true' else False
    ma_pb = request.args.get('ma_phong_ban')
    
    # Chuyển đổi ma_pb sang int hoặc None
    ma_pb = int(ma_pb) if (ma_pb and ma_pb != '' and ma_pb != 'None') else None

    try:
        # 3. Tìm bản ghi cũ hoặc tạo mới
        lock_rec = ChamCongLock.query.filter_by(
            thang=month, 
            nam=year, 
            ma_phong_ban=ma_pb
        ).first()

        if not lock_rec:
            lock_rec = ChamCongLock(
                thang=month, 
                nam=year, 
                ma_phong_ban=ma_pb
            )
            db.session.add(lock_rec)

        # 4. Cập nhật trạng thái
        lock_rec.is_locked = new_status
        lock_rec.updated_by = current_user.ma_nhan_vien
        db.session.commit()

        msg = "Đã KHÓA bảng công thành công!" if new_status else "Đã MỞ KHÓA bảng công!"
        flash(msg, "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Lỗi hệ thống: {str(e)}", "danger")

    return redirect(url_for('cham_cong', month=month, year=year, ma_phong_ban=ma_pb))

@app.route('/get_phong_ban_by_don_vi/<ma_hieu_2>')
@login_required
def get_phong_ban_by_don_vi(ma_hieu_2):
    # Lấy các phòng ban thuộc đơn vị ma_hieu_2
    phong_bans = PhongBan.query.filter_by(ma_hieu_2=ma_hieu_2).all()
    
    # Chuyển thành danh sách dict để trả về JSON
    results = [
        {"id": pb.id, "ten_phong_ban": pb.ten_phong_ban} 
        for pb in phong_bans
    ]
    return jsonify(results)

@app.route('/cham-cong', methods=['GET', 'POST'])
@login_required
def cham_cong():
    # 1. LẤY THÔNG TIN NGƯỜI DÙNG ĐANG ĐĂNG NHẬP
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()

    # Ép kiểu String để so sánh chính xác với mã phòng '2'
    user_ma_pb = str(user_info.ma_phong_ban) if user_info and user_info.ma_phong_ban else None
    user_ma_hieu_2 = user_info.ma_hieu_2 if user_info else None
    req_ma_pb = request.args.get('ma_phong_ban', '')
    
    # Quyền đặc biệt: Admin hệ thống (is_admin=1) 
    is_system_admin = (current_user.is_admin == 1)
    # Quyền Phòng Tổng hợp (mã 2): Thấy hết mọi phòng trong ĐƠN VỊ của mình
    is_phong_tong_hop = (user_ma_pb == '2')
    # Biến gộp để dùng cho các logic kiểm tra quyền chung
    is_admin_or_th = (is_system_admin or is_phong_tong_hop)

    #####print(f"DEBUG: ma_hieu_2={request.args.get('ma_hieu_2')}, ma_phong_ban={request.args.get('ma_phong_ban')}")
    
    # 2. XỬ LÝ THAM SỐ THỜI GIAN
    try:
        month = int(request.args.get('month', datetime.now().month))
        year = int(request.args.get('year', datetime.now().year))
    except (ValueError, TypeError):
        month, year = datetime.now().month, datetime.now().year

    # Xác định số ngày trong tháng
    _, num_days = calendar.monthrange(year, month)

    # 3. LOGIC PHÂN QUYỀN LỌC PHÒNG BAN
    # Lấy giá trị từ request
    req_ma_hieu_2 = request.args.get('ma_hieu_2')
    req_ma_pb = request.args.get('ma_phong_ban')
    if is_system_admin:
        # Admin thấy tất cả, mặc định lấy từ request args
        ma_hieu_2_filter = req_ma_hieu_2 if req_ma_hieu_2 else ''
        ma_pb_filter = req_ma_pb if req_ma_pb else ''
    elif is_phong_tong_hop:
        # Phòng Tổng hợp: Luôn bị khóa chặt ở đơn vị của mình
        ma_hieu_2_filter = user_ma_hieu_2
        # Nhưng có quyền chọn xem bất kỳ phòng ban nào trong đơn vị đó
        ma_pb_filter = req_ma_pb if req_ma_pb else ''
    else:
        # User thường: Khóa chặt cả đơn vị và phòng ban
        ma_hieu_2_filter = user_ma_hieu_2
        # Đảm bảo lấy ID từ database nếu người dùng không truyền params
        ma_pb_filter = str(user_info.ma_phong_ban) if user_info.ma_phong_ban else ''

    # 4. KIỂM TRA TRẠNG THÁI KHÓA BẢNG CHẤM CÔNG (Sửa lại logic filter)
    # Tìm bản ghi khóa khớp với Năm, Tháng, Đơn vị và Phòng ban
    lock_query = ChamCongLock.query.filter(
        ChamCongLock.nam == year,
        ChamCongLock.thang == month,
        ChamCongLock.is_locked == 1
    )
    
    #if ma_hieu_2_filter:
    #    # Nếu đang xem 1 đơn vị cụ thể, kiểm tra xem đơn vị đó có bị khóa không
    #    lock_query = lock_query.filter(
    #        (ChamCongLock.ma_hieu_2 == ma_hieu_2_filter) | (ChamCongLock.ma_hieu_2 == None)
    #    )
    #    if ma_pb_filter:
    #        # Nếu xem 1 phòng cụ thể, kiểm tra khóa phòng đó hoặc khóa toàn đơn vị
    #        lock_query = lock_query.filter(
    #            (ChamCongLock.ma_phong_ban == ma_pb_filter) | (ChamCongLock.ma_phong_ban == None)
    #        )
    # Nếu cần kiểm tra khóa toàn đơn vị HOẶC khóa cụ thể phòng
    if ma_hieu_2_filter and ma_pb_filter:
        lock_query = lock_query.filter(
            (ChamCongLock.ma_hieu_2 == ma_hieu_2_filter) & 
            ((ChamCongLock.ma_phong_ban == ma_pb_filter) | (ChamCongLock.ma_phong_ban.is_(None)))
        )
    
    lock_record = lock_query.first()
    is_locked = True if lock_record else False

    # 5. TRUY VẤN DANH SÁCH NGÀY LỄ
    ngay_le_records = DanhMucNgayLe.query.filter(
        db.extract('month', DanhMucNgayLe.ngay) == month,
        db.extract('year', DanhMucNgayLe.ngay) == year
    ).all()
    holidays = [r.ngay.day for r in ngay_le_records]

    # 6. XỬ LÝ LƯU DỮ LIỆU (POST)
    if request.method == 'POST':
        # Chỉ Admin hoặc Phòng Tổng hợp mới được lưu cho người khác
        # User thường chỉ được lưu nếu ma_pb_filter khớp với chính mình
        can_edit = is_system_admin or is_phong_tong_hop or (str(ma_pb_filter) == str(user_ma_pb))
        
        # CHẶN LƯU NẾU ĐÃ KHÓA (Ngoại trừ Admin có quyền can thiệp)
        if is_locked and not is_system_admin:
            flash(f"Dữ liệu tháng {month}/{year} đã bị khóa. Vui lòng liên hệ Admin!", "danger")
            return redirect(url_for('cham_cong', month=month, year=year, ma_hieu_2=ma_hieu_2_filter, ma_phong_ban=ma_pb_filter))
        if not can_edit:
            flash("Bạn không có quyền cập nhật dữ liệu ngoài phòng ban!", "danger")
            return redirect(url_for('cham_cong', month=month, year=year))

        try:
            # Lọc nhân viên cần cập nhật công theo bộ lọc đơn vị + phòng ban
            staff_query = ThongTinNguoiLaoDong.query.filter_by(trang_thai=True)
            if ma_hieu_2_filter:
                staff_query = staff_query.filter_by(ma_hieu_2=ma_hieu_2_filter)
            if ma_pb_filter:
                staff_query = staff_query.filter_by(ma_phong_ban=ma_pb_filter)

            active_staff = staff_query.all()

            # Lấy tất cả records của tháng hiện tại vào 1 dict
            existing_records = ThongTinChamCong.query.filter_by(thang=month, nam=year).all()
            record_dict = {r.ma_nhan_vien: r for r in existing_records}
            # Sử dụng no_autoflush để tối ưu hiệu năng vòng lặp
            with db.session.no_autoflush:
                for staff in active_staff:
                    # Nếu không thấy dữ liệu d1 của nhân viên này trong form, bỏ qua (tránh lỗi khi lọc)
                    if f"d1_{staff.ma_nhan_vien}" not in request.form:
                        continue

                    record = ThongTinChamCong.query.filter_by(
                        ma_nhan_vien=staff.ma_nhan_vien, 
                        thang=month, 
                        nam=year
                    ).first()
                    
                    if not record:
                        record = ThongTinChamCong(ma_nhan_vien=staff.ma_nhan_vien, thang=month, nam=year)
                        db.session.add(record)

                    # Biến tính toán tổng
                    tc_thanh_toan = 0.0
                    t_luong = 0.0
                    t_an_ca = 0.0

                    for d in range(1, 32):
                        field_name = f"d{d}"
                        # Lấy giá trị từ form
                        val = request.form.get(f"{field_name}_{staff.ma_nhan_vien}", "").strip().upper()
                        # QUAN TRỌNG: Gán giá trị vào thuộc tính của record
                        # Nếu val là "" (chuỗi rỗng), nó sẽ ghi đè vào giá trị cũ trong DB (ví dụ từ 'X' thành '')
                        setattr(record, field_name, val)

                        if d <= num_days:
                            # Hàm get_attendance_values cần trả về (công, lương, ăn ca) dựa trên ký hiệu (X, P, SC...)
                            n_cong, n_luong, a_ca = get_attendance_values(val)
                            tc_thanh_toan += n_cong
                            t_luong += n_luong
                            t_an_ca += a_ca
                        else:
                            # Xóa dữ liệu các ngày thừa của tháng (ví dụ d31 của tháng 2)
                            # Đảm bảo ngày thừa của tháng luôn rỗng
                            setattr(record, field_name, "")
                    
                    # Gán giá trị tổng vào record
                    record.tong_cong_thanh_toan = tc_thanh_toan
                    record.luong = t_luong
                    record.an_ca = t_an_ca
            
            db.session.commit()
            flash(f"Đã lưu bảng công tháng {month}/{year} thành công!", "success")
            
        except Exception as e:
            db.session.rollback()
            flash(f"Lỗi khi lưu dữ liệu: {str(e)}", "danger")
        
        return redirect(url_for('cham_cong', month=month, year=year, ma_hieu_2=ma_hieu_2_filter, ma_phong_ban=ma_pb_filter))

    # 7. XỬ LÝ HIỂN THỊ (GET)
    # Join thêm bảng đơn vị để hiển thị đầy đủ thông tin => outerjoin để tránh mất dữ liệu nếu đơn vị/phòng ban bị thiếu
    query = db.session.query(ThongTinNguoiLaoDong, PhongBan.ten_phong_ban, DonVi.ten_ma_hieu_2)\
        .outerjoin(PhongBan, ThongTinNguoiLaoDong.ma_phong_ban == PhongBan.id)\
        .outerjoin(DonVi, ThongTinNguoiLaoDong.ma_hieu_2 == DonVi.ma_hieu_2)\
        .filter(ThongTinNguoiLaoDong.trang_thai == True)

    # Lọc theo đơn vị (ma_hieu_2)
    if ma_hieu_2_filter:
        # Ép kiểu về chuỗi để so sánh an toàn
        query = query.filter(ThongTinNguoiLaoDong.ma_hieu_2 == str(ma_hieu_2_filter))
    # Lọc theo phòng ban (ma_phong_ban)
    if ma_pb_filter:
        # Nếu ma_pb_filter có giá trị, thử ép kiểu về Int nếu DB là Integer
        # Dùng str() để đảm bảo so sánh chuỗi với chuỗi
        query = query.filter(db.cast(ThongTinNguoiLaoDong.ma_phong_ban, db.String) == str(ma_pb_filter))

    #####print(f"DEBUG: Số lượng nhân viên tìm thấy: {query.count()}")
    
    #Sắp xếp theo Họ tên
    staff_list = query.order_by(ThongTinNguoiLaoDong.ho_ten.asc()).all()

    # TỐI ƯU DROPDOWN PHÒNG BAN THEO ĐƠN VỊ
    if is_system_admin:
        ds_don_vi = DonVi.query.all()
        # Nếu đã chọn đơn vị thì chỉ hiện phòng của đơn vị đó, nếu chưa thì hiện hết
        pb_query = PhongBan.query
        if ma_hieu_2_filter:
            pb_query = pb_query.filter_by(ma_hieu_2=ma_hieu_2_filter)
        ds_phong_ban = pb_query.all()
    elif is_phong_tong_hop:
        ds_don_vi = DonVi.query.filter_by(ma_hieu_2=user_ma_hieu_2).all()
        ds_phong_ban = PhongBan.query.filter_by(ma_hieu_2=user_ma_hieu_2).all()
    else:
        ds_don_vi = DonVi.query.filter_by(ma_hieu_2=user_ma_hieu_2).all()
        # User thường chỉ thấy 1 phòng của họ
        ds_phong_ban = PhongBan.query.filter_by(id=user_info.ma_phong_ban).all()
    
    # Lấy dữ liệu chấm công đã tồn tại trong db để điền vào bảng
    staff_ids = [s[0].ma_nhan_vien for s in staff_list]
    attendance_records = ThongTinChamCong.query.filter(
        ThongTinChamCong.ma_nhan_vien.in_(staff_ids) if staff_ids else text("1=0"),
        ThongTinChamCong.thang == month,
        ThongTinChamCong.nam == year
    ).all()
    
    #att_dict = {r.ma_nhan_vien: r for r in attendance_records}
    # CHUYỂN ĐỔI MODEL THÀNH DICT ĐỂ HTML DỄ HIỂN THỊ
    att_dict = {}
    for r in attendance_records:
        # Ép kiểu ma_nhan_vien về string để tránh lỗi không khớp kiểu (int vs str)
        key = str(r.ma_nhan_vien)
        
        # Gán giá trị vào dict, bao gồm cả ngày và cột tổng
        record_data = {f'd{i}': getattr(r, f'd{i}', '') for i in range(1, 32)}
        record_data['tong_cong_thanh_toan'] = r.tong_cong_thanh_toan or 0
        record_data['luong'] = r.luong or 0
        record_data['an_ca'] = r.an_ca or 0
        
        att_dict[key] = record_data


    has_data_staff_ids = [s.ma_nhan_vien for s in attendance_records]

    return render_template('cham_cong.html', 
                           staff_list=staff_list, 
                           att_dict=att_dict, 
                           month=month, 
                           year=year,
                           don_vis=ds_don_vi, 
                           phong_bans=ds_phong_ban,
                           current_don_vi=ma_hieu_2_filter, 
                           current_pb=str(ma_pb_filter) if ma_pb_filter else '', 
                           is_admin_or_th=is_admin_or_th,
                           is_locked=is_locked,
                           holidays=holidays,
                           has_data_staff_ids=has_data_staff_ids)

# ----------------------------------------------------------------------
# Xuất Excel (từ CSDL.thong_tin_cham_cong)
# ----------------------------------------------------------------------
@app.route('/export-excel')
@login_required
def export_excel():
    try:
        # --- PHÂN QUYỀN VÀ KIỂM TRA QUYỀN ---
        user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
        user_ma_pb = user_info.ma_phong_ban if user_info else None
        user_ma_hieu_2 = user_info.ma_hieu_2 if user_info else None
        req_ma_pb = request.args.get('ma_phong_ban', '')
        
        # Quyền Admin hoặc thuộc các phòng ban đặc biệt (1: Ban Giám đốc; 2: Tổng hợp; 3: Kiểm tra, giám sát; 4: Thu thập và Xử lý thông tin)
        ##is_admin_or_th = (current_user.is_admin == 1 or user_ma_pb in [2])
        is_system_admin = (current_user.is_admin == 1)
        is_phong_tong_hop = (user_ma_pb == '2')
        is_admin_or_th = (is_system_admin or is_phong_tong_hop)

        # --- LẤY THAM SỐ TỪ REQUEST ---
        month = int(request.args.get('month', datetime.now().month))
        year = int(request.args.get('year', datetime.now().year))
        req_ma_hieu_2 = request.args.get('ma_hieu_2', '')
        req_ma_pb = request.args.get('ma_phong_ban', '')

        # LOGIC BẢO MẬT: Nếu không phải Admin/TH, cưỡng ép lọc theo mã phòng của User
        ##if not is_admin_or_th:
        ##    ma_pb_filter = str(user_ma_pb)
        # --- LOGIC PHÂN QUYỀN LỌC DỮ LIỆU ĐỒNG BỘ ---
        if is_system_admin:
            ma_hieu_2_filter = req_ma_hieu_2
            ma_pb_filter = req_ma_pb
        elif is_phong_tong_hop:
            ma_hieu_2_filter = user_ma_hieu_2
            ma_pb_filter = req_ma_pb
        else:
            ma_hieu_2_filter = user_ma_hieu_2
            ma_pb_filter = user_ma_pb

        # 1. Xác định số ngày thực tế của tháng
        _, num_days = calendar.monthrange(year, month)

        # 2. Lấy dữ liệu từ Database (Join cả 2 bảng để lấy Tên Đơn Vị và Tên Phòng Ban)
        query = db.session.query(ThongTinNguoiLaoDong, PhongBan.ten_phong_ban, DonVi.ten_ma_hieu_2)\
            .outerjoin(PhongBan, ThongTinNguoiLaoDong.ma_phong_ban == PhongBan.id)\
            .outerjoin(DonVi, ThongTinNguoiLaoDong.ma_hieu_2 == DonVi.ma_hieu_2)\
            .filter(ThongTinNguoiLaoDong.trang_thai == True)

        # Áp dụng bộ lọc Đơn vị (ma_hieu_2)
        if ma_hieu_2_filter:
            query = query.filter(ThongTinNguoiLaoDong.ma_hieu_2 == str(ma_hieu_2_filter))
        # Áp dụng bộ lọc phòng ban (ma_phong_ban)
        if ma_pb_filter:
            query = query.filter(db.cast(ThongTinNguoiLaoDong.ma_phong_ban, db.String) == str(ma_pb_filter))
        
        staff_list = query.order_by(ThongTinNguoiLaoDong.ho_ten.asc()).all()
        ##if ma_pb_filter.isdigit():
        ##    query = query.filter(ThongTinNguoiLaoDong.ma_phong_ban == int(ma_pb_filter))
        ##elif not is_admin_or_th:
        ##    # Trường hợp hy hữu không xác định được phòng ban của User
        ##    query = query.filter(ThongTinNguoiLaoDong.id == -1) 
        ##staff_list = query.all()
        
        # Lấy dữ liệu chấm công của các nhân viên đã lọc
        staff_ids = [s[0].ma_nhan_vien for s in staff_list]
        # Ngăn chặn crash query lỗi IN nếu không có nhân viên nào thỏa mãn bộ lọc
        if not staff_ids:
            attendance_data = []
        else:
            attendance_data = ThongTinChamCong.query.filter(
                ThongTinChamCong.thang == month,
                ThongTinChamCong.nam == year,
                ThongTinChamCong.ma_nhan_vien.in_(staff_ids)
            ).all()

        att_dict = {str(r.ma_nhan_vien): r for r in attendance_data}

        # 3. Load Template
        template_path = os.path.join(app.root_path, 'template_bang_cong.xlsx')
        if not os.path.exists(template_path):
            raise FileNotFoundError("Không tìm thấy file template_bang_cong.xlsx")
            
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # Hàm bổ trợ: Ghi giá trị an toàn cho ô bị Merge
        def safe_write(sheet, row, col, value, style_func=None):
            cell = sheet.cell(row=row, column=col)
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break
            cell.value = value
            if style_func: style_func(cell)

        # 4. Định nghĩa màu sắc & Style
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')

        # 5. Xử lý ẩn cột ngày dư (29, 30, 31)
        for d in range(29, 32):
            col_letter = get_column_letter(3 + d)
            if d > num_days:
                ws.column_dimensions[col_letter].hidden = True
            else:
                ws.column_dimensions[col_letter].hidden = False

        # 6. Ghi tiêu đề tháng năm
        safe_write(ws, 3, 17, f"Tháng {month} Năm {year}")

        # --- LẤY TÊN ĐƠN VỊ & PHÒNG BAN ĐỂ GHI VÀO TIÊU ĐỀ FILE ---
        ten_don_vi_ghi_excel = "Tất cả"
        ten_phong_ban_ghi_excel = "Tất cả"
        # Nếu có chọn Đơn vị cụ thể
        if ma_hieu_2_filter:
            don_vi_obj = DonVi.query.filter_by(ma_hieu_2=ma_hieu_2_filter).first()
            if don_vi_obj:
                ten_don_vi_ghi_excel = don_vi_obj.ten_ma_hieu_2
        # Nếu có chọn Phòng ban cụ thể
        if ma_pb_filter:
            phong_ban_obj = PhongBan.query.filter_by(id=ma_pb_filter).first()
            if phong_ban_obj:
                ten_phong_ban_ghi_excel = phong_ban_obj.ten_phong_ban
        # Trường hợp xem "Tất cả" nhưng danh sách nhân viên có phòng ban (lấy phòng ban của nhân viên đầu tiên làm đại diện nếu cần)
        elif staff_list and len(staff_list) > 0:
            # Nếu chỉ có 1 phòng ban duy nhất trong danh sách xuất ra
            cac_phong = set([s[1] for s in staff_list if s[1]])
            if len(cac_phong) == 1:
                ten_phong_ban_ghi_excel = list(cac_phong)[0]
        # Ghi đè vào ô A3 và A4 của Template
        safe_write(ws, 3, 1, f"Đơn vị: {ten_don_vi_ghi_excel}")
        safe_write(ws, 4, 1, f"Phòng Ban: {ten_phong_ban_ghi_excel}")
        # ---------------------------------------------------------

        # 7. Ghi danh sách nhân viên và công
        start_row = 7
        current_row = start_row

        # Khởi tạo tên phòng ban mặc định để đặt tên file xuất ra sau này
        ten_file_phong_ban = ""
        
        for index, (staff, ten_pb, ten_dv) in enumerate(staff_list):
            current_row = start_row + index
            if not ten_file_phong_ban and ten_pb:
                ten_file_phong_ban = ten_pb
            
            # Ghi thông tin nhân viên
            safe_write(ws, current_row, 1, staff.ma_nhan_vien)
            safe_write(ws, current_row, 2, staff.ho_ten)
            safe_write(ws, current_row, 3, ten_pb if ten_pb else "") # Thay đổi hiển thị bộ phận

            record = att_dict.get(staff.ma_nhan_vien)
            
            # Điền dữ liệu ngày
            for d in range(1, 32):
                if d <= num_days:
                    val = getattr(record, f"d{d}", "") if record else ""
                    
                    def apply_style(c):
                        c.alignment = center_align
                        # Logic màu sắc giữ nguyên
                        if val in ['X', 'SC', 'B']:
                            c.fill = green_fill
                        elif val in ['P', 'V', 'C', 'T', 'S', 'M', 'H']:
                            c.fill = red_fill
                        elif val in ['D', 'N', 'K']:
                            c.font = Font(color="FF0000", bold=True)
                    
                    safe_write(ws, current_row, 3 + d, val, apply_style)
                else:
                    safe_write(ws, current_row, 3 + d, "")

            # Điền các cột tổng kết
            if record:
                safe_write(ws, current_row, 35, record.tong_cong_thanh_toan, lambda c: setattr(c, 'alignment', center_align))
                safe_write(ws, current_row, 36, record.luong, lambda c: setattr(c, 'alignment', center_align))
                safe_write(ws, current_row, 37, record.an_ca, lambda c: setattr(c, 'alignment', center_align))

        # --- BẮT ĐẦU: THÊM KHU VỰC CHỮ KÝ ---
        # Xác định dòng bắt đầu ghi chữ ký (Cách danh sách nhân viên 2 dòng trống)
        sign_row = current_row + 3

        # Style định dạng cho chữ ký
        bold_font = Font(name="Arial", size=11, bold=True)
        italic_font = Font(name="Arial", size=10, italic=True)
        left_align = Alignment(horizontal='left', vertical='center')

        # Thêm dòng Ngày... Tháng... Năm... ở góc phải (cột Công TT / cột 35)
        now = datetime.now()
        date_str = f"Ngày {now.day} tháng {now.month} năm {now.year}"
        safe_write(ws, sign_row, 35, date_str, lambda c: setattr(c, 'alignment', center_align))
        ws.cell(row=sign_row, column=35).font = italic_font

        # Chuyển sang dòng tiếp theo để ghi Chức danh
        sign_title_row = sign_row + 1

        # Cột 2 (Họ tên): Người lập bảng
        safe_write(ws, sign_title_row, 2, "NGƯỜI LẬP BẢNG", lambda c: setattr(c, 'alignment', center_align))
        ws.cell(row=sign_title_row, column=2).font = bold_font

        # Cột 18 (Khoảng giữa bảng): Người kiểm soát
        safe_write(ws, sign_title_row, 18, "NGƯỜI KIỂM SOÁT", lambda c: setattr(c, 'alignment', center_align))
        ws.cell(row=sign_title_row, column=18).font = bold_font

        # Cột 35 (Tổng hợp): Giám đốc
        safe_write(ws, sign_title_row, 35, "GIÁM ĐỐC", lambda c: setattr(c, 'alignment', center_align))
        ws.cell(row=sign_title_row, column=35).font = bold_font

        # Thêm dòng ghi chú phụ dưới chức danh (Ký, ghi rõ họ tên)
        sign_note_row = sign_title_row + 1
        safe_write(ws, sign_note_row, 2, "(Ký, họ tên)", lambda c: setattr(c, 'alignment', center_align))
        ws.cell(row=sign_note_row, column=2).font = italic_font

        safe_write(ws, sign_note_row, 18, "(Ký, họ tên)", lambda c: setattr(c, 'alignment', center_align))
        ws.cell(row=sign_note_row, column=18).font = italic_font

        safe_write(ws, sign_note_row, 35, "(Ký, họ tên, đóng dấu)", lambda c: setattr(c, 'alignment', center_align))
        ws.cell(row=sign_note_row, column=35).font = italic_font
        # --- KẾT THÚC: THÊM KHU VỰC CHỮ KÝ ---

        # 8. Trả về file
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Đặt tên file xuất ra linh hoạt
        if is_system_admin and not ma_hieu_2_filter and not ma_pb_filter:
            file_name = f"Bang_Cong_Toan_He_Thong_{month}_{year}.xlsx"
        else:
            ten_file = ten_file_phong_ban if ma_pb_filter else "Tong_Hop"
            file_name = f"Bang_Cong_{ten_file}_{month}_{year}.xlsx"
        
        return send_file(output, as_attachment=True, download_name=file_name)

    except Exception as e:
        print(traceback.format_exc())
        flash(f"Lỗi xuất file: {str(e)}", "danger")
        return redirect(url_for('cham_cong'))

# ----------------------------------------------------------------------
# Xuất Excel Bảng chấm công dành cho Khách
# ----------------------------------------------------------------------
@app.route('/export-excel-guest', methods=['POST'])
def export_excel_guest():
    try:
        # 1. Lấy tham số thời gian
        month = int(request.args.get('month', datetime.now().month))
        year = int(request.args.get('year', datetime.now().year))
        ma_pb_filter = request.args.get('ma_phong_ban', '')
        _, num_days = calendar.monthrange(year, month)

        # 2. Lấy danh sách nhân viên
        query = db.session.query(ThongTinNguoiLaoDong, PhongBan.ten_phong_ban)\
            .join(PhongBan, ThongTinNguoiLaoDong.ma_phong_ban == PhongBan.id)\
            .filter(ThongTinNguoiLaoDong.trang_thai == True)
        
        if ma_pb_filter and ma_pb_filter.isdigit():
            query = query.filter(ThongTinNguoiLaoDong.ma_phong_ban == int(ma_pb_filter))
        staff_list = query.all()

        # 3. Load Template & Cấu hình Style
        template_path = os.path.join(app.root_path, 'template_bang_cong.xlsx')
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')

        def safe_write(sheet, row, col, value, style_func=None):
            cell = sheet.cell(row=row, column=col)
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break
            cell.value = value
            if style_func: style_func(cell)

        # Định nghĩa quy tắc tính toán giống hệt trong file HTML của bạn
        rules = {
            'X':  {'cong': 1.0, 'luong': 1.0, 'anca': 1.0},
            'SC': {'cong': 0.5, 'luong': 0.5, 'anca': 0.5},
            'B':  {'cong': 1.0, 'luong': 1.0, 'anca': 1.0},
            'P':  {'cong': 0.0, 'luong': 1.0, 'anca': 0.0},
            'V':  {'cong': 0.0, 'luong': 1.0, 'anca': 0.0},
            'C':  {'cong': 0.0, 'luong': 1.0, 'anca': 0.0},
            'T':  {'cong': 0.0, 'luong': 1.0, 'anca': 0.0},
            'S':  {'cong': 0.0, 'luong': 1.0, 'anca': 0.0},
            'M':  {'cong': 0.0, 'luong': 1.0, 'anca': 0.0},
            'H':  {'cong': 0.0, 'luong': 1.0, 'anca': 0.0}
        }

        # 4. Ghi tiêu đề & ẩn cột dư
        safe_write(ws, 3, 17, f"Tháng {month} Năm {year}")
        for d in range(29, 32):
            ws.column_dimensions[get_column_letter(3 + d)].hidden = (d > num_days)

        # 5. Xử lý dữ liệu từng nhân viên
        start_row = 7  
        for index, (staff, ten_pb) in enumerate(staff_list):
            current_row = start_row + index
            ma_nv = staff.ma_nhan_vien
            
            safe_write(ws, current_row, 1, ma_nv)
            safe_write(ws, current_row, 2, staff.ho_ten)
            safe_write(ws, current_row, 3, ten_pb)

            # Khởi tạo biến tích lũy tổng kết
            tong_cong, tong_luong, tong_anca = 0.0, 0.0, 0.0

            for d in range(1, 32):
                if d <= num_days:
                    # Lấy giá trị từ Form (dữ liệu khách đã sửa trên web)
                    val = request.form.get(f"d{d}_{ma_nv}", "").upper().strip()
                    
                    # Logic tính toán tích lũy
                    if val in rules:
                        tong_cong += rules[val]['cong']
                        tong_luong += rules[val]['luong']
                        tong_anca += rules[val]['anca']
                    elif val.replace('.', '', 1).isdigit(): # Nếu nhập số trực tiếp (ví dụ 0.5)
                        v_num = float(val)
                        tong_cong += v_num
                        tong_luong += v_num
                        tong_anca += 1.0 if v_num >= 0.5 else 0.0

                    # Định dạng màu sắc ô ngày
                    def apply_style(c):
                        c.alignment = center_align
                        if val in ['X', 'SC', 'B']: c.fill = green_fill
                        elif val in ['P', 'V', 'C', 'T', 'S', 'M', 'H']: c.fill = red_fill
                        elif val in ['D', 'N', 'K']: c.font = Font(color="FF0000", bold=True)
                    
                    safe_write(ws, current_row, 3 + d, val, apply_style)
                else:
                    safe_write(ws, current_row, 3 + d, "")

            # 6. Ghi các cột tổng kết (Cột 35: AI, 36: AJ, 37: AK)
            style_total = lambda c: setattr(c, 'alignment', center_align)
            safe_write(ws, current_row, 35, tong_cong, style_total)
            safe_write(ws, current_row, 36, tong_luong, style_total)
            safe_write(ws, current_row, 37, tong_anca, style_total)

        # 7. Xuất file
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Bang_Cong_KHACH_{month}_{year}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename)

    except Exception as e:
        traceback.print_exc()
        return f"Lỗi hệ thống: {str(e)}", 500

# ----------------------------------------------------------------------
# Kiểm tra việc khai báo 12 tháng về mức lương cơ bản v1; mức lương v2
# ----------------------------------------------------------------------
@app.route('/config_salary')
@login_required
def config_salary():
    # 1. Kiểm tra quyền Admin
    if not current_user.role or current_user.role.upper() != 'ADMIN':
        flash("Bạn không có quyền truy cập trang này!", "danger")
        return redirect(url_for('index'))

    now = datetime.now(HANOI_TZ) if 'HANOI_TZ' in globals() else datetime.now()
    nam = now.year

    # 2. Đảm bảo luôn có dòng dữ liệu cho năm hiện tại (Dùng vòng lặp cho gọn)
    tables = {
        'v1': 'thong_tin_muc_luong_co_ban',
        'v2': 'thong_tin_muc_luong_v2'
    }
    
    for key, table in tables.items():
        check = db.session.execute(
            text(f"SELECT 1 FROM {table} WHERE nam = :nam"), {'nam': nam}
        ).fetchone()
        
        if not check:
            db.session.execute(
                text(f"INSERT INTO {table} (nam) VALUES (:nam)"), {'nam': nam}
            )
            db.session.commit()

    # 3. Lấy dữ liệu cuối cùng để hiển thị
    v1_data = db.session.execute(
        text("SELECT * FROM thong_tin_muc_luong_co_ban WHERE nam = :nam"), {'nam': nam}
    ).fetchone()
    v2_data = db.session.execute(
        text("SELECT * FROM thong_tin_muc_luong_v2 WHERE nam = :nam"), {'nam': nam}
    ).fetchone()

    return render_template('config_salary.html', 
                           v1_data=v1_data, 
                           v2_data=v2_data, 
                           selected_year=nam)

@app.route('/save_salary_config', methods=['POST'])
@login_required
def save_salary_config():
    # 1. Kiểm tra quyền Admin
    if not current_user.role or current_user.role.upper() != 'ADMIN':
        flash("Bạn không có quyền thực hiện thao tác này!", "danger")
        return redirect(url_for('index'))

    nam = request.form.get('nam')
    if not nam:
        flash("Năm không hợp lệ!", "warning")
        return redirect(url_for('index'))
    
    # 2. Cập nhật dữ liệu cho từng bảng
    tables = [
        ('thong_tin_muc_luong_co_ban', 'v1_'),
        ('thong_tin_muc_luong_v2', 'v2_')
    ]
    
    try:
        for table_name, prefix in tables:
            # Xây dựng câu SQL động
            update_parts = [f"ky_{i} = :k{i}" for i in range(1, 13)]
            sql = text(f"UPDATE {table_name} SET {', '.join(update_parts)} WHERE nam = :nam")
            
            # Chuẩn bị tham số và ép kiểu số sạch
            params = {'nam': nam}
            for i in range(1, 13):
                # Lấy giá trị, xóa dấu phẩy (nếu có), ép về float, mặc định 0 nếu lỗi
                raw_val = request.form.get(f"{prefix}ky_{i}", "0")
                clean_val = raw_val.replace(',', '').strip()
                params[f"k{i}"] = float(clean_val) if clean_val else 0.0
            
            db.session.execute(sql, params)
        
        db.session.commit()
        flash(f"Đã cập nhật cấu hình lương năm {nam} thành công!", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"Lỗi hệ thống: {str(e)}", "danger")
    
    return redirect(url_for('index'))

# ----------------------------------------------------------------------
# Bảng Hệ số lương V1, Tiền lương V1
# ----------------------------------------------------------------------
@app.route('/bang-luong-v1')
@login_required
def bang_luong_v1():
    # 1. Lấy tham số lọc
    year = request.args.get('year', datetime.now().year, type=int)
    selected_pb = request.args.get('phong_ban', '')
    ma_nv_filter = request.args.get('ma_nv', '')

    # 2. Lấy danh sách phòng ban cho dropdown
    ds_phong_ban = PhongBan.query.all()

    # 3. Xây dựng câu lệnh SQL có JOIN để lấy thông tin tổng hợp
    # Sử dụng v1.* để lấy cột approved_by và approved_at
    sql_text = """
        SELECT 
            nv.ma_nhan_vien, 
            nv.ho_ten, 
            pb.ten_phong_ban, 
            v1.*,
            u_app.fullname as ten_nguoi_duyet -- Lấy tên đầy đủ người duyệt
        FROM thong_tin_nguoi_lao_dong nv
        JOIN thong_tin_v1 v1 ON nv.ma_nhan_vien = v1.ma_nhan_vien
        LEFT JOIN phong_ban pb ON nv.ma_phong_ban = pb.id
        LEFT JOIN users u_app ON v1.approved_by = u_app.ma_nhan_vien -- Join để lấy tên người duyệt
        WHERE v1.nam = :year
    """
    
    params = {"year": year}

    if selected_pb:
        sql_text += " AND nv.ma_phong_ban = :pb"
        params["pb"] = selected_pb
    
    if ma_nv_filter:
        sql_text += " AND nv.ma_nhan_vien LIKE :ma"
        params["ma"] = f"%{ma_nv_filter}%"

    results = db.session.execute(text(sql_text), params).fetchall()

    return render_template('bang_luong_v1.html', 
                           results=results, 
                           year=year, 
                           ds_phong_ban=ds_phong_ban,
                           selected_pb=selected_pb,
                           ma_nv_filter=ma_nv_filter)

# Route xử lý phê duyệt (Dành cho Admin/KIEM_SOAT)
@app.route('/approve-v1', methods=['POST'])
@login_required
def approve_v1():
    # 1. Kiểm tra quyền hạn
    if current_user.role not in ['ADMIN', 'KIEM_SOAT']:
        return jsonify({'success': False, 'message': 'Bạn không có quyền phê duyệt!'}), 403

    data = request.get_json()
    try:
        year = int(data.get('year'))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': 'Năm không hợp lệ!'}), 400

    selected_pb = data.get('phong_ban')
    ma_nv_filter = data.get('ma_nv')
    updates = data.get('updates', []) # Nhận danh sách hệ số lương v1 mới

    try:
        # BƯỚC A: LƯU DỮ LIỆU HỆ SỐ MỚI (Nếu có gửi kèm updates)
        # Chỉ lưu cho những bản ghi chưa được duyệt
        if updates:
            for item in updates:
                ma_nv_item = item.get('ma_nv')
                # Tạo câu lệnh SET động cho 12 tháng
                set_clause = ", ".join([f"t{m} = :t{m}" for m in range(1, 13)])
                save_sql = f"""
                    UPDATE thong_tin_v1 
                    SET {set_clause} 
                    WHERE ma_nhan_vien = :ma_nv 
                      AND nam = :year 
                      AND (approved_by IS NULL OR approved_by = '' OR approved_by = 'None')
                """
                params_save = {f"t{m}": item.get(f't{m}', 0) for m in range(1, 13)}
                params_save.update({'ma_nv': ma_nv_item, 'year': year})
                
                db.session.execute(text(save_sql), params_save)

                # GỌI PROCEDURE TỔNG V1 ĐỂ TÍNH TIỀN
                db.session.execute(text("CALL sp_ChayTinhLuongV1_ToanCoQuan(:year)"), {"year": year})

                db.session.commit()

        # BƯỚC B: THỰC HIỆN PHÊ DUYỆT
        approve_query = """
            UPDATE thong_tin_v1 v1 
            JOIN thong_tin_nguoi_lao_dong nv ON v1.ma_nhan_vien = nv.ma_nhan_vien 
            SET v1.approved_by = :user, v1.approved_at = :now 
            WHERE v1.nam = :year
            AND (v1.approved_by IS NULL OR v1.approved_by = '' OR v1.approved_by = 'None')
        """
        params_approve = {
            "user": current_user.ma_nhan_vien, 
            "now": datetime.now(HANOI_TZ), 
            "year": year
        }

        # Áp dụng bộ lọc từ giao diện (nếu có)
        if selected_pb and str(selected_pb).strip():
            approve_query += " AND nv.ma_phong_ban = :pb"
            params_approve["pb"] = selected_pb
        if ma_nv_filter and str(ma_nv_filter).strip():
            approve_query += " AND v1.ma_nhan_vien = :ma"
            params_approve["ma"] = ma_nv_filter

        result = db.session.execute(text(approve_query), params_approve)
        
        # Lưu tất cả thay đổi vào DB
        db.session.commit()
        
        row_count = result.rowcount
        if row_count == 0:
            return jsonify({'success': True, 'message': 'Không tìm thấy dòng dữ liệu mới nào cần phê duyệt.'})
            
        log_event("Phê duyệt V1", f"Đã lưu và phê duyệt {row_count} bản ghi cho năm {year}")
        return jsonify({'success': True, 'message': f'Đã lưu và phê duyệt thành công {row_count} dòng!'})
    
    except Exception as e:
        db.session.rollback()
        print(f"Error in approve_v1: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/unapprove-v1', methods=['POST'])
@login_required
def unapprove_v1():
    # Chỉ Admin mới có quyền gỡ phê duyệt
    if not getattr(current_user, 'is_admin', False) and session.get('role') != 'ADMIN':
        return jsonify({'success': False, 'message': 'Chỉ Admin mới có quyền gỡ phê duyệt!'}), 403

    data = request.get_json()
    year = data.get('year')
    ma_nv = data.get('ma_nv') # Định danh dòng cụ thể để gỡ

    if not year or not ma_nv:
        return jsonify({'success': False, 'message': 'Thiếu thông tin năm hoặc mã nhân viên!'}), 400
    
    try:
        # Cập nhật bảng dữ liệu V1 (thong_tin_v1)
        query = """
            UPDATE thong_tin_v1 
            SET approved_by = NULL, approved_at = NULL 
            WHERE nam = :year AND ma_nhan_vien = :ma
        """
        db.session.execute(text(query), {"year": year, "ma": ma_nv})
        db.session.commit()
        
        log_event("Gỡ phê duyệt V1", f"Admin {current_user.ma_nhan_vien} đã gỡ phê duyệt cho NV {ma_nv} năm {year}")
        return jsonify({'success': True, 'message': 'Đã gỡ trạng thái phê duyệt!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    

@app.route('/v1-report')
@login_required  # Bắt buộc đăng nhập để biết user là ai
def v1_report():
    # 1. LẤY THÔNG TIN QUYỀN CỦA USER ĐANG ĐĂNG NHẬP
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    # Lấy mã phòng ban (để so sánh ID)
    user_ma_pb = user_info.ma_phong_ban if user_info else None
    # Lấy tên phòng ban (để hiển thị/lọc theo tên)
    # Dòng này sẽ KHÔNG còn lỗi AttributeError nữa
    user_ten_pb = user_info.phong_ban.ten_phong_ban if user_info and user_info.phong_ban else "" #user_ten_pb = ""
    
    if user_info and user_info.phong_ban:
        user_ten_pb = user_info.phong_ban.ten_phong_ban

    # Quyền Admin hoặc thuộc các phòng ban đặc biệt (1: Ban Giám đốc; 2: Tổng hợp; 3: Kiểm tra, giám sát; 4: Thu thập và Xử lý thông tin)
    is_admin_or_th = (current_user.is_admin == 1 or user_ma_pb in [2])

    # 2. XỬ LÝ THAM SỐ LỌC
    year = request.args.get('year', datetime.now().year, type=int)
    ma_nv_filter = request.args.get('ma_nv', '')
    
    # LOGIC PHÂN QUYỀN LỌC PHÒNG BAN:
    if is_admin_or_th:
        # Admin/Tổng hợp: Có thể xem phòng ban tùy chọn từ URL
        selected_pb = request.args.get('phong_ban', '')
    else:
        # User thường: Luôn ép về tên phòng ban của chính họ
        selected_pb = user_ten_pb

    # 3. TỐI ƯU: Tính công chuẩn 12 tháng một lần duy nhất
    cong_chuan_12_thang = {m: get_cong_chuan(year, m) for m in range(1, 13)}
    
    # 4. LẤY DANH SÁCH PHÒNG BAN CHO DROPDOWN (Chỉ hiển thị cho Admin/TH)
    ds_phong_ban = []
    if is_admin_or_th:
        ds_pb_query = db.session.execute(text("SELECT DISTINCT ten_phong_ban FROM phong_ban")).fetchall()
        ds_phong_ban = [r[0] for r in ds_pb_query if r[0]]

    # 5. TRUY VẤN DỮ LIỆU
    sql = """
        SELECT 
            nv.ma_nhan_vien, nv.ho_ten, pb.ten_phong_ban,
            tl.*
        FROM thong_tin_nguoi_lao_dong nv
        LEFT JOIN phong_ban pb ON nv.ma_phong_ban = pb.id
        LEFT JOIN tien_luong_v1 tl ON nv.ma_nhan_vien = tl.ma_nhan_vien AND tl.nam = :year
        WHERE nv.trang_thai = 1
    """
    
    params = {"year": year}
    
    # Ép điều kiện lọc phòng ban đã qua kiểm tra quyền
    if selected_pb:
        sql += " AND pb.ten_phong_ban = :pb"
        params["pb"] = selected_pb
    elif not is_admin_or_th:
        # Phòng trường hợp user không có phòng ban trong hệ thống
        sql += " AND 1=0" 

    if ma_nv_filter:
        sql += " AND nv.ma_nhan_vien LIKE :ma"
        params["ma"] = f"%{ma_nv_filter}%"

    sql += " ORDER BY pb.ten_phong_ban, nv.ma_nhan_vien"
    
    results = db.session.execute(text(sql), params).fetchall()

    # 6. CẤU TRÚC DỮ LIỆU CHO TEMPLATE
    report_data = []
    for r in results:
        emp_months = []
        for m in range(1, 13):
            emp_months.append({
                'hsl': getattr(r, f'hsl_v1_t{m}') or 0,
                'cong_tt': float(getattr(r, f'cong_t{m}') or 0),
                'tien': float(getattr(r, f'tien_v1_t{m}') or 0),
                'cong_chuan': cong_chuan_12_thang[m]
            })

        report_data.append({
            'ma_nv': r.ma_nhan_vien,
            'ho_ten': r.ho_ten,
            'phong_ban': r.ten_phong_ban or "N/A",
            'months': emp_months,
            'tong_nam': float(r.tien_v1_tong_nam or 0)
        })

    return render_template('v1_report.html', 
                           data=report_data, 
                           year=year, 
                           ds_phong_ban=ds_phong_ban, 
                           selected_pb=selected_pb, 
                           ma_nv_filter=ma_nv_filter,
                           is_admin_or_th=is_admin_or_th)


@app.route('/export-v1-excel/<int:year>')
@login_required
def export_v1_excel(year):
    # 1. PHÂN QUYỀN: Xác định phạm vi dữ liệu được phép xuất
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    user_ma_pb = user_info.ma_phong_ban if user_info else None
    
    # Quyền Admin hoặc thuộc các phòng ban đặc biệt (1: Ban Giám đốc; 2: Tổng hợp; 3: Kiểm tra, giám sát; 4: Thu thập và Xử lý thông tin)
    is_admin_or_th = (current_user.is_admin == 1 or user_ma_pb in [2])

    # TỐI ƯU: Tính công chuẩn của 12 tháng một lần duy nhất
    try:
        dict_cong_chuan = {m: get_cong_chuan(year, m) for m in range(1, 13)}
    except Exception as e:
        flash(f"Lỗi khi tính công chuẩn: {e}", "danger")
        return redirect(url_for('v1_report'))

    try:
        # 2. Truy vấn dữ liệu có phân quyền
        # Nếu không phải Admin/TH, thêm điều kiện lọc theo ma_phong_ban của user
        extra_filter = ""
        params = {"year": year}
        
        if not is_admin_or_th:
            extra_filter = " AND nv.ma_phong_ban = :ma_pb"
            params["ma_pb"] = user_ma_pb

        sql = text(f"""
            SELECT 
                nv.ma_nhan_vien, nv.ho_ten, pb.ten_phong_ban,
                tl.*
            FROM thong_tin_nguoi_lao_dong nv
            LEFT JOIN phong_ban pb ON nv.ma_phong_ban = pb.id
            LEFT JOIN tien_luong_v1 tl ON nv.ma_nhan_vien = tl.ma_nhan_vien AND tl.nam = :year
            WHERE nv.trang_thai = 1 {extra_filter}
            ORDER BY pb.ten_phong_ban, nv.ma_nhan_vien
        """)
        
        results = db.session.execute(sql, params).fetchall()

        # 3. Khởi tạo Workbook và Styles
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Luong V1_{year}"

        # Định nghĩa các mẫu định dạng
        fill_header_main = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") 
        fill_month_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") 
        fill_warning = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") 
        
        font_white = Font(color="FFFFFF", bold=True)
        font_bold = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 4. Tạo Header (Hàng 1 và 2)
        headers_fixed = [("Mã NV", "A"), ("Họ và Tên", "B"), ("Phòng ban", "C")]
        for text_val, col_let in headers_fixed:
            ws.merge_cells(f'{col_let}1:{col_let}2')
            cell = ws[f'{col_let}1']
            cell.value = text_val
            cell.fill = fill_header_main
            cell.font = font_white
            cell.alignment = center_align
            cell.border = thin_border

        # Các cột tháng
        col_idx = 4
        for m in range(1, 13):
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 2)
            cell_m = ws.cell(row=1, column=col_idx, value=f"Tháng {m}")
            cell_m.fill = fill_month_header
            cell_m.font = font_bold
            cell_m.alignment = center_align
            cell_m.border = thin_border

            sub_titles = ["HSL", "Công", "Tiền V1"]
            for i, title in enumerate(sub_titles):
                c = ws.cell(row=2, column=col_idx + i, value=title)
                c.font = font_bold
                c.alignment = center_align
                c.border = thin_border
            col_idx += 3

        # Header Tổng năm
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
        cell_total = ws.cell(row=1, column=col_idx, value="Tổng Năm")
        cell_total.fill = fill_header_main
        cell_total.font = font_white
        cell_total.alignment = center_align
        cell_total.border = thin_border

        # 5. Đổ dữ liệu
        row_idx = 3
        for r in results:
            ws.cell(row=row_idx, column=1, value=r.ma_nhan_vien).border = thin_border
            ws.cell(row=row_idx, column=2, value=r.ho_ten).border = thin_border
            ws.cell(row=row_idx, column=3, value=r.ten_phong_ban or "N/A").border = thin_border
            
            data_col = 4
            for m in range(1, 13):
                hsl = float(getattr(r, f'hsl_v1_t{m}') or 0)
                cong_tt = float(getattr(r, f'cong_t{m}') or 0)
                tien = float(getattr(r, f'tien_v1_t{m}') or 0)
                cong_chuan = dict_cong_chuan[m]

                ws.cell(row=row_idx, column=data_col, value=hsl).border = thin_border
                
                c_cong = ws.cell(row=row_idx, column=data_col + 1, value=cong_tt)
                c_cong.border = thin_border
                if cong_tt < cong_chuan:
                    c_cong.fill = fill_warning
                
                c_tien = ws.cell(row=row_idx, column=data_col + 2, value=tien)
                c_tien.number_format = '#,##0'
                c_tien.border = thin_border
                
                data_col += 3
            
            c_tong = ws.cell(row=row_idx, column=data_col, value=float(r.tien_v1_tong_nam or 0))
            c_tong.number_format = '#,##0'
            c_tong.font = font_bold
            c_tong.border = thin_border
            
            row_idx += 1

        # 6. Chỉnh độ rộng cột tự động
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        for c in range(4, col_idx + 1):
            ws.column_dimensions[get_column_letter(c)].width = 11

        # 7. Xuất file
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Tên file linh hoạt theo phạm vi dữ liệu
        suffix = "Toan_Cong_Ty" if is_admin_or_th else (user_info.phong_ban.ten_phong_ban if user_info.phong_ban else "Ca_Nhan")
        
        return send_file(
            output, 
            as_attachment=True, 
            download_name=f"Luong_V1_{year}_{suffix}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        db.session.rollback()
        print(f"Excel Export Error: {traceback.format_exc()}")
        flash(f"Lỗi khi tạo file Excel: {str(e)}", "danger")
        return redirect(url_for('v1_report'))
    
# ----------------------------------------------------------------------
# Bảng Hệ số lương V2, Tiền lương V2
# ----------------------------------------------------------------------
@app.route('/bang-luong-v2')
@login_required
def bang_luong_v2():
    year = request.args.get('year', datetime.now().year, type=int)
    selected_pb = request.args.get('phong_ban', '')
    ma_nv_filter = request.args.get('ma_nv', '')

    # Lấy danh sách phòng ban để hiển thị trong bộ lọc dropdown
    ds_phong_ban = db.session.execute(text("SELECT id, ten_phong_ban FROM phong_ban ORDER BY ten_phong_ban")).fetchall()

    # JOIN với thong_tin_v2 để nhập liệu hệ số
    sql_text = """
        SELECT 
            nv.ma_nhan_vien, 
            nv.ho_ten, 
            pb.ten_phong_ban, 
            v2.*,
            u_app.fullname as ten_nguoi_duyet
        FROM thong_tin_nguoi_lao_dong nv
        JOIN thong_tin_v2 v2 ON nv.ma_nhan_vien = v2.ma_nhan_vien
        LEFT JOIN phong_ban pb ON nv.ma_phong_ban = pb.id
        LEFT JOIN users u_app ON v2.approved_by = u_app.ma_nhan_vien
        WHERE v2.nam = :year
    """
    
    params = {"year": year}
    if selected_pb:
        sql_text += " AND nv.ma_phong_ban = :pb"
        params["pb"] = selected_pb
    if ma_nv_filter:
        sql_text += " AND nv.ma_nhan_vien LIKE :ma"
        params["ma"] = f"%{ma_nv_filter}%"

    results = db.session.execute(text(sql_text + " ORDER BY pb.ten_phong_ban, nv.ma_nhan_vien"), params).fetchall()

    return render_template('bang_luong_v2.html', 
                           results=results, 
                           year=year, 
                           ds_phong_ban=ds_phong_ban,
                           selected_pb=selected_pb,
                           ma_nv_filter=ma_nv_filter)

@app.route('/approve-v2', methods=['POST'])
@login_required
def approve_v2():
    # Kiểm tra quyền hạn
    if current_user.role not in ['ADMIN', 'KIEM_SOAT']:
        return jsonify({'success': False, 'message': 'Bạn không có quyền phê duyệt!'}), 403

    data = request.get_json()
    # Ép kiểu int cho year để Procedure không bị lỗi tham số
    try:
        year = int(data.get('year'))
    except:
        return jsonify({'success': False, 'message': 'Năm không hợp lệ!'}), 400
    
    selected_pb = data.get('phong_ban')
    ma_nv_filter = data.get('ma_nv')
    updates = data.get('updates', [])

    try:
        # 1. Lưu cập nhật hệ số V2 cho những dòng chưa phê duyệt
        if updates:
            for item in updates:
                # Tạo câu lệnh SET động t1=:t1, t2=:t2...
                set_clause = ", ".join([f"t{m} = :t{m}" for m in range(1, 13)])
                save_sql = f"""
                    UPDATE thong_tin_v2 
                    SET {set_clause} 
                    WHERE ma_nhan_vien = :ma_nv AND nam = :year
                    AND (approved_by IS NULL OR approved_by = '' OR approved_by = 'None')
                """
                params_save = {f"t{m}": item.get(f't{m}', 0) for m in range(1, 13)}
                params_save.update({'ma_nv': item.get('ma_nv'), 'year': year})
                db.session.execute(text(save_sql), params_save)

        # 2. Thực hiện Phê duyệt (Đánh dấu người duyệt)
        # Lưu ý: Sử dụng HANOI_TZ để đồng nhất thời gian duyệt với V1
        now_hanoi = datetime.now(pytz.timezone('Asia/Ho_Chi_Minh'))
        approve_query = """
            UPDATE thong_tin_v2 v2 
            JOIN thong_tin_nguoi_lao_dong nv ON v2.ma_nhan_vien = nv.ma_nhan_vien 
            SET v2.approved_by = :user, v2.approved_at = :now 
            WHERE v2.nam = :year
            AND (v2.approved_by IS NULL OR v2.approved_by = '' OR v2.approved_by = 'None')
        """
        params_approve = {"user": current_user.ma_nhan_vien, "now": now_hanoi, "year": year}

        if selected_pb and str(selected_pb).strip(): #if selected_pb and selected_pb != "":
            approve_query += " AND nv.ma_phong_ban = :pb"
            params_approve["pb"] = selected_pb
        if ma_nv_filter and str(ma_nv_filter).strip(): #if ma_nv_filter and ma_nv_filter != "":
            approve_query += " AND v2.ma_nhan_vien = :ma"
            params_approve["ma"] = ma_nv_filter

        result = db.session.execute(text(approve_query), params_approve)
        count_approved = result.rowcount
        
        # 3. Chỉ gọi Procedure nếu thực sự có người được duyệt mới hoặc có updates
        if count_approved > 0 or updates:
            # Lưu ý: Cần quyền EXECUTE cho user database
            db.session.execute(text("CALL sp_ChayTinhLuongV2_ToanCoQuan(:year)"), {"year": year})
        
        db.session.commit()
        return jsonify({
            'success': True, 
            'message': f'Đã phê duyệt và tính toán lương thành công cho {count_approved} nhân viên!'
        })
    except Exception as e:
        db.session.rollback()
        # In chi tiết lỗi ra console để debug nếu cần
        print(f"Approve V2 Error: {str(e)}")
        return jsonify({'success': False, 'message': f"Lỗi: {str(e)}"}), 500

@app.route('/unapprove-v2', methods=['POST'])
@login_required
def unapprove_v2():
    # Kiểm tra quyền Admin
    if current_user.role != 'ADMIN':
        return jsonify({'success': False, 'message': 'Chỉ Admin mới có quyền gỡ phê duyệt!'}), 403

    data = request.get_json()
    year = data.get('year')
    ma_nv = data.get('ma_nv')

    if not year or not ma_nv:
        return jsonify({'success': False, 'message': 'Thiếu thông tin năm hoặc mã nhân viên!'}), 400

    try:
        # Cập nhật bảng dữ liệu V2 (thong_tin_v2)
        query = """
            UPDATE thong_tin_v2 
            SET approved_by = NULL, approved_at = NULL 
            WHERE nam = :year AND ma_nhan_vien = :ma
        """
        db.session.execute(text(query), {"year": year, "ma": ma_nv})
        db.session.commit()
        
        # Ghi log sự kiện
        log_event("Gỡ phê duyệt V2", f"Admin {current_user.ma_nhan_vien} đã gỡ phê duyệt cho NV {ma_nv} năm {year}")
        
        return jsonify({'success': True, 'message': 'Đã gỡ trạng thái phê duyệt bảng lương V2!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/v2-report')
@login_required
def v2_report(): # Tên hàm này phải viết đúng là v2_report
    # 1. LẤY THÔNG TIN QUYỀN CỦA USER
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    user_ma_pb = user_info.ma_phong_ban if user_info else None
    
    # Lấy tên phòng ban từ relationship (Đã thêm db.relationship vào Model)
    user_ten_pb = user_info.phong_ban.ten_phong_ban if user_info and user_info.phong_ban else ""

    # Quyền Admin hoặc thuộc các phòng ban đặc biệt (1: Ban Giám đốc; 2: Tổng hợp; 3: Kiểm tra, giám sát; 4: Thu thập và Xử lý thông tin)
    is_admin_or_th = (current_user.is_admin == 1 or user_ma_pb in [2])

    # 2. XỬ LÝ THAM SỐ LỌC
    year = request.args.get('year', datetime.now().year, type=int)
    ma_nv_filter = request.args.get('ma_nv', '')

    # LOGIC PHÂN QUYỀN LỌC PHÒNG BAN:
    if is_admin_or_th:
        selected_pb = request.args.get('phong_ban', '')
    else:
        selected_pb = user_ten_pb

    # 3. TỐI ƯU: Tính công chuẩn 12 tháng một lần duy nhất
    cong_chuan_12_thang = {m: get_cong_chuan(year, m) for m in range(1, 13)}

    # 4. DANH SÁCH PHÒNG BAN (Chỉ dành cho Admin/TH chọn)
    ds_phong_ban = []
    if is_admin_or_th:
        ds_pb_rows = db.session.execute(text("SELECT ten_phong_ban FROM phong_ban ORDER BY ten_phong_ban")).fetchall()
        ds_phong_ban = [r[0] for r in ds_pb_rows]

    # 5. TRUY VẤN DỮ LIỆU LƯƠNG V2
    sql = """
        SELECT 
            nv.ma_nhan_vien, nv.ho_ten, pb.ten_phong_ban,
            tl.*
        FROM thong_tin_nguoi_lao_dong nv
        LEFT JOIN phong_ban pb ON nv.ma_phong_ban = pb.id
        LEFT JOIN tien_luong_v2 tl ON nv.ma_nhan_vien = tl.ma_nhan_vien AND tl.nam = :year
        WHERE nv.trang_thai = 1
    """
    params = {"year": year}
    
    if selected_pb:
        sql += " AND pb.ten_phong_ban = :pb"
        params["pb"] = selected_pb
    elif not is_admin_or_th:
        # Nếu không có quyền và không có tên phòng, ép kết quả rỗng
        sql += " AND 1=0"

    if ma_nv_filter:
        sql += " AND nv.ma_nhan_vien LIKE :ma"
        params["ma"] = f"%{ma_nv_filter}%"

    # Thực thi truy vấn
    results = db.session.execute(text(sql + " ORDER BY pb.ten_phong_ban, nv.ma_nhan_vien"), params).fetchall()

    # 6. CẤU TRÚC DỮ LIỆU
    report_data = []
    for r in results:
        emp_months = []
        for m in range(1, 13):
            emp_months.append({
                'hsl': float(getattr(r, f'hsl_v2_t{m}') or 0),
                'cong_tt': float(getattr(r, f'cong_t{m}') or 0),
                'tien': float(getattr(r, f'tien_v2_t{m}') or 0),
                'cong_chuan': cong_chuan_12_thang[m]
            })
        report_data.append({
            'ma_nv': r.ma_nhan_vien,
            'ho_ten': r.ho_ten,
            'phong_ban': r.ten_phong_ban or "N/A",
            'months': emp_months,
            'tong_nam': float(r.tien_v2_tong_nam or 0)
        })

    return render_template('v2_report.html', 
                           data=report_data, 
                           year=year, 
                           ds_phong_ban=ds_phong_ban, 
                           selected_pb=selected_pb, 
                           ma_nv_filter=ma_nv_filter,
                           is_admin_or_th=is_admin_or_th)    

@app.route('/export-v2-excel/<int:year>')
@login_required
def export_v2_excel(year):
    # 1. KIỂM TRA QUYỀN TRUY CẬP
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    user_ma_pb = user_info.ma_phong_ban if user_info else None
    
    # Quyền Admin hoặc thuộc các phòng ban đặc biệt (1: Ban Giám đốc; 2: Tổng hợp; 3: Kiểm tra, giám sát; 4: Thu thập và Xử lý thông tin)
    is_admin_or_th = (current_user.is_admin == 1 or user_ma_pb in [2])

    # TỐI ƯU: Tính công chuẩn 12 tháng một lần duy nhất
    try:
        dict_cong_chuan = {m: get_cong_chuan(year, m) for m in range(1, 13)}
    except Exception as e:
        flash(f"Lỗi tính toán lịch: {e}", "danger")
        return redirect(url_for('v2_report'))

    try:
        # 2. Xây dựng truy vấn dữ liệu có phân quyền
        extra_filter = ""
        params = {"year": year}
        
        # Nếu không có quyền quản trị, cưỡng ép lọc theo phòng ban của user
        if not is_admin_or_th:
            extra_filter = " AND nv.ma_phong_ban = :ma_pb"
            params["ma_pb"] = user_ma_pb

        sql = text(f"""
            SELECT 
                nv.ma_nhan_vien, nv.ho_ten, pb.ten_phong_ban,
                tl.*
            FROM thong_tin_nguoi_lao_dong nv
            LEFT JOIN phong_ban pb ON nv.ma_phong_ban = pb.id
            LEFT JOIN tien_luong_v2 tl ON nv.ma_nhan_vien = tl.ma_nhan_vien AND tl.nam = :year
            WHERE nv.trang_thai = 1 {extra_filter}
            ORDER BY pb.ten_phong_ban, nv.ma_nhan_vien
        """)
        results = db.session.execute(sql, params).fetchall()

        # 3. Khởi tạo Workbook và Định dạng (Styles)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Luong V2_{year}"

        # Định nghĩa màu sắc và kiểu chữ
        fill_main_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Xanh đậm
        fill_sub_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Xanh nhạt
        fill_warning = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # Hồng cảnh báo
        
        font_white_bold = Font(color="FFFFFF", bold=True)
        font_bold = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 4. Xây dựng Header (Hàng 1 & 2)
        fixed_cols = [("Mã NV", "A"), ("Họ và Tên", "B"), ("Phòng ban", "C")]
        for label, col_letter in fixed_cols:
            ws.merge_cells(f'{col_letter}1:{col_letter}2')
            cell = ws[f'{col_letter}1']
            cell.value = label
            cell.fill = fill_main_header
            cell.font = font_white_bold
            cell.alignment = center_align
            cell.border = thin_border

        current_col = 4
        for m in range(1, 13):
            # Header Tháng (Merge 3 cột)
            ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + 2)
            cell_m = ws.cell(row=1, column=current_col, value=f"Tháng {m}")
            cell_m.fill = fill_sub_header
            cell_m.font = font_bold
            cell_m.alignment = center_align
            cell_m.border = thin_border

            # Sub-header: HSL, Công, Tiền
            sub_labels = ["HSL V2", "Công", "Tiền V2"]
            for i, label in enumerate(sub_labels):
                c = ws.cell(row=2, column=current_col + i, value=label)
                c.font = font_bold
                c.alignment = center_align
                c.border = thin_border
            current_col += 3

        # Header Tổng năm
        ws.merge_cells(start_row=1, start_column=current_col, end_row=2, end_column=current_col)
        cell_total = ws.cell(row=1, column=current_col, value="Tổng Năm V2")
        cell_total.fill = fill_main_header
        cell_total.font = font_white_bold
        cell_total.alignment = center_align
        cell_total.border = thin_border

        # 5. Đổ dữ liệu nhân viên
        row_idx = 3
        for r in results:
            ws.cell(row=row_idx, column=1, value=r.ma_nhan_vien).border = thin_border
            ws.cell(row=row_idx, column=2, value=r.ho_ten).border = thin_border
            ws.cell(row=row_idx, column=3, value=r.ten_phong_ban or "N/A").border = thin_border
            
            data_col = 4
            for m in range(1, 13):
                hsl = float(getattr(r, f'hsl_v2_t{m}') or 0)
                cong_tt = float(getattr(r, f'cong_t{m}') or 0)
                tien = float(getattr(r, f'tien_v2_t{m}') or 0)
                cong_chuan = dict_cong_chuan[m]

                ws.cell(row=row_idx, column=data_col, value=hsl).border = thin_border
                
                cell_cong = ws.cell(row=row_idx, column=data_col + 1, value=cong_tt)
                cell_cong.border = thin_border
                if cong_tt < cong_chuan:
                    cell_cong.fill = fill_warning
                
                cell_tien = ws.cell(row=row_idx, column=data_col + 2, value=tien)
                cell_tien.number_format = '#,##0'
                cell_tien.border = thin_border
                
                data_col += 3
            
            cell_sum = ws.cell(row=row_idx, column=data_col, value=float(r.tien_v2_tong_nam or 0))
            cell_sum.number_format = '#,##0'
            cell_sum.font = font_bold
            cell_sum.border = thin_border
            
            row_idx += 1

        # 6. Điều chỉnh độ rộng cột
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        for c in range(4, current_col + 1):
            ws.column_dimensions[get_column_letter(c)].width = 11

        # 7. Xuất file trả về client
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Đặt tên file theo ngữ cảnh
        dept_name = user_info.phong_ban.ten_phong_ban if user_info and user_info.phong_ban else "Export"
        file_suffix = "Toan_Cong_Ty" if is_admin_or_th else dept_name.replace(" ", "_")
        
        return send_file(
            output, 
            as_attachment=True, 
            download_name=f"Bao_Cao_Luong_V2_{year}_{file_suffix}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"Lỗi Export V2: {traceback.format_exc()}")
        flash(f"Lỗi hệ thống khi xuất file: {str(e)}", "danger")
        return redirect(url_for('v2_report'))

# ----------------------------------------------------------------------
# Đổi mật khẩu người dùng
# ----------------------------------------------------------------------
@app.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        old_password = request.form.get('old_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')

        # 1. Kiểm tra mật khẩu cũ
        if not current_user.check_password(old_password):
            flash('Mật khẩu cũ không chính xác.', 'danger')
            return redirect(url_for('change_password'))

        # 2. Kiểm tra mật khẩu mới trùng khớp
        if new_password != confirm_password:
            flash('Mật khẩu mới và xác nhận mật khẩu không khớp.', 'danger')
            return redirect(url_for('change_password'))
        
        # 3. Độ dài mật khẩu (tùy chọn)
        if len(new_password) < 6:
            flash('Mật khẩu mới phải có ít nhất 6 ký tự.', 'danger')
            return redirect(url_for('change_password'))

        # 4. Cập nhật mật khẩu
        try:
            current_user.set_password(new_password)
            current_user.force_password_change = False  # Đã đổi mật khẩu thành công
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash('Có lỗi xảy ra khi lưu vào cơ sở dữ liệu.', 'danger')
            return redirect(url_for('change_password'))
        
        flash('Đổi mật khẩu thành công!', 'success')
        return redirect(url_for('index'))

    return render_template('change_password.html')

    
@app.route('/admin/reset_password/<ma_nv>', methods=['POST'])
@login_required
def reset_password_admin(ma_nv):
    # 1. Kiểm tra quyền Admin (Dựa trên role bạn đã định nghĩa)
    if not current_user.role or current_user.role.upper() != 'ADMIN':
        return jsonify({'success': False, 'message': 'Bạn không có quyền thực hiện hành động này.'}), 403

    # 2. Tìm user cần reset
    user = User.query.filter_by(ma_nhan_vien=ma_nv).first()
    if not user:
        return jsonify({'success': False, 'message': 'Không tìm thấy người dùng.'}), 404

    try:
        # 3. Đặt mật khẩu về mặc định (Mã nhân viên)
        user.set_password(ma_nv)
        user.force_password_change = True  # Bắt buộc đổi ở lần đăng nhập tới
        db.session.commit()
        return jsonify({'success': True, 'message': f'Đã reset mật khẩu cho NV {ma_nv} về mặc định. Tài khoản này sẽ bắt buộc phải đổi mật khẩu ở lần đăng nhập tiếp theo.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# ----------------------------------------------------------------------
# Quản trị user, thông tin người lao động
# ----------------------------------------------------------------------
@app.route('/admin/import_nhan_vien', methods=['POST'])
@login_required
def import_nhan_vien():
    # 1. Kiểm tra quyền Admin
    if not current_user.role or current_user.role.upper() != 'ADMIN':
        flash('Bạn không có quyền thực hiện thao tác này.', 'danger')
        return redirect(url_for('danh_sach_nhan_vien'))

    file = request.files.get('file')
    if not file or file.filename == '':
        flash('Vui lòng chọn file Excel!', 'warning')
        return redirect(url_for('danh_sach_nhan_vien'))

    try:
        # Đọc dữ liệu từ Excel
        df = pd.read_excel(file)
        
        # Làm sạch tên cột: xóa khoảng trắng thừa
        df.columns = [str(c).strip() for c in df.columns]

        # QUAN TRỌNG: Thay thế toàn bộ nan (của numpy) bằng None (để MySQL hiểu là NULL)
        # np lúc này đã được định nghĩa ở bước 1
        df = df.replace({np.nan: None, pd.NaT: None})
        
        success_count = 0
        update_count = 0

        for _, row in df.iterrows():
            # Lấy mã nhân viên và kiểm tra hợp lệ
            ma_nv_raw = row.get('ma_nhan_vien')
            if ma_nv_raw is None: continue
            
            ma_nv = str(ma_nv_raw).strip()
            if not ma_nv or ma_nv.lower() == 'none': continue
                
            ho_ten = str(row.get('ho_ten', '')).strip()

            # Hàm xử lý ngày tháng an toàn để tránh lỗi Timestamp/NaT
            def clean_date(val):
                if val is None: return None
                if isinstance(val, datetime): return val
                try:
                    res = pd.to_datetime(val)
                    return res.to_pydatetime() if not pd.isna(res) else None
                except:
                    return None

            # Hàm làm sạch số điện thoại (tránh bị 913368536.0)
            def clean_phone(val):
                if val is None: return None
                return str(val).split('.')[0].strip()

            params_nv = {
                'ma': ma_nv,
                'ten': ho_ten,
                'ns': clean_date(row.get('ngay_sinh')),
                'gt': row.get('gioi_tinh'),
                'dt': clean_phone(row.get('so_dien_thoai')),
                'mail': row.get('mail_Agribank'),
                'dc': row.get('dia_chi'),
                'np_phep': clean_date(row.get('ngay_tinh_phep')), # Đổi tên tránh trùng np của numpy
                'nv_agri': clean_date(row.get('ngay_vao_Agribank')),
                'pb': row.get('ma_phong_ban'),
                'dv': row.get('ma_hieu_2'),
                'cv': row.get('chuc_vu')
            }

            # Kiểm tra tồn tại để quyết định INSERT hay UPDATE
            check_nv = db.session.execute(
                text("SELECT id FROM thong_tin_nguoi_lao_dong WHERE ma_nhan_vien = :ma"),
                {'ma': ma_nv}
            ).fetchone()

            if not check_nv:
                # INSERT
                db.session.execute(text("""
                    INSERT INTO thong_tin_nguoi_lao_dong 
                    (ma_nhan_vien, ho_ten, ngay_sinh, gioi_tinh, so_dien_thoai, mail_Agribank, 
                     dia_chi, ngay_tinh_phep, ngay_vao_Agribank, ma_phong_ban, ma_hieu_2, chuc_vu)
                    VALUES (:ma, :ten, :ns, :gt, :dt, :mail, :dc, :np_phep, :nv_agri, :pb, :dv, :cv)
                """), params_nv)
                success_count += 1
            else:
                # UPDATE
                db.session.execute(text("""
                    UPDATE thong_tin_nguoi_lao_dong SET 
                    ho_ten=:ten, ngay_sinh=:ns, gioi_tinh=:gt, so_dien_thoai=:dt, mail_Agribank=:mail, 
                    dia_chi=:dc, ngay_tinh_phep=:np_phep, ngay_vao_Agribank=:nv_agri, 
                    ma_phong_ban=:pb, ma_hieu_2=:dv, chuc_vu=:cv
                    WHERE ma_nhan_vien = :ma
                """), params_nv)
                update_count += 1

            # Đồng bộ tài khoản người dùng
            check_user = db.session.execute(
                text("SELECT ma_nhan_vien FROM users WHERE ma_nhan_vien = :ma"),
                {'ma': ma_nv}
            ).fetchone()

            if not check_user:
                hashed_pw = bcrypt.generate_password_hash(ma_nv).decode('utf-8')
                db.session.execute(text("""
                    INSERT INTO users (ma_nhan_vien, password_hash, fullname, role, is_active)
                    VALUES (:ma, :pw, :ten, 'LAP_BANG', 1)
                """), {'ma': ma_nv, 'pw': hashed_pw, 'ten': ho_ten})

        db.session.commit()
        flash(f'Thành công: Thêm mới {success_count}, Cập nhật {update_count}.', 'success')

    except Exception as e:
        db.session.rollback()
        print(f"Import Error Detail: {str(e)}")
        flash(f'Lỗi xử lý file: {str(e)}', 'danger')

    return redirect(url_for('danh_sach_nhan_vien'))


@app.route('/admin/nhan_vien')
@login_required
def danh_sach_nhan_vien():
    # 1. Kiểm tra quyền Admin
    if not current_user.role or current_user.role.upper() != 'ADMIN':
        flash('Bạn không có quyền truy cập trang này.', 'danger')
        return redirect(url_for('index'))

    # 2. Lấy tham số lọc và phân trang
    search_query = request.args.get('search', '').strip()
    selected_dept = request.args.get('phong_ban', '').strip()
    selected_don_vi = request.args.get('don_vi', '').strip() # Thêm tiêu chí đơn vị
    page = request.args.get('page', 1, type=int)
    per_page = 12  
    offset = (page - 1) * per_page

    # 3. Lấy danh sách phòng ban cho dropdown
    depts_query = db.session.execute(text("SELECT id, ten_phong_ban FROM phong_ban ORDER BY ten_phong_ban")).fetchall()
    departments = [dict(row._mapping) for row in depts_query]

    # Lấy thông tin đơn vị đang chọn (để hiển thị tên trên Select2 khi load trang)
    don_vi_info = None
    if selected_don_vi:
        dv_row = db.session.execute(
            text("SELECT ma_hieu_2, ten_ma_hieu_2 FROM don_vi WHERE ma_hieu_2 = :ma"),
            {'ma': selected_don_vi}
        ).fetchone()
        if dv_row:
            don_vi_info = dict(dv_row._mapping)

    # 4. Xây dựng SQL
    base_sql = "FROM thong_tin_nguoi_lao_dong e LEFT JOIN phong_ban p ON e.ma_phong_ban = p.id"
    where_clauses = []
    params = {}

    if search_query:
        where_clauses.append("(e.ma_nhan_vien LIKE :search OR e.ho_ten LIKE :search OR e.so_dien_thoai LIKE :search)")
        params['search'] = f"%{search_query}%"
    
    if selected_dept:
        where_clauses.append("e.ma_phong_ban = :dept_id")
        params['dept_id'] = selected_dept

    if selected_don_vi:
        where_clauses.append("e.ma_hieu_2 = :dv_id")
        params['dv_id'] = selected_don_vi

    where_clause = " WHERE " + " AND ".join(where_clauses) if where_clauses else ""

    # 5. Tính toán phân trang
    count_sql = text(f"SELECT COUNT(*) {base_sql} {where_clause}")
    total_records = db.session.execute(count_sql, params).scalar() or 0
    total_pages = math.ceil(total_records / per_page) if total_records > 0 else 1

    # 6. Truy vấn dữ liệu chi tiết
    data_sql = text(f"""
        SELECT e.ma_nhan_vien, e.ho_ten, e.ngay_sinh, e.gioi_tinh, e.so_gttt, 
               e.so_dien_thoai, e.mail_Agribank, e.ma_hieu_2, e.ngay_tinh_phep, 
               e.ngay_vao_Agribank, p.ten_phong_ban
        {base_sql} {where_clause}
        ORDER BY e.ho_ten COLLATE utf8mb4_vietnamese_ci ASC
        LIMIT :limit OFFSET :offset
    """)
    params.update({'limit': per_page, 'offset': offset})
    
    result = db.session.execute(data_sql, params).fetchall()
    employees = [dict(row._mapping) for row in result]

    return render_template('admin_nhan_vien.html', 
                           employees=employees, 
                           page=page, 
                           total_pages=total_pages, 
                           total_records=total_records,
                           search_query=search_query,
                           departments=departments,     
                           selected_dept=selected_dept,
                           selected_don_vi=selected_don_vi,
                           don_vi_info=don_vi_info) # Truyền thông tin đơn vị để Select2 hiển thị

@app.route('/admin/nhan_vien/edit/<string:ma_nv>', methods=['GET', 'POST'])
@login_required
def edit_nhan_vien(ma_nv):
    # 1. Kiểm tra quyền Admin
    if not current_user.role or current_user.role.upper() != 'ADMIN':
        flash('Bạn không có quyền thực hiện thao tác này.', 'danger')
        return redirect(url_for('index'))

    if request.method == 'POST':
        # Xử lý trạng thái Checkbox từ Form
        trang_thai_cham_cong = 1 if request.form.get('trang_thai') == 'on' else 0
        is_active_user = 1 if request.form.get('is_active') == 'on' else 0
        
        # Lấy thông tin đơn vị và chức vụ để xác định Role
        ma_hieu_2_moi = request.form.get('ma_hieu_2')
        chuc_vu_moi = request.form.get('chuc_vu')
        ma_phong_ban_moi = request.form.get('ma_phong_ban') or None

        # --- LOGIC TỰ ĐỘNG CẬP NHẬT ROLE KHI LUÂN CHUYỂN ---
        # Danh sách mã đơn vị quản lý hoặc điều kiện chức vụ trưởng/phó
        ADMIN_UNITS = ['100', 'ORG_ADMIN'] 
        new_role = 'USER'
        
        if ma_hieu_2_moi in ADMIN_UNITS:
            new_role = 'QUAN_LY'
        elif chuc_vu_moi and any(keyword in chuc_vu_moi for keyword in ['Trưởng', 'Phó', 'Giám đốc']):
            new_role = 'QUAN_LY'

        data_update = {
            'ho_ten': request.form.get('ho_ten'),
            'ngay_sinh': request.form.get('ngay_sinh') or None,
            'gioi_tinh': request.form.get('gioi_tinh'),
            'so_gttt': request.form.get('so_gttt'),
            'so_dien_thoai': request.form.get('so_dien_thoai'),
            'mail_Agribank': request.form.get('mail_Agribank'),
            'dia_chi': request.form.get('dia_chi'),
            'ngay_tinh_phep': request.form.get('ngay_tinh_phep') or None,
            'ngay_vao_Agribank': request.form.get('ngay_vao_Agribank') or None,
            'ma_phong_ban': ma_phong_ban_moi,
            'ma_hieu_2': ma_hieu_2_moi,
            'chuc_vu': chuc_vu_moi,
            'trang_thai': trang_thai_cham_cong,
            'ma_nv': ma_nv
        }

        try:
            # 2. Cập nhật bảng thông tin nhân sự (Bao gồm Đơn vị và Phòng ban mới)
            sql_nv = text("""
                UPDATE thong_tin_nguoi_lao_dong 
                SET ho_ten = :ho_ten, ngay_sinh = :ngay_sinh, gioi_tinh = :gioi_tinh, 
                    so_gttt = :so_gttt, so_dien_thoai = :so_dien_thoai, mail_Agribank = :mail_Agribank, 
                    dia_chi = :dia_chi, ngay_tinh_phep = :ngay_tinh_phep, 
                    ngay_vao_Agribank = :ngay_vao_Agribank, ma_phong_ban = :ma_phong_ban,
                    ma_hieu_2 = :ma_hieu_2, chuc_vu = :chuc_vu, trang_thai = :trang_thai
                WHERE ma_nhan_vien = :ma_nv
            """)
            db.session.execute(sql_nv, data_update)

            # 3. Cập nhật bảng users (Cập nhật trạng thái và Role tự động)
            sql_user = text("""
                UPDATE users 
                SET is_active = :is_active, 
                    role = :role 
                WHERE ma_nhan_vien = :ma_nv
            """)
            db.session.execute(sql_user, {
                'is_active': is_active_user, 
                'role': new_role, 
                'ma_nv': ma_nv
            })

            db.session.commit()
            flash(f'Luân chuyển và cập nhật nhân viên {ma_nv} thành công!', 'success')
            return redirect(url_for('danh_sach_nhan_vien'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Lỗi khi cập nhật dữ liệu: {str(e)}', 'danger')

    # --- PHẦN GET: LẤY DỮ LIỆU HIỂN THỊ ---
    
    # Lấy thông tin nhân viên, trạng thái user và tên đơn vị hiện tại
    query_data = text("""
        SELECT nv.*, u.is_active, u.role, dv.ten_ma_hieu_2 as ten_don_vi
        FROM thong_tin_nguoi_lao_dong nv
        LEFT JOIN users u ON nv.ma_nhan_vien = u.ma_nhan_vien
        LEFT JOIN don_vi dv ON nv.ma_hieu_2 = dv.ma_hieu_2
        WHERE nv.ma_nhan_vien = :ma_nv
    """)
    result = db.session.execute(query_data, {'ma_nv': ma_nv}).fetchone()
    
    if not result:
        flash('Nhân viên không tồn tại trong hệ thống!', 'warning')
        return redirect(url_for('danh_sach_nhan_vien'))

    nv = dict(result._mapping)
    
    # Định dạng ngày tháng cho thẻ <input type="date">
    for field in ['ngay_sinh', 'ngay_tinh_phep', 'ngay_vao_Agribank']:
        if nv.get(field) and hasattr(nv[field], 'strftime'):
            nv[field] = nv[field].strftime('%Y-%m-%d')

    # Lấy danh sách PHÒNG BAN hiện có của Đơn vị này để hiển thị Dropdown ban đầu
    depts = []
    if nv.get('ma_hieu_2'):
        depts = db.session.execute(
            text("SELECT id, ten_phong_ban FROM phong_ban WHERE ma_hieu_2 = :ma_dv ORDER BY ten_phong_ban"),
            {'ma_dv': nv['ma_hieu_2']}
        ).fetchall()
    
    return render_template('admin_edit_nhan_vien.html', nv=nv, departments=depts)

# ----------------------------------------------------------------------
# Xuất Excel danh sách Users người dùng
# ----------------------------------------------------------------------
@app.route('/admin/export_nhan_vien')
@login_required
@admin_required
def export_nhan_vien():
    # 1. Kiểm tra quyền Admin
    if not current_user.role or current_user.role.upper() != 'ADMIN':
        return "Unauthorized", 403

    # 2. Lấy tham số tìm kiếm từ URL
    search_query = request.args.get('search', '').strip()
    ma_hieu_2 = request.args.get('don_vi', '').strip()
    ma_phong_ban = request.args.get('phong_ban', '').strip()
    
    # 3. Truy vấn dữ liệu với đầy đủ các trường và Join để lấy tên phòng ban
    # Sử dụng LEFT JOIN để không mất dữ liệu nếu nhân viên chưa xếp phòng ban
    query_str = """
        SELECT 
            e.ma_nhan_vien, e.ho_ten, e.ngay_sinh, e.gioi_tinh, e.so_gttt, 
            e.so_dien_thoai, e.mail_Agribank, e.dia_chi, e.ngay_tinh_phep, 
            e.ngay_vao_Agribank, e.ma_hieu_2, p.ten_phong_ban, e.chuc_vu,
            CASE WHEN e.trang_thai = 1 THEN 'Hoạt động' ELSE 'Khóa' END as trang_thai_text,
            e.created_at
        FROM thong_tin_nguoi_lao_dong e
        LEFT JOIN phong_ban p ON e.ma_phong_ban = p.id
        WHERE 1=1
    """
    params = {}
    if search_query:
        query_str += " AND (e.ma_nhan_vien LIKE :search OR e.ho_ten LIKE :search)"
        params['search'] = f"%{search_query}%"

    if ma_hieu_2:
        query_str += " AND e.ma_hieu_2 = :ma_hieu_2"
        params['ma_hieu_2'] = ma_hieu_2

    if ma_phong_ban:
        query_str += " AND e.ma_phong_ban = :ma_phong_ban"
        params['ma_phong_ban'] = ma_phong_ban
    
    employees = db.session.execute(text(query_str), params).fetchall()

    # 4. Tạo file Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh_sach_nhan_vien"

    # Header đầy đủ tương ứng với các trường trong SQL
    headers = [
        'Mã NV', 'Họ tên', 'Ngày sinh', 'Giới tính', 'Số GTTT', 
        'Điện thoại', 'Email Agribank', 'Địa chỉ', 'Ngày tính phép', 
        'Ngày vào Agribank', 'Mã Đơn vị (Mã hiệu 2)', 'Phòng ban', 'Chức vụ', 
        'Trạng thái', 'Ngày tạo hệ thống'
    ]
    ws.append(headers)

    # Định dạng Header cho đẹp
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # 1. Định nghĩa kiểu đường viền (mảnh, màu xám nhẹ)
    thin_border = Border(
        left=Side(style='thin', color="B2B2B2"),
        right=Side(style='thin', color="B2B2B2"),
        top=Side(style='thin', color="B2B2B2"),
        bottom=Side(style='thin', color="B2B2B2")
    )
    
    # Đổ dữ liệu và định dạng
    for row_idx, emp in enumerate(employees, start=2): # Bắt đầu từ dòng 2 vì dòng 1 là Header
        for col_idx, value in enumerate(emp, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Áp dụng đường viền cho mọi ô dữ liệu
            cell.border = thin_border
            # Căn lề mặc định cho dữ liệu là ở giữa theo chiều dọc
            cell.alignment = Alignment(vertical='center')
            
            # Kiểm tra nếu dữ liệu là kiểu ngày tháng (date hoặc datetime)
            if isinstance(value, (date, datetime)):
                cell.number_format = 'DD/MM/YYYY'
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Căn giữa cho Mã hiệu 2 và Trạng thái
            if col_idx in [1, 11, 14]: # Cột Mã NV, Mã hiệu 2, Trạng thái
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # 3. Áp dụng Border cho cả hàng Header (vì Header đã có Fill màu nên cần Border riêng)
    for cell in ws[1]:
        cell.border = thin_border
    
    # Tự động chỉnh độ rộng cột (Cải tiến để chính xác hơn với font chữ)
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    # Tính toán độ dài dựa trên chuỗi đã định dạng
                    val_str = cell.value.strftime('%d/%m/%Y') if isinstance(cell.value, (date, datetime)) else str(cell.value)
                    max_length = max(max_length, len(val_str))
            except: pass
        ws.column_dimensions[column_letter].width = min(max_length + 4, 50)

    # Trả file về trình duyệt
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"DS_NhanVien_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        output, 
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, 
        download_name=filename
    )

# ----------------------------------------------------------------------
# SAO LƯU DỮ LIỆU mySQL
# ----------------------------------------------------------------------
@app.route('/admin/full-backup')
@login_required
def full_database_backup():
    # 1. Kiểm tra quyền Admin
    if not current_user.is_authenticated or not current_user.is_admin:
        flash("Bạn không có quyền thực hiện thao tác này!", "danger")
        return redirect(url_for('index'))

    try:
        # 2. Tìm đường dẫn thực thi của mysqldump trên macOS
        possible_paths = [
            'mysqldump',
            '/usr/local/mysql/bin/mysqldump',
            '/usr/local/bin/mysqldump',
            '/usr/local/opt/mysql-client/bin/mysqldump',
            '/Applications/XAMPP/xamppfiles/bin/mysqldump',
            '/Applications/MAMP/Library/bin/mysqldump'
        ]

        mysqldump_path = None
        for path in possible_paths:
            if os.path.exists(path) or subprocess.run(['which', path], capture_output=True).returncode == 0:
                mysqldump_path = path
                break

        if not mysqldump_path:
            raise Exception("Không tìm thấy lệnh 'mysqldump'. Vui lòng cài đặt MySQL Client.")

        # 3. Lấy cấu hình từ file .env (thông qua app.config)
        db_uri = current_app.config.get('SQLALCHEMY_DATABASE_URI')
        if not db_uri:
            raise Exception("Không tìm thấy DATABASE_URL trong file .env")

        # Tách thông tin từ chuỗi kết nối
        parsed_uri = urlparse(db_uri.replace('mysql+pymysql://', 'http://'))
        db_user = parsed_uri.username
        db_pass = parsed_uri.password
        db_host = parsed_uri.hostname
        db_name = parsed_uri.path.lstrip('/')

        # 4. Thiết lập tên file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        sql_filename = f"full_dump_{db_name}_{timestamp}.sql"
        zip_filename = f"Full_Backup_{db_name}_{timestamp}.zip"

        # 5. Xây dựng lệnh (Sửa đổi phần password)
        dump_cmd = [
            mysqldump_path,
            f"--user={db_user}",
            f"--password={db_pass}", # Truyền trực tiếp vào đây lần nữa
            f"--host={db_host or 'localhost'}",
            "--databases", db_name,
            "--routines",
            "--triggers",
            "--single-transaction", # Quan trọng: Giúp tránh lỗi LOCK TABLES nếu user thiếu quyền
            "--default-character-set=utf8mb4"
        ]

        # 6. Truyền mật khẩu qua biến môi trường của tiến trình con
        env_copy = os.environ.copy()
        env_copy['MYSQL_PWD'] = db_pass  # Đây là chìa khóa để sửa lỗi Access Denied

        # 7. Thực thi lệnh
        result = subprocess.run(
            dump_cmd, 
            capture_output=True, 
            text=True, 
            encoding='utf-8', 
            errors='ignore',
            env=env_copy # Kết hợp cả hai
        )

        # Nếu có lỗi từ mysqldump
        if result.returncode != 0:
            raise Exception(f"Lỗi MySQL Dump: {result.stderr}")

        # 8. Nén ZIP và gửi file cho người dùng
        memory_file = io.BytesIO()
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(sql_filename, result.stdout)
        
        memory_file.seek(0)

        return send_file(
            memory_file,
            mimetype="application/zip",
            as_attachment=True,
            download_name=zip_filename
        )

    except Exception as e:
        print(f"Lỗi hệ thống: {str(e)}")
        flash(f"Lỗi khi sao lưu hệ thống: {str(e)}", "danger")
        return redirect(url_for('index'))

# ----------------------------------------------------------------------
# Route xử lý tính toán lương V1/V2 toàn cơ quan
# Gọi Procedure: sp_ChayTinhLuongV1_ToanCoQuan hoặc sp_ChayTinhLuongV2_ToanCoQuan
# ----------------------------------------------------------------------
@app.route('/calculate-salary-global', methods=['POST'])
@login_required
def calculate_salary_global():
    # 1. Kiểm tra quyền Admin (Chặt chẽ hơn)
    if not current_user.is_authenticated or not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Bạn không có quyền thực hiện thao tác này!'}), 403

    data = request.get_json()
    year = data.get('year')
    salary_type = data.get('type') # 'V1' hoặc 'V2'

    # Kiểm tra đầu vào
    if not year:
        return jsonify({'success': False, 'message': 'Vui lòng chọn năm tính toán.'})
    if salary_type not in ['V1', 'V2']:
        return jsonify({'success': False, 'message': 'Loại lương không hợp lệ (Phải là V1 hoặc V2).'})

    try:
        # 2. Thực thi Procedure
        if salary_type == 'V1':
            # Thực thi và lấy kết quả trả về từ lệnh SELECT trong Procedure
            result = db.session.execute(text("CALL sp_ChayTinhLuongV1_ToanCoQuan(:year)"), {'year': year})
        else:
            result = db.session.execute(text("CALL sp_ChayTinhLuongV2_ToanCoQuan(:year)"), {'year': year})
        
        # Lấy câu thông báo từ Procedure (ví dụ: "Đã hoàn thành... các kỳ đã khóa được giữ nguyên")
        # Do Procedure có lệnh SELECT cuối cùng, ta fetch nó ra
        row = result.fetchone()
        db_message = row[0] if row else f"Đã thực hiện tính toán lương {salary_type} năm {year}."

        # Lưu thay đổi vào DB
        db.session.commit()
        
        # 3. Trả về phản hồi cho giao diện
        return jsonify({
            'success': True, 
            'message': db_message
        })

    except Exception as e:
        db.session.rollback()
        # In lỗi chi tiết ra terminal để Admin/Dev kiểm tra
        print(f"--- LỖI TÍNH LƯƠNG {salary_type} TỔNG THỂ ---")
        print(f"Năm: {year}")
        print(f"Chi tiết: {str(e)}")
        return jsonify({
            'success': False, 
            'message': f"Lỗi hệ thống khi gọi Procedure: {str(e)}"
        })


@app.route('/report-v1-v2')
@login_required
def report_v1_v2():
    # 1. LẤY THÔNG TIN QUYỀN CỦA USER
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    user_ma_pb = user_info.ma_phong_ban if user_info else None
    user_ten_pb = user_info.phong_ban.ten_phong_ban if user_info and user_info.phong_ban else ""

    # Quyền Admin hoặc thuộc các phòng ban đặc biệt (1: Ban Giám đốc; 2: Tổng hợp; 3: Kiểm tra, giám sát; 4: Thu thập và Xử lý thông tin)
    is_admin_or_th = (current_user.is_admin == 1 or user_ma_pb in [2])

    # 2. XỬ LÝ THAM SỐ LỌC
    nam_hien_tai = datetime.now().year
    nam_chon = request.args.get('nam', default=nam_hien_tai, type=int)
    search_ma_nv = request.args.get('ma_nv', default='').strip()
    
    # LOGIC PHÂN QUYỀN LỌC PHÒNG BAN:
    if is_admin_or_th:
        search_phong_ban = request.args.get('phong_ban', default='').strip()
    else:
        search_phong_ban = user_ten_pb  # Ép buộc lấy phòng ban của chính user

    try:
        # 3. XÂY DỰNG CÂU LỆNH SQL
        query_str = "SELECT * FROM view_tong_hop_v1_v2 WHERE nam = :nam"
        params = {'nam': nam_chon}

        if search_ma_nv:
            query_str += " AND ma_nhan_vien LIKE :ma_nv"
            params['ma_nv'] = f"%{search_ma_nv}%"
        
        if search_phong_ban:
            query_str += " AND phong_ban = :phong_ban"
            params['phong_ban'] = search_phong_ban
        elif not is_admin_or_th:
            # Trường hợp user không có phòng ban và không có quyền admin -> không thấy gì
            query_str += " AND 1=0"

        # Thực thi lấy dữ liệu chính
        data = db.session.execute(text(query_str), params).mappings().all() 
        
        # 4. LẤY DỮ LIỆU CHO DROPDOWNS
        # Chỉ Admin mới cần lấy toàn bộ danh sách phòng ban
        all_depts = []
        if is_admin_or_th:
            all_depts = [row[0] for row in db.session.execute(text("SELECT ten_phong_ban FROM phong_ban ORDER BY ten_phong_ban"))]
        
        # Lấy danh sách năm
        all_years = [row[0] for row in db.session.execute(text("SELECT DISTINCT nam FROM view_tong_hop_v1_v2 ORDER BY nam DESC"))]
        if nam_chon not in all_years:
            all_years.append(nam_chon)
            all_years.sort(reverse=True)

        return render_template('bao_cao_v1_v2.html', 
                               data=data, 
                               nam_chon=nam_chon, 
                               all_years=all_years,
                               all_depts=all_depts,
                               search_ma_nv=search_ma_nv,
                               search_phong_ban=search_phong_ban,
                               is_admin_or_th=is_admin_or_th) # Truyền biến quyền sang HTML
                               
    except Exception as e:
        db.session.rollback()
        flash(f"Lỗi truy vấn dữ liệu: {str(e)}", "danger")
        return redirect(url_for('index'))

@app.route('/export-report-v1-v2')
@login_required
def export_report_v1_v2():
    # --- PHÂN QUYỀN: Kiểm tra quyền người dùng ---
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    user_ma_pb = user_info.ma_phong_ban if user_info else None
    user_ten_pb = user_info.phong_ban.ten_phong_ban if user_info and user_info.phong_ban else ""
    
    # Quyền Admin hoặc thuộc các phòng ban đặc biệt (1: Ban Giám đốc; 2: Tổng hợp; 3: Kiểm tra, giám sát; 4: Thu thập và Xử lý thông tin)
    is_admin_or_th = (current_user.is_admin == 1 or user_ma_pb in [2])

    nam_chon = request.args.get('nam', type=int)
    search_ma_nv = request.args.get('ma_nv', default='')
    search_phong_ban = request.args.get('phong_ban', default='')

    # LOGIC BẢO MẬT: Cưỡng ép lọc theo phòng ban nếu không có quyền Admin/TH
    if not is_admin_or_th:
        search_phong_ban = user_ten_pb

    try:
        # 1. Xây dựng Query dữ liệu
        query_str = "SELECT * FROM view_tong_hop_v1_v2 WHERE nam = :nam"
        params = {'nam': nam_chon}
        
        if search_ma_nv:
            query_str += " AND ma_nhan_vien LIKE :ma_nv"
            params['ma_nv'] = f"%{search_ma_nv}%"
            
        if search_phong_ban:
            # Lọc chính xác theo tên phòng ban từ view
            query_str += " AND phong_ban = :phong_ban"
            params['phong_ban'] = search_phong_ban

        data = db.session.execute(text(query_str), params).mappings().all()

        # 2. Tạo file Excel bằng openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Luong V1-V2 Nam {nam_chon}"

        # Định dạng Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Xanh đậm
        month_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Xanh nhạt
        font_white = Font(color="FFFFFF", bold=True)
        font_bold = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Tạo Header 2 tầng
        # Cột cố định
        headers_fixed = [("Mã NV", "A"), ("Họ và Tên", "B"), ("Phòng Ban", "C")]
        for text_val, col_let in headers_fixed:
            ws.merge_cells(f'{col_let}1:{col_let}2')
            cell = ws[f'{col_let}1']
            cell.value = text_val
            cell.fill = header_fill
            cell.font = font_white
            cell.alignment = center_align
            cell.border = thin_border

        # Các cột tháng (1 -> 12)
        col_idx = 4
        for i in range(1, 13):
            start_col = get_column_letter(col_idx)
            end_col = get_column_letter(col_idx + 2)
            ws.merge_cells(f'{start_col}1:{end_col}1')
            
            # Header tháng
            cell_m = ws[f'{start_col}1']
            cell_m.value = f"Tháng {i}"
            cell_m.fill = month_fill
            cell_m.font = font_bold
            cell_m.alignment = center_align
            cell_m.border = thin_border
            
            # Sub-headers
            sub_titles = ["V1", "V2", "V1+V2"]
            for j, title in enumerate(sub_titles):
                c = ws.cell(row=2, column=col_idx + j, value=title)
                c.font = font_bold
                c.alignment = center_align
                c.border = thin_border
            col_idx += 3

        # Header Tổng năm
        start_total = get_column_letter(col_idx)
        end_total = get_column_letter(col_idx + 2)
        ws.merge_cells(f'{start_total}1:{end_total}1')
        cell_t = ws[f'{start_total}1']
        cell_t.value = "TỔNG NĂM"
        cell_t.fill = header_fill
        cell_t.font = font_white
        cell_t.alignment = center_align
        cell_t.border = thin_border

        sub_totals = ["Tổng V1", "Tổng V2", "TỔNG CỘNG"]
        for j, title in enumerate(sub_totals):
            c = ws.cell(row=2, column=col_idx + j, value=title)
            c.font = font_bold
            c.alignment = center_align
            c.border = thin_border

        # 3. Điền dữ liệu
        for r_idx, row in enumerate(data, start=3):
            # Ghi thông tin cơ bản
            ws.cell(row=r_idx, column=1, value=row.ma_nhan_vien).border = thin_border
            ws.cell(row=r_idx, column=2, value=row.ho_ten).border = thin_border
            ws.cell(row=r_idx, column=3, value=row.phong_ban).border = thin_border
            
            c_idx = 4
            # Ghi dữ liệu 12 tháng
            for i in range(1, 13):
                # V1
                c1 = ws.cell(row=r_idx, column=c_idx, value=float(row[f'tien_v1_t{i}'] or 0))
                c1.number_format = '#,##0'
                c1.border = thin_border
                # V2
                c2 = ws.cell(row=r_idx, column=c_idx+1, value=float(row[f'tien_v2_t{i}'] or 0))
                c2.number_format = '#,##0'
                c2.border = thin_border
                # Tổng tháng
                c3 = ws.cell(row=r_idx, column=c_idx+2, value=float(row[f'tong_tien_t{i}'] or 0))
                c3.number_format = '#,##0'
                c3.font = font_bold
                c3.border = thin_border
                c_idx += 3
            
            # Ghi tổng năm
            ct1 = ws.cell(row=r_idx, column=c_idx, value=float(row.tong_tien_v1_nam or 0))
            ct1.number_format = '#,##0'
            ct1.border = thin_border
            
            ct2 = ws.cell(row=r_idx, column=c_idx+1, value=float(row.tong_tien_v2_nam or 0))
            ct2.number_format = '#,##0'
            ct2.border = thin_border
            
            ct3 = ws.cell(row=r_idx, column=c_idx+2, value=float(row.tong_tien_v1_v2_all_nam or 0))
            ct3.number_format = '#,##0'
            ct3.font = font_bold
            ct3.fill = month_fill # Làm nổi bật cột tổng cuối cùng
            ct3.border = thin_border

        # Tự động chỉnh độ rộng cột
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        for col in range(4, col_idx + 3):
            ws.column_dimensions[get_column_letter(col)].width = 13

        # Lưu vào bộ nhớ và gửi về trình duyệt
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Tên file linh hoạt theo phòng ban
        suffix = search_phong_ban.replace(" ", "_") if search_phong_ban else "Toan_Cong_Ty"
        filename = f"Bao_cao_V1_V2_{nam_chon}_{suffix}.xlsx"
        
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        db.session.rollback()
        import traceback
        print(traceback.format_exc())
        flash(f"Lỗi xuất Excel: {str(e)}", "danger")
        return redirect(url_for('report_v1_v2', nam=nam_chon))

# ----------------------------------------------------------------------
# Bảng lương V2 có tính đến mức đề nghị Hệ số điều chỉnh => Luong_DC = HSL_V2 * 800.000 * HSL_DC
# ----------------------------------------------------------------------
@app.route('/bang-luong-dc')
@login_required
def bang_luong_dc():
    # 1. Kiểm tra quyền hạn (Chỉ Admin, Kiểm soát, Phê duyệt mới được vào)
    if not (current_user.is_admin or current_user.role in ['KIEM_SOAT', 'PHE_DUYET']):
        flash("Bạn không có quyền truy cập mục này.", "danger")
        return redirect(url_for('index'))

    # 2. Lấy tham số lọc từ URL
    year = request.args.get('year', datetime.now().year, type=int)
    selected_pb = request.args.get('phong_ban', '')
    ma_nv_filter = request.args.get('ma_nv', '')

    # 3. Lấy danh sách phòng ban cho dropdown
    ds_phong_ban = PhongBan.query.all()

    # 4. Xây dựng câu lệnh SQL JOIN để lấy thông tin tổng hợp
    # Sử dụng LEFT JOIN với thong_tin_dc để hiện cả những người chưa có dữ liệu điều chỉnh
    sql_text = """
        SELECT 
            nv.ma_nhan_vien, 
            nv.ho_ten, 
            pb.ten_phong_ban, 
            dc.t1, dc.t2, dc.t3, dc.t4, dc.t5, dc.t6,
            dc.t7, dc.t8, dc.t9, dc.t10, dc.t11, dc.t12,
            dc.approved_by,
            dc.approved_at,
            u_app.fullname as ten_nguoi_duyet
        FROM thong_tin_nguoi_lao_dong nv
        LEFT JOIN thong_tin_dc dc ON nv.ma_nhan_vien = dc.ma_nhan_vien AND dc.nam = :year
        LEFT JOIN phong_ban pb ON nv.ma_phong_ban = pb.id
        LEFT JOIN users u_app ON dc.approved_by = u_app.ma_nhan_vien
        WHERE 1=1
    """
    
    params = {"year": year}

    # Thêm điều kiện lọc theo phòng ban
    if selected_pb:
        sql_text += " AND nv.ma_phong_ban = :pb"
        params["pb"] = selected_pb
    
    # Thêm điều kiện lọc theo mã nhân viên
    if ma_nv_filter:
        sql_text += " AND nv.ma_nhan_vien LIKE :ma"
        params["ma"] = f"%{ma_nv_filter}%"

    # Sắp xếp theo mã nhân viên
    sql_text += " ORDER BY nv.ma_nhan_vien ASC"

    # Thực thi truy vấn
    results = db.session.execute(text(sql_text), params).fetchall()

    return render_template('bang_luong_dc.html', 
                           results=results, 
                           year=year, 
                           ds_phong_ban=ds_phong_ban,
                           selected_pb=selected_pb,
                           ma_nv_filter=ma_nv_filter)


@app.route('/approve-dc', methods=['POST'])
@login_required
def approve_dc():
    # 1. Kiểm tra quyền hạn
    if current_user.role not in ['ADMIN', 'KIEM_SOAT']:
        return jsonify({'success': False, 'message': 'Bạn không có quyền phê duyệt!'}), 403

    data = request.get_json()
    try:
        year = int(data.get('year'))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': 'Năm không hợp lệ!'}), 400

    selected_pb = data.get('phong_ban')
    ma_nv_filter = data.get('ma_nv')
    updates = data.get('updates', []) # Nhận danh sách hệ số DC mới từ giao diện

    try:
        # BƯỚC A: LƯU DỮ LIỆU HỆ SỐ MỚI
        if updates:
            for item in updates:
                ma_nv_item = item.get('ma_nv')
                
                # Kiểm tra xem đã có bản ghi trong bảng thong_tin_dc chưa
                check_sql = text("SELECT COUNT(*) FROM thong_tin_dc WHERE ma_nhan_vien = :ma AND nam = :yr")
                exists = db.session.execute(check_sql, {"ma": ma_nv_item, "yr": year}).scalar()

                if not exists:
                    # Nếu chưa có thì chèn mới bản ghi với giá trị mặc định là 1.0
                    insert_sql = text("INSERT INTO thong_tin_dc (ma_nhan_vien, nam) VALUES (:ma, :yr)")
                    db.session.execute(insert_sql, {"ma": ma_nv_item, "yr": year})

                # Cập nhật giá trị 12 tháng (Chỉ cập nhật nếu chưa được duyệt)
                set_clause = ", ".join([f"t{m} = :t{m}" for m in range(1, 13)])
                save_sql = f"""
                    UPDATE thong_tin_dc 
                    SET {set_clause} 
                    WHERE ma_nhan_vien = :ma_nv 
                      AND nam = :year 
                      AND (approved_by IS NULL OR approved_by = '' OR approved_by = 'None')
                """
                # Mặc định hệ số là 1.00 nếu không có dữ liệu truyền lên
                params_save = {f"t{m}": item.get(f't{m}', 1.0) for m in range(1, 13)}
                params_save.update({'ma_nv': ma_nv_item, 'year': year})
                
                db.session.execute(text(save_sql), params_save)

            # (Tùy chọn) Gọi Procedure tính toán lại báo cáo V2_DC nếu bạn đã viết
            db.session.execute(text("CALL sp_ChayTinhLuongV2DC_ToanCoQuan(:year)"), {"year": year})
            db.session.commit()

        # BƯỚC B: THỰC HIỆN PHÊ DUYỆT HÀNG LOẠT THEO BỘ LỌC
        approve_query = """
            UPDATE thong_tin_dc dc
            JOIN thong_tin_nguoi_lao_dong nv ON dc.ma_nhan_vien = nv.ma_nhan_vien 
            SET dc.approved_by = :user, dc.approved_at = :now 
            WHERE dc.nam = :year
            AND (dc.approved_by IS NULL OR dc.approved_by = '' OR dc.approved_by = 'None')
        """
        params_approve = {
            "user": current_user.ma_nhan_vien, 
            "now": datetime.now(HANOI_TZ), 
            "year": year
        }

        # Áp dụng bộ lọc từ giao diện giống hệt V1
        if selected_pb and str(selected_pb).strip():
            approve_query += " AND nv.ma_phong_ban = :pb"
            params_approve["pb"] = selected_pb
        if ma_nv_filter and str(ma_nv_filter).strip():
            approve_query += " AND dc.ma_nhan_vien = :ma"
            params_approve["ma"] = ma_nv_filter

        result = db.session.execute(text(approve_query), params_approve)
        db.session.commit()
        
        row_count = result.rowcount
        log_event("Phê duyệt HSL_DC", f"Đã lưu và phê duyệt cho năm {year}")
        
        return jsonify({
            'success': True, 
            'message': f'Đã lưu hệ số và phê duyệt thành công {row_count} nhân viên!'
        })
    
    except Exception as e:
        db.session.rollback()
        print(f"Error in approve_dc: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/unapprove-dc', methods=['POST'])
@login_required
def unapprove_dc():
    # 1. Kiểm tra quyền hạn: Chỉ Admin hoặc KIEM_SOAT mới có quyền gỡ phê duyệt
    # Dựa theo logic v1 bạn gửi: Kiểm tra is_admin hoặc role ADMIN trong session
    is_admin = getattr(current_user, 'is_admin', False) or current_user.role == 'ADMIN'
    
    if not is_admin:
        return jsonify({'success': False, 'message': 'Chỉ Admin mới có quyền gỡ phê duyệt!'}), 403

    data = request.get_json()
    year = data.get('year')
    ma_nv = data.get('ma_nv') # Định danh nhân viên cụ thể để gỡ khóa

    # 2. Kiểm tra dữ liệu đầu vào
    if not year or not ma_nv:
        return jsonify({'success': False, 'message': 'Thiếu thông tin năm hoặc mã nhân viên!'}), 400
    
    try:
        # 3. Thực hiện cập nhật bảng thong_tin_dc bằng SQL thuần
        query = """
            UPDATE thong_tin_dc 
            SET approved_by = NULL, approved_at = NULL 
            WHERE nam = :year AND ma_nhan_vien = :ma
        """
        
        result = db.session.execute(text(query), {"year": year, "ma": ma_nv})
        db.session.commit()
        
        # Kiểm tra xem có dòng nào được cập nhật không
        if result.rowcount == 0:
            return jsonify({'success': False, 'message': 'Không tìm thấy dữ liệu phù hợp để gỡ phê duyệt!'})

        # 4. Ghi log sự kiện hệ thống (nếu bạn có hàm log_event)
        if 'log_event' in globals():
            log_event("Gỡ phê duyệt DC", f"Admin {current_user.ma_nhan_vien} đã gỡ phê duyệt HSL_DC cho NV {ma_nv} năm {year}")
            
        return jsonify({'success': True, 'message': 'Đã gỡ trạng thái phê duyệt Hệ số điều chỉnh!'})

    except Exception as e:
        db.session.rollback()
        print(f"Error in unapprove_dc: {str(e)}")
        return jsonify({'success': False, 'message': f'Lỗi hệ thống: {str(e)}'}), 500


@app.route('/v2-dc-report')
@login_required
def v2_dc_report():
    # 1. LẤY THÔNG TIN QUYỀN CỦA USER
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    user_ma_pb = user_info.ma_phong_ban if user_info else None
    
    # Lấy tên phòng ban từ relationship (Đã sửa ở bước trước)
    user_ten_pb = user_info.phong_ban.ten_phong_ban if user_info and user_info.phong_ban else ""

    # Quyền Admin hoặc thuộc các phòng ban đặc biệt (1: Ban Giám đốc; 2: Tổng hợp; 3: Kiểm tra, giám sát; 4: Thu thập và Xử lý thông tin)
    is_admin_or_th = (current_user.is_admin == 1 or user_ma_pb in [2])

    # 2. XỬ LÝ THAM SỐ LỌC
    current_year = datetime.now().year
    year = request.args.get('year', current_year, type=int)
    ma_nv_filter = request.args.get('ma_nv', '')

    # LOGIC PHÂN QUYỀN LỌC PHÒNG BAN:
    if is_admin_or_th:
        selected_pb = request.args.get('phong_ban', '')
    else:
        selected_pb = user_ten_pb # Ép buộc user thường chỉ xem phòng mình

    # 3. LẤY DANH SÁCH PHÒNG BAN (Chỉ dành cho Admin/TH chọn)
    ds_phong_ban = []
    if is_admin_or_th:
        ds_pb_rows = db.session.execute(text("SELECT ten_phong_ban FROM phong_ban ORDER BY ten_phong_ban")).fetchall()
        ds_phong_ban = [r[0] for r in ds_pb_rows]

    # 4. TRUY VẤN DỮ LIỆU
    sql = """
        SELECT 
            nv.ma_nhan_vien AS nv_ma, 
            nv.ho_ten AS nv_ten, 
            pb.ten_phong_ban AS pb_ten,
            tl.*
        FROM thong_tin_nguoi_lao_dong nv
        LEFT JOIN phong_ban pb ON nv.ma_phong_ban = pb.id
        LEFT JOIN tien_luong_v2_dc tl ON nv.ma_nhan_vien = tl.ma_nhan_vien AND tl.nam = :year
        WHERE nv.trang_thai = 1
    """
    params = {"year": year}
    
    if selected_pb:
        sql += " AND pb.ten_phong_ban = :pb"
        params["pb"] = selected_pb
    elif not is_admin_or_th:
        # Nếu ko có quyền và ko tìm thấy phòng ban user, không trả về dữ liệu
        sql += " AND 1=0"

    if ma_nv_filter:
        sql += " AND nv.ma_nhan_vien LIKE :ma"
        params["ma"] = f"%{ma_nv_filter}%"

    # Sắp xếp theo mã nhân viên
    results = db.session.execute(text(sql + " ORDER BY nv.ma_nhan_vien ASC"), params).fetchall()

    # 5. XỬ LÝ DỮ LIỆU SANG FORMAT CỦA TEMPLATE
    report_data = []
    for r in results:
        emp_months = []
        for m in range(1, 13):
            hsl_v2 = getattr(r, f'hsl_v2_t{m}', 0) or 0
            hsl_dc = getattr(r, f'hsl_dc_t{m}', 1.0) or 1.0
            tien = getattr(r, f'tien_v2dc_t{m}', 0) or 0
            
            emp_months.append({
                'hsl_v2': float(hsl_v2),
                'hsl_dc': float(hsl_dc),
                'tien': float(tien)
            })
        
        report_data.append({
            'ma_nv': r.nv_ma,
            'ho_ten': r.nv_ten,
            'phong_ban': r.pb_ten or "N/A",
            'months': emp_months,
            'tong_nam': float(getattr(r, 'tien_v2dc_tong_nam', 0) or 0)
        })

    return render_template('v2_dc_report.html', 
                           data=report_data, 
                           year=year, 
                           ds_phong_ban=ds_phong_ban, 
                           selected_pb=selected_pb, 
                           ma_nv_filter=ma_nv_filter,
                           is_admin_or_th=is_admin_or_th, # Gửi biến này sang HTML
                           now=datetime.now())

@app.route('/calculate-v2dc-global', methods=['POST'])
@login_required
def calculate_v2dc_global():
    data = request.get_json()
    year = data.get('year')
    try:
        # SỬA TẠI ĐÂY: Thay tên Procedure đúng mà bạn đã tạo trong MySQL
        # Thông thường là sp_ChayTinhLuongV2DC_ToanCoQuan hoặc tương đương
        db.session.execute(text("CALL sp_ChayTinhLuongV2DC_ToanCoQuan(:year)"), {"year": year})
        db.session.commit()
        return jsonify({"success": True, "message": f"Đã tính toán xong lương V2_DC cho năm {year}"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)})


@app.route('/export-v2-dc-excel/<int:year>')
@login_required
def export_v2_dc_excel(year):
    try:
        # 1. Truy vấn dữ liệu từ bảng tien_luong_v2_dc
        sql = text("""
            SELECT 
                nv.ma_nhan_vien, nv.ho_ten, pb.ten_phong_ban,
                tl.*
            FROM thong_tin_nguoi_lao_dong nv
            LEFT JOIN phong_ban pb ON nv.ma_phong_ban = pb.id
            LEFT JOIN tien_luong_v2_dc tl ON nv.ma_nhan_vien = tl.ma_nhan_vien AND tl.nam = :year
            WHERE nv.trang_thai = 1
            ORDER BY pb.ten_phong_ban, nv.ma_nhan_vien
        """)
        results = db.session.execute(sql, {"year": year}).fetchall()

        # 2. Khởi tạo Workbook và Định dạng
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Luong V2_DC_{year}"

        # Styles
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Xanh đậm
        fill_sub_header = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Xanh lá nhạt
        font_white_bold = Font(color="FFFFFF", bold=True)
        font_bold = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))

        # 3. Xây dựng Header (Hàng 1 & 2)
        fixed_cols = [("Mã NV", "A"), ("Họ và Tên", "B"), ("Phòng ban", "C")]
        for label, col_letter in fixed_cols:
            ws.merge_cells(f'{col_letter}1:{col_letter}2')
            cell = ws[f'{col_letter}1']
            cell.value = label
            cell.fill = fill_header
            cell.font = font_white_bold
            cell.alignment = center_align
            cell.border = thin_border

        # Duyệt 12 tháng
        current_col = 4
        for m in range(1, 13):
            # Header Tháng (Merge 3 cột: HSL V2, HSL DC, Số tiền)
            ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + 2)
            cell_m = ws.cell(row=1, column=current_col, value=f"Tháng {m}")
            cell_m.fill = fill_sub_header
            cell_m.font = font_bold
            cell_m.alignment = center_align
            cell_m.border = thin_border

            # Sub-labels
            sub_labels = ["HSL V2", "HSL DC", "Thành tiền DC"]
            for i, label in enumerate(sub_labels):
                c = ws.cell(row=2, column=current_col + i, value=label)
                c.font = Font(size=9, bold=True)
                c.alignment = center_align
                c.border = thin_border
            current_col += 3

        # Header Tổng năm
        ws.merge_cells(start_row=1, start_column=current_col, end_row=2, end_column=current_col)
        cell_total = ws.cell(row=1, column=current_col, value="Tổng Năm V2_DC")
        cell_total.fill = fill_header
        cell_total.font = font_white_bold
        cell_total.alignment = center_align
        cell_total.border = thin_border

        # 4. Đổ dữ liệu
        row_idx = 3
        for r in results:
            ws.cell(row=row_idx, column=1, value=r.ma_nhan_vien).border = thin_border
            ws.cell(row=row_idx, column=2, value=r.ho_ten).border = thin_border
            ws.cell(row=row_idx, column=3, value=r.ten_phong_ban or "N/A").border = thin_border
            
            data_col = 4
            for m in range(1, 13):
                h_v2 = float(getattr(r, f'hsl_v2_t{m}') or 0)
                h_dc = float(getattr(r, f'hsl_dc_t{m}') or 1.0)
                tien = float(getattr(r, f'tien_v2dc_t{m}') or 0)

                ws.cell(row=row_idx, column=data_col, value=h_v2).border = thin_border
                ws.cell(row=row_idx, column=data_col + 1, value=h_dc).border = thin_border
                
                c_tien = ws.cell(row=row_idx, column=data_col + 2, value=tien)
                c_tien.number_format = '#,##0'
                c_tien.border = thin_border
                data_col += 3
            
            # Tổng năm
            c_sum = ws.cell(row=row_idx, column=data_col, value=float(r.tien_v2dc_tong_nam or 0))
            c_sum.number_format = '#,##0'
            c_sum.font = font_bold
            c_sum.border = thin_border
            row_idx += 1

        # 5. Căn chỉnh độ rộng cột
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        for c in range(4, current_col + 1):
            ws.column_dimensions[get_column_letter(c)].width = 11

        # 6. Trả file
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output, 
            as_attachment=True, 
            download_name=f"Bao_Cao_V2_DC_{year}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"Lỗi Export V2_DC: {traceback.format_exc()}")
        flash(f"Lỗi hệ thống khi xuất file: {str(e)}", "danger")
        return redirect(url_for('v2_dc_report'))


@app.route('/report-v1-v2-dc')
@login_required
def report_v1_v2_dc():
    # 1. LẤY THÔNG TIN QUYỀN CỦA USER
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    user_ma_pb = user_info.ma_phong_ban if user_info else None
    user_ten_pb = user_info.phong_ban.ten_phong_ban if user_info and user_info.phong_ban else ""

    # Quyền Admin hoặc thuộc các phòng ban đặc biệt (1: Ban Giám đốc; 2: Tổng hợp; 3: Kiểm tra, giám sát; 4: Thu thập và Xử lý thông tin)
    is_admin_or_th = (current_user.is_admin == 1 or user_ma_pb in [2])

    # 2. XỬ LÝ THAM SỐ LỌC
    nam = request.args.get('nam', datetime.now().year, type=int)
    
    # LOGIC PHÂN QUYỀN LỌC PHÒNG BAN:
    if is_admin_or_th:
        search_phong_ban = request.args.get('phong_ban', '')
    else:
        search_phong_ban = user_ten_pb # Ép buộc user thường chỉ xem phòng mình

    # 3. LẤY DỮ LIỆU CHO DROPDOWN
    all_years = [r[0] for r in db.session.execute(text("SELECT DISTINCT nam FROM tien_luong_v1 ORDER BY nam DESC")).fetchall()]
    if nam not in all_years: 
        all_years.insert(0, nam)
    
    # Chỉ lấy danh sách phòng ban nếu là Admin/TH
    all_depts = []
    if is_admin_or_th:
        all_depts = [r[0] for r in db.session.execute(text("SELECT ten_phong_ban FROM phong_ban ORDER BY ten_phong_ban")).fetchall()]

    # 4. TRUY VẤN LẤY CHI TIẾT TỪNG THÁNG CỦA V1 VÀ V2_DC
    # Sử dụng INNER JOIN với phong_ban để đảm bảo lọc chính xác theo quyền
    sql = """
        SELECT 
            nv.ma_nhan_vien, nv.ho_ten, pb.ten_phong_ban,
            v1.tien_v1_t1, v1.tien_v1_t2, v1.tien_v1_t3, v1.tien_v1_t4, v1.tien_v1_t5, v1.tien_v1_t6,
            v1.tien_v1_t7, v1.tien_v1_t8, v1.tien_v1_t9, v1.tien_v1_t10, v1.tien_v1_t11, v1.tien_v1_t12,
            v1.tong_v1,
            v2.tien_v2dc_t1, v2.tien_v2dc_t2, v2.tien_v2dc_t3, v2.tien_v2dc_t4, v2.tien_v2dc_t5, v2.tien_v2dc_t6,
            v2.tien_v2dc_t7, v2.tien_v2dc_t8, v2.tien_v2dc_t9, v2.tien_v2dc_t10, v2.tien_v2dc_t11, v2.tien_v2dc_t12,
            v2.tien_v2dc_tong_nam AS tong_v2_dc,
            (COALESCE(v1.tong_v1, 0) + COALESCE(v2.tien_v2dc_tong_nam, 0)) AS tong_thu_nhap
        FROM thong_tin_nguoi_lao_dong nv
        INNER JOIN phong_ban pb ON nv.ma_phong_ban = pb.id
        LEFT JOIN (
            SELECT ma_nhan_vien, 
            tien_v1_t1, tien_v1_t2, tien_v1_t3, tien_v1_t4, tien_v1_t5, tien_v1_t6,
            tien_v1_t7, tien_v1_t8, tien_v1_t9, tien_v1_t10, tien_v1_t11, tien_v1_t12,
            (COALESCE(tien_v1_t1,0)+COALESCE(tien_v1_t2,0)+COALESCE(tien_v1_t3,0)+COALESCE(tien_v1_t4,0)+
             COALESCE(tien_v1_t5,0)+COALESCE(tien_v1_t6,0)+COALESCE(tien_v1_t7,0)+COALESCE(tien_v1_t8,0)+
             COALESCE(tien_v1_t9,0)+COALESCE(tien_v1_t10,0)+COALESCE(tien_v1_t11,0)+COALESCE(tien_v1_t12,0)) as tong_v1
            FROM tien_luong_v1 WHERE nam = :nam
        ) v1 ON nv.ma_nhan_vien = v1.ma_nhan_vien
        LEFT JOIN (
            SELECT ma_nhan_vien, 
            tien_v2dc_t1, tien_v2dc_t2, tien_v2dc_t3, tien_v2dc_t4, tien_v2dc_t5, tien_v2dc_t6,
            tien_v2dc_t7, tien_v2dc_t8, tien_v2dc_t9, tien_v2dc_t10, tien_v2dc_t11, tien_v2dc_t12,
            tien_v2dc_tong_nam 
            FROM tien_luong_v2_dc WHERE nam = :nam
        ) v2 ON nv.ma_nhan_vien = v2.ma_nhan_vien
        WHERE nv.trang_thai = 1
    """
    params = {"nam": nam}
    
    if search_phong_ban:
        sql += " AND pb.ten_phong_ban = :pb"
        params["pb"] = search_phong_ban
    elif not is_admin_or_th:
        # Nếu không có quyền và không lấy được phòng ban, chặn hiển thị
        sql += " AND 1=0"
        
    sql += " ORDER BY nv.ma_nhan_vien ASC"

    results = db.session.execute(text(sql), params).fetchall()

    return render_template('bao_cao_v1_v2_dc.html', 
                           results=results, 
                           all_years=all_years, 
                           nam_chon=nam, 
                           all_depts=all_depts, 
                           search_phong_ban=search_phong_ban,
                           is_admin_or_th=is_admin_or_th)


@app.route('/export-v1-v2-dc')
@login_required
def export_v1_v2_dc():
    # --- PHÂN QUYỀN: Kiểm tra quyền người dùng ---
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    user_ma_pb = user_info.ma_phong_ban if user_info else None
    user_ten_pb = user_info.phong_ban.ten_phong_ban if user_info and user_info.phong_ban else ""
    
    # Quyền Admin hoặc thuộc các phòng ban đặc biệt (1: Ban Giám đốc; 2: Tổng hợp; 3: Kiểm tra, giám sát; 4: Thu thập và Xử lý thông tin)
    is_admin_or_th = (current_user.is_admin == 1 or user_ma_pb in [2])

    nam = request.args.get('nam', datetime.now().year, type=int)
    search_phong_ban = request.args.get('phong_ban', '')

    # LOGIC BẢO MẬT: Nếu không phải Admin/TH, cưỡng ép lọc theo tên phòng ban của User
    if not is_admin_or_th:
        search_phong_ban = user_ten_pb

    # 1. Truy vấn đầy đủ chi tiết 12 tháng của cả V1 và V2_DC
    sql = """
        SELECT 
            nv.ma_nhan_vien, nv.ho_ten, pb.ten_phong_ban,
            v1.tien_v1_t1, v1.tien_v1_t2, v1.tien_v1_t3, v1.tien_v1_t4, v1.tien_v1_t5, v1.tien_v1_t6,
            v1.tien_v1_t7, v1.tien_v1_t8, v1.tien_v1_t9, v1.tien_v1_t10, v1.tien_v1_t11, v1.tien_v1_t12,
            v1.tong_v1,
            v2.tien_v2dc_t1, v2.tien_v2dc_t2, v2.tien_v2dc_t3, v2.tien_v2dc_t4, v2.tien_v2dc_t5, v2.tien_v2dc_t6,
            v2.tien_v2dc_t7, v2.tien_v2dc_t8, v2.tien_v2dc_t9, v2.tien_v2dc_t10, v2.tien_v2dc_t11, v2.tien_v2dc_t12,
            v2.tien_v2dc_tong_nam AS tong_v2_dc,
            (COALESCE(v1.tong_v1, 0) + COALESCE(v2.tien_v2dc_tong_nam, 0)) AS tong_thu_nhap
        FROM thong_tin_nguoi_lao_dong nv
        LEFT JOIN phong_ban pb ON nv.ma_phong_ban = pb.id
        LEFT JOIN (
            SELECT ma_nhan_vien, 
            tien_v1_t1, tien_v1_t2, tien_v1_t3, tien_v1_t4, tien_v1_t5, tien_v1_t6,
            tien_v1_t7, tien_v1_t8, tien_v1_t9, tien_v1_t10, tien_v1_t11, tien_v1_t12,
            (COALESCE(tien_v1_t1,0)+COALESCE(tien_v1_t2,0)+COALESCE(tien_v1_t3,0)+COALESCE(tien_v1_t4,0)+
             COALESCE(tien_v1_t5,0)+COALESCE(tien_v1_t6,0)+COALESCE(tien_v1_t7,0)+COALESCE(tien_v1_t8,0)+
             COALESCE(tien_v1_t9,0)+COALESCE(tien_v1_t10,0)+COALESCE(tien_v1_t11,0)+COALESCE(tien_v1_t12,0)) as tong_v1
            FROM tien_luong_v1 WHERE nam = :nam
        ) v1 ON nv.ma_nhan_vien = v1.ma_nhan_vien
        LEFT JOIN (
            SELECT ma_nhan_vien, 
            tien_v2dc_t1, tien_v2dc_t2, tien_v2dc_t3, tien_v2dc_t4, tien_v2dc_t5, tien_v2dc_t6,
            tien_v2dc_t7, tien_v2dc_t8, tien_v2dc_t9, tien_v2dc_t10, tien_v2dc_t11, tien_v2dc_t12,
            tien_v2dc_tong_nam 
            FROM tien_luong_v2_dc WHERE nam = :nam
        ) v2 ON nv.ma_nhan_vien = v2.ma_nhan_vien
        WHERE nv.trang_thai = 1
    """
    params = {"nam": nam}
    if search_phong_ban:
        sql += " AND pb.ten_phong_ban = :pb"
        params["pb"] = search_phong_ban
    sql += " ORDER BY nv.ma_nhan_vien ASC"
    results = db.session.execute(text(sql), params).fetchall()

    # 2. Khởi tạo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"V1_V2DC_{nam}"

    # Styles
    header_fill_v1 = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Xanh đậm
    header_fill_v2 = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid") # Vàng đậm
    header_fill_total = PatternFill(start_color="375623", end_color="375623", fill_type="solid") # Xanh lá đậm
    font_white = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal="center", vertical="center")

    # 3. Tạo Header 2 tầng
    # Hàng 1: Định nghĩa các vùng Merge
    ws.merge_cells('A1:C2')
    ws['A1'] = "Thông tin nhân viên"
    
    ws.merge_cells('D1:P1')
    ws['D1'] = "CHI TIẾT LƯƠNG V1"
    
    ws.merge_cells('Q1:AC1')
    ws['Q1'] = "CHI TIẾT LƯƠNG V2 ĐIỀU CHỈNH"
    
    ws.merge_cells('AD1:AD2')
    ws['AD1'] = "TỔNG THU NHẬP"

    # Hàng 2: Ghi tiêu đề cột cụ thể
    v1_cols = [f"T{i}" for i in range(1, 13)] + ["Tổng V1"]
    v2_cols = [f"T{i}" for i in range(1, 13)] + ["Tổng V2_DC"]
    
    # Ghi đè vào hàng 2 các giá trị sub-header
    sub_headers_row = ["Mã NV", "Họ tên", "Phòng ban"] + v1_cols + v2_cols + [""]
    for idx, val in enumerate(sub_headers_row, 1):
        # Vì cột A,B,C và AD đã merge từ hàng 1 xuống hàng 2, ta chỉ ghi vào hàng 2 cho các cột chưa merge hoàn toàn
        cell = ws.cell(row=2, column=idx)
        if val: cell.value = val

    # Format Headers (Hàng 1 & Hàng 2)
    for r in [1, 2]:
        for col in range(1, 31):
            cell = ws.cell(row=r, column=col)
            cell.alignment = align_center
            cell.border = thin_border
            if col <= 3: 
                cell.fill = PatternFill(start_color="D9D9D9", fill_type="solid") # Xám cho thông tin NV
            elif 4 <= col <= 16: 
                cell.fill = header_fill_v1
                cell.font = font_white
            elif 17 <= col <= 29:
                cell.fill = header_fill_v2
                cell.font = font_white
            else:
                cell.fill = header_fill_total
                cell.font = font_white

    # 4. Đổ dữ liệu
    for r in results:
        row_data = [
            r.ma_nhan_vien, r.ho_ten, r.ten_phong_ban,
            # 12 tháng V1
            float(r.tien_v1_t1 or 0), float(r.tien_v1_t2 or 0), float(r.tien_v1_t3 or 0),
            float(r.tien_v1_t4 or 0), float(r.tien_v1_t5 or 0), float(r.tien_v1_t6 or 0),
            float(r.tien_v1_t7 or 0), float(r.tien_v1_t8 or 0), float(r.tien_v1_t9 or 0),
            float(r.tien_v1_t10 or 0), float(r.tien_v1_t11 or 0), float(r.tien_v1_t12 or 0),
            float(r.tong_v1 or 0),
            # 12 tháng V2_DC
            float(r.tien_v2dc_t1 or 0), float(r.tien_v2dc_t2 or 0), float(r.tien_v2dc_t3 or 0),
            float(r.tien_v2dc_t4 or 0), float(r.tien_v2dc_t5 or 0), float(r.tien_v2dc_t6 or 0),
            float(r.tien_v2dc_t7 or 0), float(r.tien_v2dc_t8 or 0), float(r.tien_v2dc_t9 or 0),
            float(r.tien_v2dc_t10 or 0), float(r.tien_v2dc_t11 or 0), float(r.tien_v2dc_t12 or 0),
            float(r.tong_v2_dc or 0),
            # Tổng cuối
            float(r.tong_thu_nhap or 0)
        ]
        ws.append(row_data)
        
        # Format dòng dữ liệu vừa append
        curr_row = ws.max_row
        for col_idx in range(1, 31):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.border = thin_border
            if col_idx > 3:
                cell.number_format = '#,##0' # Định dạng số có dấu phân cách
                cell.alignment = Alignment(horizontal="right")

    # 5. Căn chỉnh độ rộng cột
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    # Chỉnh các cột số (D đến AC)
    for col_idx in range(4, 30):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11
    ws.column_dimensions['AD'].width = 18

    # Xuất file
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Tên file linh hoạt
    suffix = search_phong_ban.replace(" ", "_") if search_phong_ban else "Toan_Cong_Ty"
    filename = f"Bao_cao_tong_hop_DC_{nam}_{suffix}.xlsx"
    
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ----------------------------------------------------------------------
# Chấm công ngoài giờ các ngày trong tháng, trong tháng đó có bao nhiêu giờ của: Ngày Thường; Ngày Nghỉ; Ngày Lễ
# ----------------------------------------------------------------------
@app.route('/cham-cong-ngoai-gio', methods=['GET', 'POST'])
@login_required
def cham_cong_ngoai_gio():
    # 1. PHÂN QUYỀN VÀ LẤY THÔNG TIN PHÒNG BAN USER
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    user_ma_pb = str(user_info.ma_phong_ban) if user_info and user_info.ma_phong_ban else None
    user_ma_hieu_2 = user_info.ma_hieu_2 if user_info else None
    
    is_system_admin = (current_user.is_admin == 1)
    is_phong_tong_hop = (user_ma_pb == '2')
    is_admin_or_th = (is_system_admin or is_phong_tong_hop)

    # 2. THỜI GIAN & BỘ LỌC
    try:
        month = int(request.args.get('month', datetime.now().month))
        year = int(request.args.get('year', datetime.now().year))
    except (ValueError, TypeError):
        month, year = datetime.now().month, datetime.now().year

    # Logic xác định phạm vi hiển thị và lọc
    if is_system_admin:
        ma_hieu_2_filter = request.args.get('ma_hieu_2', '')
        ma_pb_filter = request.args.get('ma_phong_ban', '')
    elif is_phong_tong_hop:
        ma_hieu_2_filter = user_ma_hieu_2
        ma_pb_filter = request.args.get('ma_phong_ban', '')
    else:
        ma_hieu_2_filter = user_ma_hieu_2
        ma_pb_filter = str(user_info.ma_phong_ban) if user_info.ma_phong_ban else ''

    # Lấy danh sách ngày lễ để tính toán hệ số
    holiday_records = DanhMucNgayLe.query.filter(
        extract('month', DanhMucNgayLe.ngay) == month,
        extract('year', DanhMucNgayLe.ngay) == year
    ).all()
    holiday_days = {h.ngay.day for h in holiday_records} 

    # 3. LƯU DỮ LIỆU (POST)
    if request.method == 'POST':
        # Kiểm tra quyền ghi dữ liệu
        can_edit = is_system_admin or is_phong_tong_hop or (str(ma_pb_filter) == str(user_ma_pb))
        if not can_edit:
            flash("Bạn không có quyền cập nhật dữ liệu ngoài phạm vi quản lý!", "danger")
            return redirect(url_for('cham_cong_ngoai_gio', month=month, year=year))

        try:
            # Lọc nhân viên theo phạm vi đã chọn
            staff_query = ThongTinNguoiLaoDong.query.filter_by(trang_thai=True)
            if ma_hieu_2_filter:
                staff_query = staff_query.filter_by(ma_hieu_2=ma_hieu_2_filter)
            if ma_pb_filter:
                staff_query = staff_query.filter_by(ma_phong_ban=ma_pb_filter)
            
            active_staff = staff_query.all()

            with db.session.no_autoflush:
                for staff in active_staff:
                    # Bỏ qua nếu nhân viên không có trong dữ liệu submit (do phân trang hoặc lọc)
                    if f"d1_{staff.ma_nhan_vien}" not in request.form:
                        continue

                    record = ThongTinChamCongNgoaiGio.query.filter_by(
                        ma_nhan_vien=staff.ma_nhan_vien, thang=month, nam=year
                    ).first()

                    if not record:
                        record = ThongTinChamCongNgoaiGio(ma_nhan_vien=staff.ma_nhan_vien, thang=month, nam=year)
                        db.session.add(record)

                    t_thuong, t_nghi, t_le = 0.0, 0.0, 0.0

                    # Duyệt qua 31 ngày: Cập nhật từng ngày từ d1 -> d31
                    for d in range(1, 32):
                        val = request.form.get(f"d{d}_{staff.ma_nhan_vien}", "0").strip()
                        try:
                            num_val = float(val) if val and val.strip() != "" else 0.0
                        except ValueError:
                            num_val = 0.0
                        
                        setattr(record, f"d{d}", num_val)

                        # Phân loại giờ làm thêm để tính tổng
                        if num_val > 0:
                            try:
                                curr_date = datetime(year, month, d)
                                if d in holiday_days:
                                    t_le += num_val
                                elif curr_date.weekday() >= 5: # 5=Thứ 7, 6=Chủ Nhật
                                    t_nghi += num_val
                                else:
                                    t_thuong += num_val
                            except ValueError: # Ngày không tồn tại trong tháng (ví dụ 31/2) Ngày không hợp lệ (ví dụ ngày 31 của tháng có 30 ngày)
                                setattr(record, f"d{d}", 0.0)
                                pass 
                    
                    record.tong_gio_ngay_thuong = t_thuong
                    record.tong_gio_ngay_nghi = t_nghi
                    record.tong_gio_ngay_le = t_le
                    record.tong_gio_ngoai_gio = t_thuong + t_nghi + t_le
                    
                    # Quan trọng: Flush để Procedure đọc được giá trị mới lưu của nhân viên này
                    db.session.flush()

                    # Gọi Procedure 1: Cập nhật cột thuong_t[X], nghi_t[X], le_t[X] cho bảng năm; Cập nhật chi tiết bảng năm cho từng nhân viên
                    db.session.execute(
                        text("CALL sp_CapNhatTongHopNgoaiGio_ChiTiet(:ma, :na, :th)"),
                        {'ma': staff.ma_nhan_vien, 'na': year, 'th': month}
                    )

                # Sau khi chạy hết vòng lặp các nhân viên, gọi Procedure 2 để tính tổng cộng cả năm
                db.session.execute(
                    text("CALL sp_ChayTongHopNgoaiGio_ToanCoQuan(:na)"),
                    {'na': year}
                )
                
                db.session.commit()
                flash(f"Đã lưu thành công dữ liệu ngoài giờ tháng và tự động cập nhật bảng tổng hợp năm {month}/{year}!", "success")
            
        except Exception as e:
            db.session.rollback()
            flash(f"Lỗi hệ thống: {str(e)}", "danger")

        return redirect(url_for('cham_cong_ngoai_gio', month=month, year=year, ma_hieu_2=ma_hieu_2_filter, ma_phong_ban=ma_pb_filter))

    # 4. HIỂN THỊ (GET)
    # Join 3 bảng: Nhân viên, Phòng ban, Đơn vị # Join để lấy tên Phòng ban và tên Đơn vị (ma_hieu_2)
    query = db.session.query(ThongTinNguoiLaoDong, PhongBan.ten_phong_ban, DonVi.ten_ma_hieu_2)\
        .join(PhongBan, ThongTinNguoiLaoDong.ma_phong_ban == PhongBan.id)\
        .join(DonVi, ThongTinNguoiLaoDong.ma_hieu_2 == DonVi.ma_hieu_2)\
        .filter(ThongTinNguoiLaoDong.trang_thai == True)

    if ma_hieu_2_filter:
        query = query.filter(ThongTinNguoiLaoDong.ma_hieu_2 == ma_hieu_2_filter)
    if ma_pb_filter:
        query = query.filter(ThongTinNguoiLaoDong.ma_phong_ban == ma_pb_filter)
    
    staff_list = query.order_by(ThongTinNguoiLaoDong.ho_ten.asc()).all()

    # Tối ưu dữ liệu Dropdown (Tối ưu theo quyền)
    if is_system_admin:
        # Lấy tất cả và sắp xếp theo tên đơn vị
        ds_don_vi = DonVi.query.order_by(DonVi.ten_ma_hieu_2.asc()).all()
        # Sắp xếp luôn cả phòng ban để người dùng dễ tìm
        ds_phong_ban = (PhongBan.query.filter_by(ma_hieu_2=ma_hieu_2_filter).order_by(PhongBan.ten_phong_ban.asc()).all() 
                        if ma_hieu_2_filter else PhongBan.query.order_by(PhongBan.ten_phong_ban.asc()).all())
    elif is_phong_tong_hop:
        # Lọc theo đơn vị của user và sắp xếp
        ds_don_vi = DonVi.query.filter_by(ma_hieu_2=user_ma_hieu_2).order_by(DonVi.ten_ma_hieu_2.asc()).all()
        ds_phong_ban = PhongBan.query.filter_by(ma_hieu_2=user_ma_hieu_2).order_by(PhongBan.ten_phong_ban.asc()).all()
    else:
        ds_don_vi = DonVi.query.filter_by(ma_hieu_2=user_ma_hieu_2).order_by(DonVi.ten_ma_hieu_2.asc()).all()
        ds_phong_ban = PhongBan.query.filter_by(id=user_ma_pb).all()
        
    # Lấy dữ liệu chấm công đã có
    staff_ids = [s[0].ma_nhan_vien for s in staff_list]
    records = ThongTinChamCongNgoaiGio.query.filter(
        ThongTinChamCongNgoaiGio.ma_nhan_vien.in_(staff_ids) if staff_ids else text("1=0"),
        ThongTinChamCongNgoaiGio.thang == month,
        ThongTinChamCongNgoaiGio.nam == year
    ).all()
    
    att_dict = {r.ma_nhan_vien: r for r in records}

    # Xác định tên phòng ban hiển thị trên giao diện
    selected_pb_name = "Tất cả phòng ban"
    if ma_pb_filter:
        curr_pb_obj = next((p for p in ds_phong_ban if str(p.id) == str(ma_pb_filter)), None)
        if curr_pb_obj:
            selected_pb_name = curr_pb_obj.ten_phong_ban

    return render_template('cham_cong_ngoai_gio.html', 
                           staff_list=staff_list,
                           att_dict=att_dict,
                           month=month,
                           year=year,
                           don_vis=ds_don_vi, 
                           phong_bans=ds_phong_ban,
                           current_don_vi=ma_hieu_2_filter,
                           current_pb=str(ma_pb_filter) if ma_pb_filter else '',
                           current_pb_name=selected_pb_name,
                           is_admin_or_th=is_admin_or_th,
                           holiday_days=list(holiday_days))

# ----------------------------------------------------------------------
# Bảng tổng hợp Tổng số ngoài giờ theo từng tháng của mỗi nhân viên: Ngày Thường; Ngày Nghỉ; Ngày Lễ
# ----------------------------------------------------------------------
@app.route('/tong-hop-ngoai-gio')
@login_required
def tong_hop_ngoai_gio():
    # --- PHẦN 1: KIỂM TRA QUYỀN TRUY CẬP MENU ---
    if not current_user.is_admin:
        has_perm = UserMenuPermission.query.filter_by(
            ma_nhan_vien=current_user.ma_nhan_vien, 
            menu_slug='tong-hop-ngoai-gio'
        ).first()
        if not has_perm:
            flash("Bạn không có quyền xem Báo cáo tổng hợp ngoài giờ!", "danger")
            return redirect(url_for('index'))
    
    # --- PHẦN 2: LẤY THÔNG TIN ĐƠN VỊ CỦA USER ĐỂ PHÂN PHẠM VI ---
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    user_ma_pb = str(user_info.ma_phong_ban) if user_info and user_info.ma_phong_ban else None
    user_ma_hieu_2 = user_info.ma_hieu_2 if user_info else None
    
    is_phong_tong_hop = (user_ma_pb == '2')
    is_system_admin = (current_user.is_admin == 1)

    # --- PHẦN 3: XỬ LÝ THAM SỐ LỌC ---
    try:
        month = int(request.args.get('month', datetime.now().month))
        year = int(request.args.get('year', datetime.now().year))
    except (ValueError, TypeError):
        month, year = datetime.now().month, datetime.now().year

    ma_hieu_2_filter = request.args.get('ma_hieu_2', '').strip()
    ma_pb_filter = request.args.get('ma_phong_ban', '').strip()
    ma_nv_filter = request.args.get('ma_nhan_vien', '').strip()

    # Lấy danh sách Đơn vị (Mã hiệu 2) cho dropdown nếu là Admin hệ thống (Giống hàm chi tiết)
    don_vis_list = []
    if is_system_admin:
        don_vis_list = [row[0] for row in db.session.query(distinct(ThongTinNguoiLaoDong.ma_hieu_2))\
                        .filter(ThongTinNguoiLaoDong.ma_hieu_2 != None)\
                        .order_by(ThongTinNguoiLaoDong.ma_hieu_2).all()]

    # --- PHẦN 4: XỬ LÝ LOGIC DROPDOWN PHÒNG BAN THEO PHÂN QUYỀN ---
    phong_bans = []
    # Xác định đơn vị thực tế để load phòng ban
    effective_don_vi = ma_hieu_2_filter if is_system_admin else user_ma_hieu_2
    
    if effective_don_vi:
        phong_bans = PhongBan.query.filter_by(ma_hieu_2=effective_don_vi).order_by(PhongBan.ten_phong_ban).all()

    # --- PHẦN 5: XÂY DỰNG TRUY VẤN DỮ LIỆU ---
    query = db.session.query(
        ThongTinNguoiLaoDong.ma_nhan_vien,
        ThongTinNguoiLaoDong.ho_ten,
        ThongTinNguoiLaoDong.ma_hieu_2,
        PhongBan.ten_phong_ban,
        ThongTinChamCongNgoaiGio.tong_gio_ngay_thuong,
        ThongTinChamCongNgoaiGio.tong_gio_ngay_nghi,
        ThongTinChamCongNgoaiGio.tong_gio_ngay_le,
        ThongTinChamCongNgoaiGio.tong_gio_ngoai_gio
    ).join(PhongBan, ThongTinNguoiLaoDong.ma_phong_ban == PhongBan.id)\
     .outerjoin(ThongTinChamCongNgoaiGio, (ThongTinNguoiLaoDong.ma_nhan_vien == ThongTinChamCongNgoaiGio.ma_nhan_vien) & 
                (ThongTinChamCongNgoaiGio.thang == month) & 
                (ThongTinChamCongNgoaiGio.nam == year))\
     .filter(ThongTinNguoiLaoDong.trang_thai == True)

    # --- PHẦN 6: ÁP DỤNG BỘ LỌC PHÂN QUYỀN DỮ LIỆU (Logic giống hàm chi tiết) ---
    if is_system_admin:
        if ma_hieu_2_filter:
            query = query.filter(ThongTinNguoiLaoDong.ma_hieu_2 == ma_hieu_2_filter)
    elif is_phong_tong_hop:
        # Ép buộc chỉ xem được đơn vị của chính mình
        query = query.filter(ThongTinNguoiLaoDong.ma_hieu_2 == user_ma_hieu_2)
    else:
        # User thường: Chỉ thấy bản thân họ
        query = query.filter(ThongTinNguoiLaoDong.ma_nhan_vien == current_user.ma_nhan_vien)

    # Lọc theo Phòng ban (nếu có chọn)
    if ma_pb_filter:
        query = query.filter(ThongTinNguoiLaoDong.ma_phong_ban == ma_pb_filter)
        
    # Lọc theo tên hoặc mã nhân viên (nếu có tìm kiếm)
    if ma_nv_filter:
        query = query.filter(
            (ThongTinNguoiLaoDong.ho_ten.like(f"%{ma_nv_filter}%")) | 
            (ThongTinNguoiLaoDong.ma_nhan_vien.like(f"%{ma_nv_filter}%"))
        )

    results = query.order_by(ThongTinNguoiLaoDong.ma_hieu_2, PhongBan.id, ThongTinNguoiLaoDong.ma_nhan_vien).all()

    # --- PHẦN 7: TRẢ VỀ TEMPLATE ---
    return render_template('tong_hop_ngoai_gio.html', 
                           results=results, 
                           month=month, 
                           year=year,
                           don_vis_list=don_vis_list,
                           phong_bans=phong_bans,
                           current_don_vi=effective_don_vi,
                           current_pb=ma_pb_filter,
                           current_nv=ma_nv_filter,
                           is_system_admin=is_system_admin,
                           is_phong_tong_hop=is_phong_tong_hop)

@app.route('/export-tong-hop-ngoai-gio')
@login_required
def export_tong_hop_ngoai_gio():
    # --- 1. KIỂM TRA QUYỀN TRUY CẬP MENU ---
    if not current_user.is_admin:
        has_perm = UserMenuPermission.query.filter_by(
            ma_nhan_vien=current_user.ma_nhan_vien, 
            menu_slug='tong-hop-ngoai-gio'
        ).first()
        if not has_perm:
            return "Bạn không có quyền thực hiện chức năng này", 403

    # --- 2. LẤY THÔNG TIN USER VÀ THAM SỐ ---
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    user_ma_pb = str(user_info.ma_phong_ban) if user_info and user_info.ma_phong_ban else None
    user_ma_hieu_2 = user_info.ma_hieu_2 if user_info else None
    
    is_admin_system = (current_user.is_admin == 1)
    is_phong_tong_hop = (user_ma_pb == '2')
    
    try:
        month = int(request.args.get('month', datetime.now().month))
        year = int(request.args.get('year', datetime.now().year))
    except (ValueError, TypeError):
        month, year = datetime.now().month, datetime.now().year
    
    # Lấy filter từ request
    ma_hieu_2_filter = request.args.get('ma_hieu_2', '').strip()
    ma_pb_filter = request.args.get('ma_phong_ban', '').strip()
    ma_nv_filter = request.args.get('ma_nhan_vien', '').strip()

    # --- 3. XÂY DỰNG TRUY VẤN (GIỐNG HÀM HIỂN THỊ) ---
    query = db.session.query(
        ThongTinNguoiLaoDong.ma_nhan_vien,
        ThongTinNguoiLaoDong.ho_ten,
        ThongTinNguoiLaoDong.ma_hieu_2,
        PhongBan.ten_phong_ban,
        ThongTinChamCongNgoaiGio.tong_gio_ngay_thuong,
        ThongTinChamCongNgoaiGio.tong_gio_ngay_nghi,
        ThongTinChamCongNgoaiGio.tong_gio_ngay_le,
        ThongTinChamCongNgoaiGio.tong_gio_ngoai_gio
    ).join(PhongBan, ThongTinNguoiLaoDong.ma_phong_ban == PhongBan.id)\
     .outerjoin(ThongTinChamCongNgoaiGio, (ThongTinNguoiLaoDong.ma_nhan_vien == ThongTinChamCongNgoaiGio.ma_nhan_vien) & 
                (ThongTinChamCongNgoaiGio.thang == month) & 
                (ThongTinChamCongNgoaiGio.nam == year))\
     .filter(ThongTinNguoiLaoDong.trang_thai == True)

    # --- 4. ÁP DỤNG BỘ LỌC PHÂN QUYỀN (CHẶN CHÉO DỮ LIỆU) ---
    if is_admin_system:
        if ma_hieu_2_filter:
            query = query.filter(ThongTinNguoiLaoDong.ma_hieu_2 == ma_hieu_2_filter)
    elif is_phong_tong_hop:
        # Ép buộc theo đơn vị của User
        query = query.filter(ThongTinNguoiLaoDong.ma_hieu_2 == user_ma_hieu_2)
    else:
        # User thường chỉ thấy chính mình
        query = query.filter(ThongTinNguoiLaoDong.ma_nhan_vien == current_user.ma_nhan_vien)

    # Bộ lọc bổ sung từ giao diện
    if ma_pb_filter:
        query = query.filter(ThongTinNguoiLaoDong.ma_phong_ban == ma_pb_filter)
    if ma_nv_filter:
        query = query.filter(
            (ThongTinNguoiLaoDong.ho_ten.like(f"%{ma_nv_filter}%")) | 
            (ThongTinNguoiLaoDong.ma_nhan_vien.like(f"%{ma_nv_filter}%"))
        )

    results = query.order_by(ThongTinNguoiLaoDong.ma_hieu_2, PhongBan.id, ThongTinNguoiLaoDong.ma_nhan_vien).all()

    # --- 5. TẠO DATAFRAME VÀ XUẤT EXCEL ---
    data = []
    for r in results:
        data.append({
            "Mã NV": r.ma_nhan_vien,
            "Họ và Tên": r.ho_ten,
            "Đơn vị": r.ma_hieu_2,
            "Phòng Ban": r.ten_phong_ban,
            "Ngày Thường (h)": float(r.tong_gio_ngay_thuong or 0),
            "Ngày Nghỉ (h)": float(r.tong_gio_ngay_nghi or 0),
            "Ngày Lễ (h)": float(r.tong_gio_ngay_le or 0),
            "Tổng Cộng (h)": float(r.tong_gio_ngoai_gio or 0)
        })

    df = pd.DataFrame(data)
    if df.empty:
        # Tạo DF trống với các cột chuẩn để tránh lỗi openpyxl
        df = pd.DataFrame(columns=["Mã NV", "Họ và Tên", "Đơn vị", "Phòng Ban", "Ngày Thường (h)", "Ngày Nghỉ (h)", "Ngày Lễ (h)", "Tổng Cộng (h)"])

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='NgoaiGio')
        worksheet = writer.sheets['NgoaiGio']
        
        # Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Định dạng Header
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Định dạng nội dung (Dòng 2 trở đi)
        for row in worksheet.iter_rows(min_row=2, max_row=len(data)+1):
            for cell in row:
                cell.border = thin_border
                if cell.column >= 5: # Các cột số lượng
                    cell.alignment = center_align
                    cell.number_format = '0.0'

        # Auto-adjust column width (Sửa lỗi column_letter biến chưa định nghĩa)
        for col in worksheet.columns:
            max_length = 0
            column_letter = col[0].column_letter # Lấy chữ cái tên cột (A, B, C...)
            for cell in col:
                try:
                    val_len = len(str(cell.value))
                    if val_len > max_length: max_length = val_len
                except: pass
            worksheet.column_dimensions[column_letter].width = min(max_length + 3, 50)
    
    output.seek(0)

    # --- 6. ĐẶT TÊN FILE VÀ GỬI ---
    # Tên file linh hoạt theo tiêu chí lọc
    prefix = ma_hieu_2_filter if ma_hieu_2_filter else (user_ma_hieu_2 if not is_admin_system else "HE_THONG")
    filename = f"BC_NgoaiGio_{prefix}_T{month}_{year}.xlsx"

    return send_file(
        output, 
        as_attachment=True, 
        download_name=filename, 
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
# ----------------------------------------------------------------------
# Bảng tổng hợp Tổng số ngoài giờ theo năm gồm 12 tháng của mỗi nhân viên: Ngày Thường; Ngày Nghỉ; Ngày Lễ
# ----------------------------------------------------------------------
@app.route('/tong-hop-ngoai-gio-nam')
@login_required
def tong_hop_ngoai_gio_nam():
    # --- PHẦN 1: KIỂM TRA QUYỀN TRUY CẬP MENU ---
    if not current_user.is_admin:
        has_perm = UserMenuPermission.query.filter_by(
            ma_nhan_vien=current_user.ma_nhan_vien, 
            menu_slug='tong-hop-ngoai-gio-nam'
        ).first()
        if not has_perm:
            flash("Bạn không có quyền xem Báo cáo tổng hợp ngoài giờ năm!", "danger")
            return redirect(url_for('index'))

    # --- PHẦN 2: LẤY THÔNG TIN ĐƠN VỊ CỦA USER ĐỂ PHÂN PHẠM VI ---
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    user_ma_pb = str(user_info.ma_phong_ban) if user_info and user_info.ma_phong_ban else None
    user_ma_hieu_2 = user_info.ma_hieu_2 if user_info else None
    
    is_phong_tong_hop = (user_ma_pb == '2')
    is_system_admin = (current_user.is_admin == 1)

    # --- PHẦN 3: XỬ LÝ THAM SỐ LỌC ---
    year = request.args.get('year', default=datetime.now().year, type=int)
    ma_hieu_2_filter = request.args.get('ma_hieu_2', '').strip()
    ma_pb_filter = request.args.get('ma_phong_ban', '')
    ma_nv_filter = request.args.get('ma_nhan_vien', '').strip()

    # Lấy danh sách Đơn vị rút gọn cho dropdown (Dùng cho Admin)
    don_vis_list = []
    if is_system_admin:
        don_vis_list = [row[0] for row in db.session.query(distinct(ThongTinNguoiLaoDong.ma_hieu_2))\
                        .filter(ThongTinNguoiLaoDong.ma_hieu_2 != None)\
                        .order_by(ThongTinNguoiLaoDong.ma_hieu_2).all()]

    # Xác định đơn vị thực tế để load phòng ban và dữ liệu
    effective_don_vi = ma_hieu_2_filter if is_system_admin else user_ma_hieu_2

    # Tải danh sách phòng ban theo đơn vị hiệu dụng
    phong_bans = []
    if effective_don_vi:
        phong_bans = PhongBan.query.filter_by(ma_hieu_2=effective_don_vi).order_by(PhongBan.ten_phong_ban).all()

    # --- PHẦN 4: TRUY VẤN DỮ LIỆU ---
    query = db.session.query(TongHopNgoaiGioNam, ThongTinNguoiLaoDong.ho_ten, ThongTinNguoiLaoDong.ma_hieu_2)\
        .join(ThongTinNguoiLaoDong, TongHopNgoaiGioNam.ma_nhan_vien == ThongTinNguoiLaoDong.ma_nhan_vien)\
        .filter(TongHopNgoaiGioNam.nam == year)

    # Áp dụng logic phân quyền (Security)
    if is_system_admin:
        if ma_hieu_2_filter:
            query = query.filter(ThongTinNguoiLaoDong.ma_hieu_2 == ma_hieu_2_filter)
    elif is_phong_tong_hop:
        query = query.filter(ThongTinNguoiLaoDong.ma_hieu_2 == user_ma_hieu_2)
    else:
        # User thường: Chỉ thấy chính mình
        query = query.filter(TongHopNgoaiGioNam.ma_nhan_vien == current_user.ma_nhan_vien)

    # Bộ lọc bổ sung
    if ma_pb_filter:
        query = query.filter(ThongTinNguoiLaoDong.ma_phong_ban == ma_pb_filter)
    if ma_nv_filter:
        query = query.filter(
            (TongHopNgoaiGioNam.ma_nhan_vien.like(f"%{ma_nv_filter}%")) |
            (ThongTinNguoiLaoDong.ho_ten.like(f"%{ma_nv_filter}%"))
        )
    results = query.order_by(ThongTinNguoiLaoDong.ma_hieu_2, ThongTinNguoiLaoDong.ma_phong_ban, TongHopNgoaiGioNam.ma_nhan_vien).all()

    # --- PHẦN 5: ĐỊNH DẠNG DỮ LIỆU HIỂN THỊ ---
    formatted_data = []
    for item, ho_ten, ma_hieu_2 in results:
        row_dict = {
            'ma_nhan_vien': item.ma_nhan_vien,
            'ho_ten': ho_ten,
            'ma_hieu_2': ma_hieu_2,
            'months': [],
            'tong_nam_thuong': float(item.tong_nam_thuong or 0),
            'tong_nam_nghi': float(item.tong_nam_nghi or 0),
            'tong_nam_le': float(item.tong_nam_le or 0),
            'tong_tat_ca': float(item.tong_tat_ca or 0)
        }
        
        for i in range(1, 13):
            t = float(getattr(item, f'thuong_t{i}') or 0)
            n = float(getattr(item, f'nghi_t{i}') or 0)
            l = float(getattr(item, f'le_t{i}') or 0)
            row_dict['months'].append({
                'thuong': t, 'nghi': n, 'le': l, 'total': round(t + n + l, 1)
            })
        formatted_data.append(row_dict)

    return render_template('tong_hop_ngoai_gio_nam.html', 
                           data=formatted_data, 
                           year=year, 
                           don_vis_list=don_vis_list,
                           phong_bans=phong_bans, 
                           is_system_admin=is_system_admin,
                           is_phong_tong_hop=is_phong_tong_hop,
                           is_admin_or_th=(is_system_admin or is_phong_tong_hop),
                           current_don_vi=effective_don_vi,
                           current_pb=ma_pb_filter, 
                           current_nv=ma_nv_filter)

@app.route('/export-tong-hop-ngoai-gio-nam')
@login_required
def export_tong_hop_ngoai_gio_nam():
    # --- 1. KIỂM TRA QUYỀN TRUY CẬP MENU ---
    if not current_user.is_admin:
        has_perm = UserMenuPermission.query.filter_by(
            ma_nhan_vien=current_user.ma_nhan_vien, 
            menu_slug='tong-hop-ngoai-gio-nam'
        ).first()
        if not has_perm:
            return "Bạn không có quyền thực hiện chức năng này", 403

    # --- 2. LẤY THÔNG TIN USER VÀ THAM SỐ ---
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    user_ma_pb = str(user_info.ma_phong_ban) if user_info and user_info.ma_phong_ban else None
    user_ma_hieu_2 = user_info.ma_hieu_2 if user_info else None
    
    is_admin_system = (current_user.is_admin == 1)
    is_phong_tong_hop = (user_ma_pb == '2')
    
    year = int(request.args.get('year', datetime.now().year))
    ma_hieu_2_filter = request.args.get('ma_hieu_2', '').strip()
    ma_pb_filter = request.args.get('ma_phong_ban', '')
    ma_nv_filter = request.args.get('ma_nhan_vien', '').strip()

    # --- 3. TRUY VẤN DỮ LIỆU ---
    query = db.session.query(TongHopNgoaiGioNam, ThongTinNguoiLaoDong.ho_ten, ThongTinNguoiLaoDong.ma_hieu_2)\
        .join(ThongTinNguoiLaoDong, TongHopNgoaiGioNam.ma_nhan_vien == ThongTinNguoiLaoDong.ma_nhan_vien)\
        .filter(TongHopNgoaiGioNam.nam == year)

    # Phân quyền dữ liệu (Chặn chéo dữ liệu giữa các đơn vị)
    if is_admin_system:
        if ma_hieu_2_filter:
            query = query.filter(ThongTinNguoiLaoDong.ma_hieu_2 == ma_hieu_2_filter)
    elif is_phong_tong_hop:
        query = query.filter(ThongTinNguoiLaoDong.ma_hieu_2 == user_ma_hieu_2)
    else:
        query = query.filter(TongHopNgoaiGioNam.ma_nhan_vien == current_user.ma_nhan_vien)

    # Bộ lọc từ giao diện
    if ma_pb_filter:
        query = query.filter(ThongTinNguoiLaoDong.ma_phong_ban == ma_pb_filter)
    if ma_nv_filter:
        query = query.filter(
            (TongHopNgoaiGioNam.ma_nhan_vien.like(f"%{ma_nv_filter}%")) |
            (ThongTinNguoiLaoDong.ho_ten.like(f"%{ma_nv_filter}%"))
        )

    results = query.order_by(ThongTinNguoiLaoDong.ma_hieu_2, ThongTinNguoiLaoDong.ma_phong_ban, TongHopNgoaiGioNam.ma_nhan_vien).all()

    # --- 4. TẠO DỮ LIỆU EXCEL ---
    data_excel = []
    for row, ho_ten, ma_hieu_2 in results:
        item = {
            "Mã NV": row.ma_nhan_vien,
            "Họ Tên": ho_ten,
            "Đơn vị": ma_hieu_2
        }
        # Dữ liệu từng tháng
        for m in range(1, 13):
            # Lưu ý: Các cột tháng vẫn là thuong_t1, nghi_t1... theo file CSV
            t_val = getattr(row, f'thuong_t{m}') or 0
            n_val = getattr(row, f'nghi_t{m}') or 0
            l_val = getattr(row, f'le_t{m}') or 0
            item[f"T{m}_Thường"] = float(t_val)
            item[f"T{m}_Nghỉ"] = float(n_val)
            item[f"T{m}_Lễ"] = float(l_val)
            item[f"T{m}_Tổng"] = float(t_val + n_val + l_val)
        
        item["Tổng_Thường_Năm"] = float(row.tong_nam_thuong or 0)
        item["Tổng_Nghỉ_Năm"] = float(row.tong_nam_nghi or 0)
        item["Tổng_Lễ_Năm"] = float(row.tong_nam_le or 0)
        item["TỔNG CẢ NĂM"] = float(row.tong_tat_ca or 0)
        data_excel.append(item)

    df = pd.DataFrame(data_excel)
    if df.empty:
        # Tạo cấu trúc cột mặc định nếu không có dữ liệu
        cols = ["Mã NV", "Họ Tên", "Đơn vị"] + [f"T{m}_{x}" for m in range(1,13) for x in ["Thường", "Nghỉ", "Lễ", "Tổng"]] + ["Tổng_Thường_Năm", "Tổng_Nghỉ_Năm", "Tổng_Lễ_Năm", "TỔNG CẢ NĂM"]
        df = pd.DataFrame(columns=cols)
        
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=f'Nam_{year}')
        workbook = writer.book
        worksheet = writer.sheets[f'Nam_{year}']
        
        # Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Định dạng Header và Border cho toàn bộ bảng
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for row_idx in range(2, len(data_excel) + 2):
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                if col_idx > 3: # Các cột số liệu
                    cell.number_format = '0.0'
                    cell.alignment = center_align

        # Auto-adjust column width
        for col in worksheet.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 30)
    
    output.seek(0)

    prefix = ma_hieu_2_filter if ma_hieu_2_filter else (user_ma_hieu_2 if not is_admin_system else "HE_THONG")
    filename = f"BC_NgoaiGio_Nam_{prefix}_{year}.xlsx"

    return send_file(
        output, 
        as_attachment=True, 
        download_name=filename, 
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )






# Hàm này làm tôi loạn cả lên, chỉ là lấy kết quả tính toán từ bảng ket_qua_ngoai_gio_nam_chot
# Hàm này làm tôi loạn cả lên, chỉ là lấy kết quả tính toán từ bảng ket_qua_ngoai_gio_nam_chot
# Hàm này làm tôi loạn cả lên, chỉ là lấy kết quả tính toán từ bảng ket_qua_ngoai_gio_nam_chot
# Hàm này làm tôi loạn cả lên, chỉ là lấy kết quả tính toán từ bảng ket_qua_ngoai_gio_nam_chot
# Hàm này làm tôi loạn cả lên, chỉ là lấy kết quả tính toán từ bảng ket_qua_ngoai_gio_nam_chot
@app.route('/ket-qua-ngoai-gio-nam-chot')
@login_required
def ket_qua_ngoai_gio_nam_chot():
    # --- 1. KIỂM TRA QUYỀN TRUY CẬP MENU ---
    if not current_user.is_admin:
        has_perm = UserMenuPermission.query.filter_by(
            ma_nhan_vien=current_user.ma_nhan_vien, 
            menu_slug='ket-qua-ngoai-gio-nam-chot'
        ).first()
        if not has_perm:
            flash("Bạn không có quyền xem báo cáo chốt tiền ngoài giờ năm!", "danger")
            return redirect(url_for('index'))
    
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    if not user_info:
        return "Không tìm thấy thông tin nhân viên", 404

    user_ma_pb = str(user_info.ma_phong_ban) if user_info.ma_phong_ban else None
    user_ma_hieu_2 = user_info.ma_hieu_2 if user_info else None
    
    is_admin_system = (current_user.is_admin == 1)
    is_phong_tong_hop = (user_ma_pb == '2')
    is_admin_or_th = is_admin_system or is_phong_tong_hop
       
    # Lấy tham số từ URL và làm sạch
    year = request.args.get('year', default=datetime.now().year, type=int)
    ma_hieu_2_filter = request.args.get('ma_hieu_2', '').strip()
    ma_pb_raw = request.args.get('ma_phong_ban', '').strip()
    ma_nv_filter = request.args.get('ma_nhan_vien', '').strip()

    # --- 3. XÁC ĐỊNH PHẠM VI DỮ LIỆU (EFFECTIVE UNIT) ---
    effective_don_vi = ma_hieu_2_filter if is_admin_system else user_ma_hieu_2
    
    # Lấy danh sách Đơn vị rút gọn cho Admin
    don_vis_list = []
    if is_admin_system:
        don_vis_list = [row[0] for row in db.session.query(distinct(ThongTinNguoiLaoDong.ma_hieu_2))\
                        .filter(ThongTinNguoiLaoDong.ma_hieu_2 != None)\
                        .order_by(ThongTinNguoiLaoDong.ma_hieu_2).all()]

    # Tải danh sách phòng ban theo đơn vị hiệu dụng
    phong_bans = []
    if effective_don_vi:
        phong_bans = PhongBan.query.filter_by(ma_hieu_2=effective_don_vi).order_by(PhongBan.ten_phong_ban).all()
    else:
        phong_bans = PhongBan.query.all()

    # TRUY VẤN
    # --- 4. XÂY DỰNG TRUY VẤN ---
    query = db.session.query(KetQuaNgoaiGioNamChot)\
        .join(ThongTinNguoiLaoDong, KetQuaNgoaiGioNamChot.ma_nhan_vien == ThongTinNguoiLaoDong.ma_nhan_vien)\
        .filter(KetQuaNgoaiGioNamChot.nam == year)

    # Áp dụng logic lọc Đơn vị (Mã hiệu 2)
    if is_admin_system:
        if ma_hieu_2_filter:
            query = query.filter(ThongTinNguoiLaoDong.ma_hieu_2 == ma_hieu_2_filter)
    else:
        # Ép buộc theo đơn vị của user (Bảo mật chéo)
        query = query.filter(ThongTinNguoiLaoDong.ma_hieu_2 == user_ma_hieu_2)

    # Áp dụng logic lọc Phòng ban
    current_pb = "all"
    if is_admin_or_th:
        if ma_pb_raw and ma_pb_raw != "all":
            current_pb = ma_pb_raw
            query = query.filter(cast(ThongTinNguoiLaoDong.ma_phong_ban, String) == current_pb)
    else:
        # User thường chỉ thấy chính mình hoặc trong phòng mình
        current_pb = user_ma_pb
        query = query.filter(cast(ThongTinNguoiLaoDong.ma_phong_ban, String) == current_pb)
        if not is_phong_tong_hop:
             query = query.filter(KetQuaNgoaiGioNamChot.ma_nhan_vien == current_user.ma_nhan_vien)

    if ma_nv_filter:
        query = query.filter(
            (KetQuaNgoaiGioNamChot.ma_nhan_vien.like(f"%{ma_nv_filter}%")) |
            (KetQuaNgoaiGioNamChot.ho_ten.like(f"%{ma_nv_filter}%"))
        )

    # --- 5. FORMAT DỮ LIỆU ---
    results = query.order_by(ThongTinNguoiLaoDong.ma_hieu_2, ThongTinNguoiLaoDong.ma_phong_ban, KetQuaNgoaiGioNamChot.ma_nhan_vien).all()
    
    # ... [Phần vòng lặp formatted_data giữ nguyên] ...
    formatted_data = []
    for item in results:
        row_dict = {
            'ma_nhan_vien': item.ma_nhan_vien, 'ho_ten': item.ho_ten, 'months': [],
            'tong_gio_thuong': float(item.tong_gio_thuong or 0),
            'tong_gio_nghi': float(item.tong_gio_nghi or 0),
            'tong_gio_le': float(item.tong_gio_le or 0),
            'tong_tien_nam': float(item.tong_tien_nam or 0)
        }
        for i in range(1, 13):
            row_dict['months'].append({
                'tien_thuong': float(getattr(item, f'tien_ngoai_gio_thuong_t{i}') or 0),
                'tien_nghi': float(getattr(item, f'tien_ngoai_gio_nghi_t{i}') or 0),
                'tien_le': float(getattr(item, f'tien_ngoai_gio_le_t{i}') or 0)
            })
        formatted_data.append(row_dict)


    # Thêm tên phòng ban hiện tại để hiển thị lên tiêu đề
    ten_phong_hien_tai = "Tất cả phòng ban"
    if current_pb != "all":
        pb_obj = next((p for p in phong_bans if str(p.id) == current_pb), None)
        if pb_obj: 
            ten_phong_hien_tai = pb_obj.ten_phong_ban

    return render_template('ket_qua_ngoai_gio_nam_chot.html', 
                           data=formatted_data, 
                           year=year, 
                           don_vis_list=don_vis_list,
                           phong_bans=phong_bans, 
                           is_admin_or_th=is_admin_or_th,
                           is_system_admin=is_admin_system,
                           current_don_vi=effective_don_vi,
                           current_pb=current_pb, 
                           current_nv=ma_nv_filter,
                           ten_phong_hien_tai=ten_phong_hien_tai)

#Trong mysql là bảng ket_qua_ngoai_gio_nam_chot
@app.route('/kich-hoat-tinh-ngoai-gio/<int:year>')
@login_required
def kich_hoat_tinh_ngoai_gio(year):
    # 1. Lấy thông tin đơn vị muốn xử lý (từ URL hoặc từ chính người dùng)
    # Nếu admin truyền ?ma_hieu_2=... thì lấy giá trị đó, nếu không lấy của chính họ
    requested_ma_hieu_2 = request.args.get('ma_hieu_2')

    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    if not user_info:
        flash("Không tìm thấy thông tin người dùng trong cơ sở dữ liệu nhân sự.", "danger")
        return redirect(url_for('index'))

    user_ma_hieu_2 = user_info.ma_hieu_2
    user_ma_pb = str(user_info.ma_phong_ban).strip()

    # 2. KIỂM TRA QUYỀN TRUY CẬP (Access Control)
    is_admin_system = (current_user.is_admin == 1)
    is_tong_hop_don_vi = (user_ma_pb == '2')

    # Xác định đơn vị mục tiêu
    target_ma_hieu_2 = requested_ma_hieu_2 if (is_admin_system and requested_ma_hieu_2) else user_ma_hieu_2

    # Ngăn chặn trường hợp user thường cố tình nhập ma_hieu_2 của đơn vị khác trên URL
    if not is_admin_system and requested_ma_hieu_2 and requested_ma_hieu_2 != user_ma_hieu_2:
        flash("Bạn không có quyền can thiệp vào dữ liệu của đơn vị khác.", "danger")
        return redirect(url_for('ket_qua_ngoai_gio_nam_chot', year=year))

    if not (is_admin_system or is_tong_hop_don_vi):
        flash("Chỉ bộ phận Tổng hợp mới có quyền kích hoạt tính toán.", "danger")
        return redirect(url_for('ket_qua_ngoai_gio_nam_chot', year=year))

    try:
        # Bước 1: Sao lưu (Lưu ý: Nếu sp_Backup... chạy quá lâu, nên xem xét lọc theo đơn vị)
        db.session.execute(text("CALL sp_BackupKetQuaNgoaiGio()"))
        
        # Bước 2: Tổng hợp giờ (Đã lọc theo target_ma_hieu_2)
        db.session.execute(
            text("CALL sp_ChayTongHopNgoaiGio_TheoDonVi(:p_nam, :p_ma_hieu_2)"),
            {'p_nam': year, 'p_ma_hieu_2': target_ma_hieu_2}
        )

        # Bước 3: Tính tiền (Đã lọc theo target_ma_hieu_2)
        db.session.execute(
            text("CALL sp_CapNhatKetQuaNgoaiGio_TheoDonVi(:p_nam, :p_ma_hieu_2)"),
            {'p_nam': year, 'p_ma_hieu_2': target_ma_hieu_2}
        )
        
        db.session.commit()
        flash(f"Thành công: Đã cập nhật dữ liệu cho đơn vị {target_ma_hieu_2} năm {year}.", "success")
        
    except Exception as e:
        db.session.rollback()
        # Không hiển thị lỗi SQL thuần cho người dùng để bảo mật, chỉ hiện thông báo chung
        error_log = str(e)
        flash(f"Lỗi khi xử lý đơn vị {target_ma_hieu_2}. Vui lòng liên hệ Admin kỹ thuật.", "danger")
        print(f"FAILED: Unit {target_ma_hieu_2}, Year {year}. Error: {error_log}")
        # print(traceback.format_exc()) # Dùng để debug
        
    return redirect(url_for('ket_qua_ngoai_gio_nam_chot', year=year, ma_hieu_2=target_ma_hieu_2))
# ----------------------------------------------------------------------
# Theo dõi lớp học, theo dõi đào tạo
# ----------------------------------------------------------------------
@app.route('/admin/ma-hieu-lop', methods=['GET', 'POST'])
@login_required
def admin_ma_hieu_lop():
    # 1. Xác định danh sách user đặc biệt được phép truy cập
    allowed_users = ['200905691', '200900615'] # Thêm mã nhân viên của user vào đây: Đinh Thị Ánh Hồng; Phạm Quỳnh Anh

    # 2. Kiểm tra quyền: Phải là Admin HOẶC nằm trong danh sách được phép
    is_allowed = current_user.is_admin or current_user.ma_nhan_vien in allowed_users

    if not is_allowed:
        flash("Bạn không có quyền truy cập trang này!", "danger")
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            ma_hieu = request.form.get('ma_hieu')
            ten_lop = request.form.get('ten_lop_hoc')
            noi_dung = request.form.get('noi_dung_hoc')
            
            # Kiểm tra trùng mã hiệu
            ton_tai = MaHieuLopHoc.query.filter_by(ma_hieu=ma_hieu).first()
            if ton_tai:
                flash(f"Mã hiệu {ma_hieu} đã tồn tại!", "warning")
            else:
                moi = MaHieuLopHoc(ma_hieu=ma_hieu, ten_lop_hoc=ten_lop, noi_dung_hoc=noi_dung)
                db.session.add(moi)
                db.session.commit()
                flash("Thêm mã hiệu lớp học thành công!", "success")
                
        elif action == 'delete':
            # Chỉ Admin mới được quyền xóa thực sự
            if not current_user.is_admin:
                flash("Chỉ Admin mới có quyền xóa mã hiệu!", "danger")
            else:
                id_xoa = request.form.get('id')
                item = MaHieuLopHoc.query.get(id_xoa)
                if item:
                    db.session.delete(item)
                    db.session.commit()
                    flash("Đã xóa mã hiệu lớp học!", "info")
        
        return redirect(url_for('admin_ma_hieu_lop'))

    danh_sach = MaHieuLopHoc.query.order_by(MaHieuLopHoc.created_at.desc()).all()
    #danh_sach = MaHieuLopHoc.query.order_by(MaHieuLopHoc.ma_hieu).all()
    return render_template('admin_ma_hieu_lop.html', danh_sach=danh_sach)


@app.route('/theo_doi_lop_hoc')
@login_required
def theo_doi_lop_hoc():
    # 1. LẤY THÔNG TIN CƠ BẢN CỦA USER ĐANG ĐĂNG NHẬP
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    user_ma_pb = str(user_info.ma_phong_ban) if user_info and user_info.ma_phong_ban else None
    user_ma_hieu_2 = user_info.ma_hieu_2 #if user_info else None
    
    is_system_admin = (current_user.is_admin == 1)
    is_phong_tong_hop = (user_ma_pb == '2') # Phòng Tổng hợp có mã phòng ban là '2'
    is_admin_or_th = (is_system_admin or is_phong_tong_hop)

    all_employees = []
    ds_don_vi = []
    ds_phong_ban = []
    ds_lop_hoc = []

    filter_tu_ngay = request.args.get('tu_ngay')
    filter_den_ngay = request.args.get('den_ngay')
        
    try:
        # 2. THIẾT LẬP PHẠM VI NHÂN VIÊN (Cho dropdown Thêm mới)
        if is_system_admin:
            all_employees = ThongTinNguoiLaoDong.query.filter_by(trang_thai=True).order_by(ThongTinNguoiLaoDong.ho_ten.asc()).all()
            ds_don_vi = DonVi.query.order_by(DonVi.ten_ma_hieu_2.asc()).all()
        elif is_phong_tong_hop:
            all_employees = ThongTinNguoiLaoDong.query.filter_by(ma_hieu_2=user_ma_hieu_2, trang_thai=True).order_by(ThongTinNguoiLaoDong.ho_ten.asc()).all()
        else:
            all_employees = ThongTinNguoiLaoDong.query.filter_by(ma_phong_ban=user_ma_pb, trang_thai=True).order_by(ThongTinNguoiLaoDong.ho_ten.asc()).all()

        # 3. TRUY VẤN DANH SÁCH LỚP HỌC (ÁP DỤNG PHÂN QUYỀN HIỂN THỊ)
        query_lop = db.session.query(
            DanhSachLopHoc, 
            ThongTinNguoiLaoDong.ho_ten,
            ThongTinNguoiLaoDong.ma_hieu_2,
            ThongTinNguoiLaoDong.ma_phong_ban
        ).join(
            ThongTinNguoiLaoDong, 
            DanhSachLopHoc.ma_nhan_vien == ThongTinNguoiLaoDong.ma_nhan_vien
        )

        # --- BẮT ĐẦU PHẦN LỌC THEO THỜI GIAN ---
        if filter_tu_ngay:
            # Lọc các lớp có ngày bắt đầu lớn hơn hoặc bằng tu_ngay
            query_lop = query_lop.filter(DanhSachLopHoc.tu_ngay >= filter_tu_ngay)
        
        if filter_den_ngay:
            # Lọc các lớp có ngày bắt đầu nhỏ hơn hoặc bằng den_ngay
            # (Hoặc lọc theo den_ngay của lớp tùy theo nghiệp vụ của bạn)
            query_lop = query_lop.filter(DanhSachLopHoc.tu_ngay <= filter_den_ngay)
        # --- KẾT THÚC PHẦN LỌC ---
        
        # Lọc dữ liệu hiển thị dựa trên role
        if is_system_admin:
            # Admin: Không thêm filter (thấy toàn bộ)
            pass
        #elif is_phong_tong_hop:
        #    # Phòng Tổng hợp (User tổng hợp): Thấy toàn bộ nhân viên thuộc cùng đơn vị (ma_hieu_2)
        #    query_lop = query_lop.filter(ThongTinNguoiLaoDong.ma_hieu_2 == user_ma_hieu_2)
        else:
            # TẤT CẢ CÁC USER CÒN LẠI (bao gồm Lập bảng, Tổng hợp): 
            # BẮT BUỘC phải cùng đơn vị (ma_hieu_2)
            query_lop = query_lop.filter(ThongTinNguoiLaoDong.ma_hieu_2 == user_ma_hieu_2)
            # Nếu không phải Phòng Tổng hợp (tức là user đơn vị/phòng ban bình thường)
            # thì lọc thêm theo mã phòng ban
            if not is_phong_tong_hop:
                if user_ma_pb:
                    query_lop = query_lop.filter(ThongTinNguoiLaoDong.ma_phong_ban == user_ma_pb)
                else:
                    query_lop = query_lop.filter(False) # Không có phòng ban thì không thấy gì

        results = query_lop.order_by(DanhSachLopHoc.tu_ngay.desc()).all()
        # VÒNG LẶP XỬ LÝ DỮ LIỆU ĐÃ ĐƯỢC FIX LỖI NONETYPE
        for item in results:
            # Kiểm tra nếu bản ghi trống hoặc đối tượng lớp học bị None thì bỏ qua (bảo vệ code)
            if not item or item[0] is None:
                continue
                
            # Unpack dữ liệu sau khi đảm bảo an toàn
            lop, ho_ten, nv_ma_hieu_2, nv_ma_pb = item
            
            lop.fullname_display = ho_ten if ho_ten else "N/A"
            
            # Gán thêm cờ can_edit để Frontend ẩn/hiện nút Sửa/Xóa
            if is_system_admin:
                lop.can_edit = True
            elif is_phong_tong_hop:
                lop.can_edit = (nv_ma_hieu_2 == user_ma_hieu_2)
            else:
                lop.can_edit = (str(nv_ma_pb) == user_ma_pb)
                
            ds_lop_hoc.append(lop)

        # 4. LẤY DỮ LIỆU DANH MỤC BỔ TRỢ
        all_ma_hieu = MaHieuLopHoc.query.order_by(MaHieuLopHoc.ma_hieu).all()
        all_linh_vuc = LinhVuc.query.order_by(LinhVuc.ten_linh_vuc.asc()).all()
        
        return render_template('theo_doi_lop_hoc.html', 
                               ds_lop_hoc=ds_lop_hoc, 
                               all_employees=all_employees,
                               don_vis=ds_don_vi,
                               phong_bans=ds_phong_ban,
                               all_ma_hieu=all_ma_hieu,
                               all_linh_vuc=all_linh_vuc,
                               is_admin_or_th=is_admin_or_th,
                               user_ma_pb=user_ma_pb,
                               user_ma_hieu_2=user_ma_hieu_2,
                               tu_ngay_val=filter_tu_ngay,
                               den_ngay_val=filter_den_ngay)

    except Exception as e:
        db.session.rollback()
        print(f"Lỗi hiển thị danh sách lớp học: {traceback.format_exc()}")
        flash(f"Lỗi hệ thống: {str(e)}", "danger")
        return redirect(url_for('index'))


@app.route('/add_lop_hoc', methods=['POST'])
@login_required
def add_lop_hoc():
    try:
        # 1. Lấy danh sách mã nhân viên từ form
        # 'ma_nhan_vien[]' khớp với thuộc tính 'name' của thẻ select trong HTML
        danh_sach_ma_nv = request.form.getlist('ma_nhan_vien[]')
        
        if not danh_sach_ma_nv:
            flash("Vui lòng chọn ít nhất một nhân viên!", "warning")
            return redirect(url_for('theo_doi_lop_hoc'))
        
        # 2. Kiểm tra quyền hạn của người đang thao tác
        user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
        if not user_info:
            flash("Không tìm thấy thông tin tài khoản người dùng!", "danger")
            return redirect(url_for('theo_doi_lop_hoc'))

        is_system_admin = (current_user.is_admin == 1)
        user_ma_pb = str(user_info.ma_phong_ban) if user_info.ma_phong_ban else None
        user_ma_hieu_2 = user_info.ma_hieu_2
        
        # Phòng Tổng hợp có mã phòng ban là '2'
        is_phong_tong_hop = (user_ma_pb == '2')

        # 3. Lấy thông tin mã hiệu lớp học chung cho cả nhóm
        ma_hieu_da_chon = request.form.get('ma_hieu')
        thong_tin_goc = MaHieuLopHoc.query.filter_by(ma_hieu=ma_hieu_da_chon).first()
        
        if not thong_tin_goc:
            flash("Mã hiệu lớp học không hợp lệ hoặc không tồn tại!", "danger")
            return redirect(url_for('theo_doi_lop_hoc'))

        # 4. Xử lý các giá trị thời gian và thông tin chung
        tu_ngay_str = request.form.get('tu_ngay')
        if not tu_ngay_str:
            flash("Vui lòng nhập ngày bắt đầu!", "warning")
            return redirect(url_for('theo_doi_lop_hoc'))
        tu_ngay_val = datetime.strptime(tu_ngay_str, '%Y-%m-%d').date()
        
        so_ngay_str = request.form.get('so_ngay', '1.0')
        so_ngay_val = float(request.form.get('so_ngay', 1.0))
        # Logic tính Đến ngày: tu_ngay + ceil(so_ngay) - 1
        days_to_add = max(0, math.ceil(so_ngay_val) - 1)
        den_ngay_val = tu_ngay_val + timedelta(days=days_to_add)

        # 5. Lặp qua từng nhân viên để kiểm tra an ninh và tạo bản ghi
        count_success = 0
        for ma_nv in danh_sach_ma_nv:
            ma_nv = ma_nv.strip().upper()
            nv_target = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=ma_nv).first()
            
            if not nv_target:
                continue

            # Kiểm tra an ninh: Không cho phép thêm nhân viên ngoài phạm vi quản lý
            if not is_system_admin:
                if is_phong_tong_hop:
                    # Phòng Tổng hợp chỉ được thêm người trong cùng ĐƠN VỊ (ma_hieu_2)
                    if nv_target.ma_hieu_2 != user_ma_hieu_2:
                        flash(f"Lỗi: Nhân viên {nv_target.ho_ten} không thuộc đơn vị của bạn!", "danger")
                        return redirect(url_for('theo_doi_lop_hoc'))
                else:
                    # User thường chỉ được thêm người trong cùng PHÒNG BAN (ma_phong_ban)
                    if str(nv_target.ma_phong_ban) != user_ma_pb:
                        flash(f"Lỗi: Nhân viên {nv_target.ho_ten} không thuộc phòng ban của bạn!", "danger")
                        db.session.rollback()
                        return redirect(url_for('theo_doi_lop_hoc'))

            # Tạo bản ghi mới
            new_record = DanhSachLopHoc(
                ma_nhan_vien=ma_nv.strip().upper(),
                ma_hieu=ma_hieu_da_chon,
                ten_lop_hoc=thong_tin_goc.ten_lop_hoc,
                noi_dung_hoc=thong_tin_goc.noi_dung_hoc,
                hinh_thuc_hoc=request.form.get('hinh_thuc_hoc'),
                linh_vuc=request.form.get('linh_vuc'),
                don_vi_dau_moi=request.form.get('don_vi_dau_moi'),
                don_vi_to_chuc=request.form.get('don_vi_to_chuc'),
                nguon_kinh_phi=request.form.get('nguon_kinh_phi'),
                tu_ngay=tu_ngay_val,
                den_ngay=den_ngay_val,
                so_ngay=so_ngay_val,
                dia_diem=request.form.get('dia_diem'),
                ghi_chu=request.form.get('ghi_chu'),
                trang_thai='Tham gia' # Mặc định khi thêm mới
            )
            db.session.add(new_record)
            count_success += 1

        # 6. Lưu tất cả thay đổi vào cơ sở dữ liệu
        if count_success > 0:
            db.session.commit()
            flash(f"Đã lưu thành công thông tin đào tạo cho {count_success} nhân viên.", "success")
        else:
            flash("Không có nhân viên hợp lệ để lưu!", "warning")

        
    except Exception as e:
        db.session.rollback()
        # In lỗi chi tiết ra console để debug
        print(f"Lỗi tại add_lop_hoc: {traceback.format_exc()}")
        flash("Đã xảy ra lỗi trong quá trình lưu dữ liệu. Vui lòng kiểm tra lại định dạng nhập liệu!", "danger")
        
    return redirect(url_for('theo_doi_lop_hoc'))


@app.route('/edit_lop_hoc/<int:id>', methods=['POST'])
@login_required
def edit_lop_hoc(id):
    # Lớp 1: Kiểm tra vai trò cơ bản
    if not (current_user.is_admin == 1 or current_user.role == 'LAP_BANG'):
        flash("Bạn không có quyền thực hiện chức năng này!", "danger")
        return redirect(url_for('theo_doi_lop_hoc'))

    # Lấy bản ghi cần sửa hoặc 404
    record = DanhSachLopHoc.query.get_or_404(id)

    try:
        # 1. LẤY THÔNG TIN NGƯỜI THỰC HIỆN (Gộp chung để dùng xuyên suốt)
        user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
        if not user_info:
            flash("Không tìm thấy thông tin tài khoản!", "danger")
            return redirect(url_for('theo_doi_lop_hoc'))

        is_system_admin = (current_user.is_admin == 1)
        user_ma_hieu_2 = user_info.ma_hieu_2
        user_ma_pb = str(user_info.ma_phong_ban) if user_info.ma_phong_ban else None
        is_phong_tong_hop = (user_ma_pb == '2')

        # 2. KIỂM TRA QUYỀN TRÊN BẢN GHI GỐC (Chống sửa chéo đơn vị qua ID)
        if not is_system_admin:
            # Truy vấn thông tin nhân viên cũ đang gắn với lớp học này
            nv_hien_tai = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=record.ma_nhan_vien).first()
            if not nv_hien_tai:
                flash("Dữ liệu gốc của lớp học không hợp lệ!", "danger")
                return redirect(url_for('theo_doi_lop_hoc'))

            if is_phong_tong_hop:
                if nv_hien_tai.ma_hieu_2 != user_ma_hieu_2:
                    flash("Bạn không có quyền sửa dữ liệu của đơn vị khác!", "danger")
                    return redirect(url_for('theo_doi_lop_hoc'))
            else:
                if str(nv_hien_tai.ma_phong_ban) != user_ma_pb:
                    flash("Bạn không có quyền sửa dữ liệu của phòng ban khác!", "danger")
                    return redirect(url_for('theo_doi_lop_hoc'))

        # 3. KIỂM TRA THÔNG TIN MỚI TỪ FORM
        ma_hieu_moi = request.form.get('ma_hieu')
        thong_tin_goc = MaHieuLopHoc.query.filter_by(ma_hieu=ma_hieu_moi).first()
        if not thong_tin_goc:
            flash("Mã hiệu lớp học không hợp lệ!", "danger")
            return redirect(url_for('theo_doi_lop_hoc'))

        # Kiểm tra nhân viên mục tiêu mới (nv_target)
        ma_nv_moi = request.form.get('ma_nhan_vien', '').strip().upper()
        nv_target = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=ma_nv_moi, trang_thai=True).first()
        if not nv_target:
            flash("Nhân viên mới không tồn tại hoặc đã nghỉ việc!", "danger")
            return redirect(url_for('theo_doi_lop_hoc'))

        # Lớp 3: Kiểm tra nhân viên mới có thuộc phạm vi quản lý không
        if not is_system_admin:
            if is_phong_tong_hop:
                if nv_target.ma_hieu_2 != user_ma_hieu_2:
                    flash(f"Nhân viên {nv_target.ho_ten} không thuộc đơn vị của bạn!", "danger")
                    return redirect(url_for('theo_doi_lop_hoc'))
            else:
                if str(nv_target.ma_phong_ban) != user_ma_pb:
                    flash(f"Nhân viên {nv_target.ho_ten} không thuộc phòng ban của bạn!", "danger")
                    return redirect(url_for('theo_doi_lop_hoc'))

        # 4. LOGIC TÍNH TOÁN VÀ CẬP NHẬT
        tu_ngay_str = request.form.get('tu_ngay')
        so_ngay_val = float(request.form.get('so_ngay', 1.0))
        tu_ngay_val = datetime.strptime(tu_ngay_str, '%Y-%m-%d').date()
        
        days_to_add = max(0, math.ceil(so_ngay_val) - 1)
        den_ngay_val = tu_ngay_val + timedelta(days=days_to_add)

        # Gán dữ liệu
        record.trang_thai = request.form.get('trang_thai', record.trang_thai)
        record.ly_do_vang = request.form.get('ly_do_vang', record.ly_do_vang)
        
        record.ma_nhan_vien = ma_nv_moi
        record.ma_hieu = ma_hieu_moi
        record.ten_lop_hoc = thong_tin_goc.ten_lop_hoc
        record.noi_dung_hoc = thong_tin_goc.noi_dung_hoc
        record.hinh_thuc_hoc = request.form.get('hinh_thuc_hoc')
        record.tu_ngay = tu_ngay_val
        record.den_ngay = den_ngay_val
        record.so_ngay = so_ngay_val
        record.dia_diem = request.form.get('dia_diem')
        record.ghi_chu = request.form.get('ghi_chu')
        record.linh_vuc = request.form.get('linh_vuc')
        record.don_vi_dau_moi = request.form.get('don_vi_dau_moi')
        record.don_vi_to_chuc = request.form.get('don_vi_to_chuc')
        record.nguon_kinh_phi = request.form.get('nguon_kinh_phi')

        db.session.commit()
        flash("Cập nhật thông tin lớp học thành công!", "success")

    except Exception as e:
        db.session.rollback()
        print(f"Lỗi cập nhật lớp học: {traceback.format_exc()}")
        flash(f"Lỗi hệ thống khi cập nhật: {str(e)}", "danger")
        
    return redirect(url_for('theo_doi_lop_hoc'))

@app.route('/delete_lop_hoc/<int:id>')
@login_required
def delete_lop_hoc(id):
    if not (current_user.is_admin):
        flash("Bạn không có quyền xóa!", "danger")
        return redirect(url_for('theo_doi_lop_hoc'))

    record = DanhSachLopHoc.query.get_or_404(id)
    try:
        db.session.delete(record)
        db.session.commit()
        flash("Đã xóa bản ghi thành công!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Lỗi khi xóa: {str(e)}", "danger")
        
    return redirect(url_for('theo_doi_lop_hoc'))


# --- ROUTE MỚI: XỬ LÝ NHANH TRẠNG THÁI VẮNG MẶT ---
@app.route('/update_trang_thai_lop/<int:id>', methods=['POST'])
@login_required
def update_trang_thai_lop(id):
    # 1. Lấy bản ghi hoặc trả về 404
    record = DanhSachLopHoc.query.get_or_404(id)
    
    try:
        # 2. KIỂM TRA QUYỀN HẠN (Security Check tương tự edit_lop_hoc)
        user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
        if not user_info:
            return jsonify({"success": False, "message": "Không tìm thấy thông tin tài khoản!"}), 403

        is_system_admin = (current_user.is_admin == 1)
        user_ma_hieu_2 = user_info.ma_hieu_2
        user_ma_pb = str(user_info.ma_phong_ban) if user_info.ma_phong_ban else None
        is_phong_tong_hop = (user_ma_pb == '2')

        if not is_system_admin:
            # Truy vấn nhân viên gắn với bản ghi này để kiểm tra chéo đơn vị
            nv_hien_tai = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=record.ma_nhan_vien).first()
            if not nv_hien_tai:
                return jsonify({"success": False, "message": "Dữ liệu nhân viên không hợp lệ!"}), 400

            if is_phong_tong_hop:
                if nv_hien_tai.ma_hieu_2 != user_ma_hieu_2:
                    return jsonify({"success": False, "message": "Bạn không có quyền sửa dữ liệu của đơn vị khác!"}), 403
            else:
                if str(nv_hien_tai.ma_phong_ban) != user_ma_pb:
                    return jsonify({"success": False, "message": "Bạn không có quyền sửa dữ liệu của phòng ban khác!"}), 403

        # 3. CẬP NHẬT DỮ LIỆU
        trang_thai_moi = request.form.get('trang_thai')
        ly_do_vang = request.form.get('ly_do_vang', '').strip()

        record.trang_thai = trang_thai_moi
        record.ly_do_vang = ly_do_vang

        # Nếu vắng mặt, có thể reset số ngày học về 0 (tùy nhu cầu nghiệp vụ của bạn)
        if trang_thai_moi == 'Vắng mặt':
            record.so_ngay = 0
        else:
            # Nếu chuyển lại thành Tham gia, bạn có thể cần tính lại so_ngay từ tu_ngay/den_ngay
            # Hoặc giữ nguyên giá trị cũ nếu không muốn làm phức tạp thêm logic
            pass

        db.session.commit()
        return jsonify({"success": True, "message": "Đã cập nhật trạng thái tham gia thành công!"})

    except Exception as e:
        db.session.rollback()
        print(f"Lỗi update_trang_thai_lop: {traceback.format_exc()}")
        return jsonify({"success": False, "message": f"Lỗi hệ thống: {str(e)}"}), 500

@app.route('/export_lop_hoc_excel')
@login_required
def export_lop_hoc_excel():
    # Thêm dòng này để debug
    print(f"DEBUG: tu_ngay={request.args.get('tu_ngay')}, den_ngay={request.args.get('den_ngay')}")
    try:
        # 1. LẤY CÁC THAM SỐ LỌC TỪ URL (Nếu có)
        # Các tham số này nên khớp với tên 'name' của các thẻ input/select trên giao diện
        filter_ma_hieu_2 = request.args.get('ma_hieu_2')
        filter_ma_pb = request.args.get('ma_pb')
        filter_ma_hieu = request.args.get('ma_hieu')
        filter_linh_vuc = request.args.get('linh_vuc')
        filter_tu_ngay = request.args.get('tu_ngay')
        filter_den_ngay = request.args.get('den_ngay')
        
        # 2. LẤY THÔNG TIN USER ĐỂ PHÂN QUYỀN TRUY XUẤT
        user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
        if not user_info:
            flash("Không tìm thấy thông tin người dùng!", "danger")
            return redirect(url_for('theo_doi_lop_hoc'))

        is_system_admin = (current_user.is_admin == 1)
        user_ma_pb = str(user_info.ma_phong_ban) if user_info.ma_phong_ban else None
        user_ma_hieu_2 = user_info.ma_hieu_2
        is_phong_tong_hop = (user_ma_pb == '2')

        # 3. KHỞI TẠO QUERY GỐC
        query = db.session.query(
            DanhSachLopHoc, 
            ThongTinNguoiLaoDong.ho_ten,
            ThongTinNguoiLaoDong.ma_hieu_2,
            ThongTinNguoiLaoDong.ma_phong_ban
        ).outerjoin(
            ThongTinNguoiLaoDong, 
            DanhSachLopHoc.ma_nhan_vien == ThongTinNguoiLaoDong.ma_nhan_vien
        )

        # 4. ÁP DỤNG BỘ LỌC BẢO MẬT PHÂN QUYỀN
        if not is_system_admin:
            if is_phong_tong_hop:
                # Phòng Tổng hợp: Chỉ xuất dữ liệu nhân viên cùng ĐƠN VỊ # Chỉ được xuất trong đơn vị của mình
                query = query.filter(ThongTinNguoiLaoDong.ma_hieu_2 == user_ma_hieu_2)
            else:
                # User thường: Chỉ xuất dữ liệu nhân viên cùng PHÒNG BAN
                query = query.filter(ThongTinNguoiLaoDong.ma_phong_ban == user_ma_pb)

        #####results = query.order_by(DanhSachLopHoc.tu_ngay.desc()).all()

        # 5. ÁP DỤNG CÁC BỘ LỌC TÙY CHỌN TỪ GIAO DIỆN ; CHUẨN BỊ DỮ LIỆU CHO DATAFRAME
        if filter_ma_hieu_2:
            query = query.filter(ThongTinNguoiLaoDong.ma_hieu_2 == filter_ma_hieu_2)
        if filter_ma_pb:
            query = query.filter(ThongTinNguoiLaoDong.ma_phong_ban == filter_ma_pb)
        if filter_ma_hieu:
            query = query.filter(DanhSachLopHoc.ma_hieu == filter_ma_hieu)
        if filter_linh_vuc:
            query = query.filter(DanhSachLopHoc.linh_vuc == filter_linh_vuc)

        if filter_tu_ngay:
            query = query.filter(DanhSachLopHoc.tu_ngay >= filter_tu_ngay)
        if filter_den_ngay:
            query = query.filter(DanhSachLopHoc.tu_ngay <= filter_den_ngay)

        results = query.order_by(DanhSachLopHoc.tu_ngay.desc()).all()

        # 6. CHUẨN BỊ DỮ LIỆU CHO DATAFRAME
        data = []
        for i, (lop, ho_ten, ma_hieu_2, ma_pb) in enumerate(results, 1):
            data.append({
                "TT": i,
                "Mã Nhân viên": lop.ma_nhan_vien,
                "Họ và Tên": ho_ten or "N/A",
                "Đơn vị": ma_hieu_2 or "",
                "Mã Phòng ban": ma_pb or "",
                "Trạng thái": lop.trang_thai, 
                "Lý do vắng": lop.ly_do_vang or "", 
                "Tên chương trình": lop.ten_lop_hoc,
                "Nội dung đào tạo": lop.noi_dung_hoc or "",
                "Thời lượng (ngày)": lop.so_ngay or 0,
                "Từ ngày": lop.tu_ngay.strftime('%d/%m/%Y') if lop.tu_ngay else "",
                "Đến ngày": lop.den_ngay.strftime('%d/%m/%Y') if lop.den_ngay else "",
                "Địa điểm": lop.dia_diem or "",
                "Hình thức đào tạo": lop.hinh_thuc_hoc,
                "Lĩnh vực": lop.linh_vuc or "",
                "Ghi chú": lop.ghi_chu or ""
            })

        if not data:
            flash("Không tìm thấy dữ liệu phù hợp với bộ lọc để xuất file!", "warning")
            return redirect(url_for('theo_doi_lop_hoc'))

        df = pd.DataFrame(data)
        output = BytesIO()

        # 7. GHI FILE EXCEL VỚI XLSXWRITER
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='BaoCaoDaoTao')
            
            workbook  = writer.book
            worksheet = writer.sheets['BaoCaoDaoTao']

            # Định dạng Header chuyên nghiệp
            header_format = workbook.add_format({
                'bold': True, 'text_wrap': True, 'valign': 'vcenter',
                'align': 'center', 'fg_color': '#D7E4BC', 'border': 1
            })
            
            # Định dạng Body
            cell_format = workbook.add_format({'border': 1, 'valign': 'vcenter'})
            num_format = workbook.add_format({'border': 1, 'valign': 'vcenter', 'align': 'center'})

            # Áp dụng định dạng cho Header và set độ rộng cột
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
                
                # Tùy chỉnh độ rộng riêng cho một số cột quan trọng
                if value in ["Tên chương trình", "Nội dung đào tạo"]:
                    worksheet.set_column(col_num, col_num, 40, cell_format)
                elif value in ["Họ và Tên", "Ghi chú", "Đơn vị tổ chức"]:
                    worksheet.set_column(col_num, col_num, 25, cell_format)
                elif value == "TT":
                    worksheet.set_column(col_num, col_num, 5, num_format)
                else:
                    worksheet.set_column(col_num, col_num, 15, cell_format)

        output.seek(0)

        # 8. PHẢN HỒI FILE VỀ TRÌNH DUYỆT
        file_name = f"Bao_cao_dao_tao_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output, 
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            download_name=file_name, 
            as_attachment=True
        )
                            
    except Exception as e:
        db.session.rollback() # Tránh treo transaction nếu lỗi
        print(f"Lỗi xuất Excel chi tiết: {traceback.format_exc()}")
        flash(f"Lỗi xuất Excel: {str(e)}", "danger")
        return redirect(url_for('theo_doi_lop_hoc'))


# ----------------------------------------------------------------------
# Menu Báo cáo đào tạo theo: Người học; Chức vụ (dòng là tên chương trình học lớp học, các cột là các chức vụ)
# ----------------------------------------------------------------------
@app.route('/bao_cao_tong_hop_dao_tao')
@login_required
def bao_cao_tong_hop_dao_tao():
    try:
        tu_ngay = request.args.get('tu_ngay')
        den_ngay = request.args.get('den_ngay')
        export = request.args.get('export')

        # --- 0. XÁC ĐỊNH QUYỀN HẠN VÀ PHẠM VI (SECURITY LAYER) ---
        user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
        if not user_info:
            flash("Không tìm thấy thông tin tài khoản người dùng!", "danger")
            return redirect(url_for('theo_doi_lop_hoc'))

        is_system_admin = (current_user.is_admin == 1)
        user_ma_pb = str(user_info.ma_phong_ban) if user_info.ma_phong_ban else None
        user_ma_hieu_2 = user_info.ma_hieu_2
        is_phong_tong_hop = (user_ma_pb == '2')

        # --- 1. LẤY DANH SÁCH LỚP HỌC DUY NHẤT (THEO PHẠM VI QUẢN LÝ) ---
        # Chỉ hiển thị các cột lớp học mà nhân viên trong phạm vi quản lý đã từng học
        query_classes = db.session.query(
            DanhSachLopHoc.ten_lop_hoc, 
            DanhSachLopHoc.ma_hieu,
            DanhSachLopHoc.hinh_thuc_hoc
        ).join(
            ThongTinNguoiLaoDong, DanhSachLopHoc.ma_nhan_vien == ThongTinNguoiLaoDong.ma_nhan_vien
        ).distinct()
        
        # Lọc lớp học theo thời gian
        if tu_ngay: query_classes = query_classes.filter(DanhSachLopHoc.tu_ngay >= tu_ngay)
        if den_ngay: query_classes = query_classes.filter(DanhSachLopHoc.tu_ngay <= den_ngay)
        
        # Lọc lớp học theo phạm vi đơn vị/phòng ban
        if not is_system_admin:
            if is_phong_tong_hop:
                query_classes = query_classes.filter(ThongTinNguoiLaoDong.ma_hieu_2 == user_ma_hieu_2)
            else:
                query_classes = query_classes.filter(ThongTinNguoiLaoDong.ma_phong_ban == user_ma_pb)
        
        results_classes = query_classes.all()
        classes_truc_tiep = sorted([(c[0], c[1]) for c in results_classes if c[2] == 'Trực tiếp'])
        classes_truc_tuyen = sorted([(c[0], c[1]) for c in results_classes if c[2] == 'Trực tuyến'])

        # --- 2. LẤY DỮ LIỆU CHI TIẾT (CÓ LỌC PHÂN QUYỀN) ---
        query_data = db.session.query(
            ThongTinNguoiLaoDong.ma_nhan_vien,
            ThongTinNguoiLaoDong.ho_ten,
            ThongTinNguoiLaoDong.chuc_vu, 
            PhongBan.ten_phong_ban,
            DanhSachLopHoc.ten_lop_hoc,
            DanhSachLopHoc.ma_hieu,
            DanhSachLopHoc.hinh_thuc_hoc,
            DanhSachLopHoc.so_ngay
        ).join(
            DanhSachLopHoc, ThongTinNguoiLaoDong.ma_nhan_vien == DanhSachLopHoc.ma_nhan_vien
        ).join(
            PhongBan, ThongTinNguoiLaoDong.ma_phong_ban == PhongBan.id
        )

        # Lọc dữ liệu chi tiết theo thời gian
        if tu_ngay: query_data = query_data.filter(DanhSachLopHoc.tu_ngay >= tu_ngay)
        if den_ngay: query_data = query_data.filter(DanhSachLopHoc.tu_ngay <= den_ngay)
        
        # ÉP BUỘC: Lọc nhân viên theo đơn vị (Nguyên tắc cốt lõi)
        if not is_system_admin:
            if is_phong_tong_hop:
                # Phòng Tổng hợp thấy toàn bộ đơn vị
                query_data = query_data.filter(ThongTinNguoiLaoDong.ma_hieu_2 == user_ma_hieu_2)
            else:
                # User thường chỉ thấy phòng ban mình
                query_data = query_data.filter(ThongTinNguoiLaoDong.ma_phong_ban == user_ma_pb)
        
        raw_data = query_data.all()

        # --- 3. PIVOT DỮ LIỆU (NHÂN VIÊN -> CÁC LỚP ĐÃ HỌC) ---
        pivot_dict = {}
        for ma_nv, ho_ten, chuc_vu, phong_ban, ten_lop, ma_hieu, hinh_thuc, so_ngay in raw_data:
            if ma_nv not in pivot_dict:
                pivot_dict[ma_nv] = {
                    'ma_nv': ma_nv, 'ho_ten': ho_ten, 'phong_ban': phong_ban, 'chuc_vu': chuc_vu,
                    'user_classes': set(),
                    'tong_lop_ttiep': 0, 'tong_lop_ttuyen': 0,
                    'tong_ngay_ttiep': 0.0, 'tong_ngay_ttuyen': 0.0
                }
            
            class_identity = (ten_lop, ma_hieu)
            if class_identity not in pivot_dict[ma_nv]['user_classes']:
                pivot_dict[ma_nv]['user_classes'].add(class_identity)
                if hinh_thuc == 'Trực tiếp':
                    pivot_dict[ma_nv]['tong_lop_ttiep'] += 1
                    pivot_dict[ma_nv]['tong_ngay_ttiep'] += float(so_ngay or 0)
                else:
                    pivot_dict[ma_nv]['tong_lop_ttuyen'] += 1
                    pivot_dict[ma_nv]['tong_ngay_ttuyen'] += float(so_ngay or 0)

        # --- 4. XUẤT EXCEL (GIỮ NGUYÊN CODE ĐỊNH DẠNG CỦA BẠN) ---
        if export == 'excel':
            fixed_cols = ['STT', 'Mã NV', 'Họ tên', 'Phòng ban', 'Chức vụ']
            calc_cols = ['Tổng lớp Trực tiếp', 'Tổng lớp Trực tuyến', 'Tổng lớp học',
                         'Tổng ngày Trực tiếp', 'Tổng ngày Trực tuyến', 'Tổng ngày học']
            
            excel_data = []
            for idx, person in enumerate(pivot_dict.values(), 1):
                row = [idx, person['ma_nv'], person['ho_ten'], person['phong_ban'], person['chuc_vu']]
                
                for c_tuple in classes_truc_tiep:
                    row.append('V' if c_tuple in person['user_classes'] else '-')
                for c_tuple in classes_truc_tuyen:
                    row.append('V' if c_tuple in person['user_classes'] else '-')
                
                t_lop = person['tong_lop_ttiep'] + person['tong_lop_ttuyen']
                t_ngay = person['tong_ngay_ttiep'] + person['tong_ngay_ttuyen']
                row.extend([person['tong_lop_ttiep'], person['tong_lop_ttuyen'], t_lop,
                            person['tong_ngay_ttiep'], person['tong_ngay_ttuyen'], t_ngay])
                excel_data.append(row)

            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df = pd.DataFrame(excel_data)
                df.to_excel(writer, index=False, header=False, startrow=2, sheet_name='Ma Tran')
                ws = writer.sheets['Ma Tran']

                # Styles
                header_font = Font(bold=True, color="FFFFFF")
                center_style = Alignment(horizontal='center', vertical='center', wrap_text=True)
                vertical_style = Alignment(horizontal='center', vertical='center', textRotation=90, wrap_text=True)
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                # A. Header cố định
                for col_idx in range(1, len(fixed_cols) + 1):
                    ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
                    cell = ws.cell(row=1, column=col_idx, value=fixed_cols[col_idx-1])
                    cell.fill = PatternFill(start_color="858796", end_color="858796", fill_type="solid")
                    cell.font = header_font
                    cell.alignment = center_style

                curr_col = len(fixed_cols) + 1

                # B & C. Header Lớp học (Tên + Mã)
                for group_name, group_data, color in [("TRỰC TIẾP", classes_truc_tiep, "4E73DF"), 
                                                       ("TRỰC TUYẾN", classes_truc_tuyen, "36B9CC")]:
                    if group_data:
                        start_c = curr_col
                        end_c = curr_col + len(group_data) - 1
                        ws.merge_cells(start_row=1, start_column=start_c, end_row=1, end_column=end_c)
                        cell = ws.cell(row=1, column=start_c, value=group_name)
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                        cell.font = header_font
                        cell.alignment = center_style
                        
                        for i, (ten, ma) in enumerate(group_data):
                            c = ws.cell(row=2, column=start_c + i, value=f"{ten}\n({ma})")
                            c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                            c.font = Font(bold=True, color="FFFFFF", size=9)
                            c.alignment = vertical_style
                        curr_col = end_c + 1

                # D. Header Tính toán
                for i, col_name in enumerate(calc_cols):
                    col_idx = curr_col + i
                    ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    cell.fill = PatternFill(start_color="F6C23E", end_color="F6C23E", fill_type="solid")
                    cell.font = Font(bold=True)
                    cell.alignment = center_style

                # Định dạng độ rộng và khung
                last_col = curr_col + len(calc_cols) - 1
                for col_idx in range(1, last_col + 1):
                    letter = get_column_letter(col_idx)
                    if col_idx <= 5: ws.column_dimensions[letter].width = 15
                    elif col_idx < curr_col: ws.column_dimensions[letter].width = 6
                    else: ws.column_dimensions[letter].width = 12
                
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=last_col):
                    for cell in row:
                        cell.border = border
                        if cell.row > 2: cell.alignment = center_style

            output.seek(0)
            return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True, download_name=f"Bao_cao_dao_tao_tong_hop.xlsx")

        return render_template('bao_cao_dao_tao_tong_hop.html', 
                               classes_truc_tiep=classes_truc_tiep,
                               classes_truc_tuyen=classes_truc_tuyen,
                               pivot_data=pivot_dict.values(),
                               tu_ngay=tu_ngay, den_ngay=den_ngay)

    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"Lỗi Bao cao tong hop: {traceback.format_exc()}")
        flash(f"Lỗi hệ thống: {str(e)}", "danger")
        return redirect(url_for('theo_doi_lop_hoc'))
# ----------------------------------------------------------------------
# Báo cáo đào tạo theo chức vụ (dòng là tên chương trình học lớp học, các cột là các chức vụ
# ----------------------------------------------------------------------
@app.route('/bao-cao-dao-tao-chuc-vu')
@login_required
def bao_cao_dao_tao_chuc_vu():
    tu_ngay = request.args.get('tu_ngay', '')
    den_ngay = request.args.get('den_ngay', '')
    hinh_thuc = request.args.get('hinh_thuc', '')
    page = request.args.get('page', 1, type=int)
    per_page = 15
    offset = (page - 1) * per_page

    # --- 0. XÁC ĐỊNH PHẠM VI TRUY XUẤT (SECURITY) ---
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    is_system_admin = (current_user.is_admin == 1)
    user_ma_pb = str(user_info.ma_phong_ban) if user_info and user_info.ma_phong_ban else None
    user_ma_hieu_2 = user_info.ma_hieu_2 if user_info else None
    is_phong_tong_hop = (user_ma_pb == '2')

    # 1. Điều kiện lọc cơ bản
    where_clauses = ["1=1"]
    sql_params = {}
    
    if tu_ngay:
        where_clauses.append("l.tu_ngay >= :tu_ngay")
        sql_params['tu_ngay'] = tu_ngay
    if den_ngay:
        where_clauses.append("l.den_ngay <= :den_ngay")
        sql_params['den_ngay'] = den_ngay
    if hinh_thuc:
        where_clauses.append("l.hinh_thuc_hoc = :hinh_thuc")
        sql_params['hinh_thuc'] = hinh_thuc

    # --- BỔ SUNG LỌC PHÂN QUYỀN VÀO SQL ---
    if not is_system_admin:
        if is_phong_tong_hop:
            # Phòng tổng hợp: Chỉ tính toán trên nhân viên cùng đơn vị (ma_hieu_2)
            where_clauses.append("n.ma_hieu_2 = :user_ma_hieu_2")
            sql_params['user_ma_hieu_2'] = user_ma_hieu_2
        else:
            # User thường: Chỉ tính toán trên nhân viên cùng phòng ban
            where_clauses.append("n.ma_phong_ban = :user_ma_pb")
            sql_params['user_ma_pb'] = user_ma_pb

    where_sql = " AND ".join(where_clauses)

    # 2. Câu lệnh SQL đầy đủ cho 10 cột
    # Sử dụng COALESCE để tránh giá trị NULL khi SUM
    # Tách riêng phần lọc phạm vi để không làm mất dữ liệu lớp học
    # Giả sử where_sql_lop là lọc theo thời gian/hình thức của lớp
    # Giả sử where_sql_nv là lọc theo đơn vị/phòng ban của nhân viên

    core_sql = f"""
        SELECT 
            l.ten_lop_hoc, 
            CAST(SUM(CASE WHEN LOWER(n.chuc_vu) LIKE '%trưởng đơn vị%' 
                            OR LOWER(n.chuc_vu) LIKE '%phó đơn vị%' THEN 1 ELSE 0 END) AS SIGNED) AS lanh_dao_don_vi,
            CAST(SUM(CASE WHEN LOWER(n.chuc_vu) LIKE '%trưởng phòng%' 
                            OR LOWER(n.chuc_vu) LIKE '%phó trưởng phòng%' 
                            OR LOWER(n.chuc_vu) LIKE '%tương đương%' THEN 1 ELSE 0 END) AS SIGNED) AS lanh_dao_phong,
            CAST(SUM(CASE WHEN LOWER(n.chuc_vu) = 'lao động chuyên môn, nghiệp vụ' 
                            OR n.chuc_vu IS NULL 
                            OR n.chuc_vu = '' THEN 1 ELSE 0 END) AS SIGNED) AS chuyen_vien,
            COUNT(n.ma_nhan_vien) AS tong_cong,
            MAX(l.so_ngay) AS so_ngay_hoc, -- Lấy số ngày của lớp học đó
            MIN(l.tu_ngay) AS tu_ngay_dt,
            MAX(l.den_ngay) AS den_ngay_dt,
            MAX(l.dia_diem) AS dia_diem,
            MAX(l.hinh_thuc_hoc) AS hinh_thuc_hoc
        FROM danh_sach_lop_hoc l
        INNER JOIN thong_tin_nguoi_lao_dong n ON l.ma_nhan_vien = n.ma_nhan_vien
        WHERE {where_sql} 
        GROUP BY l.ten_lop_hoc 
        ORDER BY tu_ngay_dt DESC
    """

    # 3. Xử lý Xuất Excel
    if request.args.get('export') == 'excel':
        results_all = db.session.execute(text(core_sql), sql_params).fetchall()
        data_all = [dict(row._mapping) for row in results_all]
        
        export_data = []
        for idx, r in enumerate(data_all, 1):
            export_data.append({
                'STT': idx,
                'Tên Lớp Học': r['ten_lop_hoc'],
                'Lãnh đạo ĐV': r['lanh_dao_don_vi'],
                'Lãnh đạo Phòng': r['lanh_dao_phong'],
                'Chuyên viên': r['chuyen_vien'],
                'Tổng cộng': r['tong_cong'],
                'Số ngày': r['so_ngay_hoc'],
                'Từ ngày': r['tu_ngay_dt'].strftime('%d/%m/%Y') if r['tu_ngay_dt'] else "",
                'Đến ngày': r['den_ngay_dt'].strftime('%d/%m/%Y') if r['den_ngay_dt'] else "",
                'Địa điểm học': r['dia_diem'],
                'Hình thức học': r['hinh_thuc_hoc']
            })
        
        df = pd.DataFrame(export_data)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='BaoCaoChucVu')
        output.seek(0)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=f"Bao_cao_chuc_vu_{datetime.now().strftime('%Y%m%d')}.xlsx")

    # 4. Truy vấn hiển thị Web (có phân trang)
    # Cần join với bảng nhân viên trong count_sql để kết quả phân trang khớp với bảo mật
    count_sql = text(f"""
        SELECT COUNT(DISTINCT l.ten_lop_hoc) 
        FROM danh_sach_lop_hoc l 
        LEFT JOIN thong_tin_nguoi_lao_dong n ON l.ma_nhan_vien = n.ma_nhan_vien
        WHERE {where_sql}
    """)
    total_records = db.session.execute(count_sql, sql_params).scalar() or 0
    total_pages = math.ceil(total_records / per_page)

    data_query = text(core_sql + " LIMIT :limit OFFSET :offset")
    sql_params.update({'limit': per_page, 'offset': offset})
    results = db.session.execute(data_query, sql_params).fetchall()
    
    # Chuyển results thành danh sách dict để template xử lý ổn định
    results_as_dict = [dict(row._mapping) for row in results]

    return render_template('bao_cao_dao_tao_chuc_vu.html', 
                           results=results_as_dict, 
                           tu_ngay=tu_ngay, 
                           den_ngay=den_ngay, 
                           hinh_thuc=hinh_thuc,
                           current_page=page, 
                           total_pages=total_pages, 
                           total_records=total_records)

# ----------------------------------------------------------------------
# Báo cáo tổng hợp chi tiết chấm công (bao_cao_cong.html) từ ngày đến ngày, mỗi ma_nhan_vien có tổng cộng bao nhiêu X; SC; B; P; V; C; T; S; M; H; D; N; K; L
# ----------------------------------------------------------------------
@app.route('/bao-cao-cong-chi-tiet', methods=['GET'])
@login_required
def bao_cao_cong_chi_tiet():
    # --- PHẦN 1: KIỂM TRA QUYỀN TRUY CẬP MENU ---
    # Nếu là Admin hệ thống thì cho qua, nếu không phải thì check bảng UserMenuPermission
    if not current_user.is_admin:
        has_perm = UserMenuPermission.query.filter_by(
            ma_nhan_vien=current_user.ma_nhan_vien, 
            menu_slug='bao-cao-cong-chi-tiet'
        ).first()
        if not has_perm:
            flash("Bạn không có quyền xem Báo cáo chấm công tổng hợp!", "danger")
            return redirect(url_for('index'))
    
    # --- PHẦN 2: LẤY THÔNG TIN ĐƠN VỊ CỦA USER ĐỂ PHÂN PHẠM VI ---
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    user_ma_pb = str(user_info.ma_phong_ban) if user_info and user_info.ma_phong_ban else None
    user_ma_hieu_2 = user_info.ma_hieu_2 if user_info else None
    
    # Quyền Phòng Tổng hợp (mã 2): Chỉ thấy người trong Đơn vị mình
    is_phong_tong_hop = (user_ma_pb == '2')
    is_system_admin = (current_user.is_admin == 1)

    # --- PHẦN 3: XỬ LÝ THAM SỐ THỜI GIAN ---
    tu_ngay_str = request.args.get('tu_ngay')
    den_ngay_str = request.args.get('den_ngay')
    ho_ten_search = request.args.get('ho_ten', '').strip()
    ma_hieu_2_filter = request.args.get('ma_hieu_2', '').strip()

    today = datetime.now()
    if not tu_ngay_str:
        tu_ngay_str = today.replace(day=1).strftime('%Y-%m-%d')
    if not den_ngay_str:
        den_ngay_str = today.strftime('%Y-%m-%d')

    # Lấy danh sách Đơn vị (Mã hiệu 2) cho dropdown nếu là Admin hệ thống
    don_vis_list = []
    if is_system_admin:
        # Lấy duy nhất các Mã hiệu 2 hiện có để tránh tải hàng nghìn bản ghi thừa
        don_vis_list = [row[0] for row in db.session.query(distinct(ThongTinNguoiLaoDong.ma_hieu_2)).filter(ThongTinNguoiLaoDong.ma_hieu_2 != None).all()]

    # --- PHẦN 4: XÂY DỰNG TRUY VẤN SQL --- # Tạo chuỗi UNION cho 31 ngày
    union_parts = []
    for i in range(1, 32):
        # Dùng STR_TO_DATE (cho MySQL) để tạo ngày và lọc ngay từ đầu cho hiệu năng cao
        part = f"""
            SELECT ma_nhan_vien, Nam, Thang, D{i} AS ky_hieu,
            STR_TO_DATE(CONCAT(Nam, '-', Thang, '-', {i}), '%Y-%m-%d') AS ngay_thuc_te
            FROM thong_tin_cham_cong
        """
        union_parts.append(part)
    sql_subquery = " UNION ALL ".join(union_parts)

    # Trong SQL chính, ta lấy thêm MIN/MAX của tháng năm để hiển thị
    sql_text = f"""
        SELECT 
            BT.ma_nhan_vien, 
            NV.ho_ten,
            NV.ma_hieu_2,
            MIN(BT.Thang) as Thang_Min,
            MAX(BT.Thang) as Thang_Max,
            MIN(BT.Nam) as Nam_Min,
            SUM(CASE WHEN BT.ky_hieu = 'X' THEN 1 ELSE 0 END) AS Tong_X,
            SUM(CASE WHEN BT.ky_hieu = 'SC' THEN 0.5 ELSE 0 END) AS Tong_SC,
            SUM(CASE WHEN BT.ky_hieu = 'B' THEN 1 ELSE 0 END) AS Tong_B,
            SUM(CASE WHEN BT.ky_hieu = 'P' THEN 1 ELSE 0 END) AS Tong_P,
            SUM(CASE WHEN BT.ky_hieu = 'V' THEN 1 ELSE 0 END) AS Tong_V,
            SUM(CASE WHEN BT.ky_hieu = 'C' THEN 1 ELSE 0 END) AS Tong_C,
            SUM(CASE WHEN BT.ky_hieu = 'T' THEN 1 ELSE 0 END) AS Tong_T,
            SUM(CASE WHEN BT.ky_hieu = 'S' THEN 1 ELSE 0 END) AS Tong_S,
            SUM(CASE WHEN BT.ky_hieu = 'M' THEN 1 ELSE 0 END) AS Tong_M,
            SUM(CASE WHEN BT.ky_hieu = 'H' THEN 1 ELSE 0 END) AS Tong_H,
            SUM(CASE WHEN BT.ky_hieu = 'D' THEN 1 ELSE 0 END) AS Tong_D,
            SUM(CASE WHEN BT.ky_hieu = 'N' THEN 1 ELSE 0 END) AS Tong_N,
            SUM(CASE WHEN BT.ky_hieu = 'K' THEN 1 ELSE 0 END) AS Tong_K,
            SUM(CASE WHEN BT.ky_hieu = 'L' THEN 1 ELSE 0 END) AS Tong_L
        FROM ({sql_subquery}) AS BT
        LEFT JOIN thong_tin_nguoi_lao_dong AS NV ON BT.ma_nhan_vien = NV.ma_nhan_vien
        WHERE BT.ngay_thuc_te BETWEEN :tu_ngay AND :den_ngay
          AND BT.ky_hieu IS NOT NULL AND BT.ky_hieu != ''
    """

    params = {'tu_ngay': tu_ngay_str, 'den_ngay': den_ngay_str}

    # --- PHẦN 5: ÁP DỤNG BỘ LỌC PHÂN QUYỀN DỮ LIỆU ---
    # Nếu không phải Admin tối cao nhưng là Phòng Tổng hợp -> Chỉ lấy nhân viên cùng đơn vị
    # Logic: Admin thấy hết hoặc lọc theo dropdown. Phòng TH chỉ thấy đơn vị mình.
    if is_system_admin:
        if ma_hieu_2_filter:
            sql_text += " AND NV.ma_hieu_2 = :ma_hieu_2 "
            params['ma_hieu_2'] = ma_hieu_2_filter
    elif is_phong_tong_hop:
        # Ép buộc chỉ xem được đơn vị của chính mình
        sql_text += " AND NV.ma_hieu_2 = :ma_hieu_2 "
        params['ma_hieu_2'] = user_ma_hieu_2
    else:
        # Trường hợp user thường được cấp quyền (nếu có) nhưng không thuộc phòng TH, 
        # có thể chỉ cho xem chính họ hoặc trả về rỗng tùy chính sách.
        sql_text += " AND BT.ma_nhan_vien = :me "
        params['me'] = current_user.ma_nhan_vien

    # Lọc theo tên hoặc mã nhân viên nếu có tìm kiếm
    if ho_ten_search:
        sql_text += " AND (NV.ho_ten LIKE :q OR NV.ma_nhan_vien LIKE :q) "
        params['q'] = f"%{ho_ten_search}%"

    sql_text += " GROUP BY BT.ma_nhan_vien, NV.ho_ten, NV.ma_hieu_2 ORDER BY NV.ho_ten ASC"

    results = db.session.execute(text(sql_text), params).fetchall()

    return render_template('bao_cao_cong.html', 
                           results=results, 
                           tu_ngay=tu_ngay_str, 
                           den_ngay=den_ngay_str,
                           ho_ten=ho_ten_search,
                           don_vis_list=don_vis_list,
                           current_don_vi=ma_hieu_2_filter if is_system_admin else user_ma_hieu_2,
                           is_system_admin=is_system_admin)


@app.route('/export-bao-cao-cong', methods=['GET'])
@login_required
def export_bao_cao_cong():
    if not current_user.is_admin:
        has_perm = UserMenuPermission.query.filter_by(
            ma_nhan_vien=current_user.ma_nhan_vien, 
            menu_slug='bao-cao-cong-chi-tiet'
        ).first()
        if not has_perm:
            return "Bạn không có quyền thực hiện chức năng này", 403

    # --- PHẦN 2: LẤY THÔNG TIN ĐƠN VỊ VÀ THAM SỐ LỌC ---
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    user_ma_hieu_2 = user_info.ma_hieu_2 if user_info else None
    user_ma_pb = str(user_info.ma_phong_ban) if user_info and user_info.ma_phong_ban else None
    
    is_system_admin = (current_user.is_admin == 1)
    is_phong_tong_hop = (user_ma_pb == '2')
    
    # 1. Lấy tham số ngày từ request
    tu_ngay = request.args.get('tu_ngay')
    den_ngay = request.args.get('den_ngay')
    ho_ten_search = request.args.get('ho_ten', '').strip()
    ma_hieu_2_filter = request.args.get('ma_hieu_2', '').strip()
    
    # Thiết lập mặc định nếu thiếu tham số
    today = datetime.now()
    if not tu_ngay: tu_ngay = today.replace(day=1).strftime('%Y-%m-%d')
    if not den_ngay: den_ngay = today.strftime('%Y-%m-%d')
    
    try:
        tu_ngay_obj = datetime.strptime(tu_ngay, '%Y-%m-%d')
        den_ngay_obj = datetime.strptime(den_ngay, '%Y-%m-%d')
        tu_period = tu_ngay_obj.year * 100 + tu_ngay_obj.month
        den_period = den_ngay_obj.year * 100 + den_ngay_obj.month
    except ValueError:
        return "Định dạng ngày không hợp lệ", 400

    # --- PHẦN 3: XÂY DỰNG SQL VỚI PHÂN QUYỀN DỮ LIỆU ---
    # Sử dụng Union tối ưu
    union_parts = [f"SELECT ma_nhan_vien, Nam, Thang, D{i} AS ky_hieu FROM thong_tin_cham_cong" for i in range(1, 32)]
    sql_subquery = " UNION ALL ".join(union_parts)

    # Khởi tạo SQL và params
    params = {'tu_period': tu_period, 'den_period': den_period}

    # Filter theo quyền: Nghìn đơn vị vẫn đảm bảo tốc độ nhờ lọc ngay tại WHERE
    where_clauses = [
        "(BT.Nam * 100 + BT.Thang) >= :tu_period",
        "(BT.Nam * 100 + BT.Thang) <= :den_period",
        "BT.ky_hieu IS NOT NULL",
        "BT.ky_hieu != ''"
    ]

    # Phân quyền đơn vị
    if is_system_admin:
        if ma_hieu_2_filter:
            where_clauses.append("NV.ma_hieu_2 = :ma_hieu_2")
            params['ma_hieu_2'] = ma_hieu_2_filter
    elif is_phong_tong_hop:
        where_clauses.append("NV.ma_hieu_2 = :user_dv")
        params['user_dv'] = user_ma_hieu_2
    else:
        where_clauses.append("BT.ma_nhan_vien = :me")
        params['me'] = current_user.ma_nhan_vien

    # Lọc theo tên/mã NV
    if ho_ten_search:
        where_clauses.append("(NV.ho_ten LIKE :q OR NV.ma_nhan_vien LIKE :q)")
        params['q'] = f"%{ho_ten_search}%"

    # Chạy SQL lấy dữ liệu (Đầy đủ từ D1 đến D31)
    sql_text = f"""
        SELECT 
            BT.ma_nhan_vien, NV.ho_ten, BT.Thang, BT.Nam,
            SUM(CASE WHEN BT.ky_hieu = 'X' THEN 1 ELSE 0 END) AS X,
            SUM(CASE WHEN BT.ky_hieu = 'SC' THEN 1 ELSE 0 END) AS SC,
            SUM(CASE WHEN BT.ky_hieu = 'B' THEN 1 ELSE 0 END) AS B,
            SUM(CASE WHEN BT.ky_hieu = 'P' THEN 1 ELSE 0 END) AS P,
            SUM(CASE WHEN BT.ky_hieu = 'V' THEN 1 ELSE 0 END) AS V,
            SUM(CASE WHEN BT.ky_hieu = 'C' THEN 1 ELSE 0 END) AS C,
            SUM(CASE WHEN BT.ky_hieu = 'T' THEN 1 ELSE 0 END) AS T,
            SUM(CASE WHEN BT.ky_hieu = 'S' THEN 1 ELSE 0 END) AS S,
            SUM(CASE WHEN BT.ky_hieu = 'M' THEN 1 ELSE 0 END) AS M,
            SUM(CASE WHEN BT.ky_hieu = 'H' THEN 1 ELSE 0 END) AS H,
            SUM(CASE WHEN BT.ky_hieu = 'D' THEN 1 ELSE 0 END) AS D,
            SUM(CASE WHEN BT.ky_hieu = 'N' THEN 1 ELSE 0 END) AS N,
            SUM(CASE WHEN BT.ky_hieu = 'K' THEN 1 ELSE 0 END) AS K,
            SUM(CASE WHEN BT.ky_hieu = 'L' THEN 1 ELSE 0 END) AS L
        FROM ({sql_subquery}) AS BT
        LEFT JOIN thong_tin_nguoi_lao_dong AS NV ON BT.ma_nhan_vien = NV.ma_nhan_vien
        WHERE {" AND ".join(where_clauses)}
        GROUP BY BT.ma_nhan_vien, NV.ho_ten, NV.ma_hieu_2, BT.Nam, BT.Thang
        ORDER BY BT.Nam DESC, BT.Thang DESC, NV.ma_hieu_2 ASC, NV.ho_ten ASC
    """
    
    results = db.session.execute(text(sql_text), params).fetchall()

    # 3. Tạo file Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Báo cáo công tổng hợp"

    # Định nghĩa tiêu đề
    headers = ['Mã NV', 'Họ Tên', 'Đơn vị', 'Tháng', 'Năm', 'X', 'SC', 'B', 'P', 'V', 'C', 'T', 'S', 'M', 'H', 'D', 'N', 'K', 'L']
    ws.append(headers)

    # Định dạng Header
    header_fill = PatternFill(start_color="004C99", end_color="004C99", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Đổ dữ liệu và định dạng dòng
    for row_idx, row_data in enumerate(results, start=2):
        data_list = list(row_data)
        ws.append(data_list)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx >= 4: # Từ cột Tháng trở đi thì căn giữa
                cell.alignment = center_align

    # Tự động căn chỉnh độ rộng cột
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 50)

    # 4. Gửi file về trình duyệt
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Đặt tên file gợi nhớ: Bao_cao_DonVi_ThoiGian.xlsx
    prefix = ma_hieu_2_filter if ma_hieu_2_filter else "TongHop"
    filename = f"BC_tong_hop_Cong_{prefix}_{tu_ngay}_{den_ngay}.xlsx"
    
    return send_file(
        output, 
        as_attachment=True, 
        download_name=filename, 
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ----------------------------------------------------------------------
# Theo dõi nghỉ phép
# ----------------------------------------------------------------------
@app.route('/theo-doi-nghi-phep')
@login_required
def theo_doi_nghi_phep():
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    user_ma_pb = str(user_info.ma_phong_ban) if user_info and user_info.ma_phong_ban else None
    user_ma_hieu_2 = user_info.ma_hieu_2 if user_info else None

    is_system_admin = (current_user.is_admin == 1)
    is_phong_tong_hop = (user_ma_pb == '2') # Phòng Tổng hợp mã là '2'
    
    # 2. Lấy các tham số lọc và phân trang
    year = request.args.get('year', datetime.now().year, type=int)
    search_query = request.args.get('search', '', type=str).strip()
    page = request.args.get('page', 1, type=int)
    per_page = 20  # Số nhân viên mỗi trang

    # 3. LOGIC KHÓA BỘ LỌC THEO QUYỀN
    if is_system_admin:
        # Admin: Thấy tất cả, có thể chọn đơn vị bất kỳ
        ma_hieu_2_filter = request.args.get('don_vi', '').strip()
    elif is_phong_tong_hop:
        # Phòng TH: Chỉ xem trong ĐƠN VỊ của mình (không được đổi đơn vị)
        ma_hieu_2_filter = user_ma_hieu_2
    else:
        # User thường: Chỉ xem chính mình (hoặc có thể cấu hình xem cả phòng)
        # Ở đây nếu muốn user thường chỉ thấy mình, ta lọc theo ma_nhan_vien
        ma_hieu_2_filter = user_ma_hieu_2

    # 4. Xây dựng điều kiện lọc SQL
    extra_filters = ""
    params = {'y': year}
    
    if search_query:
        extra_filters += " AND (v.ma_nhan_vien LIKE :s OR n.ho_ten LIKE :s)"
        params['s'] = f"%{search_query}%"
    # Áp dụng filter đơn vị đã qua xử lý quyền
    if ma_hieu_2_filter:
        extra_filters += " AND n.ma_hieu_2 = :dv"
        params['dv'] = ma_hieu_2_filter

    # Ràng buộc cho User thường: Chỉ thấy bản ghi của chính mình
    if not is_system_admin and not is_phong_tong_hop:
        extra_filters += " AND v.ma_nhan_vien = :me"
        params['me'] = current_user.ma_nhan_vien

    # 3. Đếm tổng số bản ghi để tính toán phân trang
    count_sql = text(f"SELECT COUNT(*) FROM v_theo_doi_nghi_phep v JOIN thong_tin_nguoi_lao_dong n ON v.ma_nhan_vien = n.ma_nhan_vien WHERE v.nam = :y {extra_filters}")
    total_records = db.session.execute(count_sql, params).scalar() or 0
    total_pages = math.ceil(total_records / per_page)
    offset = (page - 1) * per_page

    # 4. Truy vấn dữ liệu chính (có LIMIT/OFFSET) ORDER BY v.ma_nhan_vien ASC
    sql = text(f"""
        SELECT v.*, n.ho_ten, n.ngay_tinh_phep, n.ma_hieu_2 
        FROM v_theo_doi_nghi_phep v
        JOIN thong_tin_nguoi_lao_dong n ON v.ma_nhan_vien = n.ma_nhan_vien
        WHERE v.nam = :y {extra_filters}
        ORDER BY n.ho_ten ASC
        LIMIT :limit OFFSET :offset
    """)
    params.update({'limit': per_page, 'offset': offset})
    # Dùng mappings() để truy cập row['column_name'] chính xác
    raw_data = db.session.execute(sql, params).mappings().all()
    
    results = []
    today = datetime.now().date()
    moc_tinh_hien_thi = today if year == today.year else date(year, 12, 31)
    
    for row in raw_data:
        so_thang = 0
        if row['ngay_tinh_phep']:
            diff = relativedelta(moc_tinh_hien_thi, row['ngay_tinh_phep'])
            so_thang = diff.years * 12 + diff.months
        
        results.append({
            'phep': row,
            'nv': row,
            'so_thang': so_thang
        })
        
    return render_template('theo_doi_nghi_phep.html', 
                           results=results, 
                           current_year=year, 
                           search_query=search_query,
                           selected_don_vi=ma_hieu_2_filter,
                           is_system_admin=is_system_admin,
                           is_phong_tong_hop=is_phong_tong_hop,
                           page=page,
                           total_pages=total_pages)


@app.route('/chot-phep-nam/<int:current_year>')
@login_required
@admin_or_staff_required  #tham số gọi hàm quy định users được phép sử dụng
def chot_phep_nam(current_year):
    last_year = current_year - 1
    try:
        # 1. Cập nhật dữ liệu năm ngoái lần cuối để lấy số dư chính xác nhất
        update_nghi_phep_logic(last_year)
        
        # 2. Lấy số dư THỰC TẾ cuối cùng từ VIEW của năm ngoái
        # Sử dụng .mappings() để đảm bảo truy cập được tên cột từ View
        sql = text("SELECT ma_nhan_vien, tong_phep_con_lai FROM v_theo_doi_nghi_phep WHERE nam = :ly")
        result = db.session.execute(sql, {'ly': last_year})
        last_year_data = result.mappings().all() 
        
        for row in last_year_data:
            # Truy cập bằng key của dictionary thay vì thuộc tính đối tượng
            ma_nv = row['ma_nhan_vien']
            ton_cu = row['tong_phep_con_lai']
            
            # Tìm bản ghi năm nay của nhân viên
            new_record = TheoDoiNghiPhep.query.filter_by(ma_nhan_vien=ma_nv, nam=current_year).first()
            
            # Số dư năm ngoái (nếu < 0 thì coi như bằng 0)
            # Số dư năm ngoái kết chuyển sang (Nếu nhân viên nghỉ lạm phép < 0 thì đưa về 0)
            phep_ton_chuyen_sang = max(0, float(ton_cu or 0))
            
            if new_record:
                new_record.phep_ton_nam_truoc_goc = phep_ton_chuyen_sang
            else:
                # Khi tạo mới ở đây, không nên gán cứng 12.0
                new_record = TheoDoiNghiPhep(
                    ma_nhan_vien=ma_nv, 
                    nam=current_year,
                    phep_ton_nam_truoc_goc=phep_ton_chuyen_sang,
                    phep_huong_trong_nam=0.0 # Tạm thời để 0, hàm update_logic phía dưới sẽ tính lại con số đúng
                )
                db.session.add(new_record)
        
        db.session.commit()
        
        # 3. Tính toán lại thâm niên cho năm mới
        # 3. Sau khi chốt, tính toán lại thâm niên cho năm mới luôn
        update_nghi_phep_logic(current_year)
        
        flash(f"Đã kết chuyển số dư từ năm {last_year} sang {current_year}!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Lỗi hệ thống khi chốt phép: {str(e)}", "danger")
    return redirect(url_for('theo_doi_nghi_phep', year=current_year))


# Đảm bảo tên hàm là update_nghi_phep và có nhận biến year
@app.route('/update-nghi-phep/<int:year>')
@login_required
@admin_or_staff_required  #tham số gọi hàm quy định users được phép sử dụng
def update_nghi_phep(year): # Tên hàm này phải khớp với url_for
    try:
        update_nghi_phep_logic(year)
        flash(f"Cập nhật thành công dữ liệu nghỉ phép năm {year}", "success")
    except Exception as e:
        flash(f"Lỗi: {str(e)}", "danger")
    return redirect(url_for('theo_doi_nghi_phep', year=year))


def update_nghi_phep_logic(year):
    nhan_viens = ThongTinNguoiLaoDong.query.all()
    today = date.today()
    
    # Xác định mốc tính để tính thâm niên chính xác
    moc_tinh = date(year, 12, 31) if year < today.year else today

    for nv in nhan_viens:
        ngay_bat_dau = nv.ngay_tinh_phep or nv.ngay_vao_Agribank
        if not ngay_bat_dau: continue

        # --- TỰ ĐỘNG TÍNH PHÉP HƯỞNG (12 ngày/năm, tỷ lệ theo tháng) ---
        phep_huong = 12.0
        if ngay_bat_dau.year == year:
            # Nếu vào làm trong năm xét: Phép = Số tháng làm việc thực tế
            phep_huong = float(12 - ngay_bat_dau.month + 1)
        elif ngay_bat_dau.year > year:
            phep_huong = 0.0

        # --- TỰ ĐỘNG TÍNH THÂM NIÊN (Cứ đủ 5 năm tặng 1 ngày) ---
        tham_nien = 0
        if ngay_bat_dau <= moc_tinh:
            # relativedelta tính chính xác số năm từ ngày vào đến mốc tính
            tham_nien = relativedelta(moc_tinh, ngay_bat_dau).years // 5

        # --- QUÉT CHẤM CÔNG ---
        cham_cong = ThongTinChamCong.query.filter_by(ma_nhan_vien=nv.ma_nhan_vien, nam=year).all()
        q1 = sum(1 for row in cham_cong if row.thang <= 3 for i in range(1, 32) if getattr(row, f'd{i}') == 'P')
        sau_q1 = sum(1 for row in cham_cong if row.thang > 3 for i in range(1, 32) if getattr(row, f'd{i}') == 'P')

        # --- CẬP NHẬT DATABASE ---
        record = TheoDoiNghiPhep.query.filter_by(ma_nhan_vien=nv.ma_nhan_vien, nam=year).first()
        if not record:
            record = TheoDoiNghiPhep(ma_nhan_vien=nv.ma_nhan_vien, nam=year)
            db.session.add(record)
        
        record.phep_huong_trong_nam = phep_huong
        record.phep_tham_nien = tham_nien
        record.da_nghi_p_q1 = q1
        record.da_nghi_p_sau_q1 = sau_q1
            
    db.session.commit()


@app.route('/export-nghi-phep/<int:year>')
@login_required
def export_nghi_phep(year):
    try:
        # 1. Truy vấn dữ liệu từ View giống như trang xem báo cáo
        sql = text("""
            SELECT v.*, n.ho_ten 
            FROM v_theo_doi_nghi_phep v
            JOIN thong_tin_nguoi_lao_dong n ON v.ma_nhan_vien = n.ma_nhan_vien
            WHERE v.nam = :y
            ORDER BY v.ma_nhan_vien ASC
        """)
        results = db.session.execute(sql, {'y': year}).mappings().all()

        # 2. Tạo file Excel mới
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Nghi Phep {year}"

        # 3. Định dạng tiêu đề chính
        title = f"BẢNG THEO DÕI NGHỈ PHÉP NĂM {year}"
        ws.merge_cells('A1:J1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        # 4. Định dạng Header bảng
        headers = [
            "Mã NV", "Họ Tên", "Tồn Gốc", "Tồn Hiện Tại", 
            "Phép Năm", "Thâm Niên", "Nghỉ Q1", "Nghỉ Sau Q1", 
            "Tổng Còn Lại"
        ]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))

        for cell in ws[2]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # 5. Đổ dữ liệu vào bảng
        for row in results:
            data_row = [
                row['ma_nhan_vien'],
                row['ho_ten'],
                float(row['phep_ton_nam_truoc_goc'] or 0),
                float(row['phep_ton_kha_dung'] or 0),
                float(row['phep_huong_trong_nam'] or 0),
                float(row['phep_tham_nien'] or 0),
                float(row['da_nghi_p_q1'] or 0),
                float(row['da_nghi_p_sau_q1'] or 0),
                float(row['tong_phep_con_lai'] or 0)
            ]
            ws.append(data_row)

        # 6. Kẻ khung cho toàn bộ bảng và căn giữa các cột số
        for r_idx, row in enumerate(ws.iter_rows(min_row=3), 3):
            for c_idx, cell in enumerate(row, 1):
                cell.border = border
                if c_idx != 2: # Trừ cột Họ tên
                    cell.alignment = Alignment(horizontal='center')

        # Tự động chỉnh độ rộng cột
        for col in ws.columns:
            max_length = 0
            column = col[1].column_letter
            for cell in col:
                if cell.value: max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = max_length + 2

        # 7. Xuất file trả về trình duyệt
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Theo_doi_nghi_phep_{year}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename, 
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        flash(f"Lỗi xuất Excel: {str(e)}", "danger")
        return redirect(url_for('theo_doi_nghi_phep', year=year))


# ----------------------------------------------------------------------
# Lịch Đoàn công tác theo Lịch tháng; Lịch công tác đoàn
# ----------------------------------------------------------------------
@app.route('/lich-cong-tac')
@login_required
def lich_cong_tac():
    # Lấy tháng và năm từ tham số URL, mặc định là hiện tại
    now = datetime.now(HANOI_TZ)
    month = request.args.get('month', now.month, type=int)
    year = request.args.get('year', now.year, type=int)
    
    # 1. LẤY THÔNG TIN USER ĐỂ PHÂN QUYỀN
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    # Quyền đặc biệt: Admin (is_admin=1) hoặc bạn có thể thêm các mã phòng ban cụ thể được phép
    is_system_admin = (current_user.is_admin == 1)

    # Mặc định lấy đơn vị của user
    user_ma_hieu_2 = user_info.ma_hieu_2 if user_info else None
    user_ten_ma_hieu_2 = user_info.don_vi.ten_ma_hieu_2 if (user_info and user_info.don_vi) else ""

    # 2. LẤY THAM SỐ TÌM KIẾM TỪ URL
    search_tu_ngay = request.args.get('search_tu_ngay')
    search_den_ngay = request.args.get('search_den_ngay')
    search_truong_doan = request.args.get('search_truong_doan', '').strip()
    search_don_vi = request.args.get('search_don_vi') # Lọc theo mã hiệu 2
    search_linh_vuc = request.args.get('search_linh_vuc', type=int)

    # 3. Xây dựng Query lọc dữ liệu
    query = LichDoanCongTac.query

    # --- BẮT ĐẦU PHÂN QUYỀN TẠI ĐÂY ---
    if not is_system_admin:
        # Nếu không phải admin, chỉ lấy những lịch mà đơn vị của user có tham gia
        # Sử dụng join vào bảng đơn vị để lọc
        query = query.join(LichDoanCongTac.danh_sach_don_vi).filter(DonVi.ma_hieu_2 == user_ma_hieu_2)
    # --- KẾT THÚC PHÂN QUYỀN ---
    
    # Logic lọc nâng cao
    if search_tu_ngay:
        query = query.filter(LichDoanCongTac.doan_vao_tu_ngay >= search_tu_ngay)
    if search_den_ngay:
        query = query.filter(LichDoanCongTac.doan_vao_den_ngay <= search_den_ngay)
    if search_truong_doan:
        query = query.filter(LichDoanCongTac.truong_doan.icontains(search_truong_doan))
    if search_don_vi:
        query = query.join(LichDoanCongTac.danh_sach_don_vi).filter(DonVi.ma_hieu_2 == search_don_vi)
    if search_linh_vuc:
        query = query.join(LichDoanCongTac.danh_sach_linh_vuc).filter(LinhVuc.id == search_linh_vuc)

    # Nếu KHÔNG có bất kỳ tham số tìm kiếm nào, mặc định hiển thị theo tháng (Calendar mode)
    is_searching = any([search_tu_ngay, search_den_ngay, search_truong_doan, search_don_vi, search_linh_vuc])
    
    if not is_searching:
        # Xác định ngày bắt đầu và kết thúc tháng
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
            
        # Lọc lịch trình giao thoa với tháng hiện tại
        query = query.filter(
            and_(
                LichDoanCongTac.doan_vao_tu_ngay <= end_date,
                LichDoanCongTac.doan_vao_den_ngay >= start_date
            )
        )

    # Thực thi lấy dữ liệu
    lich_trinh = query.order_by(LichDoanCongTac.doan_vao_tu_ngay).all()

    # 4. QUAN TRỌNG: Không query DonVi.query.all() ở đây nữa
    # Chúng ta sẽ để Select2 gọi API api_search_don_vi khi người dùng gõ tìm kiếm
    danh_sach_don_vi = DonVi.query.order_by(DonVi.ten_ma_hieu_2.asc()).all()
    danh_sach_linh_vuc = LinhVuc.query.order_by(LinhVuc.ten_linh_vuc.asc()).all()

    # 5. Ma trận lịch và Lĩnh vực (Lĩnh vực thường ít nên có thể giữ lại)
    cal = calendar.Calendar(firstweekday=0) # 0 = Thứ Hai
    month_days = cal.monthdayscalendar(year, month)

    return render_template('lich_cong_tac.html', 
                           month=month, 
                           year=year, 
                           month_days=month_days,
                           today=now.date(),
                           lich_trinh=lich_trinh,
                           danh_sach_linh_vuc=danh_sach_linh_vuc,
                           danh_sach_don_vi=danh_sach_don_vi,
                           is_system_admin=is_system_admin,
                           user_ma_hieu_2=user_ma_hieu_2,
                           user_ten_ma_hieu_2=user_ten_ma_hieu_2,
                           search_params=request.args, # Gửi lại tham số để điền vào form search
                            is_searching=is_searching   # Để template biết đang ở chế độ tìm kiếm hay lịch tháng
                           )

@app.route('/api/search_don_vi', methods=['GET'])
@login_required
def api_search_don_vi():
    try:
        # 1. Lấy từ khóa từ tham số 'q' (do Select2 gửi lên)
        search_text = request.args.get('q', '').strip()

        # 2. Luôn bắt đầu bằng việc sắp xếp (Order by trước khi Limit)
        db_query = DonVi.query.order_by(DonVi.ten_ma_hieu_2.asc())

        # Khởi tạo biến search_results là danh sách rỗng để tránh lỗi "referenced before assignment"
        search_results = []
        
        # 3. Nếu có từ khóa thì lọc, nếu không có trả về danh sách trống hoặc mặc định
        if search_text:
            # Trường hợp có từ khóa tìm kiếm
            search_results = db_query.filter(
                (DonVi.ten_ma_hieu_2.icontains(search_text)) | 
                (DonVi.ma_hieu_2.icontains(search_text))
            ).limit(50).all()
        else:
            # Nếu Admin vừa mới Click vào ô (chưa gõ gì):
            # CHỈ trả về 10-20 đơn vị tiêu biểu để mồi dữ liệu, không load hết 1000 cái.
            # Điều này giúp Admin thấy có danh sách để chọn mà không làm treo máy.
            # Trường hợp gõ trống hoặc mới click vào
            search_results = db_query.limit(20).all()
        
        # 5. Định dạng dữ liệu trả về cho Select2
        # id: giá trị lưu vào DB (ma_hieu_2), text: nội dung hiển thị trên giao diện
        formatted_results = [
            {
                "id": dv.ma_hieu_2, 
                "text": f"{dv.ten_ma_hieu_2} ({dv.ma_hieu_2})"
            } 
            for dv in search_results
        ]
        
        return jsonify({"results": formatted_results})

    except Exception as e:
        # In log lỗi ra console để debug khi cần
        print(f"API Error (search_don_vi): {str(e)}")
        traceback.print_exc()
        return jsonify({"results": [], "error": str(e)}), 500

@app.route('/api/get_phong_ban/<ma_don_vi>')
@login_required
def get_phong_ban(ma_don_vi):
    # Lấy các phòng ban thuộc mã đơn vị được chọn
    query = text("SELECT id, ten_phong_ban FROM phong_ban WHERE ma_hieu_2 = :ma ORDER BY ten_phong_ban")
    result = db.session.execute(query, {'ma': ma_don_vi}).fetchall()
    
    phong_bans = [{'id': row.id, 'ten': row.ten_phong_ban} for row in result]
    return jsonify(phong_bans)

@app.route('/api/get_nhan_vien_dynamic')
@login_required
def get_nhan_vien_dynamic():
    ma_pb_req = request.args.get('ma_pb')
    ma_hieu_2_req = request.args.get('ma_hieu_2')
    
    # 1. Lấy thông tin User hiện tại
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    if not user_info:
        return jsonify([])

    user_ma_hieu_2 = user_info.ma_hieu_2
    user_ma_pb = str(user_info.ma_phong_ban) if user_info.ma_phong_ban else None
    is_system_admin = (current_user.is_admin == 1)
    is_phong_tong_hop = (user_ma_pb == '2') # mã '2' là Phòng Tổng Hợp

    # 2. Khởi tạo Query
    query = ThongTinNguoiLaoDong.query.filter_by(trang_thai=True)

    # 3. Logic phân quyền chặt chẽ
    if not is_system_admin:
        # HÀNG RÀO 1: Luôn khóa chặt theo Đơn vị (ma_hieu_2) của User
        query = query.filter(ThongTinNguoiLaoDong.ma_hieu_2 == user_ma_hieu_2)
        
        # HÀNG RÀO 2: Phân cấp trong đơn vị
        if not is_phong_tong_hop:
            # Nếu chỉ là User phòng ban (quyền lập bảng), ép lọc theo đúng phòng của họ
            # Bỏ qua tham số ma_pb_req từ trình duyệt gửi lên để tránh fake request
            query = query.filter(ThongTinNguoiLaoDong.ma_phong_ban == user_ma_pb)
        else:
            # Nếu là Phòng Tổng Hợp của đơn vị đó, cho phép lọc theo ma_pb_req 
            # nhưng vẫn nằm trong phạm vi ma_hieu_2 đã filter ở trên
            if ma_pb_req:
                query = query.filter(ThongTinNguoiLaoDong.ma_phong_ban == ma_pb_req)
    else:
        # Admin hệ thống: Lọc linh hoạt
        if ma_pb_req:
            query = query.filter_by(ma_phong_ban=ma_pb_req)
        elif ma_hieu_2_req:
            query = query.filter_by(ma_hieu_2=ma_hieu_2_req)

    results = query.order_by(ThongTinNguoiLaoDong.ho_ten.asc()).all()
    return jsonify([{'ma': n.ma_nhan_vien, 'ten': n.ho_ten} for n in results])


@app.route('/api/search_linh_vuc', methods=['GET'])
@login_required
def api_search_linh_vuc():
    try:
        q = request.args.get('q', '').strip()
        query = LinhVuc.query
        if q:
            query = query.filter(LinhVuc.ten_linh_vuc.icontains(q))
        results = query.order_by(LinhVuc.ten_linh_vuc.asc()).limit(20).all()
        return jsonify({
            "results": [{"id": lv.id, "text": lv.ten_linh_vuc} for lv in results]
        })
    except Exception as e:
        return jsonify({"results": []})
    
@app.route('/add-lich-cong-tac', methods=['POST'])
@login_required
def add_lich_cong_tac():
    # Khởi tạo ngày mặc định từ form hoặc ngày hiện tại để redirect chính xác
    now = datetime.now()
    try:
        tu_ngay_str = request.form.get('doan_vao_tu_ngay')
        tu_ngay = datetime.strptime(tu_ngay_str, '%Y-%m-%d').date() if tu_ngay_str else now.date()
    except Exception:
        tu_ngay = now.date()
    
    try:
        # --- BẮT ĐẦU PHẦN KIỂM TRA QUYỀN VÀ XÁC ĐỊNH ĐƠN VỊ ---
        # Lấy thông tin chi tiết của người dùng đang đăng nhập
        user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
        
        # Kiểm tra nếu là Admin (is_admin=1) hoặc bạn có thể chỉ định thêm mã phòng ban đặc biệt (ví dụ: '2')
        is_system_admin = (current_user.is_admin == 1)
        
        # Khởi tạo biến chứa danh sách mã đơn vị sẽ được gán vào lịch
        selected_ma_hieu_list = []

        if is_system_admin:
            # Nếu là Admin: Lấy danh sách từ Select2 Multiple (cho phép chọn nhiều đơn vị)
            selected_ma_hieu_list = request.form.getlist('don_vi_id[]')
        else:
            # Nếu là User thường: Lấy chính mã đơn vị (ma_hieu_2) của user đó từ DB
            # Điều này đảm bảo user không thể "fake" dữ liệu từ trình duyệt
            if user_info and user_info.ma_hieu_2:
                selected_ma_hieu_list = [user_info.ma_hieu_2]
            else:
                # Nếu user chưa được gán đơn vị, báo lỗi để tránh dữ liệu mồ côi
                flash("Tài khoản của bạn chưa được gán đơn vị. Vui lòng liên hệ Admin!", "danger")
                return redirect(url_for('lich_cong_tac', month=tu_ngay.month, year=tu_ngay.year))
        # --- KẾT THÚC PHẦN KIỂM TRA QUYỀN ---
        
        # 1. Lấy dữ liệu cơ bản từ Form
        ten_doan = request.form.get('ten_doan')
        den_ngay_str = request.form.get('doan_vao_den_ngay')
        den_ngay = datetime.strptime(den_ngay_str, '%Y-%m-%d').date()
        
        # 2. Lấy danh sách Lĩnh vực từ Form (Mối quan hệ Nhiều-Nhiều)
        selected_linh_vuc_ids = request.form.getlist('linh_vuc_id[]')
        
        # 3. Tạo đối tượng Lịch (Các trường thông tin cơ bản)
        new_lich = LichDoanCongTac(
            ten_doan=ten_doan,
            doan_vao_tu_ngay=tu_ngay,
            doan_vao_den_ngay=den_ngay,
            truong_doan=request.form.get('truong_doan'),
            noi_dung=request.form.get('noi_dung'),
            thanh_phan=request.form.get('thanh_phan'),
            ghi_chu=request.form.get('ghi_chu')
        )
        
        # --- BƯỚC QUAN TRỌNG ĐỂ SỬA LỖI SAWarning ---
        # Thêm đối tượng vào session và flush để nó có danh tính (Identity) 
        # trước khi gán các mối quan hệ Many-to-Many
        db.session.add(new_lich)
        db.session.flush() 
        # -------------------------------------------

        # 4. Xử lý gán LĨNH VỰC (Mối quan hệ Nhiều-Nhiều)
        if selected_linh_vuc_ids:
            # Lọc bỏ các giá trị trống và truy vấn
            linh_vuc_objects = LinhVuc.query.filter(LinhVuc.id.in_(selected_linh_vuc_ids)).all()
            new_lich.danh_sach_linh_vuc = linh_vuc_objects

        # 5. Xử lý gán ĐƠN VỊ (Mối quan hệ Nhiều-Nhiều)
        if selected_ma_hieu_list:
            # Lưu ý: Nếu Select2 của bạn trả về ID của Đơn vị thì dùng DonVi.id.in_
            # Nếu trả về mã hiệu thì dùng DonVi.ma_hieu_2.in_ như cũ
            don_vi_objects = DonVi.query.filter(
                DonVi.ma_hieu_2.in_(selected_ma_hieu_list)
            ).order_by(DonVi.ten_ma_hieu_2.asc()).all()
            
            new_lich.danh_sach_don_vi = don_vi_objects
        
        # 6. Commit cuối cùng sau khi đã gán xong các mối quan hệ
        db.session.commit()
        flash("Đã thêm lịch công tác thành công!", "success")
        
    except Exception as e:
        db.session.rollback()
        # In chi tiết lỗi ra Console để kiểm tra nếu còn lỗi
        traceback.print_exc()
        flash(f"Lỗi khi thêm lịch: {str(e)}", "danger")
        
    return redirect(url_for('lich_cong_tac', month=tu_ngay.month, year=tu_ngay.year))


@app.route('/api/get_lich_detail/<int:id>')
@login_required
def get_lich_detail(id):
    lich = LichDoanCongTac.query.get_or_404(id)
    
    # Định dạng dữ liệu đơn vị cho Select2
    don_vi_data = [
        {"id": dv.ma_hieu_2, "text": f"{dv.ten_ma_hieu_2} ({dv.ma_hieu_2})"} 
        for dv in lich.danh_sach_don_vi
    ]
    
    # Định dạng dữ liệu lĩnh vực cho Select2
    linh_vuc_data = [
        {"id": lv.id, "text": lv.ten_linh_vuc} 
        for lv in lich.danh_sach_linh_vuc
    ]

    return jsonify({
        "id": lich.id,
        "ten_doan": lich.ten_doan,
        "truong_doan": lich.truong_doan,
        "tu_ngay": lich.doan_vao_tu_ngay.strftime('%Y-%m-%d'),
        "den_ngay": lich.doan_vao_den_ngay.strftime('%Y-%m-%d'),
        "noi_dung": lich.noi_dung,
        "thanh_phan": lich.thanh_phan,
        "ghi_chu": lich.ghi_chu,
        "danh_sach_don_vi": don_vi_data,
        "danh_sach_linh_vuc": linh_vuc_data
    })


@app.route('/delete-lich-cong-tac/<int:id>', methods=['POST'])
@login_required
def delete_lich_cong_tac(id):
    try:
        # 1. Lấy thông tin người dùng và đối tượng lịch
        user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
        lich = LichDoanCongTac.query.get_or_404(id)
        
        # Quyền Admin hệ thống
        is_system_admin = (current_user.is_admin == 1)
        
        # 2. KIỂM TRA QUYỀN XÓA
        if not is_system_admin:
            # Nếu không phải Admin, kiểm tra xem lịch này có thuộc đơn vị của user không
            # Lấy danh sách mã đơn vị liên kết với lịch này
            ma_hieu_don_vi_cua_lich = [dv.ma_hieu_2 for dv in lich.danh_sach_don_vi]
            
            # Nếu đơn vị của user không nằm trong danh sách đơn vị của lịch, từ chối xóa
            if user_info.ma_hieu_2 not in ma_hieu_don_vi_cua_lich:
                flash("Bạn không có quyền xóa lịch của đơn vị khác!", "danger")
                return redirect(url_for('lich_cong_tac', 
                                        month=lich.doan_vao_tu_ngay.month, 
                                        year=lich.doan_vao_tu_ngay.year))

        # 3. Lưu lại thông tin thời gian để điều hướng chính xác sau khi xóa
        target_month = lich.doan_vao_tu_ngay.month
        target_year = lich.doan_vao_tu_ngay.year
        
        # 4. Xử lý các mối quan hệ Many-to-Many (nếu cần thiết)
        # Thông thường SQLAlchemy với cascade delete sẽ lo việc này, 
        # nhưng làm thủ công sẽ giúp tránh các lỗi "orphaned" hoặc lỗi session.
        # Xóa các liên kết trong bảng trung gian trước để đảm bảo tính toàn vẹn dữ liệu
        lich.danh_sach_don_vi = []
        lich.danh_sach_linh_vuc = []
        
        # 5. Thực hiện xóa
        db.session.delete(lich)
        db.session.commit()
        
        flash("Đã xóa lịch công tác thành công!", "success")
        return redirect(url_for('lich_cong_tac', month=target_month, year=target_year))
        
    except Exception as e:
        db.session.rollback()
        # In log chi tiết để dễ debug
        traceback.print_exc()
        
        flash(f"Lỗi hệ thống khi xóa: {str(e)}", "danger")
        # Nếu lỗi, quay về trang lịch hiện tại (không kèm tham số tháng/năm cụ thể để tránh lỗi redirect)
        return redirect(url_for('lich_cong_tac'))

@app.route('/export-lich-cong-tac')
@login_required
def export_excel_lich():
    # 1. LẤY LẠI CÁC THAM SỐ LỌC (Giống hệt hàm lich_cong_tac)
    search_tu_ngay = request.args.get('search_tu_ngay')
    search_den_ngay = request.args.get('search_den_ngay')
    search_truong_doan = request.args.get('search_truong_doan', '').strip()
    search_don_vi = request.args.get('search_don_vi')
    search_linh_vuc = request.args.get('search_linh_vuc', type=int)
    
    # Lấy thêm month/year để trường hợp không search thì xuất theo tháng đang xem
    now = datetime.now(HANOI_TZ)
    month = request.args.get('month', now.month, type=int)
    year = request.args.get('year', now.year, type=int)

    # 2. XÂY DỰNG QUERY LỌC DỮ LIỆU
    query = LichDoanCongTac.query

    if search_tu_ngay:
        query = query.filter(LichDoanCongTac.doan_vao_tu_ngay >= search_tu_ngay)
    if search_den_ngay:
        query = query.filter(LichDoanCongTac.doan_vao_den_ngay <= search_den_ngay)
    if search_truong_doan:
        query = query.filter(LichDoanCongTac.truong_doan.icontains(search_truong_doan))
    if search_don_vi:
        query = query.join(LichDoanCongTac.danh_sach_don_vi).filter(DonVi.ma_hieu_2 == search_don_vi)
    if search_linh_vuc:
        query = query.join(LichDoanCongTac.danh_sach_linh_vuc).filter(LinhVuc.id == search_linh_vuc)

    # Nếu không có tham số tìm kiếm cụ thể, xuất dữ liệu của tháng hiện tại
    if not any([search_tu_ngay, search_den_ngay, search_truong_doan, search_don_vi, search_linh_vuc]):
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
        query = query.filter(and_(LichDoanCongTac.doan_vao_tu_ngay <= end_date, LichDoanCongTac.doan_vao_den_ngay >= start_date))

    results = query.order_by(LichDoanCongTac.doan_vao_tu_ngay).all()

    # 3. CHUYỂN ĐỔI DỮ LIỆU SANG PANDAS DATAFRAME
    data = []
    for index, item in enumerate(results, start=1):
        data.append({
            "STT": index,
            "Tên đoàn công tác": item.ten_doan,
            "Từ ngày": item.doan_vao_tu_ngay.strftime('%d/%m/%Y') if item.doan_vao_tu_ngay else "",
            "Đến ngày": item.doan_vao_den_ngay.strftime('%d/%m/%Y') if item.doan_vao_den_ngay else "",
            "Trưởng đoàn": item.truong_doan,
            "Đơn vị đến làm việc": ", ".join([dv.ten_ma_hieu_2 for dv in item.danh_sach_don_vi]),
            "Lĩnh vực": ", ".join([lv.ten_linh_vuc for lv in item.danh_sach_linh_vuc]),
            "Thành phần tham gia": item.thanh_phan
        })

    df = pd.DataFrame(data)

    # 4. TẠO FILE EXCEL TRONG BỘ NHỚ (Memory-stream)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Lich_Cong_Tac')
        
        # Tùy chỉnh định dạng (độ rộng cột)
        workbook  = writer.book
        worksheet = writer.sheets['Lich_Cong_Tac']
        header_format = workbook.add_format({'bold': True, 'bg_color': '#D7E4BC', 'border': 1})
        
        # Set độ rộng cột cơ bản
        worksheet.set_column('B:B', 40) # Tên đoàn
        worksheet.set_column('C:D', 15) # Ngày
        worksheet.set_column('E:G', 25) # Trưởng đoàn, đơn vị, lĩnh vực
        worksheet.set_column('H:H', 50) # Thành phần

    output.seek(0)
    
    file_name = f"Lich_Cong_Tac_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=file_name
    )

# ----------------------------------------------------------------------
# Module Trắc nghiệm
# ----------------------------------------------------------------------
@app.route('/trac_nghiem')
def trac_nghiem():
    return render_template('trac_nghiem.html')

@app.route('/get_topics')
@login_required
def get_topics():
    try:
        # Lấy danh sách topic không trùng lặp từ bảng questions
        topics_query = db.session.query(Question.topic).distinct().all()
        topics = [row[0] for row in topics_query if row[0]]
        return jsonify(topics)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/get_questions', methods=['POST'])
@login_required
def get_questions():
    try:
        data = request.json
        selected_topics = data.get('topics', {}) # Ví dụ: {"Chủ đề A": 5, "Chủ đề B": 10}
        print(f"DEBUG: Các chủ đề nhận được: {selected_topics}")
        max_total = int(data.get('max_total', 500))
        
        all_selected_questions = []

        # Duyệt qua từng chủ đề người dùng đã chọn
        for topic_name, count in selected_topics.items():
            num_to_take = int(count)
            if num_to_take <= 0:
                continue
                
            # 1. Lấy ngẫu nhiên câu hỏi từ bảng 'questions' theo topic
            # Sử dụng func.rand() cho MySQL hoặc func.random() cho SQLite
            # Sử dụng % xung quanh topic_name nếu muốn tìm kiếm tương đối thay vì questions = Question.query.filter_by(topic=topic_name)\
            questions = Question.query.filter(Question.topic.like(f"%{topic_name}%"))\
                .order_by(func.rand())\
                .limit(num_to_take).all()

            print(f"DEBUG: Chủ đề '{topic_name}' tìm thấy {len(questions)} câu hỏi")
            
            for q in questions:
                if len(all_selected_questions) >= max_total:
                    break
                    
                # 2. Lấy danh sách các lựa chọn từ bảng 'options'
                db_options = Option.query.filter_by(question_id=q.id).all()
                # KHỞI TẠO BIẾN: options_text_list
                options_text_list = [opt.option_text.strip() for opt in db_options if opt.option_text]

                # Nội dung đáp án đúng từ bảng questions
                correct_content = q.correct_answer.strip()
                
                # Kiểm tra an toàn: Nếu đáp án đúng chưa có trong list options thì thêm vào; Đảm bảo đáp án đúng luôn có mặt trong list
                if correct_content not in options_text_list:
                    options_text_list.append(correct_content)
                
                # Đảo đáp án, Xáo trộn đáp án (Cần import random ở đầu file)
                random.shuffle(options_text_list)
                
                # 4. Xác định nhãn mới (A, B, C, D) cho đáp án đúng
                labels = ['A', 'B', 'C', 'D']
                new_correct_label = ""
                for i, text in enumerate(options_text_list):
                    if text == correct_content:
                        new_correct_label = labels[i] # Gán nhãn A, B, C hoặc D
                        break

                all_selected_questions.append({
                    'id': q.id,
                    'content': q.question_text,
                    'options': options_text_list, 
                    'answer': new_correct_label, # Gửi nhãn (VD: "B") về cho Client
                    'explanation': q.explanation
                })

        # Đưa đoạn check rỗng xuống SAU KHI đã chạy vòng lặp tìm kiếm
        # Nếu không có câu hỏi nào, hãy trả về lỗi rõ ràng để Frontend biết
        if not all_selected_questions:
            print("DEBUG: Kết quả cuối cùng rỗng!")
            return jsonify({"error": "Không tìm thấy câu hỏi nào trong CSDL cho các chủ đề đã chọn!"}), 404
        
        # Xáo trộn thứ tự các câu hỏi
        random.shuffle(all_selected_questions)

        return jsonify(all_selected_questions)
        
    except Exception as e:
        print(f"ERROR: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/submit_quiz', methods=['POST'])
@login_required
def submit_quiz():
    try:
        data = request.json
        user_answers = data.get('answers', [])
        duration_seconds = data.get('duration_seconds', 0)
        total_in_quiz = data.get('total_in_quiz') or len(user_answers)
        
        emp_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
        if not emp_info:
            return jsonify({"status": "error", "error": "Không tìm thấy thông tin người lao động"}), 404

        new_result = QuizResult(
            user_id=current_user.ma_nhan_vien,
            ho_ten=emp_info.ho_ten,
            ngay_sinh=emp_info.ngay_sinh,
            gioi_tinh=emp_info.gioi_tinh,
            mail_Agribank=emp_info.mail_Agribank,
            ma_hieu_2=emp_info.ma_hieu_2,
            ma_phong_ban=str(emp_info.ma_phong_ban) if emp_info.ma_phong_ban else None,
            score=0,
            total_questions=total_in_quiz,
            correct_count=0,
            duration_seconds=duration_seconds,
            ip_address=request.remote_addr,
            mac_address=data.get('mac_address', 'N/A')
        )
        
        db.session.add(new_result)
        db.session.flush() 

        correct_count = 0
        detailed_records = []

        for item in user_answers:
            q_id = item.get('question_id')
            if q_id is None: # Sửa lỗi NULL primary key
                continue

            # u_ans: Nhãn người dùng chọn (A, B, C...)
            u_ans = str(item.get('user_answer_text', '')).strip().upper()
            # c_ans: Nhãn đúng gửi từ Frontend (A, B, C...)
            c_ans = str(item.get('correct_answer_label', '')).strip().upper()
            
            # Sửa lỗi LegacyAPIWarning bằng cách dùng Session.get()
            q_origin = db.session.get(Question, q_id)
            if not q_origin: continue

            # SO SÁNH NHÃN VỚI NHÃN
            is_correct = 1 if u_ans == c_ans else 0
            if is_correct:
                correct_count += 1

            detail = QuizQuestionDetail(
                quiz_result_id=new_result.id,
                question_id=q_id,
                question_text=q_origin.question_text,
                user_answer=u_ans, # Lưu nhãn (A, B...)
                correct_answer=c_ans, # Lưu nhãn đúng (A, B...)
                is_correct=is_correct,
                time_spent_on_question=item.get('time_spent', 0)
            )
            detailed_records.append(detail)

        # --- TÍNH TOÁN ĐIỂM VÀ PHẦN TRĂM ---
        if total_in_quiz > 0:
            final_score = round((correct_count / total_in_quiz) * 100, 1)
            # Tính phần trăm (thường giống thang điểm 100 nhưng định dạng là số nguyên hoặc 1 chữ số thập phân)
            percentage = round((correct_count / total_in_quiz) * 100, 1)
        else:
            final_score = 0
            percentage = 0
        
        new_result.score = final_score
        new_result.correct_count = correct_count
        
        db.session.add_all(detailed_records)
        db.session.commit()

        # Trả về thêm field "percentage" để Frontend hiển thị thanh tiến trình hoặc vòng tròn kết quả
        return jsonify({
            "status": "success",
            "score": final_score,
            "percentage": percentage,
            "correct": correct_count,
            "total": total_in_quiz,
            "ho_ten": emp_info.ho_ten,
            "ma_nhan_vien": current_user.ma_nhan_vien,
            "ma_hieu_2": emp_info.ma_hieu_2, 
            "result_id": new_result.id
        })

    except Exception as e:
        db.session.rollback()
        print(f"Lỗi: {str(e)}")
        return jsonify({"status": "error", "error": str(e)}), 500

@app.route('/print_certificate/<int:result_id>')
@login_required
def print_certificate(result_id):
    # Truy vấn kết quả từ Database
    result = QuizResult.query.get_or_404(result_id)
    # Đảm bảo người dùng chỉ in được bài của chính mình
    if result.user_id != current_user.ma_nhan_vien:
        abort(403)
        
    return render_template('certificate_template.html', result=result)

@app.route('/admin/questions')
@login_required
@admin_required
def admin_questions():
    page = request.args.get('page', 1, type=int)
    # Mỗi trang chỉ lấy 50 câu hỏi
    pagination = Question.query.order_by(Question.id.desc()).paginate(page=page, per_page=50)
    questions = pagination.items
    return render_template('admin/questions.html', questions=questions, pagination=pagination)

@app.route('/admin/question/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_question(id):
    try:
        # Sử dụng get() để kiểm tra sự tồn tại #q = Question.query.get_or_404(id)
        q = Question.query.get(id)
        if not q:
            return jsonify({"status": "error", "message": "Câu hỏi không tồn tại"}), 404
        
        # Chỉ cần xóa Question, MySQL sẽ tự động xóa các Options nhờ CASCADE
        db.session.delete(q)
        ## Đã khai báo ON DELETE CASCADE trong MySQL, do đó, không làm theo cách: thực hiện theo cách thủ công là xóa Option trước rồi mới xóa Question
        ## Xóa các options liên quan trước
        ##Option.query.filter_by(question_id=id).delete()
        ##db.session.delete(q)

        db.session.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/admin/export_template')
@login_required
@admin_required
def export_template():
    # Tạo file mẫu để người dùng nhập liệu
    df = pd.DataFrame(columns=['question_text', 'correct_answer', 'topic', 'explanation', 'option_1', 'option_2', 'option_3', 'option_4'])
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Mau_Import')
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="mau_import_cau_hoi.xlsx")

@app.route('/admin/import_questions', methods=['POST'])
@login_required
@admin_required
def import_questions():
    if 'file' not in request.files:
        flash("Không tìm thấy file!", "danger")
        return redirect(url_for('admin_questions'))
    
    file = request.files['file']
    if file.filename == '':
        flash("Chưa chọn file!", "warning")
        return redirect(url_for('admin_questions'))

    try:
        # Đọc file excel
        df = pd.read_excel(file)
        
        # Loại bỏ các dòng hoàn toàn trống để tránh import nhầm dòng trắng
        df = df.dropna(how='all')
        
        # Chuyển đổi các giá trị NaN thành None để Database hiểu là NULL
        df = df.where(pd.notnull(df), None)

        count = 0
        batch_size = 100 # Cứ mỗi 100 câu hỏi sẽ commit một lần để tránh treo server
        
        for index, row in df.iterrows():
            # Kiểm tra dữ liệu bắt buộc (ví dụ câu hỏi không được trống)
            if not row.get('question_text'):
                continue

            # 1. Tạo câu hỏi mới
            new_q = Question(
                question_text=str(row['question_text']).strip(),
                correct_answer=str(row['correct_answer']).strip(),
                topic=str(row['topic']).strip() if row['topic'] else "Chưa phân loại",
                explanation=str(row.get('explanation', '') or '').strip()
            )
            db.session.add(new_q)
            
            # Buộc phải flush để lấy ID cho các Options
            db.session.flush() 

            # 2. Thêm các options (duyệt các cột option_1, 2, 3, 4)
            for i in range(1, 5):
                opt_val = row.get(f'option_{i}')
                if opt_val is not None and str(opt_val).strip() != "":
                    new_opt = Option(
                        question_id=new_q.id, 
                        option_text=str(opt_val).strip()
                    )
                    db.session.add(new_opt)
            
            count += 1
            
            # 3. Batch commit: Cứ 100 dòng thì lưu vào DB một lần
            if count % batch_size == 0:
                db.session.commit()

        # Commit phần dư còn lại
        db.session.commit()
        
        flash(f"Import thành công tổng cộng {count} câu hỏi!", "success")
        
    except Exception as e:
        db.session.rollback()
        # In lỗi chi tiết ra console để bạn kiểm tra
        print(f"DEBUG ERROR: {str(e)}") 
        flash(f"Lỗi hệ thống: {str(e)}", "danger")
    
    return redirect(url_for('admin_questions'))


@app.route('/admin/export_data')
@login_required
@admin_required
def export_data():
    questions = Question.query.all()
    data = []
    for q in questions:
        opts = Option.query.filter_by(question_id=q.id).all()
        row = {
            "ID": q.id,
            "Câu hỏi": q.question_text,
            "Đáp án đúng": q.correct_answer,
            "Chủ đề": q.topic,
            "Giải thích": q.explanation
        }
        for i, o in enumerate(opts):
            row[f"Lựa chọn {i+1}"] = o.option_text
        data.append(row)
    
    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="danh_sach_cau_hoi.xlsx")

@app.route('/admin/get_question_detail/<int:id>')
@login_required
@admin_required
def get_question_detail(id):
    q = Question.query.get_or_404(id)
    opts = Option.query.filter_by(question_id=id).all()
    return jsonify({
        "id": q.id,
        "question_text": q.question_text,
        "correct_answer": q.correct_answer,
        "topic": q.topic,
        "explanation": q.explanation,
        "options": [{"id": o.id, "option_text": o.option_text} for o in opts]
    })

@app.route('/admin/question/update/<int:id>', methods=['POST'])
@login_required
@admin_required
def update_question(id):
    try:
        data = request.json
        q = Question.query.get_or_404(id)
        
        # Cập nhật thông tin câu hỏi
        q.question_text = data['question_text']
        q.topic = data['topic']
        q.correct_answer = data['correct_answer'].strip() # Loại bỏ khoảng trắng thừa
        q.explanation = data.get('explanation')

        # Cập nhật từng option
        updated_option_texts = []
        for opt_data in data['options']:
            opt = Option.query.get(opt_data['id'])
            if opt:
                new_text = opt_data['text'].strip()
                opt.option_text = new_text
                updated_option_texts.append(new_text)

        # 3. Kiểm tra tính nhất quán (Validation)
        if q.correct_answer not in updated_option_texts:
            return jsonify({
                "status": "error", 
                "message": f"Lỗi: Đáp án đúng '{q.correct_answer}' không trùng với bất kỳ phương án nào bên dưới!"
            }), 400

        db.session.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/admin/question/add', methods=['POST'])
@login_required
@admin_required
def add_question():
    try:
        data = request.json

        # Kiểm tra dữ liệu đầu vào cơ bản (Tránh lỗi KeyErrors)
        required_fields = ['question_text', 'correct_answer', 'topic', 'options']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({"status": "error", "message": f"Thiếu thông tin: {field}"}), 400
        
        # 1. Tạo đối tượng câu hỏi
        new_q = Question(
            question_text=data['question_text'],
            correct_answer=data['correct_answer'],
            topic=data['topic'],
            explanation=data.get('explanation')
        )
        db.session.add(new_q)
        db.session.flush() # Để lấy ID của câu hỏi mới cho bảng options # MySQL tạo ID cho new_q

        # 2. Lưu các tùy chọn (Options)
        options_list = data.get('options', [])
        if not isinstance(options_list, list) or len(options_list) < 2:
            return jsonify({"status": "error", "message": "Cần ít nhất 2 phương án trả lời"}), 400

        for opt_text in options_list:
            if opt_text.strip(): # Chỉ lưu nếu option không để trống
                new_opt = Option(
                    question_id=new_q.id,
                    option_text=opt_text.strip()
                )
                db.session.add(new_opt)
            
        db.session.commit()
        return jsonify({"status": "success", "id": new_q.id}) # Trả về cả ID mới
    except Exception as e:
        db.session.rollback()
        # In lỗi ra console để debug khi phát triển
        print(f"Lỗi thêm câu hỏi: {str(e)}")
        return jsonify({"status": "error", "message": str(e)}), 500

# ----------------------------------------------------------------------
# Module Theo dõi hồ sơ khen thưởng
# ----------------------------------------------------------------------
@app.route('/admin/nhom-ho-so')
@login_required
@admin_or_staff_required
def admin_nhom_ho_so():
    ds_nhom = DanhMucNhomHoSo.query.order_by(DanhMucNhomHoSo.ten_nhom_ho_so.desc()).all()
    now_date = date.today().strftime('%Y-%m-%d')
    return render_template('admin/nhom_ho_so.html', ds_nhom=ds_nhom, now_date=now_date)

@app.route('/api/admin/them-nhom-ho-so', methods=['POST'])
@login_required
@admin_or_staff_required
def api_admin_them_nhom_ho_so():
    try:
        data = request.form
        ten_nhom = data.get('ten_nhom_ho_so', '').strip()
        mo_ta = data.get('mo_ta', '').strip()

        if not ten_nhom:
            return jsonify({"status": "error", "message": "Tên nhóm hồ sơ không được để trống"}), 400

        new_nhom = DanhMucNhomHoSo(
            ten_nhom_ho_so=ten_nhom,
            mo_ta=mo_ta,
            trang_thai=True
        )
        db.session.add(new_nhom)
        db.session.commit()
        return jsonify({"status": "success", "message": "Đã tạo nhóm hồ sơ mới thành công!"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/admin/sua-nhom-ho-so/<int:id>', methods=['POST'])
@login_required
@admin_or_staff_required
def api_admin_sua_nhom_ho_so(id):
    try:
        nhom = DanhMucNhomHoSo.query.get_or_404(id)
        data = request.form
        ten_nhom = data.get('ten_nhom_ho_so', '').strip()
        if not ten_nhom:
             return jsonify({"status": "error", "message": "Tên nhóm hồ sơ không được để trống"}), 400
             
        nhom.ten_nhom_ho_so = ten_nhom
        nhom.mo_ta = data.get('mo_ta', '').strip()
        
        db.session.commit()
        return jsonify({"status": "success", "message": "Cập nhật nhóm hồ sơ thành công!"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/admin/danh-muc-giay-to')
@login_required
@admin_or_staff_required
def admin_giay_to():
    # 1. Lấy các tham số lọc từ URL
    nhom_id = request.args.get('nhom_id', type=int)
    trang_thai = request.args.get('trang_thai')
    
    # 2. Khởi tạo query gốc
    query = db.session.query(DanhMucGiayTo, DanhMucNhomHoSo).\
        join(DanhMucNhomHoSo, DanhMucGiayTo.nhom_id == DanhMucNhomHoSo.id)
    
    # 3. Áp dụng logic lọc vào đối tượng 'query'
    if nhom_id:
        query = query.filter(DanhMucGiayTo.nhom_id == nhom_id)
    
    if trang_thai == '1':
        query = query.filter(DanhMucGiayTo.trang_thai == True)
    elif trang_thai == '0':
        query = query.filter(DanhMucGiayTo.trang_thai == False)

    # 4. THỰC THI QUERY (Quan trọng: dùng query.all() thay vì tạo lại query mới)
    ds_giay_to = query.all()
    
    # Lấy danh sách nhóm để hiển thị trong bộ lọc và modal thêm mới
    ds_nhom_active = DanhMucNhomHoSo.query.filter_by(trang_thai=True).all()
    
    #now_date = date.today().strftime('%Y-%m-%d')
    
    return render_template('admin/giay_to.html', 
                           ds_giay_to=ds_giay_to, 
                           ds_nhom=ds_nhom_active, 
                           current_nhom=nhom_id,
                           current_status=trang_thai)

@app.route('/admin/export-giay-to')
@login_required
@admin_or_staff_required
def export_giay_to():
    try:
        # Lấy dữ liệu
        query = db.session.query(DanhMucGiayTo.ten_giay_to, DanhMucNhomHoSo.ten_nhom_ho_so, 
                                 DanhMucGiayTo.so_ngay_quy_dinh, DanhMucGiayTo.trang_thai).\
            join(DanhMucNhomHoSo, DanhMucGiayTo.nhom_id == DanhMucNhomHoSo.id).all()
        
        if not query:
            return "Không có dữ liệu để xuất", 404

        data = []
        for row in query:
            data.append({
                "Tên giấy tờ": row[0],
                "Thuộc nhóm": row[1],
                "Số ngày quy định": row[2],
                "Trạng thái": "Đang hiệu lực" if row[3] else "Ngừng sử dụng"
            })
        
        # Tạo file Excel
        df = pd.DataFrame(data)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='DanhMucGiayTo')
        
        output.seek(0)
        
        # Sửa lỗi attachment_filename -> download_name
        return send_file(
            output, 
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, 
            download_name=f"danh_muc_giay_to_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
    except Exception as e:
        return f"Lỗi xuất Excel: {str(e)}", 500

@app.route('/api/get-giay-to/<int:nhom_id>', methods=['GET']) # Đảm bảo URL này khớp 100% với Script
@login_required
def api_get_giay_to_theo_nhom(nhom_id):
    # Lấy các giấy tờ thuộc nhóm và phải đang hoạt động (trang_thai=True)
    giay_tos = DanhMucGiayTo.query.filter_by(nhom_id=nhom_id, trang_thai=True).all()
    
    # Trả về mảng JSON
    return jsonify([{
        "id": gt.id,
        "ten_giay_to": gt.ten_giay_to,
        "so_ngay_quy_dinh": gt.so_ngay_quy_dinh
    } for gt in giay_tos])

# 1. API Lấy thông tin chi tiết (Dùng cho chức năng Sửa)
@app.route('/api/admin/get-giay-to/<int:id>', methods=['GET'])
@login_required
@admin_or_staff_required
def api_get_giay_to(id):
    gt = DanhMucGiayTo.query.get_or_404(id)
    return jsonify({
        "id": gt.id,
        "ten_giay_to": gt.ten_giay_to,
        "nhom_id": gt.nhom_id,
        "so_ngay_quy_dinh": gt.so_ngay_quy_dinh,
        "la_bat_buoc": gt.la_bat_buoc,
        "ngay_bat_dau": gt.ngay_bat_dau.strftime('%Y-%m-%d') if gt.ngay_bat_dau else "",
        "ngay_ket_thuc": gt.ngay_ket_thuc.strftime('%Y-%m-%d') if gt.ngay_ket_thuc else "",
        "file_mau": gt.file_mau
    })

# 2. API Thêm mới giấy tờ
@app.route('/api/admin/them-giay-to', methods=['POST'])
@login_required
@admin_or_staff_required
def api_admin_them_giay_to():
    ensure_upload_dir() # Đảm bảo có chỗ để lưu
    file_mau_name = None
    try:
        data = request.form
        # 1. Xử lý File mẫu (Tối ưu hóa)
        if 'file_mau' in request.files:
            file = request.files['file_mau']
            if file and file.filename != '':
                if not allowed_file(file.filename):
                    return jsonify({"status": "error", "message": "Định dạng file không hỗ trợ"}), 400
                
                # Sử dụng tên file an toàn
                ext = file.filename.rsplit('.', 1)[1].lower()
                safe_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}"
                file_mau_name = safe_name
                
                path = os.path.join(app.config['UPLOAD_FOLDER_FILE_MAU'], file_mau_name)
                file.save(path)

        # 2. Xử lý logic ngày tháng (Tránh lỗi ValueError)
        ngay_bd_str = data.get('ngay_bat_dau')
        ngay_bd = datetime.strptime(ngay_bd_str, '%Y-%m-%d').date() if ngay_bd_str else date.today()
        
        ngay_kt_str = data.get('ngay_ket_thuc')
        ngay_kt = datetime.strptime(ngay_kt_str, '%Y-%m-%d').date() if ngay_kt_str else None

        # 3. Lưu vào Database
        new_gt = DanhMucGiayTo(
            nhom_id=data.get('nhom_id'),
            ten_giay_to=data.get('ten_giay_to'),
            so_ngay_quy_dinh=int(data.get('so_ngay_quy_dinh', 30)),
            la_bat_buoc=data.get('la_bat_buoc') in ['on', 'true', '1'],
            ngay_bat_dau=ngay_bd,
            ngay_ket_thuc=ngay_kt,
            file_mau=file_mau_name,
            trang_thai=True
        )
        db.session.add(new_gt)
        db.session.commit()
        return jsonify({"status": "success", "message": "Thêm thành công!"})
    except Exception as e:
        db.session.rollback()
        # Nếu lưu db lỗi thì xóa file đã upload để tránh rác
        if file_mau_name:
            try: os.remove(os.path.join(app.config['UPLOAD_FOLDER_FILE_MAU'], file_mau_name))
            except: pass
        return jsonify({"status": "error", "message": f"Lỗi hệ thống: {str(e)}"}), 500

# 3. API Chỉnh sửa giấy tờ
@app.route('/api/admin/sua-giay-to/<int:id>', methods=['POST'])
@login_required
@admin_or_staff_required
def api_admin_sua_giay_to(id):
    try:
        gt = DanhMucGiayTo.query.get_or_404(id)
        data = request.form
        
        # Cập nhật thông tin text
        gt.ten_giay_to = data.get('ten_giay_to')
        gt.nhom_id = data.get('nhom_id')
        gt.so_ngay_quy_dinh = int(data.get('so_ngay_quy_dinh', 30)) #Mặc định số ngày quy định: 30 ngày

        # Logic checkbox: data.get trả về 'on' nếu tích, None nếu không tích
        gt.la_bat_buoc = True if data.get('la_bat_buoc') in ['on', 'true', '1'] else False

        # Xử lý ngày tháng an toàn
        # 2. Xử lý ngày tháng an toàn (Tránh lỗi 500 nếu chuỗi trống)
        ngay_bd_str = data.get('ngay_bat_dau')
        if ngay_bd_str:
            gt.ngay_bat_dau = datetime.strptime(ngay_bd_str, '%Y-%m-%d').date()
        
        ngay_kt_str = data.get('ngay_ket_thuc')
        # Nếu người dùng xóa ngày kết thúc, set về None (Vô thời hạn)
        gt.ngay_ket_thuc = datetime.strptime(ngay_kt_str, '%Y-%m-%d').date() if ngay_kt_str else None

        # Xử lý File mới (nếu có)
        if 'file_mau' in request.files:
            file = request.files['file_mau']
            if file and file.filename != '':
                # Kiểm tra định dạng & dung lượng
                if not allowed_file(file.filename):
                    return jsonify({"status": "error", "message": "Filetype không được phép upload"}), 400
                
                # 2. Xử lý tên file an toàn (slugify_filename hoặc secure_filename)
                filename = slugify_filename(file.filename) #filename = secure_filename(file.filename)
                new_filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}"

                # 3. Đảm bảo thư mục tồn tại
                upload_path = app.config['UPLOAD_FOLDER_FILE_MAU']
                if not os.path.exists(upload_path):
                    os.makedirs(upload_path)

                # 4. Xóa file cũ để tiết kiệm bộ nhớ (Nếu có)
                if gt.file_mau:
                    old_path = os.path.join(upload_path, gt.file_mau)
                    if os.path.exists(old_path):
                        try:
                            os.remove(old_path)
                        except Exception as e:
                            print(f"Không thể xóa file cũ: {e}")

                # 5. Lưu file mới và cập nhật Database
                file.save(os.path.join(upload_path, new_filename))
                gt.file_mau = new_filename
                
        db.session.commit()
        return jsonify({"status": "success", "message": "Cập nhật thành công!"})
    except Exception as e:
        db.session.rollback()
        print(traceback.format_exc()) # In lỗi chi tiết ra console
        return jsonify({"status": "error", "message": str(e)}), 500

# 4. API Bật/Tắt trạng thái
@app.route('/api/admin/toggle-status/<string:type>/<int:id>', methods=['POST'])
@login_required
@admin_or_staff_required
def api_toggle_status(type, id):
    try:
        # Xác định Model cần tác động
        TargetModel = DanhMucNhomHoSo if type == 'nhom' else DanhMucGiayTo
        item = TargetModel.query.get_or_404(id)
        
        # Đảo ngược trạng thái
        item.trang_thai = not item.trang_thai
        db.session.commit()
        
        status_text = "Đã kích hoạt" if item.trang_thai else "Đã ngừng sử dụng"
        return jsonify({
            "status": "success", 
            "new_status": item.trang_thai, 
            "message": status_text
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

# ----------------------------------------------------------------------
# Đơn vị lập hồ sơ (Dự thảo Draft), tải lên Upload, Gửi kiểm soát => (Chờ kiểm soát Pending Control);
# Cán bộ chuyên trách rà soát hồ sơ (nút Tiếp nhận; Trả lại; Trình phê duyệt): đúng đủ => Trình phê duyệt; thiếu sót => Yêu cầu sửa đổi bổ sung Revisions Required, nhập y_kien_kiem_soat => Đơn vị nhận thông báo để sửa.
# Lãnh đạo phê duyệt: Đã duyệt hoặc Từ chối.
# ----------------------------------------------------------------------
@app.route('/admin/ho-so-permissions')
@login_required
def admin_ho_so_permissions():
    if not current_user.is_admin:
        flash("Bạn không có quyền truy cập trang này!", "danger")
        return redirect(url_for('index'))
    
    try:
        # Lấy danh sách đơn vị từ bảng ThongTinNguoiLaoDong (để lọc theo nhân viên)
        don_vis_query = db.session.query(distinct(ThongTinNguoiLaoDong.ma_hieu_2))\
                            .filter(ThongTinNguoiLaoDong.ma_hieu_2 != None)\
                            .order_by(ThongTinNguoiLaoDong.ma_hieu_2).all()
        don_vis = [dv[0] for dv in don_vis_query]

        # Lấy danh sách đầy đủ đơn vị từ bảng DonVi (để chọn trong Modal cấp quyền hồ sơ)
        # 3. Lấy danh sách Đơn vị từ bảng DonVi để hiển thị trong Modal (cấp quyền xử lý)
        # Lấy các đơn vị đang hoạt động để gán quyền chính xác
        ds_don_vi = DonVi.query.filter_by(trang_thai='Hoạt động').order_by(DonVi.ten_ma_hieu_2.asc()).all()
        
    except Exception as e:
        print(f"Lỗi truy vấn dữ liệu phân quyền hồ sơ: {e}")
        don_vis = []
        ds_don_vi = []

    # Danh sách Slugs menu (hiển thị dưới dạng checkbox trong bảng)
    managed_slugs = [
        {'code': 'theo-doi-ho-so', 'name': 'Theo dõi hồ sơ khen thưởng'},
        {'code': 'admin/nhom-ho-so', 'name': 'Quản trị Nhóm hồ sơ'},
        {'code': 'admin/danh-muc-giay-to', 'name': 'Quản trị Giấy tờ'}
    ]
    
    return render_template('admin/admin_ho_so_permissions.html',
                           ds_don_vi=ds_don_vi, 
                           don_vis=don_vis,
                           managed_slugs=managed_slugs)

def has_permission(perm_code, unit_code):
    """
    Kiểm tra quyền của user hiện tại.
    Admin hệ thống (is_admin=1) luôn có quyền.
    """
    if current_user.is_admin:
        return True
    if not current_user.is_authenticated:
        return False

    # Nếu kiểm tra quyền tổng quát (để hiện nút bấm)
    if unit_code == 'ANY':
        sql = text("""
            SELECT 1 FROM user_unit_permissions 
            WHERE ma_nhan_vien = :mnv AND permission_code = :code
        """)
        params = {'mnv': current_user.ma_nhan_vien, 'code': perm_code}
    else:
        # Kiểm tra quyền cụ thể tại 1 đơn vị hoặc quyền ALL toàn hệ thống
        sql = text("""
            SELECT 1 FROM user_unit_permissions 
            WHERE ma_nhan_vien = :mnv 
            AND permission_code = :code 
            AND (ma_hieu_2 = :unit OR ma_hieu_2 = 'ALL')
        """)
        params = {'mnv': current_user.ma_nhan_vien, 'code': perm_code, 'unit': unit_code}
    
    perm = db.session.execute(sql, params).fetchone()
    return perm is not None
    
@app.context_processor
def inject_permissions():
    def can_access_bc48():
        if not current_user.is_authenticated: return False
        if current_user.is_admin: return True
        # Kiểm tra bảng trong DB bc48
        return PhanQuyenModuleBC48.query.filter_by(
            ma_nhan_vien=current_user.ma_nhan_vien, 
            module_slug='bc48', 
            trang_thai=True
        ).first() is not None
        
    return dict(has_permission=has_permission, can_access_bc48=can_access_bc48)

# --- Modal phân quyền [HS_LAP: Lập hồ sơ & Gửi kiểm soát; HS_KS: Kiểm soát hồ sơ (Trả lại/Trình duyệt); HS_PD: Phê duyệt hồ sơ (Duyệt/Từ chối)] ---
@app.route('/api/admin/update-ho-so-permission', methods=['POST'])
@login_required
def update_ho_so_permission():
    if not current_user.is_admin:
        return jsonify({"status": "error", "message": "Không có quyền"}), 403
    
    data = request.get_json() # Lấy dữ liệu từ AJAX JSON.stringify
    if not data:
        return jsonify({"status": "error", "message": "Dữ liệu không hợp lệ"}), 400
    
    mnv = data.get('ma_nhan_vien')
    perm_code = data.get('permission_code') # HS_LAP; HS_KS; HS_PD
    unit_code = data.get('ma_hieu_2')      
    action = data.get('action')            # 'grant' hoặc 'revoke'

    try:
        if action == 'grant':
            # Thêm quyền (dùng INSERT IGNORE hoặc kiểm tra trước khi chèn)
            db.session.execute(text("""
                INSERT INTO user_unit_permissions (ma_nhan_vien, permission_code, ma_hieu_2)
                VALUES (:mnv, :code, :unit)
                ON DUPLICATE KEY UPDATE ma_hieu_2 = ma_hieu_2
            """), {'mnv': mnv, 'code': perm_code, 'unit': unit_code})
        else:
            # Thu hồi quyền (Delete)
            db.session.execute(text("""
                DELETE FROM user_unit_permissions 
                WHERE ma_nhan_vien = :mnv AND permission_code = :code AND ma_hieu_2 = :unit
            """), {'mnv': mnv, 'code': perm_code, 'unit': unit_code})
    
        db.session.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

def add_ho_so_log(ho_so_id, hanh_dong, status_before, status_after, comment):
    """Hàm tiện ích để ghi nhật ký luồng hồ sơ"""
    try:
        new_log = HoSoLog(
            ho_so_id=ho_so_id,
            ma_nhan_vien=current_user.ma_nhan_vien,
            hanh_dong=hanh_dong,
            trang_thai_truoc=status_before,
            trang_thai_sau=status_after,
            noi_dung_y_kien=comment,
            created_at=datetime.now()
        )
        db.session.add(new_log)
    except Exception as e:
        print(f"Lỗi khi ghi Log: {str(e)}")

@app.route('/api/ho-so-logs/<int:ho_so_id>')
@login_required
def get_ho_so_logs(ho_so_id):
    # Sử dụng joinedload để lấy luôn thông tin User qua relationship đã khai báo
    logs = HoSoLog.query.options(db.joinedload(HoSoLog.user))\
        .filter_by(ho_so_id=ho_so_id)\
        .order_by(HoSoLog.created_at.desc()).all()
    
    results = []
    for log in logs:
        results.append({
            "time": log.created_at.strftime('%d/%m/%Y %H:%M:%S'),
            # Truy cập fullname thông qua relationship 'user'
            "user": log.user.fullname if log.user else "Hệ thống",
            "action": log.hanh_dong,
            "comment": log.noi_dung_y_kien,
            "status_step": f"{log.trang_thai_truoc} -> {log.trang_thai_sau}"
        })
    return jsonify(results)

def send_notification(user_id, title, message, link=None):
    """Lưu thông báo vào database cho người dùng"""
    try:
        new_noti = Notification(
            ma_nhan_vien=user_id,
            tieu_de=title,
            noi_dung=message,
            duong_dan=link,
            is_read=False,
            created_at=datetime.now()
        )
        db.session.add(new_noti)
        # Lưu ý: Không commit ở đây mà để commit chung với transaction của hàm gọi nó
    except Exception as e:
        print(f"Lỗi gửi thông báo: {str(e)}")


@app.route('/api/notifications')
@login_required
def get_notifications():
    # Lấy 10 thông báo mới nhất
    notifications = Notification.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien)\
        .order_by(Notification.created_at.desc()).limit(10).all()
    
    # Đếm số thông báo chưa đọc
    unread_count = Notification.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien, is_read=False).count()
    
    return jsonify({
        "unread_count": unread_count,
        "notifications": [{
            "id": n.id,
            "tieu_de": n.tieu_de,
            "noi_dung": n.noi_dung,
            "link": n.duong_dan or "#",
            "is_read": n.is_read,
            "time": n.created_at.strftime('%d/%m %H:%M')
        } for n in notifications]
    })

@app.route('/api/notifications/read-all', methods=['POST'])
@login_required
def read_all_notifications():
    Notification.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien, is_read=False).update({"is_read": True})
    db.session.commit()
    return jsonify({"status": "success"})


@app.route('/theo-doi-ho-so')
@login_required
def theo_doi_ho_so_khen_thuong():
    # --- PHẦN 1: KIỂM TRA QUYỀN TRUY CẬP MENU ---
    # Nếu là Admin hệ thống thì cho qua, nếu không phải thì check bảng UserMenuPermission
    if not current_user.is_admin:
        has_perm = UserMenuPermission.query.filter_by(
            ma_nhan_vien=current_user.ma_nhan_vien, 
            menu_slug='theo-doi-ho-so'
        ).first()
        if not has_perm:
            flash("Bạn không có quyền sử dụng menu chức năng này!", "danger")
            return redirect(url_for('index'))
    
    # --- PHẦN 2: LẤY THÔNG TIN ĐƠN VỊ CỦA USER ĐỂ PHÂN PHẠM VI ---
    user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    user_ma_hieu_2 = user_info.ma_hieu_2 if user_info else None
    user_ma_pb = str(user_info.ma_phong_ban) if user_info and user_info.ma_phong_ban else None
    user_ten_pb = user_info.phong_ban.ten_phong_ban if user_info and user_info.phong_ban else ""

    # Lấy danh sách các mã quyền (permission_code) của user
    user_perms_query = db.session.execute(text("""
        SELECT permission_code FROM user_unit_permissions 
        WHERE ma_nhan_vien = :mnv
    """), {'mnv': current_user.ma_nhan_vien}).fetchall()
    user_perms = [r[0] for r in user_perms_query]
    
    # Quyền Admin hoặc thuộc các phòng ban đặc biệt (1: Ban Giám đốc; 2: Tổng hợp; 3: Kiểm tra, giám sát; 4: Thu thập và Xử lý thông tin)
    is_system_admin = (current_user.is_admin == 1)
    is_phong_tong_hop = (user_ma_pb == '2') # Phòng Tổng hợp có mã phòng ban là '2'
    is_admin_or_th = (is_system_admin or is_phong_tong_hop) #is_admin_or_th = (current_user.is_admin == 1 or user_ma_pb in [2])
    # LOGIC BẢO MẬT: Nếu không phải Admin/TH, cưỡng ép lọc theo tên phòng ban của User
    if not is_admin_or_th:
        search_phong_ban = user_ten_pb

    # Quyềnn: Kiểm soát hoặc Phê duyệt
    has_ks_pd = any(p in user_perms for p in ['HS_KS', 'HS_PD'])
    # Biến kiểm soát việc không giới hạn phạm vi đơn vị
    is_unrestricted = (is_system_admin or is_phong_tong_hop or has_ks_pd)
        
    try:
        # Sử dụng joinedload để tối ưu query (Eager Loading)
        query = db.session.query(TheoDoiHoSo).options(
            joinedload(TheoDoiHoSo.loai_giay_to),
            joinedload(TheoDoiHoSo.don_vi_rel)
        )

        # LOGIC BẢO MẬT: Nếu không thuộc nhóm không giới hạn, chỉ thấy hồ sơ đơn vị mình
        if not is_unrestricted:
            query = query.filter(TheoDoiHoSo.ma_hieu_2 == user_ma_hieu_2)

        danh_sach_query = query.order_by(TheoDoiHoSo.han_chot.asc()).all()

        # --- PHẦN 4: PHÂN QUYỀN DANH SÁCH ĐƠN VỊ TRÊN MODAL (Để chọn khi lập mới) ---
        if is_unrestricted:
            # Admin/KS/PD được quyền chọn tất cả đơn vị để lập hộ hoặc quản lý
            ds_don_vi = DonVi.query.filter_by(trang_thai='Hoạt động').order_by(DonVi.ten_ma_hieu_2.asc()).all()
        else:
            # User chỉ có quyền HS_LAP: Cưỡng ép chỉ được chọn chính đơn vị của mình
            ds_don_vi = DonVi.query.filter_by(ma_hieu_2=user_ma_hieu_2, trang_thai='Hoạt động').all()

        # 5. DỮ LIỆU BỔ TRỢ
        # Quan trọng: Danh sách Nhóm hồ sơ đang hoạt động để chọn trước
        # Danh sách tất cả giấy tờ (vẫn gửi kèm nếu cần dùng cho filter khác hoặc mặc định)
        ds_nhom_ho_so = DanhMucNhomHoSo.query.filter_by(trang_thai=True).all()
        ds_giay_to = DanhMucGiayTo.query.filter_by(trang_thai=True).all()

        # Lấy ngày hiện tại kiểu date object để so sánh quá hạn trong HTML
        today_obj = date.today() 
        # Lấy ngày hiện tại kiểu string để gán giá trị mặc định cho input HTML
        now_date_str = today_obj.strftime('%Y-%m-%d') #Lấy ngày hiện tại định dạng YYYY-MM-DD để gán giá trị mặc định cho input date
        
        return render_template('theo_doi_ho_so.html', 
                               danh_sach=danh_sach_query, 
                               ds_don_vi=ds_don_vi,
                               user_ma_hieu_2=user_ma_hieu_2,
                               ds_nhom_ho_so=ds_nhom_ho_so, 
                               ds_giay_to=ds_giay_to,
                               now_date=now_date_str,      # Dùng cho value của input date
                               now_date_obj=today_obj,     # Dùng để check quá hạn (màu đỏ)
                               is_admin_or_th=is_unrestricted)
    except Exception as e:
        db.session.rollback()
        print(f"Lỗi truy vấn hồ sơ: {str(e)}")
        flash(f"Lỗi: {str(e)}", "danger")
        return redirect(url_for('index'))
    

@app.route('/api/them-nhieu-ho-so', methods=['POST'])
@login_required
def api_them_nhieu_ho_so():
    try:
        data = request.form
        ma_hieu_2 = data.get('ma_hieu_2')
        # Lấy hành động từ frontend: 'draft' hoặc 'send'
        workflow_action = data.get('workflow_action', 'draft') 

        # 1. KIỂM TRA QUYỀN TRUY CẬP DỮ LIỆU (Security Check)
        user_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
        user_ma_hieu_2 = user_info.ma_hieu_2 if user_info else None

        # Kiểm tra nếu không phải Admin/KS/PD mà lại nộp đơn vị khác đơn vị mình
        is_unrestricted = current_user.is_admin or has_permission('HS_KS', ma_hieu_2) or has_permission('HS_PD', ma_hieu_2)

        if not is_unrestricted and ma_hieu_2 != user_ma_hieu_2:
            return jsonify({"status": "error", "message": "Bạn chỉ được phép lập hồ sơ cho đơn vị của chính mình."}), 403

        if not has_permission('HS_LAP', ma_hieu_2):
             return jsonify({"status": "error", "message": "Bạn không có quyền lập hồ sơ tại đơn vị này."}), 403

        # LẤY THÔNG TIN CHI TIẾT CỦA NGƯỜI DÙNG ĐANG ĐĂNG NHẬP
        ho_ten_gan = user_info.ho_ten if user_info else current_user.fullname
        chuc_vu_gan = user_info.chuc_vu if user_info else "N/A"
        
        ngay_phat_sinh_str = data.get('ngay_phat_sinh')
        ghi_chu_tong = data.get('ghi_chu_tong') 
        giay_to_ids = request.form.getlist('giay_to_id[]')
        files = request.files.getlist('file_dinh_kem[]')

        if not ma_hieu_2 or not giay_to_ids or not ngay_phat_sinh_str:
            return jsonify({"status": "error", "message": "Vui lòng nhập đầy đủ thông tin đơn vị và ít nhất một loại giấy tờ"}), 400

        ngay_phat_sinh = datetime.strptime(ngay_phat_sinh_str, '%Y-%m-%d').date()
        
        # Đảm bảo thư mục riêng của đơn vị tồn tại
        unit_path = os.path.join(app.config['UPLOAD_FOLDER_HO_SO'], ma_hieu_2)
        os.makedirs(unit_path, exist_ok=True)

        # XÁC ĐỊNH TRẠNG THÁI VÀ LOG DỰA TRÊN HÀNH ĐỘNG
        if workflow_action == 'send':
            target_status = 'Chờ kiểm soát'
            log_action = 'Khởi tạo & Trình kiểm soát'
            msg_success = "Đã lưu và gửi hồ sơ đi kiểm soát thành công."
        else:
            target_status = 'Dự thảo'
            log_action = 'Khởi tạo hồ sơ'
            msg_success = "Đã lưu bản nháp hồ sơ thành công."

        new_ids = [] 
        records_created = 0
        now_ts = datetime.now()

        # 2. Lặp qua từng loại giấy tờ được chọn
        for i, gt_id in enumerate(giay_to_ids):
            if not gt_id: continue
            loai_gt = db.session.get(DanhMucGiayTo, gt_id)
            if not loai_gt: continue
            
            # Xử lý file đính kèm
            file_db_path = None
            if i < len(files):
                file = files[i]
                if file and file.filename != '' and allowed_file(file.filename):
                    clean_name = slugify_filename(file.filename)
                    filename = f"{now_ts.strftime('%Y%m%d_%H%M%S')}_{gt_id}_{clean_name}"
                    file.save(os.path.join(unit_path, filename))
                    file_db_path = f"uploads/ho_so_khen_thuong/{ma_hieu_2}/{filename}"

            # 3. Tính hạn chót & Tạo object bản ghi
            han_chot = ngay_phat_sinh + timedelta(days=loai_gt.so_ngay_quy_dinh)
            
            new_record = TheoDoiHoSo(
                ma_hieu_2=ma_hieu_2,
                giay_to_id=gt_id,
                ngay_phat_sinh=ngay_phat_sinh,
                han_chot=han_chot,
                trang_thai=target_status, 
                ghi_chu=ghi_chu_tong,
                nguoi_nop=current_user.ma_nhan_vien, 
                ho_ten_nguoi_nop=ho_ten_gan,      
                chuc_vu_nguoi_nop=chuc_vu_gan,    
                ngay_gui_kiem_soat=now_ts if workflow_action == 'send' else None,
                meta_data={"file_dinh_kem": file_db_path} if file_db_path else {}
            )
            db.session.add(new_record)
            db.session.flush()

            # Ghi Log lịch sử
            new_log = HoSoLog(
                ho_so_id=new_record.id,
                ma_nhan_vien=current_user.ma_nhan_vien,
                hanh_dong=log_action,
                trang_thai_truoc=None,
                trang_thai_sau=target_status,
                noi_dung_y_kien=ghi_chu_tong
            )
            db.session.add(new_log)
            new_ids.append(new_record.id)
            records_created += 1

        db.session.commit()
        return jsonify({
            "status": "success", 
            "message": f"{msg_success} (Tổng cộng {records_created} loại giấy tờ)",
            "data": {"ids": new_ids, "count": records_created}
        })

    except Exception as e:
        db.session.rollback()
        print(f"Lỗi API thêm nhiều hồ sơ: {str(e)}")
        return jsonify({"status": "error", "message": f"Lỗi hệ thống: {str(e)}"}), 500


@app.route('/api/xoa-ho-so/<int:id>', methods=['DELETE'])
@login_required
def api_xoa_ho_so(id):
    ho_so = TheoDoiHoSo.query.get_or_404(id)
    
    # Lưu lại thông tin cần thiết trước khi xóa hồ sơ
    ma_nv_lap = ho_so.nguoi_nop  # Trường 'nguoi_nop' lưu mã nhân viên
    ma_ho_so_str = f"HS{ho_so.id:06d}"
    ten_loai_hoso = ho_so.loai_giay_to.ten_giay_to if ho_so.loai_giay_to else "Hồ sơ"
    
    # Kiểm tra quyền: Chỉ người lập hoặc Admin mới được xóa
    if ma_nv_lap != current_user.ma_nhan_vien and not current_user.is_admin:
        return jsonify({"status": "error", "message": "Bạn không có quyền xóa hồ sơ này"}), 403
        
    # Kiểm tra trạng thái: Chỉ cho phép xóa hồ sơ "Dự thảo"
    if ho_so.trang_thai.strip() != 'Dự thảo' and not current_user.is_admin:
        return jsonify({"status": "error", "message": "Hồ sơ đã đi vào quy trình, không thể xóa!"}), 400

    try:
        # Xóa file vật lý nếu có
        if ho_so.meta_data and ho_so.meta_data.get('file_dinh_kem'):
            file_path = os.path.join(current_app.root_path, 'static', ho_so.meta_data['file_dinh_kem'])
            if os.path.exists(file_path):
                os.remove(file_path)
                
        # 1. Gửi thông báo nếu người xóa là Admin và không phải là người tạo hồ sơ
        if current_user.is_admin and ma_nv_lap != current_user.ma_nhan_vien:
            send_notification(
                user_id=ma_nv_lap,
                title="Hồ sơ đã bị xóa bởi Admin",
                message=f"Hồ sơ {ma_ho_so_str} ({ten_loai_hoso}) của bạn đã bị quản trị viên xóa.",
                link=None  # Không để link vì hồ sơ đã bị xóa, không thể xem lại
            )

        # 2. Xóa các Log liên quan (nếu có ràng buộc khóa ngoại)
        # user 'write' không có quyền xóa trực tiếp bảng ho_so_log thì bỏ dòng này
        #HoSoLog.query.filter_by(ho_so_id=id).delete()
        
        # 3. Xóa hồ sơ chính
        # Nhờ 'ON DELETE CASCADE', MySQL sẽ tự dọn dẹp bảng ho_so_logs cho bạn.
        db.session.delete(ho_so)
        
        # 4. Commit tất cả các thay đổi (Bao gồm cả thông báo và xóa)
        db.session.commit()
        
        return jsonify({"status": "success", "message": "Đã xóa hồ sơ thành công"})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/cap-nhat-trang-thai/<int:id>', methods=['POST'])
@login_required
def api_cap_nhat_trang_thai(id):
    try:
        ho_so = TheoDoiHoSo.query.get_or_404(id)
        
        if ho_so.trang_thai == 'Đã nộp':
            return jsonify({"status": "info", "message": "Hồ sơ này đã được nộp trước đó"}), 200

        ho_so.trang_thai = 'Đã nộp'
        # Dùng .date() nếu bạn chỉ muốn lưu ngày nộp, hoặc datetime.now() cho timestamp
        ho_so.ngay_nop_thuc_te = datetime.now() 
        
        db.session.commit()
        
        # Trả về thêm thông tin để frontend cập nhật UI mà không cần reload trang
        return jsonify({
            "status": "success", 
            "message": "Đã ghi nhận nộp hồ sơ thành công!",
            "ngay_nop": ho_so.ngay_nop_thuc_te.strftime('%d/%m/%Y %H:%M')
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/gui-kiem-soat/<int:id>', methods=['POST'])
@login_required
def api_gui_kiem_soat(id):
    ho_so = TheoDoiHoSo.query.get_or_404(id)
    
    # Kiểm tra: User có phải người của đơn vị đó và có quyền Lập không?
    if not has_permission('HS_LAP', ho_so.ma_hieu_2):
        return jsonify({"status": "error", "message": "Bạn không có quyền gửi hồ sơ cho đơn vị này"}), 403

    ho_so.trang_thai = 'Chờ kiểm soát'
    ho_so.ngay_gui_kiem_soat = datetime.now()

    # Ghi Log xử lý hồ sơ
    add_ho_so_log(id, "Gửi hồ sơ kiểm soát", status_before, ho_so.trang_thai, "Đơn vị trình hồ sơ")
    
    db.session.commit()
    return jsonify({"status": "success", "message": "Đã gửi kiểm soát"})


@app.route('/api/kiem-soat-ho-so/<int:id>', methods=['POST'])
@login_required
def api_kiem_soat(id):
    ho_so = TheoDoiHoSo.query.get_or_404(id)

    # CHÈN THÊM: Kiểm tra trạng thái hồ sơ có đang ở bước Kiểm soát không
    if ho_so.trang_thai != 'Chờ kiểm soát':
        return jsonify({"status": "error", "message": "Hồ sơ không ở trạng thái chờ kiểm soát"}), 400
    
    if not has_permission('HS_KS', ho_so.ma_hieu_2):
        return jsonify({"status": "error", "message": "Bạn không được phân công kiểm soát đơn vị này"}), 403
    
    data = request.json
    status_before = ho_so.trang_thai
    y_kien = data.get('y_kien', '')
    
    if data.get('action') == 'approve':
        ho_so.trang_thai = 'Chờ phê duyệt'
        hanh_dong = 'Kiểm soát: Chấp nhận trình duyệt'
    else:
        ho_so.trang_thai = 'Yêu cầu sửa đổi' # Trả lại đơn vị
        hanh_dong = 'Kiểm soát: Trả lại yêu cầu sửa đổi'
        
    #ho_so.nguoi_kiem_soat = current_user.ma_nhan_vien
    ho_so.nguoi_kiem_soat = current_user.fullname if hasattr(current_user, 'fullname') else current_user.username
    ho_so.y_kien_kiem_soat = y_kien

    # Ghi log trước khi commit
    add_ho_so_log(id, hanh_dong, status_before, ho_so.trang_thai, y_kien)
    
    db.session.commit()
    return jsonify({"status": "success"})


@app.route('/api/phe-duyet-ho-so/<int:id>', methods=['POST'])
@login_required
def api_phe_duyet(id):
    try:
        ho_so = TheoDoiHoSo.query.get_or_404(id)

        # 1. Kiểm tra quyền Phê duyệt (HS_PD)
        if not has_permission('HS_PD', ho_so.ma_hieu_2):
            return jsonify({"status": "error", "message": "Bạn không có quyền phê duyệt hồ sơ đơn vị này"}), 403

        # 2. Kiểm tra trạng thái hồ sơ
        if ho_so.trang_thai != 'Chờ phê duyệt':
            return jsonify({"status": "error", "message": "Hồ sơ chưa được trình phê duyệt"}), 400

                    
        data = request.json
        action = data.get('action') # 'duyet' hoặc 'tu_choi'
        y_kien = data.get('y_kien', '')
        status_before = ho_so.trang_thai

        if action == 'duyet':
            ho_so.trang_thai = 'Đã duyệt'
            ho_so.ngay_nop_thuc_te = datetime.now()
            hanh_dong = "Lãnh đạo phê duyệt: ĐỒNG Ý"
        else:
            ho_so.trang_thai = 'Từ chối'
            hanh_dong = "Lãnh đạo phê duyệt: TỪ CHỐI"

        # Lưu thông tin người duyệt (Dùng fullname hoặc username tùy hệ thống của bạn)
        ho_so.nguoi_phe_duyet = current_user.fullname if hasattr(current_user, 'fullname') else current_user.username
        ho_so.ngay_phe_duyet = datetime.now()
        ho_so.y_kien_phe_duyet = y_kien

        # Ghi log xử lý hồ sơ
        add_ho_so_log(id, hanh_dong, status_before, ho_so.trang_thai, y_kien)

        db.session.commit()

        # Gửi thông báo cho người lập hồ sơ (giả sử bạn có hàm send_notification)
        if action == 'duyet':
            send_notification(ho_so.ma_nhan_vien_lap, f"Hồ sơ {ho_so.id} của bạn đã được phê duyệt!")
        else:
            send_notification(ho_so.ma_nhan_vien_lap, f"Hồ sơ {ho_so.id} bị từ chối. Lý do: {y_kien}")
        
        return jsonify({"status": "success", "message": "Đã cập nhật quyết định phê duyệt"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/stats-ho-so')
@login_required
def api_stats():
    # Lấy danh sách mã đơn vị được quyền xem
    allowed_units_query = db.session.execute(text(
        "SELECT ma_hieu_2 FROM user_unit_permissions WHERE ma_nhan_vien = :mnv"
    ), {'mnv': current_user.ma_nhan_vien}).fetchall()
    allowed_units = [r[0] for r in allowed_units_query]

    query_stats = db.session.query(TheoDoiHoSo.trang_thai, db.func.count(TheoDoiHoSo.id))
    query_overdue = TheoDoiHoSo.query.filter(TheoDoiHoSo.trang_thai != 'Đã duyệt', TheoDoiHoSo.han_chot < date.today())

    if not current_user.is_admin and 'ALL' not in allowed_units:
        query_stats = query_stats.filter(TheoDoiHoSo.ma_hieu_2.in_(allowed_units))
        query_overdue = query_overdue.filter(TheoDoiHoSo.ma_hieu_2.in_(allowed_units))

    stats = query_stats.group_by(TheoDoiHoSo.trang_thai).all()
    overdue_count = query_overdue.count()
    
    return jsonify({
        "status_summary": dict(stats),
        "overdue_count": overdue_count
    })


@app.route('/export-ho-so-excel')
@login_required
def export_ho_so_excel():
    # Lấy tham số lọc từ URL (Query String)
    ma_hieu_2 = request.args.get('ma_hieu_2')
    trang_thai = request.args.get('trang_thai')
    tu_ngay = request.args.get('tu_ngay')
    den_ngay = request.args.get('den_ngay')

    query = TheoDoiHoSo.query
    
    # Áp dụng bộ lọc
    if ma_hieu_2: query = query.filter(TheoDoiHoSo.ma_hieu_2 == ma_hieu_2)
    if trang_thai: query = query.filter(TheoDoiHoSo.trang_thai == trang_thai)
    if tu_ngay: query = query.filter(TheoDoiHoSo.ngay_phat_sinh >= tu_ngay)
    if den_ngay: query = query.filter(TheoDoiHoSo.ngay_phat_sinh <= den_ngay)

    data = query.all()
    
    # Chuyển đổi sang List Dictionary để đưa vào Pandas
    export_data = []
    for h in data:
        export_data.append({
            "Mã Hồ Sơ": f"HS{h.id:06d}",
            "Đơn Vị": h.ma_hieu_2,
            "Loại Hồ Sơ": h.loai_giay_to.ten_giay_to if h.loai_giay_to else "",
            "Ngày Phát Sinh": h.ngay_phat_sinh.strftime('%d/%m/%Y') if h.ngay_phat_sinh else "",
            "Hạn Chót": h.han_chot.strftime('%d/%m/%Y') if h.han_chot else "",
            "Trạng Thái": h.trang_thai,
            "Người Xử Lý": h.nguoi_kiem_soat or h.nguoi_phe_duyet or ""
        })

    df = pd.DataFrame(export_data)
    
    # Ghi vào buffer để gửi về trình duyệt mà không cần lưu file tạm trên server
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='DanhSachHoSo')
    output.seek(0)

    return send_file(output, 
                     attachment_filename=f"Danh_sach_ho_so_{date.today()}.xlsx", 
                     as_attachment=True)

# ----------------------------------------------------------------------
# Module NSD yêu cầu mẫu biểu, liệt kê mẫu biểu đính kèm yêu cầu
# ----------------------------------------------------------------------
@app.route('/admin/danh-muc-mau', methods=['GET', 'POST'])
@login_required
def quan_ly_danh_muc():
    if not current_user.is_admin: abort(403)
    
    if request.method == 'POST':
        ten_dm = request.form.get('ten_danh_muc')
        mo_ta = request.form.get('mo_ta')
        if ten_dm:
            # Kiểm tra trùng tên
            ton_tai = DanhMucMau.query.filter_by(ten_danh_muc=ten_dm).first()
            if ton_tai:
                flash('Tên danh mục này đã tồn tại!', 'warning')
            else:
                moi = DanhMucMau(ten_danh_muc=ten_dm, mo_ta=mo_ta)
                db.session.add(moi)
                db.session.commit()
                flash(f'Đã thêm danh mục: {ten_dm}', 'success')
    
    ds_dm = DanhMucMau.query.order_by(DanhMucMau.id.desc()).all()
    return render_template('admin/admin_danh_muc.html', ds_dm=ds_dm)

@app.route('/admin/xoa-danh-muc/<int:id>')
@login_required
def xoa_danh_muc(id):
    if not current_user.is_admin: abort(403)
    dm = DanhMucMau.query.get_or_404(id)
    # Kiểm tra xem có gói mẫu nào đang dùng danh mục này không
    co_goi_mau = YeuCauMau.query.filter_by(danh_muc_id=id).first()
    if co_goi_mau:
        flash('Không thể xóa danh mục đang có gói mẫu sử dụng!', 'danger')
    else:
        db.session.delete(dm)
        db.session.commit()
        flash('Đã xóa danh mục', 'success')
    return redirect(url_for('quan_ly_danh_muc'))


@app.route('/admin/tao-goi-mau', methods=['POST'])
@login_required
def admin_tao_goi_mau():
    if not current_user.is_admin: abort(403)
    
    ten_goi = request.form.get('ten_goi')
    # Lấy danh_muc_id từ select box thay vì string
    dm_id = request.form.get('danh_muc_id')
    selected_mau_ids = request.form.getlist('mau_bieu_ids[]')

    if not ten_goi or not selected_mau_ids:
        flash('Vui lòng nhập tên gói và chọn ít nhất một mẫu biểu!', 'warning')
        return redirect(url_for('admin_quan_ly_yeu_cau'))
    
    if ten_goi and selected_mau_ids:
        # Tạo một yêu cầu đặc biệt không thuộc về ai
        new_package = YeuCauMau(
            noi_dung_yeu_cau=ten_goi,
            danh_muc_id=int(dm_id) if dm_id else None,
            trang_thai='Gói mẫu có sẵn',
            is_template=True,
            user_id=None # Không thuộc nhân viên nào cụ thể
        )
        # Gán các mẫu biểu vào gói
        new_package.danh_sach_mau = MauBieu.query.filter(MauBieu.id.in_(selected_mau_ids)).all()
        db.session.add(new_package)
        db.session.commit()
        ten_dm = new_package.danh_muc_obj.ten_danh_muc if new_package.danh_muc_obj else "Chưa phân loại"
        flash(f'Đã tạo gói mẫu [{ten_goi}] thuộc danh mục [{ten_dm}]', 'success')
        
    return redirect(url_for('admin_quan_ly_yeu_cau'))

@app.route('/admin/mau_bieu', methods=['GET', 'POST'])
@login_required
def quan_ly_mau_bieu():
    if not current_user.is_admin: abort(403)

    # 1. Xử lý thêm mới
    if request.method == 'POST':
        ten_mau = request.form.get('ten_mau')
        ma_mau = request.form.get('ma_mau').strip().upper()
        file = request.files.get('file_mau')

        if not file or not ten_mau or not ma_mau:
            flash('Vui lòng điền đầy đủ thông tin và chọn file!', 'warning')
            return redirect(url_for('quan_ly_mau_bieu'))
        
        if file and ten_mau and ma_mau:
            # Kiểm tra mã mẫu đã tồn tại chưa
            ton_tai = MauBieu.query.filter_by(ma_mau=ma_mau).first()
            if ton_tai:
                flash(f'Lỗi: Mã mẫu {ma_mau} đã tồn tại trên hệ thống!', 'danger')
                return redirect(url_for('quan_ly_mau_bieu'))

            # TỐI ƯU TÊN FILE: Chỉ giữ lại Mã mẫu + đuôi file
            ext = os.path.splitext(file.filename)[1].lower()
            if ext != '.docx':
                flash('Chỉ chấp nhận file định dạng .docx', 'warning')
                return redirect(url_for('quan_ly_mau_bieu'))
                
            filename = f"{ma_mau}{ext}" # Ví dụ: CV_01.docx thay vì CV_01_Giay_xin_nghi_phep_dai_ngay.docx
            
            upload_path = app.config['UPLOAD_FOLDER_MAU']
            if not os.path.exists(upload_path):
                os.makedirs(upload_path)
            
            # Lưu file (ghi đè nếu cùng mã mẫu)
            file.save(os.path.join(upload_path, filename))
            
            moi = MauBieu(ten_mau=ten_mau, ma_mau=ma_mau, file_path=filename)
            db.session.add(moi)
            db.session.commit()
            flash('Khai báo mẫu biểu thành công!', 'success')
            return redirect(url_for('quan_ly_mau_bieu'))

    # XỬ LÝ PHÂN TRANG & TÌM KIẾM (Cho 1000+ mẫu)
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '').strip()
    
    query = MauBieu.query
    if search:
        query = query.filter(or_(
            MauBieu.ten_mau.icontains(search),
            MauBieu.ma_mau.icontains(search)
        ))
    
    # Mỗi trang hiển thị 20 dòng
    pagination = query.order_by(MauBieu.id.desc()).paginate(page=page, per_page=20)
    ds_mau = pagination.items

    return render_template('admin/admin_mau_bieu.html', 
                           ds_mau=ds_mau, 
                           pagination=pagination, 
                           search=search)

@app.route('/admin/xoa_mau_bieu/<int:id>')
@login_required
def xoa_mau_bieu(id):
    if not current_user.is_admin: abort(403)
    
    mau = MauBieu.query.get_or_404(id)
    # 1. Xóa file vật lý trong thư mục
    file_path = os.path.join(app.config['UPLOAD_FOLDER_MAU'], mau.file_path)
    if os.path.exists(file_path):
        os.remove(file_path)
        
    # 2. Xóa trong Database
    # 2. Xóa record (Cascading sẽ tự xử lý trong bảng trung gian nếu đã cấu hình ON DELETE CASCADE)
    db.session.delete(mau)
    db.session.commit()
    flash(f' Đã xóa mẫu {mau.ma_mau} và file đính kèm.', 'info')
    return redirect(url_for('quan_ly_mau_bieu'))


@app.route('/yeu-cau-mau-bieu')
@login_required
def yeu_cau_mau_bieu_page():
    # 1. Lấy các gói mẫu dùng chung (Admin tạo sẵn)
    goi_mau_co_san = YeuCauMau.query.filter_by(is_template=True).all()
    
    # 2. Lấy danh sách danh mục (Chỉ lấy những danh mục thực sự có gói mẫu để hiển thị bộ lọc)
    # JOIN với YeuCauMau để chắc chắn danh mục đó có dữ liệu
    danh_sach_dm = DanhMucMau.query.join(YeuCauMau).filter(YeuCauMau.is_template==True).distinct().all()

    # 3. Lấy yêu cầu cá nhân
    user_requests = YeuCauMau.query.filter_by(user_id=current_user.ma_nhan_vien)\
                                  .order_by(YeuCauMau.ngay_yeu_cau.desc()).all()
    
    return render_template('yeu_cau_mau_bieu.html', 
                           user_requests=user_requests, 
                           goi_mau_co_san=goi_mau_co_san,
                           danh_sach_dm=danh_sach_dm)

@app.route('/admin/quan-ly-yeu-cau')
@login_required
def admin_quan_ly_yeu_cau():
    if not current_user.is_admin: abort(403)

    # Lấy danh sách danh mục để đổ vào các thẻ <select> trong Modal
    ds_danh_muc = DanhMucMau.query.order_by(DanhMucMau.ten_danh_muc.asc()).all()
    
    # Lấy các yêu cầu từ nhân viên (is_template=False)
    requests = YeuCauMau.query.filter_by(is_template=False)\
                              .order_by(YeuCauMau.trang_thai.asc(), YeuCauMau.ngay_yeu_cau.desc()).all()
    
    # Lấy các gói mẫu biểu dùng chung đã tạo (is_template=True) để Admin có thể xem/xóa nếu cần
    goi_mau_da_tao = YeuCauMau.query.filter_by(is_template=True)\
                                    .order_by(YeuCauMau.danh_muc_id.asc()).all()
    
    all_available_mau = MauBieu.query.all()
    
    return render_template('admin/admin_quan_ly_yeu_cau.html', 
                           requests=requests, 
                           goi_mau_da_tao=goi_mau_da_tao,
                           all_available_mau=all_available_mau,
                           ds_danh_muc=ds_danh_muc)


@app.route('/api/search-mau-bieu')
@login_required
def api_search_mau_bieu():
    query = request.args.get('q', '').strip()
    if not query:
        return jsonify([])
        
    try:
        # Sử dụng ilike để đảm bảo tìm kiếm không phân biệt hoa thường tốt hơn trên một số DB
        results = MauBieu.query.filter(
            or_(
                MauBieu.ten_mau.ilike(f'%{query}%'),
                MauBieu.ma_mau.ilike(f'%{query}%')
            )
        ).order_by(MauBieu.ten_mau.asc()).limit(20).all()
        
        return jsonify([{'id': m.id, 'text': f"[{m.ma_mau}] {m.ten_mau}"} for m in results])
    except Exception as e:
        return jsonify([]), 500

@app.route('/admin/phe-duyet-mau/<int:id>', methods=['POST'])
@login_required
def phe_duyet_mau(id):
    if not current_user.is_admin: abort(403)
    yeu_cau = YeuCauMau.query.get_or_404(id)
    
    # 1. Cập nhật thông tin cơ bản
    yeu_cau.noi_dung_yeu_cau = request.form.get('noi_dung_edit', yeu_cau.noi_dung_yeu_cau)
    
    dm_id = request.form.get('danh_muc_id')
    if dm_id:
        yeu_cau.danh_muc_id = int(dm_id)
    
    # 2. Cập nhật danh sách mẫu biểu (Quan trọng: Sửa tên key sang 'mau_bieu_ids[]')
    selected_ids = request.form.getlist('mau_bieu_ids[]')
    
    if selected_ids:
        # Chuyển đổi sang list int để query an toàn
        id_ints = [int(i) for i in selected_ids]
        yeu_cau.danh_sach_mau = MauBieu.query.filter(MauBieu.id.in_(id_ints)).all()
        
        if not yeu_cau.is_template:
            yeu_cau.trang_thai = 'Đã đáp ứng'
        flash('Cập nhật mẫu biểu thành công!', 'success')
    else:
        # Nếu gửi lên rỗng (Admin xóa hết mẫu cũ), ta làm trống danh sách
        yeu_cau.danh_sach_mau = []
        flash('Đã cập nhật thông tin. Danh sách mẫu biểu hiện đang trống.', 'warning')
    
    # 3. Luôn commit sau khi thực hiện thay đổi
    db.session.commit()
    return redirect(url_for('admin_quan_ly_yeu_cau'))

@app.route('/gui-yeu-cau', methods=['POST'])
@login_required
def gui_yeu_cau():
    noi_dung = request.form.get('noi_dung_yeu_cau')
    if noi_dung:
        new_req = YeuCauMau(
            user_id=current_user.ma_nhan_vien, 
            noi_dung_yeu_cau=noi_dung,
            trang_thai='Chờ xử lý'
        )
        db.session.add(new_req)
        db.session.commit()
        flash('Yêu cầu của bạn đã được gửi tới Admin!', 'success')
    return redirect(url_for('yeu_cau_mau_bieu_page'))

@app.route('/admin/xoa-yeu-cau/<int:id>')
@login_required
def xoa_yeu_cau_mau(id):
    if not current_user.is_admin: abort(403)
    yeu_cau = YeuCauMau.query.get_or_404(id)
    db.session.delete(yeu_cau)
    db.session.commit()
    flash('Đã xóa dữ liệu thành công!', 'info')
    return redirect(url_for('admin_quan_ly_yeu_cau'))


@app.route('/in-mau-bieu/<string:ma_mau>')
@login_required
def in_mau_bieu(ma_mau):
    mau = MauBieu.query.filter_by(ma_mau=ma_mau).first_or_404()
    file_path = os.path.join(app.config['UPLOAD_FOLDER_MAU'], mau.file_path)
    
    if os.path.exists(file_path):
        # as_attachment=True để trình duyệt tải về thay vì cố mở file
        return send_file(file_path, as_attachment=True, download_name=f"{mau.ten_mau}.docx")
    
    flash('Tập tin không tồn tại trên hệ thống!', 'danger')
    return redirect(request.referrer or url_for('yeu_cau_mau_bieu_page'))


# "Tự động điền dữ liệu" vào file Word (Mail Merge), thư viện python-docx
# "Tự động điền dữ liệu" vào file Word (Mail Merge), thư viện python-docx
@app.route('/danh-sach-mau')
@login_required
def user_view_mau():
    # 1. Lấy thông tin chi tiết của nhân viên đang đăng nhập
    nv_info = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()
    
    # 2. Lấy danh sách mẫu biểu (Bỏ joinedload để tránh lỗi với lazy='dynamic')
    ds_mau = MauBieu.query.order_by(MauBieu.ten_mau.asc()).all()

    # Chuẩn bị dữ liệu danh mục cho từng mẫu biểu
    # Vì MauBieu.cac_yeu_cau là 'dynamic', chúng ta truy vấn lọc ngay trong vòng lặp
    mau_to_categories = {}
    for mau in ds_mau:
        # Lấy danh sách ID danh mục từ các yêu cầu được đánh dấu là template
        # Truy vấn trực tiếp từ thuộc tính dynamic để tối ưu hiệu năng
        categories = mau.cac_yeu_cau.filter_by(is_template=True).with_entities(YeuCauMau.danh_muc_id).all()
        
        # Chuyển set các ID thành chuỗi "1,2,3" để JavaScript xử lý lọc
        ids = {str(c[0]) for c in categories if c[0]}
        mau_to_categories[mau.id] = ",".join(ids)

    # 3. Lấy danh sách danh mục mẫu biểu để làm bộ lọc
    ds_danh_muc = DanhMucMau.query.order_by(DanhMucMau.ten_danh_muc.asc()).all()

    # 4. Xử lý logic hiển thị Phòng ban - Đơn vị
    phong_ban_hien_thi = ""
    if nv_info:
        # Kiểm tra tồn tại của đối tượng phòng ban và đơn vị trước khi lấy tên
        ten_pb = nv_info.phong_ban.ten_phong_ban if (hasattr(nv_info, 'phong_ban') and nv_info.phong_ban) else ""
        ten_dv = nv_info.don_vi.ten_ma_hieu_2 if (hasattr(nv_info, 'don_vi') and nv_info.don_vi) else ""
        
        if ten_pb and ten_dv:
            phong_ban_hien_thi = f"{ten_dv} - {ten_pb}"
        else:
            phong_ban_hien_thi = ten_pb or ten_dv
        
    return render_template('user_mau_bieu.html', 
                           ds_mau=ds_mau,
                           mau_to_categories=mau_to_categories,
                           ds_danh_muc=ds_danh_muc, 
                           nv_info=nv_info,
                           phong_ban_hien_thi=phong_ban_hien_thi,
                           now=datetime.now())
@app.route('/tu-dong-dien-du-lieu/<ma_mau>', methods=['POST'])
@login_required
def tu_dong_dien_du_lieu(ma_mau):
    # 1. Tìm thông tin mẫu biểu
    mau_info = MauBieu.query.filter_by(ma_mau=ma_mau).first()
    if not mau_info:
        flash("Mẫu biểu không tồn tại!", "danger")
        return redirect(url_for('user_view_mau'))

    # Lấy thông tin NV hiện tại từ database để làm đối chứng (thông tin cũ)
    nv = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=current_user.ma_nhan_vien).first()

    # Xác định đơn vị và mã hiệu từ database
    dv_day_du = "NGÂN HÀNG NÔNG NGHIỆP VÀ PTNT VIỆT NAM"
    if nv and nv.don_vi:
        dv_day_du = nv.don_vi.ten_ma_hieu_2 or dv_day_du

    # 2. Kiểm tra file mẫu trên server
    template_path = os.path.join(app.config['UPLOAD_FOLDER_MAU'], mau_info.file_path)
    if not os.path.exists(template_path):
        flash("Không tìm thấy file mẫu gốc trên máy chủ!", "warning")
        return redirect(url_for('user_view_mau'))

    try:
        # 3. Chuẩn bị Context dữ liệu tách biệt CŨ và MỚI
        fullname = request.form.get('fullname', '').strip()
        
        context = {
            # Thông tin định danh cơ bản
            'ho_ten': fullname.upper() if fullname else (nv.ho_ten.upper() if nv else ''),
            'ma_nhan_vien': request.form.get('ma_nhan_vien', current_user.ma_nhan_vien),
            
            # --- THÔNG TIN CŨ (Lấy trực tiếp từ DB hoặc Form Readonly) ---
            'phong_ban_cu': request.form.get('phong_ban_cu', ''),
            'chuc_vu_cu': request.form.get('position_cu', nv.chuc_vu if nv else ''),
            'dien_thoai_cu': request.form.get('dien_thoai_cu', nv.so_dien_thoai if nv else ''),
            
            # --- THÔNG TIN MỚI (Lấy từ input người dùng nhập) ---
            'phong_ban_moi': request.form.get('phong_ban_moi', ''),
            'chuc_vu_moi': request.form.get('position_moi', ''), # Tương ứng name="position_moi"
            'dien_thoai_moi': request.form.get('dien_thoai_moi', ''), # Tương ứng name="dien_thoai_moi"
            
            'mail': request.form.get('email', nv.mail_Agribank if nv else ''),
            'ly_do': request.form.get('ly_do', ''),

            # Dữ liệu bổ trợ
            'ten_ma_hieu_2': dv_day_du,
            'ngay': datetime.now().day,
            'thang': datetime.now().month,
            'nam': datetime.now().year,
            'dia_danh': "Hà Nội" 
        }

        # 4. Sử dụng docxtpl để điền dữ liệu
        doc = DocxTemplate(template_path)
        doc.render(context)

        # 5. Xuất file
        file_stream = io.BytesIO()
        doc.save(file_stream)
        file_stream.seek(0)

        safe_name = f"{mau_info.ma_mau}_{current_user.ma_nhan_vien}.docx"

        return send_file(
            file_stream,
            as_attachment=True,
            download_name=safe_name,
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )

    except Exception as e:
        app.logger.error(f"Error rendering Word file: {str(e)}")
        flash(f"Có lỗi xảy ra: {str(e)}", "danger")
        return redirect(url_for('user_view_mau'))



@app.context_processor
def inject_pending_requests():
    if current_user.is_authenticated and current_user.is_admin:
        count = YeuCauMau.query.filter_by(trang_thai='Chờ xử lý').count()
        return dict(pending_count=count)
    return dict(pending_count=0)

# ----------------------------------------------------------------------
# Module Chuyển đổi font .VnTime sang Times New Roman; def tcvn3_to_unicode(text); def process_dynamic_file(file_path)
# ----------------------------------------------------------------------
@app.route('/chuyen-font')
@login_required
def chuyen_font_page():
    return render_template('chuyen_font.html')

# API 1: Chuyển đổi văn bản trực tiếp (Dán text)
@app.route('/api/convert-font-text', methods=['POST'])
@login_required
def api_convert_text():
    data = request.json
    text_input = data.get('text', '')
    converted = tcvn3_to_unicode(text_input) # Đã đổi tên hàm cho khớp
    return jsonify({'result': converted})

# API 2: Chuyển đổi FILE (Hỗ trợ CSV, XLS, XLSX với cấu trúc linh hoạt; Số dòng/cột không cố định)
@app.route('/api/convert-font-file', methods=['POST'])
@login_required
def api_convert_file():
    file = request.files.get('file')
    if not file: 
        return jsonify({"error": "Không tìm thấy file"}), 400
    filename = file.filename.lower()
    try:
        # 1. Đọc file
        if filename.endswith('.csv'):
            try:
                # Dùng engine='c' để đọc CSV lớn nhanh hơn
                df = pd.read_csv(file, encoding='utf-8', engine='c')
            except:
                file.seek(0)
                df = pd.read_csv(file, encoding='cp1252', engine='c')
        elif filename.endswith('.xls'):
            df = pd.read_excel(file, engine='xlrd')
        else: # .xlsx
            # Bỏ read_only=True để tránh lỗi gán dữ liệu phía dưới, 
            # 50MB vẫn an toàn với bộ nhớ RAM của server.
            df = pd.read_excel(file, engine='openpyxl')

        # 2. Xử lý dữ liệu (Sử dụng chính biến df)
        
        # Bước 2.1: Chuyển đổi tiêu đề cột
        df.columns = [tcvn3_to_unicode(str(c)) for c in df.columns]
        
        # Bước 2.2: Chỉ xử lý các cột kiểu Object (văn bản) để tăng tốc độ cho file 50MB
        for col in df.select_dtypes(include=['object']).columns:
            # fillna('') giúp tránh lỗi khi ép kiểu str(x) trên ô trống
            df[col] = df[col].fillna('').astype(str).apply(tcvn3_to_unicode)

        # 3. Xuất file ra bộ nhớ (BytesIO)
        output = BytesIO()
        if filename.endswith('.csv'):
            # utf-8-sig giúp Excel mở CSV nhận diện đúng tiếng Việt
            df.to_csv(output, index=False, encoding='utf-8-sig')
            mimetype = 'text/csv'
        else:
            # engine='openpyxl' là tiêu chuẩn cho .xlsx
            df.to_excel(output, index=False, engine='openpyxl')
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        output.seek(0)
        
        # Tạo tên file an toàn
        safe_filename = f"Unicode_{filename}"
        
        return send_file(
            output, 
            mimetype=mimetype, 
            as_attachment=True, 
            download_name=safe_filename
        )

    except Exception as e:
        app.logger.error(f"Lỗi chuyển đổi file: {str(e)}")
        # Log chi tiết lỗi để bạn dễ debug khi file 50MB bị hỏng
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Lỗi xử lý file: {str(e)}"}), 500

# ----------------------------------------------------------------------
# Rà soát báo cáo dữ liệu điện tử gửi NHNN bc48 (http://127.0.0.1:5001/bc48_pcrt => bc48.html)
# ----------------------------------------------------------------------
@app.route('/bc48_pcrt', methods=['GET', 'POST'])
@login_required
@admin_or_user_sd_bc48 # Chỉ Admin và nhân viên trong danh sách allowed_users mới sử dụng bc48
def giaodien_bc48():
    # 1. Lấy năm hiện tại (ví dụ: 2026) để hiển thị giao diện
    current_year = datetime.now().year
    data_results = [] # Danh sách kết quả vấn tin chi tiết
    history_results = []    # Báo cáo bc48 toàn hàng: Loại Báo cáo; tên bảng; loại tiền; số dòng; Tổng số tiền; Tổng quy đổi
    history_don_vi = []     # Báo cáo bc48 theo đơn vị: Loại Báo cáo; tên bảng; ma_hieu_1; ten_ma_hieu_1; ma_hieu_2; ten_ma_hieu_2; ma_khu_vuc; ten_khu_vuc; loại tiền; số dòng; Tổng số tiền; Tổng quy đổi

    # --- BỔ SUNG BIẾN CHO SAO KÊ TXT của CTR; DWT; EFT; PTR từ ctr_yyyymm; dwt_yyyymm; eft_yyyymm; ptr_yyyymm ---
    saoke_txt_results = []  # Danh sách kết quả cho view Sao kê theo TXT
    # ----------------------------------
    
    page = request.args.get('page', 1, type=int) # Đưa lên đầu để dùng chung
    per_page = 50
    total_pages = 0 # Khởi tạo biến để tránh lỗi render
    total_rows = 0

    # Lấy tham số tìm kiếm từ cả POST (khi nhấn nút) và GET (khi lật trang)
    # Dùng .values.get để lấy từ cả hai nguồn
    selected_loai = request.values.get('loai_bc', 'CTR')
    tu_ngay = request.values.get('tu_ngay', f"{current_year}-01-01")
    den_ngay = request.values.get('den_ngay', f"{current_year}-12-31")

    # Kiểm tra xem người dùng nhấn vào menu "Báo cáo" hay menu "Nạp file"
    # Nhận các view_mode: import, report, history_toan_hang, history_don_vi
    ##view_mode = request.args.get('view', 'import')
    view_mode = request.values.get('view', 'import')
    filter_ma = request.args.get('filter_ma', '').strip() # Lấy mã cần lọc

    if request.method == 'POST':
        action = request.form.get('action')
        # --- HÀNH ĐỘNG 1: Tổng hợp dữ liệu (CHẠY PROCEDURE sp_TongHopToanBoBaoCao) Chỉ chạy khi POST---
        if action == 'run_procedure':
            # Lấy tháng năm từ form người dùng chọn
            start_month = request.form.get('start_period') # Dạng YYYY-MM
            end_month = request.form.get('end_period')     # Dạng YYYY-MM
            # Thêm flag để biết tổng hợp toàn hàng hay đơn vị (tùy menu gửi lên)
            target = request.form.get('target', 'all')
            
            try:
                # Chuyển đổi YYYY-MM -> YYYYMM (int) để truyền vào Procedure: ví dụ 202601
                start_int = int(start_month.replace('-', ''))
                end_int = int(end_month.replace('-', ''))

                if target == 'don_vi':
                    # Bạn có hàm thực thi Procedure cho đơn vị trong file bc48.py
                    proc_res = bc48.thuc_thi_tong_hop_don_vi_bc48(db, start_int, end_int)
                    redirect_view = 'history_don_vi'
                else:
                    # Gọi Procedure với khoảng thời gian chủ động trong file bc48.py
                    proc_res = bc48.thuc_thi_tong_hop_bc48(db, start_int, end_int)
                    redirect_view = 'report'
                    
                flash(proc_res['msg'], 'success' if proc_res['success'] else 'danger')
                # Quan trọng: Truyền filter_ma khi redirect nếu là đơn vị để nó tự load lại đúng mã đó
                return redirect(url_for('giaodien_bc48', view=redirect_view, filter_ma=filter_ma))
            except Exception as e:
                flash(f"Lỗi thực thi: {str(e)}", 'danger')

    # 2. Xử lý Truy vấn Dữ liệu (REPORT) - Chạy cho cả GET và POST
    # Chỉ cần check view_mode, không cần check request.method
    if view_mode == 'report':
        try:
            data_results = bc48.van_tin_bc48(db, selected_loai, tu_ngay, den_ngay, page=page, per_page=per_page)
            # Tùy chọn: Bạn nên bổ sung hàm đếm tổng số dòng để hiển thị total_pages
            total_rows = bc48.count_van_tin_bc48(db, selected_loai, tu_ngay, den_ngay) or 0
            total_pages = (total_rows + per_page - 1) // per_page if total_rows > 0 else 0
            data_results = bc48.van_tin_bc48(db, selected_loai, tu_ngay, den_ngay, page=page, per_page=per_page)
            if not data_results:
                flash(f"Không có dữ liệu {selected_loai} trong khoảng thời gian đã chọn.", "info")
        except Exception as e:
            flash(f"Lỗi Vấn tin: {e}", "danger")
    # --- BỔ SUNG XỬ LÝ CHO VIEW SAO KÊ THEO TXT ---
    elif view_mode == 'saoke_txt':
        try:
            # Ví dụ gọi hàm truy vấn riêng cho sao kê txt từ file bc48.py (nếu có)
            # Bạn có thể thay đổi hàm bc48.van_tin_saoke_txt theo cấu trúc thực tế của bạn
            if hasattr(bc48, 'van_tin_saoke_txt'):
                saoke_txt_results = bc48.van_tin_saoke_txt(db, selected_loai, tu_ngay, den_ngay, page=page, per_page=per_page)
                total_rows = bc48.count_saoke_txt(db, selected_loai, tu_ngay, den_ngay) or 0
                total_pages = (total_rows + per_page - 1) // per_page if total_rows > 0 else 0
            else:
                # Nếu chưa viết hàm riêng, tạm thời lấy chung hàm vấn tin hoặc để trống xử lý sau
                saoke_txt_results = []
        except Exception as e:
            flash(f"Lỗi truy vấn Sao kê TXT: {e}", "danger")
    # ---------------------------------------------
    
    # Xử lý phần hiển thị History
    engine_bc48 = bc48.get_bc48_engine(db) #sử dụng db bc48 khác với db Cham_cong_Ngoai_gio
    with engine_bc48.connect() as conn:
        if view_mode in ['report', 'history_toan_hang']:
            history_results = conn.execute(text("SELECT * FROM bao_cao_tong_hop_history ORDER BY ngay_cap_nhat DESC")).fetchall()

        if view_mode == 'history_don_vi':
            offset = (page - 1) * per_page
            clean_filter_ma = filter_ma # Đã strip ở trên

            if clean_filter_ma:
                search_val = f"%{clean_filter_ma}%"
                
                # SỬA TẠI ĐÂY: Sử dụng manh8so_moi làm trọng tâm tìm kiếm
                # Kết hợp TRIM để loại bỏ ký tự rác trong các cột mã cũ
                where_clause = """
                    WHERE manh8so_moi LIKE :ma 
                       OR TRIM(ma_hieu_1) LIKE :ma 
                       OR ten_ma_hieu_1 LIKE :ma
                """
                
                count_sql = text(f"SELECT COUNT(*) FROM bao_cao_tong_hop_don_vi_history {where_clause}")
                total_rows = conn.execute(count_sql, {"ma": search_val}).scalar()

                data_sql = text(f"""
                    SELECT * FROM bao_cao_tong_hop_don_vi_history 
                    {where_clause}
                    ORDER BY ngay_cap_nhat DESC 
                    LIMIT :limit OFFSET :offset
                """)
                history_don_vi = conn.execute(data_sql, {
                    "ma": search_val, 
                    "limit": per_page, 
                    "offset": offset
                }).fetchall()
            else:
                total_rows = conn.execute(text("SELECT COUNT(*) FROM bao_cao_tong_hop_don_vi_history")).scalar()
                data_sql = text("""
                    SELECT * FROM bao_cao_tong_hop_don_vi_history 
                    ORDER BY ngay_cap_nhat DESC 
                    LIMIT :limit OFFSET :offset
                """)
                history_don_vi = conn.execute(data_sql, {"limit": per_page, "offset": offset}).fetchall()

            total_pages = (total_rows + per_page - 1) // per_page if total_rows and total_rows > 0 else 0

    return render_template('bc48.html', 
                           results=data_results, 
                           history_results=history_results,
                           history_don_vi=history_don_vi,
                           current_year=current_year,
                           selected_loai=selected_loai,
                           view_mode=view_mode,
                           filter_ma=filter_ma,    # Truyền biến lọc về lại template
                           total_rows=total_rows,
                           current_page=page,
                           total_pages=total_pages)

################################################################################################################################################
# Thực hiện nạp TXT do Core trả ra hàng ngày
################################################################################################################################################
@app.route('/bc48_import', methods=['POST'])
@login_required
@admin_or_user_sd_bc48 # Chỉ Admin và nhân viên trong danh sách allowed_users mới sử dụng bc48
def import_bc48_files():
    if 'files[]' not in request.files:
        return jsonify({"error": "Không tìm thấy file"}), 400

    # 1. Lấy tên người dùng đang đăng nhập để ghi log vào bảng log_import_errors khi TXT có sai khác giữa số dòng thực tế và số dòng khai báo tại dòng 1
    # Nếu hệ thống của bạn dùng trường khác (như fullname), hãy thay đổi tương ứng
    # Lấy trực tiếp từ cột ma_nhan_vien của model User
    user_performing_upload = f"{current_user.ma_nhan_vien} - {current_user.fullname}"
    
    # Lấy danh sách file từ form
    files = request.files.getlist('files[]')
    files_data = []

    for f in files:
        # Chuyển tên file về chữ thường trước khi kiểm tra đuôi
        if f.filename.lower().endswith('.txt'):
            # Đọc nội dung file
            content = f.read() 
            if content:
                # Quan trọng: Truyền filename và content (dưới dạng bytes hoặc string)
                # để bc48.py có thể xử lý nhiều lần (check Bước 0 và nạp data)
                files_data.append((f.filename, content))

    if not files_data:
        return jsonify({"error": "Vui lòng chọn ít nhất một file .TXT hợp lệ"}), 400

    # Ép Flask trả về response ngay, sử dụng stream_with_context, không đợi đệm
    def generate():
        # Gọi hàm xử lý từ bc48.py. 
        # Đảm bảo hàm này trong bc48.py đã có logic yield json.dumps({"status": "check", ...})
        # 2. Truyền thêm tham số current_user_name vào hàm xử lý ở bc48.py
        # Tham số này sẽ được hàm bên bc48.py dùng để ghi vào bảng log_import_errors
        for chunk in bc48.process_import_files(db, files_data, current_user_name=user_performing_upload):
            yield chunk

    response = Response(stream_with_context(generate()), mimetype='text/event-stream')
    response.headers['X-Accel-Buffering'] = 'no'
    response.headers['Cache-Control'] = 'no-cache'
    return response

################################################################################################################################################
# Sao kê theo ten_file_goc từ các bảng ctr_yyyymm; dwt_yyyymm; eft_yyyymm; ptr_yyyymm
################################################################################################################################################
@app.route('/bc48_saoke_txt', methods=['GET', 'POST'])
@login_required
@admin_or_user_sd_bc48
def saoke_txt():
    current_year = datetime.now().year
    saoke_txt_results = []
    
    page = request.args.get('page', 1, type=int)
    per_page = 50

    selected_loai = request.values.get('loai_bc', 'CTR')
    tu_ngay = request.values.get('tu_ngay', f"{current_year}-01-01")
    den_ngay = request.values.get('den_ngay', f"{current_year}-12-31")
    
    # Tiếp nhận thêm tham số tên file gốc lọc từ giao diện
    ten_file_goc = request.args.get('ten_file_goc', '').strip()

    try:
        if hasattr(bc48, 'van_tin_saoke_txt'):
            # Truyền thêm tham số ten_file_goc vào hàm đếm dòng và hàm lấy dữ liệu
            total_rows = bc48.count_saoke_txt(db, selected_loai, tu_ngay, den_ngay, ten_file_goc=ten_file_goc) or 0
            #total_pages = (total_rows + per_page - 1) // per_page if total_rows > 0 else 0
            total_pages = int((total_rows + per_page - 1) // per_page) if total_rows > 0 else 0
            
            saoke_txt_results = bc48.van_tin_saoke_txt(db, selected_loai, tu_ngay, den_ngay, 
                                                       page=page, per_page=per_page, ten_file_goc=ten_file_goc)
            
            if not saoke_txt_results and (request.args.get('loai_bc') or request.method == 'POST'):
                flash(f"Không tìm thấy dữ liệu phù hợp với điều kiện hoặc tên file chọn lọc.", "info")
        else:
            total_pages = 0
            total_rows = 0
            flash("Hàm 'van_tin_saoke_txt' chưa được định nghĩa trong bc48.py", "danger")
    except Exception as e:
        total_pages = 0
        total_rows = 0
        flash(f"Lỗi hệ thống khi truy vấn: {str(e)}", "danger")

    return render_template('bc48_saoke_txt.html', 
                           saoke_txt_results=saoke_txt_results,
                           current_year=current_year,
                           selected_loai=selected_loai,
                           tu_ngay=tu_ngay,
                           den_ngay=den_ngay,
                           total_rows=total_rows,
                           current_page=page,
                           total_pages=total_pages)

################################################################################################################################################
# Thực hiện nạp CSV do Cục PCRT trả ra
################################################################################################################################################
@app.route('/bc48_import_csv', methods=['POST'])
@login_required
@admin_or_user_sd_bc48
def import_bc48_csv():
    if 'files[]' not in request.files:
        return jsonify({"error": "Không tìm thấy file"}), 400
    
    files = request.files.getlist('files[]')
    try:
        # Gọi hàm xử lý CSV từ file bc48.py
        results = bc48.process_csv_error_files(db, files)
        # TRẢ VỀ JSON HỢP LỆ
        return jsonify(results)
    except Exception as e:
        # Nếu có lỗi hệ thống, vẫn trả về JSON để frontend không bị crash
        return jsonify({"error": [str(e)]}), 500

# Tính năng "Nạp từ thư mục Server" dành cho các file csv 500MB
@app.route('/import_csv_from_server', methods=['POST'])
@login_required
@admin_or_user_sd_bc48
def import_csv_from_server():
    folder_path = os.path.join(os.getcwd(), 'uploads_manual')
    try:
        # Gọi hàm nạp csv (file 500MB) từ thư mục trên server
        results = bc48.import_csv_from_server_logic(db, folder_path)
        return jsonify(results)
    except Exception as e:
        return jsonify({"error": [str(e)]}), 500
    


# Vấn tin từ ngày đến ngày của bảng CTR/DWT/EFT/PTR và Xuất Excel CSV (Full)
@app.route('/bc48_pcrt/export', methods=['POST'])
@login_required
@admin_or_user_sd_bc48
def export_bc48():
    # Sử dụng request.values để lấy dữ liệu bất kể nó được gửi từ POST hay GET
    loai_bc = request.values.get('loai_bc')
    tu_ngay = request.values.get('tu_ngay')
    den_ngay = request.values.get('den_ngay')
    
    filename = f"BC48_{loai_bc}_{tu_ngay.replace('-', '')}_{datetime.now().strftime('%H%M%S')}.csv"

    # Định nghĩa một generator wrapper để đảm bảo ngữ cảnh luôn tồn tại khi chạy
    def generate():
        with app.app_context(): # Tạo lại context cho generator
            yield from bc48.generate_csv_stream_bc48(db, loai_bc, tu_ngay, den_ngay)

    return Response(
        generate(),
        mimetype='text/csv',
        headers={
            "Content-disposition": f"attachment; filename={filename}",
            "Cache-Control": "no-cache",
            "X-Content-Type-Options": "nosniff"
        }
    )


# là "người điều hướng" (Route), gọi hàm từ bc48.py và trả về kết quả cho trình duyệt
@app.route('/bc48_pcrt/export_history/<mode>')
@login_required
@admin_or_user_sd_bc48 # Decorator phân quyền sử dụng bc48
def export_history_bc48(mode):
    # Gọi hàm từ module bc48
    output, result = bc48.export_history_bc48_logic(db, mode)
    
    # Nếu output là None tức là có lỗi hoặc không có dữ liệu
    if output is None:
        flash(f"Thông báo: {result}", "warning" if result == "Không có dữ liệu" else "danger")
        return redirect(url_for('giaodien_bc48'))
    
    # Nếu thành công, trả về file cho trình duyệt tải xuống
    return send_file(
        output, 
        as_attachment=True, 
        download_name=result, # result lúc này là filename
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )



# Route hiển thị giao diện gọi stored procedure trong mysql sp_TongHopBaoCaoTheoDonVi_RiengLe
@app.route('/bc48/xu-ly-rieng-le')
@login_required
@admin_or_user_sd_bc48
def giaodien_xu_ly_riengle():
    # Lấy context chung và thêm biến 'view'
    context = get_common_context()
    context['view'] = 'xu_ly_rieng_le'
    context['title'] = 'Tổng hợp 01 đơn vị'
    return render_template('index.html', **context)

# Route xử lý dữ liệu (đường dẫn redirect)
@app.route('/tonghop_donvi_riengle', methods=['POST'])
@login_required
@admin_or_user_sd_bc48 # Decorator phân quyền sử dụng bc48
def tonghop_donvi_riengle():
    # Lấy mã từ Form gửi lên từ giao diện và dùng strip() để xóa mọi dấu cách thừa ở đầu/cuối
    manh8so = str(request.form.get('manh8so_moi', '')).strip()
    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')

    if not all([manh8so, start_date, end_date]):
        flash("Vui lòng nhập đầy đủ thông tin Mã ngân hàng 8 số mới!", "warning")
        return redirect(url_for('giaodien_xu_ly_riengle'))
        ##return redirect(url_for('giaodien_bc48', view='history_don_vi'))

    # Gọi hàm xử lý từ bc48.py (phần đầu app.py đã khai báo: from modules import bc48)
    # db ở đây là đối tượng quản lý kết nối của bạn
    ket_qua = bc48.thuc_thi_tong_hop_don_vi_riengle(db, start_date, end_date, manh8so)

    if ket_qua['success']:
        flash(ket_qua['msg'], "success")
        # Sau khi xong, chuyển về trang lịch sử đơn vị để xem kết quả
        return redirect(url_for('giaodien_bc48', view='history_don_vi', filter_ma=manh8so))
    else:
        flash(ket_qua['msg'], "danger")
        return redirect(url_for('giaodien_xu_ly_riengle'))

@app.route('/bc48_log_errors')
@login_required
@admin_or_user_sd_bc48 # Decorator phân quyền sử dụng bc48
def giaodien_log_errors():
    # Lấy thông số lọc từ request (nếu không có thì lấy mặc định)
    # Ví dụ: mặc định lấy lỗi trong 30 ngày qua
    engine = bc48.get_bc48_engine(db)
    
    # Truy vấn dữ liệu từ bảng log
    query = text("""
        SELECT log.*, dm.noi_dung_loi_bc48 AS mo_ta_loi 
        FROM log_import_errors log
        LEFT JOIN dm_ma_loi_bc48 dm ON log.ma_loi_bc48 = dm.ma_loi_bc48
        WHERE log.error_time >= DATE_SUB(NOW(), INTERVAL 30 DAY)
        ORDER BY log.error_time DESC
    """)
    
    with engine.connect() as conn:
        result = conn.execute(query)
        logs = [dict(row._mapping) for row in result]
        
    return render_template('bc48_log_errors.html', logs=logs)

@app.route('/bc48/log/update_status/<int:error_id>', methods=['POST'])
@login_required
@admin_or_user_sd_bc48 # Decorator phân quyền sử dụng bc48
def update_error_status(error_id):
    # Cập nhật status thành 'Đã xử lý'
    engine = bc48.get_bc48_engine(db)
    try:
        with engine.begin() as conn:
            conn.execute(text("UPDATE log_import_errors SET status = 'Đã xử lý' WHERE id = :id"), {"id": error_id})
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "msg": str(e)}), 500

####################################################################################
# 
####################################################################################
@app.route('/bc48/logs')
@login_required
@admin_or_user_sd_bc48
def bc48_logs():
    # Lấy engine của db_bc48
    engine = bc48.get_bc48_engine(db)
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    ma_loi = request.args.get('ma_loi')
    
    query = "SELECT * FROM log_import_errors WHERE 1=1"
    params = {}
    
    if start_date:
        query += " AND DATE(error_time) >= :start"
        params['start'] = start_date
    if end_date:
        query += " AND DATE(error_time) <= :end"
        params['end'] = end_date
    if ma_loi:
        query += " AND ma_loi_bc48 = :ma_loi"
        params['ma_loi'] = ma_loi
    
    # Sử dụng engine để kết nối và execute thay vì db.session
    with engine.connect() as conn:
        logs = conn.execute(text(query), params).fetchall()
    
    return render_template('bc48_log_errors.html', logs=logs)

@app.route('/bc48/export_csv')
@login_required
@admin_or_user_sd_bc48
def export_logs_csv():
    # 1. Khởi tạo engine từ bind db_bc48
    engine = bc48.get_bc48_engine(db)
    
    # 2. Lấy tham số lọc từ URL
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    ma_loi = request.args.get('ma_loi')
    
    # 3. Xây dựng câu query lọc (Thêm LEFT JOIN để lấy dm.noi_dung_loi_bc48 AS mo_ta_loi)
    query = """
        SELECT log.*, dm.noi_dung_loi_bc48 AS mo_ta_loi 
        FROM log_import_errors log
        LEFT JOIN dm_ma_loi_bc48 dm ON log.ma_loi_bc48 = dm.ma_loi_bc48
        WHERE 1=1
    """
    params = {}
    if start_date:
        query += " AND DATE(log.error_time) >= :start"
        params['start'] = start_date
    if end_date:
        query += " AND DATE(log.error_time) <= :end"
        params['end'] = end_date
    if ma_loi:
        query += " AND log.ma_loi_bc48 = :ma_loi"
        params['ma_loi'] = ma_loi
        
    query += " ORDER BY log.error_time DESC"
    
    # 4. Thực thi truy vấn và chuyển đổi thành danh sách dict để tránh lỗi thuộc tính
    with engine.connect() as conn:
        result = conn.execute(text(query), params)
        logs = [dict(row._mapping) for row in result]
    
    # 5. Tạo file CSV trong bộ nhớ
    si = io.StringIO()
    cw = csv.writer(si, quoting=csv.QUOTE_MINIMAL)
    
    # Ghi tiêu đề (Header)
    cw.writerow([
        'Thời điểm', 'Người dùng', 'Tên file', 'Mã lỗi', 
        'Mô tả', 'Khai báo', 'Thực tế', 'Dòng Header'
    ])
    
    # Ghi dữ liệu (Sử dụng cách gọi key của dict: log.get('tên_cột'))
    for log in logs:
        # Định dạng lại thời gian thành chuỗi ngày/tháng/năm giờ:phút cho giống giao diện nếu cần
        error_time_str = log.get('error_time').strftime('%d/%m/%Y %H:%M') if log.get('error_time') else ''
        
        cw.writerow([
            error_time_str, 
            log.get('user_import', ''), 
            log.get('file_name', ''), 
            log.get('ma_loi_bc48', ''), 
            log.get('mo_ta_loi', 'Chưa xác định'), # Đã lấy được nhờ LEFT JOIN
            log.get('declared_rows', 0), 
            log.get('actual_rows', 0),
            log.get('header_content', '')
        ])
    
    # 6. Chuẩn bị response gửi về client kèm theo BOM chống lỗi font mã UTF-8 trên Excel
    response = make_response("\ufeff" + si.getvalue())
    response.headers["Content-Type"] = "text/csv; charset=utf-8-sig"
    response.headers["Content-Disposition"] = "attachment; filename=log_bc48_export.csv"
    
    return response

####################################################################################
# Truy vấn lịch sử nạp dữ liệu TXT vào bc48; Nhật ký nạp TXT vào bc48
####################################################################################
@app.route('/bc48/file-logs')
@login_required
@admin_or_user_sd_bc48
def bc48_file_logs():
    engine = bc48.get_bc48_engine(db)
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    loai_bc = request.args.get('loai_bc') or "" # Nếu None thì gán thành chuỗi rỗng

    page = int(request.args.get('page', 1)) # Mặc định là trang 1
    per_page = 50
    
    # Gọi hàm từ module bc48
    query, params = bc48.get_file_logs_query(start_date, end_date, loai_bc, page=page)
    stats_q, stats_p = bc48.get_file_logs_stats(start_date, end_date, loai_bc)

    with engine.connect() as conn:
        file_logs = [dict(row._mapping) for row in conn.execute(text(query), params).fetchall()]
        stats = conn.execute(text(stats_q), stats_p).fetchone()
        
    return render_template('bc48_file_logs.html', 
                           file_logs=file_logs,
                           stats=stats, 
                           start_date=start_date, 
                           end_date=end_date, 
                           loai_bc=loai_bc,
                           page=page)

@app.route('/bc48/file-logs/export')
@login_required
@admin_or_user_sd_bc48
def export_file_logs_csv():
    # 1. Lấy tham số từ request
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    loai_bc = request.args.get('loai_bc', '')
    
    # 2. Tạo tên file động
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"BC48_{loai_bc if loai_bc else 'All'}_{start_date or 'Start'}_{end_date or 'End'}_{timestamp}.csv"

    # 3. Lấy dữ liệu
    engine = bc48.get_bc48_engine(db)
    
    # GỌI HÀM VỚI page=None ĐỂ KHÔNG DÙNG LIMIT/OFFSET
    query, params = bc48.get_file_logs_query(start_date, end_date, loai_bc, page=None)
    
    with engine.connect() as conn:
        logs = conn.execute(text(query), params).fetchall()

    # 4. Tạo file CSV trong bộ nhớ
    si = io.StringIO()
    cw = csv.writer(si)
    # Header
    cw.writerow(['File Name', 'Ma CN', 'Thoi Diem', 'Loai BC', 'Hinh Thuc', 'STT', 'So Luong', 'User', 'Ngay Nap', 'Status'])
    
    for log in logs:
        # Chuyển đổi datetime sang string
        import_date_str = log.import_date.strftime('%d/%m/%Y %H:%M') if log.import_date else ''

        cw.writerow([
            log.file_name, 
            log.macn, 
            log.thoidiem, 
            log.loai_bc, 
            getattr(log, 'hinh_thuc', ''), 
            getattr(log, 'stt', ''), 
            log.so_luong, 
            log.user_import, 
            import_date_str, 
            log.status
        ])

    # 5. Phản hồi tải file
    output = make_response("\ufeff" + si.getvalue())
    output.headers["Content-Disposition"] = f"attachment; filename={filename}"
    output.headers["Content-type"] = "text/csv; charset=utf-8-sig"
    return output

####################################################################################
# Truy vấn lịch sử nạp dữ liệu CSV vào bc48; Nhật ký nạp CSV
####################################################################################
@app.route('/bc48/csv-logs')
@login_required
@admin_or_user_sd_bc48
def bc48_csv_logs():
    # Khởi tạo engine
    engine = bc48.get_bc48_engine(db)
    
    # Lấy param từ request
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    loai_bc = request.args.get('loai_bc', "")
    ma_nhan_vien = request.args.get('user_import')
    status = request.args.get('status')
    page = int(request.args.get('page', 1))
    per_page = 50 # Giữ đồng bộ với hàm trong bc48.py
    
    # Gọi logic từ bc48.py
    raw_logs = bc48.get_csv_logs_data(engine, start_date, end_date, loai_bc, ma_nhan_vien, status, page, per_page)
    total_records = bc48.count_csv_logs(engine, start_date, end_date, loai_bc, ma_nhan_vien, status)

    total_pages = (total_records + per_page - 1) // per_page
    
    # Logic kiểm tra trang sau:
    # Nếu kết quả trả về lớn hơn per_page, nghĩa là có dữ liệu cho trang tiếp theo
    has_next = len(raw_logs) > per_page

    # 3. Logic hiển thị # Chỉ lấy đúng số lượng per_page để hiển thị
    has_next = len(raw_logs) > per_page
    csv_logs = raw_logs[:per_page]

    # Tính toán total_records
    total_records = bc48.count_csv_logs(engine, start_date, end_date, loai_bc, ma_nhan_vien, status)
    # Tính toán total_unique_files
    total_unique_files = bc48.count_unique_files(engine, start_date, end_date, loai_bc, ma_nhan_vien, status)
    
    return render_template('bc48_csv_logs.html', 
                           csv_logs=csv_logs,
                           total_unique_files=total_unique_files,
                           total_records=total_records,
                           has_next=has_next, # Biến này để template ẩn/hiện nút "Sau"
                           total_pages=total_pages,
                           start_date=start_date, 
                           end_date=end_date, 
                           loai_bc=loai_bc,
                           user_import=ma_nhan_vien,
                           status=status,
                           page=page)

@app.route('/bc48/csv-logs/export')
@login_required
@admin_or_user_sd_bc48
def export_csv_logs():
    engine = bc48.get_bc48_engine(db)
    
    # 1. Lấy tham số lọc
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    loai_bc = request.args.get('loai_bc')
    ma_nhan_vien = request.args.get('user_import')
    status = request.args.get('status')
    
    # 2. Gọi với export=True để bỏ qua phân trang
    csv_logs = bc48.get_csv_logs_data(engine, start_date, end_date, loai_bc, ma_nhan_vien, status, export=True)
    
    # 3. Tạo file CSV trong bộ nhớ
    si = io.StringIO()
    cw = csv.writer(si, quoting=csv.QUOTE_MINIMAL)
    
    # Header
    cw.writerow(['Tên file', 'Loại BC', 'Ngày báo cáo', 'Số dòng', 'Người nạp', 'Trạng thái', 'Ghi chú'])
    
    # Data
    for log in csv_logs:
        cw.writerow([
            log['file_name'], 
            log['loai_bc'], 
            log['ngay_baocao'], 
            log['so_dong_du_lieu_csv'], 
            log['user_import'], 
            log['trang_thai'], 
            log['ghi_chu']
        ])
    
    # 4. Phản hồi tải file
    # Xây dựng suffix tên file từ các điều kiện lọc
    filter_parts = []
    if loai_bc: filter_parts.append(loai_bc)
    if start_date: filter_parts.append(f"Tu{start_date.replace('-', '')}")
    if end_date: filter_parts.append(f"Den{end_date.replace('-', '')}")
    if ma_nhan_vien: filter_parts.append(ma_nhan_vien)
    if status: filter_parts.append(status)
    
    suffix = f"_{'_'.join(filter_parts)}" if filter_parts else ""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    # Kết quả sẽ có dạng: BC48_CSV_Logs_CTR_Tu20260101_Den20260131_NV001_SUCCESS_20260506_163200.csv
    filename = f"BC48_CSV_Logs{suffix}_{timestamp}.csv"
    
    output = make_response("\ufeff" + si.getvalue())
    output.headers["Content-Disposition"] = f"attachment; filename={filename}"
    output.headers["Content-type"] = "text/csv; charset=utf-8-sig"
    return output

####################################################################################
# Cục PCRT trả ra CSV, Muốn biết trong CSV theo tháng/ngày có bao nhiêu dòng "KIỂM TRA GIAO DỊCH LỖI"; "KIỂM TRA FILE THÀNH CÔNG"; chi_tiet_loi_master
####################################################################################
@app.route('/bc48/cuc-pcrt-tra-loi')
@login_required
@admin_or_user_sd_bc48
def giaodien_cuc_pcrt_tra_loi():
    engine = bc48.get_bc48_engine(db)
    
    # Chỉ load dữ liệu tổng hợp nhẹ nhàng cho Tab 1 và Tab 2
    data_hang_ngay = bc48.get_tong_hop_sodong_pcrt(engine)
    data_theo_thang = bc48.get_tong_hop_loi_theo_thang(engine)
    data_chi_tiet_ngay = bc48.get_chi_tiet_loi_theo_ngay(engine)
    
    # Lấy dữ liệu mồi cho Select Option bộ lọc của Tab 3 trực tiếp từ DB
    filter_options = bc48.get_master_filters_options(engine)
    
    return render_template('bc48_cuc_pcrt_tra_loi.html', 
                           data_daily=data_hang_ngay, 
                           data_monthly=data_theo_thang,
                           data_monthly_detail=data_chi_tiet_ngay,
                           filter_options=filter_options)

# Tab 4: Từ ket_qua_do_tim_loi thống kê loaibc, hinhthuc, namthang, slgfile, sodong_loi, tong_quydoi_loi lấy chi tiết bảng dashboard_tke_loi_chi_tiet kết quả của sp_thong_ke_bang_loi
# API cho Tab thứ tư mới bổ sung (Server-side Processing)
@app.route('/bc48/api/tke-loi-chi-tiet-serverside', methods=['GET'])
@login_required
@admin_or_user_sd_bc48
def api_tke_loi_chi_tiet_serverside():
    """API xử lý phân trang Server-side cho bảng thống kê mã lỗi chi tiết"""
    engine = bc48.get_bc48_engine(db)
    params = request.args.to_dict()
    result = bc48.get_dashboard_tke_loi_chi_tiet_serverside(engine, params)
    return jsonify(result)

# Chạy luôn procedure sp_tong_hop_sodong_bang_error để thống kê chi tiết hằng ngày cục pcrt trả ra csv có bao nhiêu dòng Kiểm tra file thành công; Kiểm tra giao dịch lỗi;
@app.route('/bc48/api/run-procedure-daily', methods=['POST'])
@login_required
@admin_or_user_sd_bc48
def api_run_procedure_daily():
    engine = bc48.get_bc48_engine(db)
    
    # Gọi hàm thực thi Procedure
    success, message = bc48.execute_procedure_daily_raw(engine)
    
    if success:
        return jsonify({"success": True, "message": "Procedure executed successfully"})
    else:
        return jsonify({"success": False, "message": message}), 500


# Chạy các lệnh "Làm mới toàn bộ (Full Refresh)
#TRUNCATE TABLE danh_sach_bang_da_kiem_tra;
#TRUNCATE TABLE log_kiem_tra_du_lieu;
#TRUNCATE TABLE danh_sach_bang_error_da_quet;
#TRUNCATE TABLE log_loi_xu_ly_bang;
#TRUNCATE TABLE bang_tonghop_ktra_gd_loi;
#TRUNCATE TABLE chi_tiet_loi_master;
#TRUNCATE TABLE log_chi_tiet_loi_phan_tach;
#TRUNCATE TABLE ket_qua_do_tim_loi;
#TRUNCATE TABLE giao_dich_error_khong_tim_thay;
#CALL sp_xuly_error_trung_tam();
#CALL sp_doi_soat_khac_phuc_loi();
#CALL sp_TuDongKiemTraDuLieu();
@app.route('/bc48/api/run-full-refresh', methods=['POST'])
@login_required
@admin_or_user_sd_bc48
def api_run_full_refresh():
    engine = bc48.get_bc48_engine(db)
    
    # Giả định bạn đã tạo hàm này trong bc48.py
    # Hàm này sẽ thực hiện chuỗi TRUNCATE và CALL các procedure bạn đã liệt kê
    success, message = bc48.execute_full_refresh_procedure(engine)
    
    if success:
        return jsonify({"success": True, "message": "Toàn bộ dữ liệu đã được làm mới thành công!"})
    else:
        return jsonify({"success": False, "message": message}), 500

# ======================================================================
# API CHUẨN HOÁ CHO SELECT2 AJAX (ĐẶT TRONG APP.PY)
# ======================================================================
@app.route('/bc48/api/get-filter-file-goc', methods=['GET'])
@login_required
@admin_or_user_sd_bc48
def api_get_filter_file_goc():
    engine = bc48.get_bc48_engine(db)
    search_term = request.args.get('q', '').strip()
    
    # Gọi hàm logic thuần túy từ module bc48
    result = bc48.get_filter_file_goc_logic(engine, search_term)
    return jsonify(result)

@app.route('/bc48/api/master-serverside', methods=['GET'])
@login_required
@admin_or_user_sd_bc48
def api_master_serverside():
    """API cổng nhận xử lý yêu cầu phân trang từ DataTables Frontend"""
    engine = bc48.get_bc48_engine(db)
    # Lấy toàn bộ tham số gửi lên từ Ajax của DataTables
    params = request.args.to_dict()
    result = bc48.get_chi_tiet_loi_master_serverside(engine, params)
    return jsonify(result)


@app.route('/bc48/master/export-csv', methods=['GET'])
@login_required
@admin_or_user_sd_bc48
def export_master_csv():
    """Route chuyên trách xuất file CSV toàn bộ theo bộ lọc trực tiếp từ DB bảng chi_tiet_loi_master"""
    engine = bc48.get_bc48_engine(db)
    
    # Đọc các bộ lọc đang chọn từ URL
    f_loai_bc = request.args.get('f_loai_bc', '')
    f_thang_nam = request.args.get('f_thang_nam', '')
    f_ngay_bc = request.args.get('f_ngay_bc', '')
    f_ten_file = request.args.get('f_ten_file', '')
    
    where_clauses = ["1=1"]
    sql_params = {}
    if f_loai_bc:
        where_clauses.append("loai_bc = :f_loai_bc")
        sql_params['f_loai_bc'] = f_loai_bc
    if f_thang_nam:
        where_clauses.append("thang_nam = :f_thang_nam")
        sql_params['f_thang_nam'] = f_thang_nam
    if f_ngay_bc:
        where_clauses.append("NGAY_BAOCAO = :f_ngay_bc")
        sql_params['f_ngay_bc'] = f_ngay_bc
    if f_ten_file: # BỔ SUNG
        where_clauses.append("ten_file_goc = :f_ten_file")
        sql_params['f_ten_file'] = f_ten_file
        
    where_str = " AND ".join(where_clauses)
    
    query = f"""
        SELECT id, ma_giao_dich, loai_bc, thang_nam, trang_thai, ma_loi, mota_loi,
               DATE_FORMAT(NGAY_BAOCAO, '%Y-%m-%d') as NGAY_BAOCAO, HINHTHUC_GUI, SOLAN_GUI, ten_file_goc
        FROM chi_tiet_loi_master
        WHERE {where_str}
        ORDER BY NGAY_BAOCAO DESC, id DESC
    """
    
    try:
        with engine.connect() as conn:
            df = pd.read_sql_query(text(query), conn, params=sql_params)
            
        # Chuyển đổi DataFrame thành file CSV dạng stream bộ nhớ RAM để tải về
        output = io.BytesIO()
        # Ghi kèm mã Byte Order Mark (BOM) để Microsoft Excel đọc chuỗi tiếng Việt không bị lỗi font
        output.write(b'\xef\xbb\xbf') 
        df.to_csv(output, index=False, encoding='utf-8', mode='ab')
        output.seek(0)
        
        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-disposition": "attachment; filename=Bao_cao_chi_tiet_loi_master.csv"}
        )
    except Exception as e:
        return f"Lỗi xuất file dữ liệu: {str(e)}", 500

@app.route('/bc48/api/tke-loi-chi-tiet/export-csv', methods=['GET'])
@login_required
@admin_or_user_sd_bc48
def export_tke_loi_csv():
    """Route xuất trọn vẹn 100% dữ liệu thống kê kết quả dò tìm lỗi chi tiết (Tab 4) ra file CSV từ Server"""
    engine = bc48.get_bc48_engine(db)
    
    try:
        # Gọi tầng nghiệp vụ lấy dữ liệu nhị phân đã build cấu trúc CSV chuẩn Excel
        output_data = bc48.export_dashboard_tke_loi_chi_tiet_csv(engine)
        
        import datetime
        current_date = datetime.datetime.now().strftime("%Y%m%d")
        filename = f"Thong_Ke_Ket_Qua_Do_Tim_Loi_Chi_Tiet_{current_date}.csv"
        
        return Response(
            output_data,
            mimetype="text/csv",
            headers={"Content-disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        return f"Lỗi hệ thống khi xuất file dữ liệu Tab 4: {str(e)}", 500
    


####################################################################################
# Truy vấn bảng Ket_qua_do_tim_loi: sao kê chi tiết của bảng Tổng Hợp, nhưng chỉ chứa những giao dịch đã được "tìm thấy xác nhận" trong bảng dữ liệu gốc (CTR, DWT, EFT, PTR)
# bang_tonghop_ktra_gd_loi (Bảng Tổng Hợp): Đóng vai trò là bảng Dashboard. Chỉ chứa con số tổng (Count) để bạn nhìn nhanh tháng nào, hệ thống nào đang có vấn đề
# chi_tiet_loi_master (Bảng Sao Kê Lỗi): Chứa danh sách tất cả các giao dịch bị báo lỗi từ các bảng _error. Đây là "danh sách chờ các đơn vị khắc phục chỉnh sửa"
####################################################################################
@app.route('/bc48/ket-qua-do-tim')
@login_required
@admin_or_user_sd_bc48
def giaodien_ket_qua_do_tim():
    # 1. Khởi tạo engine từ kết nối db chung
    engine = bc48.get_bc48_engine(db)
    
    # 2. Lấy tham số lọc và phân trang
    # Lấy tham số lọc và xử lý cắt khoảng trắng thừa (.strip())
    loai_bc = request.args.get('loai_bc', "").strip()

    # Làm sạch biến thang_nam một cách triệt để
    thang_nam_raw = request.args.get('thang_nam', "").strip()
    thang_nam = thang_nam_raw if (thang_nam_raw and thang_nam_raw.isdigit() and len(thang_nam_raw) == 6) else ""
    
    ma_hieu_1 = request.args.get('ma_hieu_1', "").strip()
    ma_hieu_2 = request.args.get('ma_hieu_2', "").strip()
    ma_giao_dich = request.args.get('ma_giao_dich', "").strip()
    mota_loi = request.args.get('mota_loi', "").strip()
    tu_ngay = request.args.get('tu_ngay', "").strip()
    den_ngay = request.args.get('den_ngay', "").strip()
    # BỔ SUNG MỚI
    ten_file_goc = request.args.get('ten_file_goc', "").strip()
    ten_file_error = request.args.get('ten_file_error', "").strip()
    
    page = int(request.args.get('page', 1))
    per_page = 50

    # Nếu loai_bc là "ALL" hoặc rỗng, xem như không lọc theo Loại BC
    loai_bc_filter = None if loai_bc in ["", "ALL"] else loai_bc
    
    # 3. Gọi logic từ bc48.py
    raw_data = bc48.get_ket_qua_do_tim_data(
        engine, 
        loai_bc=loai_bc_filter, 
        thang_nam=thang_nam, 
        ma_hieu_1=ma_hieu_1, 
        ma_hieu_2=ma_hieu_2, 
        ma_giao_dich=ma_giao_dich, 
        mota_loi=mota_loi, 
        page=page, 
        per_page=per_page, 
        tu_ngay=tu_ngay, 
        den_ngay=den_ngay,
        ten_file_goc=ten_file_goc,
        ten_file_error=ten_file_error
    )
    total_records = bc48.count_ket_qua_do_tim(engine, loai_bc_filter, thang_nam, ma_hieu_1, ma_hieu_2, ma_giao_dich, mota_loi, tu_ngay=tu_ngay, den_ngay=den_ngay,
        ten_file_goc=ten_file_goc,
        ten_file_error=ten_file_error)

    # 4. Tính toán phân trang
    total_pages = (total_records + per_page - 1) // per_page if total_records > 0 else 1
    has_next = len(raw_data) > per_page
    data_display = raw_data[:per_page]

    # Tính toán vị trí bản ghi hiển thị trên trang hiện tại
    if total_records == 0:
        start_record = 0
        end_record = 0
    else:
        start_record = (page - 1) * per_page + 1
        # Đảm bảo số bản ghi kết thúc không vượt quá tổng số bản ghi thực tế
        end_record = min(page * per_page, total_records)

    return render_template('bc48_ket_qua_do_tim.html', 
                           data=data_display,
                           has_next=has_next,
                           total_pages=total_pages,
                           total_records=total_records, # Truyền thêm tổng số bản ghi
                           start_record=start_record,   # Truyền thêm bản ghi bắt đầu
                           end_record=end_record,       # Truyền thêm bản ghi kết thúc
                           loai_bc=loai_bc,
                           thang_nam=thang_nam,
                           ma_hieu_1=ma_hieu_1,
                           ma_hieu_2=ma_hieu_2,
                           ma_giao_dich=ma_giao_dich,
                           mota_loi=mota_loi,
                           tu_ngay=tu_ngay,      
                           den_ngay=den_ngay,
                           ten_file_goc=ten_file_goc,       # BỔ SUNG MỚI
                           ten_file_error=ten_file_error,   # BỔ SUNG MỚI
                           page=page)

@app.route('/bc48/ket-qua-do-tim/export')
@login_required
@admin_or_user_sd_bc48
def export_ket_qua_do_tim():
    engine = bc48.get_bc48_engine(db)
    
    loai_bc = request.args.get('loai_bc', '')
    thang_nam = request.args.get('thang_nam', '')
    ma_hieu_1 = request.args.get('ma_hieu_1', '')
    ma_hieu_2 = request.args.get('ma_hieu_2', '')
    ma_giao_dich = request.args.get('ma_giao_dich', '')
    mota_loi = request.args.get('mota_loi', '')
    tu_ngay = request.args.get('tu_ngay', '')
    den_ngay = request.args.get('den_ngay', '')
    # BỔ SUNG MỚI
    ten_file_goc = request.args.get('ten_file_goc', '')
    ten_file_error = request.args.get('ten_file_error', '')

    # Lấy toàn bộ dữ liệu không phân trang
    data = bc48.get_ket_qua_do_tim_data(
        engine, 
        loai_bc=loai_bc, 
        thang_nam=thang_nam, 
        ma_hieu_1=ma_hieu_1, 
        ma_hieu_2=ma_hieu_2, 
        ma_giao_dich=ma_giao_dich, 
        mota_loi=mota_loi, 
        tu_ngay=tu_ngay, 
        den_ngay=den_ngay,
        ten_file_goc=ten_file_goc,
        ten_file_error=ten_file_error,
        export=True
    )
    
    # Định nghĩa tên file export
    filename = f"Ket_qua_do_tim_{loai_bc or 'ALL'}_{thang_nam or 'ALL'}_{ma_hieu_1 or 'ALL'}_{ma_hieu_2 or 'ALL'}_{datetime.now().strftime('%Y%m%d')}.csv"

    # Sử dụng generator để stream dữ liệu tiết kiệm RAM
    def generate():
        # 1. Ghi ký tự BOM để Excel nhận diện đúng UTF-8
        yield '\ufeff'
        
        # 2. Khởi tạo một proxy writer ghi vào bộ nhớ đệm tạm thời
        output = io.StringIO()
        cw = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
        
        # 3. Ghi dòng Header
        cw.writerow([
            'Mã GD', 'Loại BC', 'Tháng', 'Mã Lỗi', 'Mô tả', 'Mã CN', 'Tên KH', 'Số GT', 
            'Mã ĐV Cấp 1', 'Tên ĐV Cấp 1', 'Mã ĐV Cấp 2', 'Tên ĐV Cấp 2', 'Tên file gốc', 'Tên file error'
        ])
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)
        
        # 4. Ghi từng dòng dữ liệu (Dùng .get để tránh KeyError nếu thiếu cột)
        for row in data:
            cw.writerow([
                row.get('ma_giao_dich', ''), 
                str(row.get('loai_bc', '')).upper(), 
                row.get('thang_nam', ''),
                row.get('ma_loi', ''), 
                row.get('mota_loi', ''), 
                row.get('macn', ''),
                row.get('tenkh', ''), 
                row.get('sogt', ''), 
                row.get('ma_hieu_1', ''), 
                row.get('ten_ma_hieu_1', ''),
                row.get('ma_hieu_2', ''), 
                row.get('ten_ma_hieu_2', ''),
                row.get('ten_file_goc', ''),
                row.get('ten_file_error', '')
            ])
            yield output.getvalue()
            output.seek(0)
            output.truncate(0)

    # Trả về response dạng stream dữ liệu trực tiếp
    response = Response(stream_with_context(generate()), mimetype="text/csv")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    response.headers["Content-Type"] = "text/csv; charset=utf-8-sig"
    return response

# Xuất Excel kết quả lọc bảng ket_qua_do_tim_loi
@app.route('/export-excel-ket-qua')
@login_required
@admin_or_user_sd_bc48
def export_excel_ket_qua():
    # 1. Thu thập các tham số lọc từ URL do JavaScript gửi lên
    loai_bc = request.args.get('loai_bc', 'ALL')
    thang_nam = request.args.get('thang_nam', '')
    ma_hieu_1 = request.args.get('ma_hieu_1', '')
    ma_hieu_2 = request.args.get('ma_hieu_2', '')
    ma_giao_dich = request.args.get('ma_giao_dich', '')
    mota_loi = request.args.get('mota_loi', '')
    tu_ngay = request.args.get('tu_ngay', '')
    den_ngay = request.args.get('den_ngay', '')
    # BỔ SUNG MỚI
    ten_file_goc = request.args.get('ten_file_goc', '')
    ten_file_error = request.args.get('ten_file_error', '')

    # 2. Xây dựng câu lệnh SQL WHERE động dựa trên bộ lọc
    conditions = []
    params_list = []

    if loai_bc and loai_bc != 'ALL':
        conditions.append("loai_bc = %s")
        params_list.append(loai_bc)
    if thang_nam:
        conditions.append("thang_nam = %s")  # Thay :thang_nam bằng %s
        params_list.append(int(thang_nam))
    if ma_hieu_1:
        conditions.append("ma_hieu_1 LIKE %s")
        params_list.append(f"%{ma_hieu_1}%")
    if ma_hieu_2:
        conditions.append("ma_hieu_2 LIKE %s")
        params_list.append(f"%{ma_hieu_2}%")
    if ma_giao_dich:
        conditions.append("ma_giao_dich LIKE %s")
        params_list.append(f"%{ma_giao_dich}%")
    if mota_loi:
        conditions.append("mota_loi LIKE %s")
        params_list.append(f"%{mota_loi}%")
    if tu_ngay:
        conditions.append("ngay_baocao >= %s")
        params_list.append(tu_ngay)
    if den_ngay:
        conditions.append("ngay_baocao <= %s")
        params_list.append(den_ngay)
    # BỔ SUNG MỚI
    if ten_file_goc:
        conditions.append("ten_file_goc LIKE %s")
        params_list.append(f"%{ten_file_goc}%")
    if ten_file_error:
        conditions.append("ten_file_error LIKE %s")
        params_list.append(f"%{ten_file_error}%")

    where_clause = "WHERE " + " AND ".join(conditions) if conditions else ""

    # 3. Câu lệnh SQL bốc ĐỦ ĐÚNG 28 CỘT (Cập nhật where_clause)
    query = dedent(f"""
        SELECT 
            id AS `ID`,
            ma_giao_dich AS `Mã Giao Dịch`,
            UPPER(loai_bc) AS `Loại BC`,
            thang_nam AS `Tháng Năm`,
            ma_loi AS `Mã Lỗi`,
            ma_loi_f_ao AS `Mã Lỗi F Ảo`,
            mota_loi AS `Mô tả lỗi từ CSV`,
            ds_ma_nghiep_vu AS `DS Mã nghiệp vụ`,
            ds_ten_nghiep_vu AS `DS Tên nghiệp vụ`,
            ds_ten_cot_sql AS `DS Tên cột SQL`,
            ds_ma_quy_dinh AS `DS Mã quy định`,
            macn AS `Mã CN`,
            thoidiem AS `Thời điểm`,
            loaigd AS `Loại GD`,
            kieukh AS `Kiểu KH`,
            tenkh AS `Tên Khách Hàng`,
            loaigt AS `Loại GT`,
            sogt AS `Số GT/Hộ chiếu`,
            sothithuc AS `Số thị thực`,
            CONCAT(IFNULL(ma_hieu_1, ''), ' - ', IFNULL(ten_ma_hieu_1, '')) AS `Đơn vị Cấp 1`,
            CONCAT(IFNULL(ma_hieu_2, ''), ' - ', IFNULL(ten_ma_hieu_2, '')) AS `Đơn vị Cấp 2`,
            loaitien AS `Loại tiền`,
            sotien AS `Số tiền`,
            quydoi AS `Quy đổi`,
            bang_goc_tim_thay AS `Bảng gốc tìm thấy`,
            ngay_baocao AS `Ngày báo cáo`,
            ten_file_goc AS `Tên file gốc`,
            ten_file_error AS `Tên file error`
        FROM bc48.ket_qua_do_tim_loi
        {where_clause}
        ORDER BY id DESC
    """)

    # BỔ SUNG: Khởi tạo engine riêng của phân hệ bc48
    engine = bc48.get_bc48_engine(db)
    # 4. Thực hiện truy vấn dữ liệu bằng Pandas qua engine vừa khởi tạo
    try:
        # Chuyển params_list thành tuple(params_list) để tránh lỗi driver
        df = pd.read_sql_query(query, db.engine, params=tuple(params_list))
    except Exception as e:
        return f"Lỗi truy vấn dữ liệu: {str(e)}", 500

    # 5. Xuất dữ liệu ra luồng dữ liệu cấu trúc Excel (.xlsx)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='KetQuaDoTim')
        
        # Tự động căn chỉnh độ rộng các cột trong Excel cho đẹp mắt
        workbook = writer.book
        worksheet = writer.sheets['KetQuaDoTim']
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output.seek(0)

    # 6. Trả file về phía client
    filename = f"Ket_qua_do_tim_loi_{loai_bc}_{thang_nam if thang_nam else 'TatCa'}.xlsx"
    return Response(
        output.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Thêm nút lệnh "Chỉ bóc tách mota_loi của CSV", lấy giá trị (Fx.x); vào bc48/ket-qua-do-tim; bc48_ket_qua_do_tim.html
@app.route('/bc48/ket-qua-do-tim/extract-errors', methods=['POST'])
@login_required
@admin_or_user_sd_bc48
def extract_errors_action():
    engine = bc48.get_bc48_engine(db)
    
    # Đọc chính xác dữ liệu từ Form POST được gửi lên
    loai_bc = request.form.get('loai_bc', '').strip()
    ky_baocao_type = request.form.get('ky_baocao_type', 'by_month')
    thang_nam = request.form.get('thang_nam', '').strip()

    # Chuẩn hóa loai_bc về dạng 'ALL' nếu người dùng không chọn hoặc chọn trống
    if not loai_bc or loai_bc == "":
        loai_bc = "ALL"

    # Xử lý điều kiện thời gian raw_thang_nam dựa trên dropdown cấu hình kỳ báo cáo
    if ky_baocao_type == 'all_time':
        raw_thang_nam = 'ALL'
    else:
        if not thang_nam:
            flash("Vui lòng nhập Tháng báo cáo (YYYYMM) hoặc chọn 'Tất cả kỳ' trước khi thực hiện!", "warning")
            return redirect(url_for('giaodien_ket_qua_do_tim', loai_bc=loai_bc if loai_bc != 'ALL' else ''))
        raw_thang_nam = thang_nam

    # Gọi hàm xử lý đóng gói Transaction an toàn từ file chuyên biệt bc48.py
    result = bc48.process_extract_errors_only(engine, loai_bc, raw_thang_nam)

    if isinstance(result, str):
        # Trường hợp hàm trả về Chuỗi ký tự (Cảnh báo hoặc Lỗi hệ thống)
        flash(result, "danger" if "Lỗi" in result else "info")
    else:
        # Trường hợp bóc tách thành công hoàn toàn
        ky_hien_thi = "TẤT CẢ KỲ" if raw_thang_nam == 'ALL' else thang_nam
        flash(f"Đã làm sạch & bóc tách thành công dữ liệu báo cáo [{loai_bc.upper()}] kỳ [{ky_hien_thi}] cho {len(result)} đơn vị/chi nhánh.", "success")
        
    # Điều hướng quay lại giao diện và truyền trả lại bộ lọc cũ cho người dùng
    return redirect(url_for('giaodien_ket_qua_do_tim', 
                            loai_bc=loai_bc if loai_bc != 'ALL' else '', 
                            thang_nam=thang_nam if ky_baocao_type != 'all_time' else ''))

# Thêm nút lệnh Chạy chủ động sp_xuly_error_trung_tam để có ket_qua_do_tim_loi (toàn bộ hệ thống)
# Hàm phụ trợ chạy ngầm thực thi SQL
def bg_worker_run_procedure(engine):
    try:
        # Sử dụng một kết nối biệt lập cho luồng ngầm
        with engine.begin() as connection:
            print("[Thread Ngầm] Bắt đầu TRUNCATE các bảng kết quả...")
            connection.execute(text("TRUNCATE TABLE danh_sach_bang_error_da_quet;"))
            connection.execute(text("TRUNCATE TABLE log_loi_xu_ly_bang;"))
            connection.execute(text("TRUNCATE TABLE chi_tiet_loi_master;"))
            connection.execute(text("TRUNCATE TABLE ket_qua_do_tim_loi;"))
            connection.execute(text("TRUNCATE TABLE giao_dich_error_khong_tim_thay;"))
            connection.execute(text("TRUNCATE TABLE bang_tonghop_ktra_gd_loi;"))

            connection.execute(text("TRUNCATE TABLE ket_qua_do_tim_loi_2;"))
            
            print("[Thread Ngầm] Đang thực thi CALL sp_xuly_error_trung_tam()...")
            connection.execute(text("CALL sp_xuly_error_trung_tam();"))

            print("[Thread Ngầm] Đang thực thi CALL sp_enrich_du_lieu_loi_phat()...")
            connection.execute(text("CALL sp_enrich_du_lieu_loi_phat();"))

            print("[Thread Ngầm] Tiến trình chạy SP hoàn thành thành công!")
    except Exception as e:
        print(f"[Thread Ngầm] Xảy ra lỗi nghiêm trọng: {str(e)}")
        
@app.route('/chay-sp-trung-tam-action', methods=['POST'])
@login_required
@admin_or_user_sd_bc48
def chay_sp_trung_tam_action():
    try:
        # Khởi tạo engine kết nối DB từ cấu hình hệ thống của bạn
        engine = bc48.get_bc48_engine(db)
        
        # Tạo và kích hoạt một Thread ngầm để xử lý cơ sở dữ liệu
        thr = threading.Thread(target=bg_worker_run_procedure, args=(engine,))
        thr.daemon = True  # Đảm bảo luồng tự giải phóng khi ứng dụng tắt
        thr.start()
        
        # Trả phản hồi ngay lập tức về giao diện Web (chỉ mất ~ 0.05 giây)
        flash("🚀 Tiến trình đối chiếu trung tâm đã được kích hoạt chạy ngầm hệ thống thành công! Vui lòng đợi vài phút rồi bấm nút 'Làm mới' để xem dữ liệu mới.", "info")
        
    except Exception as e:
        # Nếu có lỗi bất kỳ trong khối 'with', hệ thống tự động ROLLBACK
        print(f"Lỗi hệ thống khi thực thi SP: {str(e)}")
        flash(f"Không thể khởi tạo tiến trình chạy ngầm: {str(e)}", "error")

    # Quay trở về trang giao diện hiển thị kết quả dò tìm lỗi ban đầu        
    return redirect(url_for('giaodien_ket_qua_do_tim'))


# Tao Procedure sp_xuly_error_trung_tam_by_date lay so luong va chi tiet trang thai (KIEM TRA GIAO DICH LOI) tu tat ca cac bang _error v12 (ket_qua_do_tim_loi), tu ngay den ngay
def bg_worker_run_procedure_by_date(app_context, engine, tu_ngay, den_ngay):
    """
    Hàm worker chạy ngầm độc lập với request-response cycle của Flask.
    Sử dụng app_context để an toàn nếu cần log dữ liệu hoặc thao tác thêm với DB Flask.
    """
    with app_context:
        try:
            print(f"[Thread Ngầm] Bắt đầu chạy sp_xuly_error_trung_tam_by_date đối chiếu từ {tu_ngay} đến {den_ngay}...")
            
            # Gọi hàm xử lý thực thi Procedure trong bc48.py
            bc48.execute_central_procedure_by_date(engine, tu_ngay, den_ngay)
            
            print(f"[Thread Ngầm] Hoàn thành chạy sp_xuly_error_trung_tam_by_date đối chiếu từ {tu_ngay} đến {den_ngay} thành công!")
        except Exception as e:
            print(f"[Thread Ngầm] LỖI hệ thống khi thực thi SP: {str(e)}")

@app.route('/bc48/chay-sp-theo-ngay', methods=['POST'])
@login_required
@admin_or_user_sd_bc48
def chay_sp_theo_ngay_action():
    try:
        # 1. Lấy dữ liệu ngày từ form gửi lên
        tu_ngay = request.form.get('tu_ngay', '').strip()
        den_ngay = request.form.get('den_ngay', '').strip()
        
        if not tu_ngay or not den_ngay:
            flash("⚠️ Vui lòng chọn đầy đủ từ ngày đến ngày!", "danger")
            return redirect(url_for('giaodien_ket_qua_do_tim'))

        # 2. Khởi tạo engine kết nối DB
        engine = bc48.get_bc48_engine(db)
        
        # 3. Lấy app_context hiện tại 
        # (Bắt buộc phải truyền vào Thread nếu trong hàm bc48 có dùng cấu hình của Flask)
        app_context = current_app.app_context()
        
        # 4. Tạo và kích hoạt Thread ngầm (truyền engine, tu_ngay, den_ngay qua args)
        thr = threading.Thread(
            target=bg_worker_run_procedure_by_date, 
            args=(app_context, engine, tu_ngay, den_ngay)
        )
        thr.daemon = True  # Tự giải phóng khi ứng dụng tắt
        thr.start()
        
        # 5. Trả phản hồi ngay lập tức cho người dùng
        flash(f"🚀 Tiến trình đối chiếu từ ngày {tu_ngay} đến {den_ngay} đã được kích hoạt chạy ngầm! Vui lòng đợi vài phút rồi bấm nút 'Làm mới' để cập nhật dữ liệu.", "info")
        
    except Exception as e:
        print(f"Lỗi khởi tạo tiến trình Thread: {str(e)}")
        flash(f"❌ Không thể khởi tạo tiến trình chạy ngầm: {str(e)}", "danger")

    # Quay trở về giao diện ban đầu
    return redirect(url_for('giaodien_ket_qua_do_tim'))

####################################################################################
# Khai báo mail chi nhánh tiếp nhận csv lỗi để xử lý cập nhật
####################################################################################
@app.route('/bc48/khai-bao-mail', methods=['GET', 'POST'])
@login_required
@admin_required #@admin_or_user_sd_bc48
def giaodien_khai_bao_mail(): # Tên hàm này phải khớp với url_for trong index.html
    engine = bc48.get_bc48_engine(db)
    
    if request.method == 'POST':
        data = {
            'ma_hieu_1': request.form.get('ma_hieu_1'),
            'ten_chi_nhanh': request.form.get('ten_chi_nhanh'),
            'email_nhan': request.form.get('email_nhan'),
            'email_cc': request.form.get('email_cc'),
            'ghi_chu': request.form.get('ghi_chu'),
            'trang_thai': 1 if request.form.get('trang_thai') else 0
        }
        
        try:
            bc48.save_mail_config(engine, data)
            flash(f"Đã cập nhật thông tin mail cho chi nhánh {data['ma_hieu_1']}", "success")
        except Exception as e:
            flash(f"Lỗi hệ thống: {str(e)}", "danger")
            
        return redirect(url_for('giaodien_khai_bao_mail'))

    danh_sach = bc48.get_danh_sach_mail(engine)
    return render_template('bc48_khai_bao_mail.html', danh_sach=danh_sach)
####################################################################################
# Theo dõi kết quả khắc phục lỗi ket_qua_do_tim_loi
####################################################################################
@app.route('/bc48/gui-mail-loi', methods=['GET', 'POST'])
@login_required
@admin_or_user_sd_bc48
def giaodien_gui_mail_chi_nhanh():
    # Lấy trang hiện tại từ URL (mặc định là 1)
    page = request.args.get('page', 1, type=int)
    
    # Lấy filters từ form hoặc từ session để giữ trạng thái khi sang trang khác
    if request.method == 'POST':
        # Lấy sạch dữ liệu từ form
        filters = request.form.to_dict()
        # Loại bỏ các giá trị trống/None trước khi lưu
        filters = {k: v for k, v in filters.items() if v and v.strip() != ""}
        # Chỉ lưu vào session nếu có điều kiện lọc (tránh lưu rỗng)
        session['last_filters'] = filters
    else:
        # Nếu là GET (chuyển trang), lấy từ session
        # Nếu không có tham số gì, lấy filters mặc định là rỗng
        filters = session.get('last_filters', {})

    # --- Lấy danh sách mô tả lỗi duy nhất từ bảng `ket_qua_do_tim_loi` ---
    engine = bc48.get_bc48_engine(db)
    ds_mota_loi = []
    try:
        with engine.connect() as conn:
            query_mota = text("""
                SELECT DISTINCT mota_loi 
                FROM ket_qua_do_tim_loi 
                WHERE mota_loi IS NOT NULL AND mota_loi != '' 
                ORDER BY mota_loi ASC
            """)
            ds_mota_loi = [row[0] for row in conn.execute(query_mota).fetchall()]
    except Exception as e:
        print(f"Lỗi lấy danh sách mô tả lỗi: {str(e)}")

    # Gọi service xử lý phân trang lấy dữ liệu chi tiết lỗi từ bảng ket_qua_do_tim_loi
    pagination_obj = bc48.get_chi_tiet_loi_new(
        engine=engine, 
        filters=filters, 
        page=page,
        per_page=30
    )
    
    return render_template(
        'bc48_ket_qua_gui_mail.html', 
        chi_tiet_data=pagination_obj.get('items', []),
        pagination=pagination_obj, 
        filters=filters,
        ds_mota_loi=ds_mota_loi
    )

# Tách riêng theo macn từ bảng ket_qua_do_tim_loi, bổ sung 3 cột: noi_dung, trang_thai_ghi_nhan, ngay_chinh_sua
@app.route('/bc48/tach-rieng-macn', methods=['POST'])
@login_required
@admin_or_user_sd_bc48
def tach_rieng_macn():
    # 1. Lấy và chuẩn hóa bộ lọc từ Form gửi lên
    filters = request.form.to_dict()
    filters = {k: v for k, v in filters.items() if v and v.strip() != ""}
    # Đồng bộ bộ lọc vào session để giữ trạng thái màn hình tìm kiếm
    session['last_filters'] = filters
    
    # 2. Gọi tầng DB thông qua engine để lấy dữ liệu tổng hợp
    engine = bc48.get_bc48_engine(db)
    try:
        df = bc48.xu_ly_tach_theo_macn(engine, filters)
    except Exception as e:
        flash(f"Lỗi khi truy vấn dữ liệu từ cơ sở dữ liệu: {str(e)}", "danger")
        return redirect(url_for('giaodien_gui_mail_chi_nhanh'))
        
    # 3. Kiểm tra xem có dữ liệu thỏa mãn bộ lọc hay không
    if df.empty:
        flash("Không tìm thấy bản ghi lỗi nào thỏa mãn điều kiện lọc để phân tách!", "warning")
        return redirect(url_for('giaodien_gui_mail_chi_nhanh'))
    
    try:
        # - Loại báo cáo (Nếu chọn tất cả thì để ALL)
        loai_bc = filters.get('loai_bc', 'ALL').strip()
        
        # - Khoảng ngày tìm kiếm (Định dạng lại từ YYYY-MM-DD sang DDMMYYYY cho ngắn gọn)
        tu_ngay = filters.get('tu_ngay', '').replace('-', '')
        den_ngay = filters.get('den_ngay', '').replace('-', '')
        # Phòng trường hợp người dùng không chọn ngày, lấy giá trị mặc định trống
        str_tungay = f"Tu{tu_ngay}" if tu_ngay else ""
        str_dengay = f"Den{den_ngay}" if den_ngay else ""
        str_khoang_ngay = f"_{str_tungay}_{str_dengay}" if (str_tungay or str_dengay) else ""
        
        # - Thời điểm xuất file (Định dạng: YYYYMMDD_HHMMSS để đảm bảo không bao giờ trùng tên)
        thoidiemxuatfile = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Định nghĩa sẵn styles dùng chung ngoài vòng lặp để tiết kiệm RAM/CPU
        highlight_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid") # Màu vàng nhạt chuẩn theo mẫu
        custom_header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid") # Vàng đậm nổi bật
        custom_font = Font(name="Arial", size=10, bold=True, color="000000")
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        comments_dict = {
            "noi_dung": "Giá trị sau chỉnh sửa; nội dung chỉnh sửa; lý giải nguyên nhân sai lệch",
            "trang_thai_ghi_nhan": "Lý do chưa chỉnh sửa; Nhập đúng nội dung sau: DANG_SUA/DA_SUA_CHO_QUET/DA_HOAN_THANH",
            "ngay_chinh_sua": "Ghi chú: Ngày thực tế chỉnh sửa dd/mm/yyyy"
        }

        macn_col = 'macn' if 'macn' in df.columns else 'MACN'
        
        # Sử dụng BytesIO duy nhất cho file ZIP tổng
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Nhóm dữ liệu (Group By) theo cột 'macn'
            # Sử dụng .get('macn', 'macn') để phòng trường hợp database trả về tên cột chữ HOA/thường
            macn_col = 'macn' if 'macn' in df.columns else 'MACN'            
            for macn, group_df in df.groupby(macn_col):
                # Làm sạch chuỗi mã chi nhánh để đặt tên file cho chuẩn
                clean_macn = str(macn).strip()
                if not clean_macn or clean_macn == 'None':
                    clean_macn = "CHUA_PHAN_LOAI"

                # 1. Khởi tạo Workbook openpyxl mới
                wb = Workbook()
                ws = wb.active
                ws.title = f'CN_{clean_macn}'
                # Bật lưới dòng (Gridlines) trong Excel
                ws.views.sheetView[0].showGridLines = True
                # 2. Đổ dữ liệu từ DataFrame vào dòng Excel
                # dataframe_to_rows tự động lấy cả Header ở dòng 1
                for r in dataframe_to_rows(group_df, index=False, header=True):
                    ws.append(r)
                
                # 3. Xác định vị trí các cột mới bổ sung
                last_col_idx = group_df.shape[1] # Số lượng cột ban đầu
                new_headers = ["noi_dung", "trang_thai_ghi_nhan", "ngay_chinh_sua"]
                
                # Thêm tiêu đề cho 3 cột mới và cấu hình Ghi chú (Comment) + Định dạng bôi màu
                for i, h_name in enumerate(new_headers, start=1):
                    col_cell = ws.cell(row=1, column=last_col_idx + i, value=h_name)
                    col_cell.fill = custom_header_fill
                    col_cell.font = custom_font
                    col_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
                    # Thêm Ghi chú (Comment) chỉ dẫn khi di chuột vào tiêu đề
                    col_cell.comment = Comment(comments_dict[h_name], "Hệ Thống BC48")
                    # Định kích thước khung hiển thị ghi chú cho dễ đọc
                    col_cell.comment.width = 300
                    col_cell.comment.height = 50
                
                # Bôi màu nền nhẹ cho toàn bộ các ô trống thuộc 3 cột mới (để người dùng biết khu vực cần nhập liệu)
                thin_border = Border(
                    left=Side(style='thin', color='D9D9D9'),
                    right=Side(style='thin', color='D9D9D9'),
                    top=Side(style='thin', color='D9D9D9'),
                    bottom=Side(style='thin', color='D9D9D9')
                )

                for row_idx in range(2, ws.max_row + 1):
                    for i in range(1, 4):
                        cell = ws.cell(row=row_idx, column=last_col_idx + i)
                        cell.fill = highlight_fill
                        cell.border = thin_border
                        
                # =========================================================================
                # ĐẶT TÊN FILE THEO ĐÚNG CHUẨN KHOA HỌC YÊU CẦU:
                # cấu trúc: macn_loaibc_tungay_dengay_thoidiemxuatfile.xlsx
                # =========================================================================
                #filename = f"{clean_macn}_{loai_bc}{str_khoang_ngay}_{thoidiemxuatfile}.xlsx"
                # Ví dụ mẫu: 79001000_CTR_Tu20260701_Den20260702_20260702_213700.xlsx
                filename = f"{clean_macn}{loai_bc}_NHNN_ERROR{str_khoang_ngay}.xlsx"
                # =========================================================================
                
                ## Ghi file Excel trực tiếp vào file ZIP
                ##zip_file.writestr(filename, excel_buffer.getvalue())
                # openpyxl hỗ trợ lưu trực tiếp vào một file-like object mở từ zip
                with zip_file.open(filename, 'w') as zip_entry:
                    wb.save(zip_entry)
        
        # Đưa con trỏ bộ nhớ ZIP về đầu để chuẩn bị gửi sang Client
        zip_buffer.seek(0)

        # Đặt tên file ZIP bên ngoài cũng theo chuẩn khoa học chung để người dùng dễ quản lý
        zip_name = f"BC48_Tach_MaCN_{loai_bc}{str_khoang_ngay}_{thoidiemxuatfile}.zip"
        
        # 5. Trả file ZIP về cho trình duyệt tự động tải xuống
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=zip_name
        )
        
    except Exception as e:
        flash(f"Lỗi trong quá trình xử lý tách file và nén ZIP: {str(e)}", "danger")
        return redirect(url_for('giaodien_gui_mail_chi_nhanh'))

# Nút lệnh "4. Tách riêng theo macn (ket_qua_do_tim_loi_2)" trên bc48_ket_qua_gui_mail.html, gọi Stored Procedure sp_select_by_loai_bc với danh sách tham số
@app.route('/bc48/tach-macn-loi2', methods=['POST'])
@login_required
@admin_or_user_sd_bc48
def tach_rieng_macn_loi2():
    try:
        engine = bc48.get_bc48_engine(db)

        # =========================================================================
        # TỰ ĐỘNG CHUYỂN FONT THÀNH BASE64 ĐỂ XỬ LÝ OFFLINE TRÊN MAC
        # =========================================================================
        current_dir = os.path.dirname(os.path.abspath(__file__))
        font_dir = os.path.join(current_dir, 'fonts')
        
        path_regular = os.path.join(font_dir, 'times.ttf')
        path_bold = os.path.join(font_dir, 'timesbd.ttf')

        base64_times_regular = ""
        base64_times_bold = ""
        
        # Đọc và chuyển đổi tự động, nếu không tìm thấy file font sẽ dự phòng log lỗi
        try:
            with open(path_regular, "rb") as f_reg:
                base64_times_regular = base64.b64encode(f_reg.read()).decode('utf-8')
            with open(path_bold, "rb") as f_bld:
                base64_times_bold = base64.b64encode(f_bld.read()).decode('utf-8')
        except FileNotFoundError:
            base64_times_regular = ""
            base64_times_bold = ""
            print("CẢNH BÁO: Không tìm thấy file font trong thư mục 'fonts/'. Kiểm tra lại đường dẫn!")
        
        # 1. Lấy dữ liệu từ bộ lọc form
        filters = request.form.to_dict()
        filters = {k: v.strip() if isinstance(v, str) else v for k, v in filters.items()}
        
        # 2. Gọi hàm xử lý lấy dữ liệu tổng hợp
        df = bc48.xu_ly_tach_theo_macn_loi2(engine, filters)
        
        if df.empty:
            flash("Không tìm thấy dữ liệu phù hợp với điều kiện lọc từ bảng kết quả lỗi 2!", "warning")
            return redirect(url_for('giaodien_gui_mail_chi_nhanh'))

        # --- ĐỊNH DANH TÊN CỘT THEO ĐÚNG ĐẦU RA CỦA HÀM CHUẨN HÓA TRONG BC48.PY ---
        col_macn = 'macn'
        col_ngay_bc = 'ngaydulieubaocao'

        if 'macn' not in df.columns:
            df['macn'] = 'CHUA_PHAN_LOAI'
        if col_ngay_bc not in df.columns:
            df[col_ngay_bc] = 'UNKNOWN_DATE'

        # 3. Phân tích tham số ngày tháng / loại báo cáo làm tên file
        filter_loai_bc = filters.get('loai_bc', 'ALL').strip()
        loai_bc = filter_loai_bc if filter_loai_bc != 'ALL' else 'ALL_BC'

        tu_ngay = filters.get('tu_ngay', '')
        den_ngay = filters.get('den_ngay', '')
        
        def format_date_utility(date_val, target_format='%d%m%Y'):
            if not date_val or str(date_val).lower() == 'nan': 
                return ""
            try:
                if hasattr(date_val, 'strftime'):
                    return date_val.strftime(target_format)
                date_str = str(date_val).split()[0].strip()
                return datetime.strptime(date_str, '%Y-%m-%d').strftime(target_format)
            except:
                return str(date_val).replace('-', '').replace('/', '').split()[0]

        str_tu = format_date_utility(tu_ngay, '%d%m%Y')
        str_den = format_date_utility(den_ngay, '%d%m%Y')
        
        str_khoang_ngay = f"_TU_{str_tu}" if str_tu else ""
        if str_den:
            str_khoang_ngay += f"_DEN_{str_den}" if str_khoang_ngay else f"_DEN_{str_den}"
        if str_khoang_ngay and not str_khoang_ngay.startswith("_"):
            str_khoang_ngay = "_" + str_khoang_ngay
            
        # 4. Thiết lập Style Openpyxl & Nội dung mô tả 3 cột bổ sung
        highlight_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        custom_header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        custom_font = Font(name="Arial", size=10, bold=True, color="000000")
        thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

        comments_dict = {
            "Giá trị sau chỉnh sửa": "noi_dung; Lý giải nội dung chỉnh sửa; nguyên nhân sai lệch",
            "Lý do chưa chỉnh sửa": "trang_thai_ghi_nhan; Nhập đúng nội dung sau: DANG_SUA/DA_SUA_CHO_QUET/DA_HOAN_THANH",
            "Ghi chú": "ngay_chinh_sua: Ngày thực tế chỉnh sửa dd/mm/yyyy"
        }
        new_headers = ["Giá trị sau chỉnh sửa", "Lý do chưa chỉnh sửa", "Ghi chú"]

        # --- KHỞI TẠO CẤU TRÚC ĐỆM ---
        macn_files_registry = defaultdict(list)
        excel_files_data = {}

        # 5. TIẾN HÀNH XỬ LÝ VÀ GOM NHÓM DỮ LIỆU EXCEL
        for (macn_code, ngay_bc_val), group_df in df.groupby(['macn', col_ngay_bc]):
            clean_macn = str(macn_code).strip()
            if not clean_macn or clean_macn.lower() == 'nan':
                clean_macn = "KHONG_MA_CN"

            ngay_du_lieu_bao_cao = format_date_utility(ngay_bc_val, '%Y%m%d')
            if not ngay_du_lieu_bao_cao:
                ngay_du_lieu_bao_cao = "UNKNOWN"

            # Tạo file Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "KetQuaDoTimLoi2"
            
            for r in dataframe_to_rows(group_df, index=False, header=True):
                ws.append(r)
                
            start_col_idx = group_df.shape[1] + 1
            
            for i, header_name in enumerate(new_headers):
                col_idx = start_col_idx + i
                cell = ws.cell(row=1, column=col_idx, value=header_name)
                cell.fill = custom_header_fill
                cell.font = custom_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if header_name in comments_dict:
                    cell.comment = Comment(comments_dict[header_name], "Hệ thống")
                    
            max_row = ws.max_row
            if max_row > 1:
                for row in range(2, max_row + 1):
                    for i in range(len(new_headers)):
                        col_idx = start_col_idx + i
                        data_cell = ws.cell(row=row, column=col_idx)
                        data_cell.value = ""
                        data_cell.fill = highlight_fill
                        data_cell.border = thin_border
                        
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            excel_filename = f"{clean_macn}_{ngay_du_lieu_bao_cao}_{loai_bc}_NHNN_ERROR.xlsx"
            excel_files_data[excel_filename] = excel_buffer.getvalue()
            
            # Lưu vết danh sách file Excel thuộc về Mã CN này
            so_ban_ghi = max(0, max_row - 1)
            stt_hien_tai = len(macn_files_registry[clean_macn]) + 1
            macn_files_registry[clean_macn].append([str(stt_hien_tai), excel_filename, f"{so_ban_ghi:,}".replace(',', '.')])

        # =========================================================================
        # 6. MỞ LUỒNG GHI ZIP DUY NHẤT ĐỂ ĐÓNG GÓI EXCEL & CÁC FILE PDF RIÊNG BIỆT
        # =========================================================================
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            
            # Bước 6a: Đóng gói toàn bộ các file Excel chi tiết
            for filename, binary_content in excel_files_data.items():
                zf.writestr(filename, binary_content)

            # Bước 6b: Duyệt qua từng Mã CN để tạo duy nhất 1 file PDF bàn giao gộp chung các ngày
            for target_macn, files_list in macn_files_registry.items():
                try:
                    # Chuỗi HTML liền mạch ở phần CSS để ép WeasyPrint nhận diện Font Base64 trên Mac
                    html_content = f"<html><head><meta charset='utf-8'><style>@font-face {{font-family: 'TimesNewRomanCustom'; src: url(data:font/truetype;charset=utf-8;base64,{base64_times_regular}) format('truetype'); font-weight: normal; font-style: normal;}} @font-face {{font-family: 'TimesNewRomanCustom'; src: url(data:font/truetype;charset=utf-8;base64,{base64_times_bold}) format('truetype'); font-weight: bold; font-style: normal;}} @page {{ size: A4; margin: 20mm; }} body {{ font-family: 'TimesNewRomanCustom', serif; font-size: 11pt; line-height: 1.5; color: #000; }} .title {{ text-align: center; font-weight: bold; font-size: 14pt; margin-bottom: 25px; text-transform: uppercase; }} .info-block {{ margin-bottom: 15px; }} table {{ width: 100%; border-collapse: collapse; margin-top: 20px; font-size: 10pt; }} th, td {{ border: 1px solid #000; padding: 8px; text-align: center; vertical-align: middle; }} th {{ background-color: #343a40; color: white; font-weight: bold; }} .text-left {{ text-align: left; }} .signatures-container {{ margin-top: 40px; width: 100%; clear: both; }} .sig-box-left {{ float: left; width: 45%; text-align: center; font-weight: bold; }} .sig-box-right {{ float: right; width: 45%; text-align: center; font-weight: bold; }} .sub-title {{ font-weight: normal; font-style: italic; font-size: 9.5pt; display: block; margin-top: 3px; }}</style></head><body>"
                    
                    html_content += f"<div class='title'>BIÊN BẢN BÀN GIAO DANH SÁCH DỮ LIỆU LỖI<br>MÃ CHI NHÁNH: {target_macn}</div>"
                    html_content += f"<div class='info-block'><p><b>Ngày lập:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</p><p><b>Loại báo cáo:</b> {loai_bc}</p></div>"
                    html_content += "<table><thead><tr><th style='width: 8%;'>STT</th><th>Tên File Dữ Liệu Chi Tiết (Excel)</th><th style='width: 25%;'>Số Lượng Dòng Lỗi</th></tr></thead><tbody>"
                    
                    for row in files_list:
                        html_content += f"""<tr><td>{row[0]}</td><td class="text-left">{row[1]}</td><td>{row[2]}</td></tr>"""
                        
                    html_content += f"""</tbody></table><div class="signatures-container"><div class="sig-box-left">BP BÁO CÁO TUÂN THỦ PCRT<span class="sub-title">(Ký, ghi rõ họ tên)</span></div><div class="sig-box-right">BP PCRT CHI NHÁNH<span class="sub-title">(Ký, đóng dấu, ghi rõ họ tên)</span></div></div></body></html>"""
                    
                    # BIỆN PHÁP ÉP KHỬ ĐỘC: Dọn dẹp tuyệt đối mọi ký tự xuống dòng rác sinh ra trong chuỗi trước khi chuyển giao cho WeasyPrint
                    html_content = "".join([line.strip() for line in html_content.splitlines()])
                    
                    # 2. Xuất trực tiếp ra PDF thông qua WeasyPrint mà không sợ lỗi Font trên Mac
                    pdf_buffer = BytesIO()
                    HTML(string=html_content).write_pdf(pdf_buffer)
                    pdf_buffer.seek(0)
                    
                    # --- ĐẶT TÊN FILE PDF GỘP THEO ĐÚNG ĐỊNH DẠNG CỦA FILE EXCEL (Bỏ phần ngày chi tiết) ---
                    # Kết quả mẫu: CN01_ALL_BC_NHNN_ERROR_TU_01012026_DEN_10012026.pdf
                    pdf_filename = f"{target_macn}_{loai_bc}_NHNN_ERROR{str_khoang_ngay}.pdf"
                    
                    # Đóng gói file PDF gộp của Mã CN này vào ZIP
                    zf.writestr(pdf_filename, pdf_buffer.getvalue())
                    
                except Exception as pdf_err:
                    print(f"Lỗi khi sinh file PDF bàn giao gộp cho CN {target_macn}: {str(pdf_err)}")
                    traceback.print_exc()

        # 7. TRẢ FILE ZIP VỀ TRÌNH DUYỆT KHÁCH HÀNG
        zip_buffer.seek(0)
        zip_download_name = f"BC48_Tach_MaCN_Ngay_Loi2_{loai_bc}{str_khoang_ngay}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        
        return send_file(
            zip_buffer,
            mimetype="application/zip",
            as_attachment=True,
            download_name=zip_download_name
        )
        
    except Exception as e:
        traceback.print_exc()
        flash(f"Có lỗi xảy ra khi bóc tách nhóm dữ liệu Mã CN lỗi 2: {str(e)}", "danger")
        return redirect(url_for('giaodien_gui_mail_chi_nhanh'))


@app.route('/bc48/xoa-loc')
@login_required
@admin_or_user_sd_bc48
def xoa_loc_bc48():
    # Xóa bộ lọc cũ trong session
    if 'last_filters' in session:
        session.pop('last_filters')
    # Chuyển hướng về trang danh sách (lúc này session đã trống nên sẽ hiện tất cả)
    return redirect(url_for('giaodien_gui_mail_chi_nhanh'))

@app.route('/bc48/tiep-nhan-file/<int:id_dot>', methods=['POST'])
@login_required
@admin_or_user_sd_bc48
def tiep_nhan_file_tu_don_vi(id_dot):
    """
    Màn hình hoặc API tiếp nhận phản hồi xác nhận đơn vị đã nhận file thành công.
    """
    ghi_chu = request.form.get('ghi_chu', '')
    engine = bc48.get_bc48_engine(db)
    with engine.begin() as conn:
        conn.execute(text("""
            UPDATE lich_su_gui_file 
            SET trang_thai_tiep_nhan = 'DA_TIEP_NHAN', 
                ngay_tiep_nhan = NOW(),
                ghi_chu_tiep_nhan = :ghi_chu
            WHERE id_dot = :id_dot
        """), {"id_dot": id_dot, "ghi_chu": ghi_chu})
    flash("Đã cập nhật trạng thái tiếp nhận tệp tin xử lý thành công!", "success")
    return redirect(url_for('giaodien_gui_mail_chi_nhanh'))

@app.route('/bc48/tong-hop-loi', methods=['POST'])
@login_required
@admin_or_user_sd_bc48
def tong_hop_loi():
    try:
        # Khởi tạo engine kết nối DB từ cấu hình hệ thống của bạn
        engine = bc48.get_bc48_engine(db)
        
        # Tạo và kích hoạt một Thread ngầm để xử lý cơ sở dữ liệu
        thr = threading.Thread(target=bg_worker_run_procedure, args=(engine,))
        thr.daemon = True  # Đảm bảo luồng tự giải phóng khi ứng dụng tắt
        thr.start()
        
        # Trả phản hồi ngay lập tức về giao diện Web (chỉ mất ~ 0.05 giây)
        flash("🚀 Tiến trình đối chiếu trung tâm đã được kích hoạt chạy ngầm hệ thống thành công! Vui lòng đợi vài phút rồi bấm nút 'Làm mới' để xem dữ liệu mới.", "info")
        
    except Exception as e:
        # Nếu có lỗi bất kỳ trong khối 'with', hệ thống tự động ROLLBACK
        print(f"Lỗi hệ thống khi thực thi SP: {str(e)}")
        flash(f"Không thể khởi tạo tiến trình chạy ngầm: {str(e)}", "error")

    # Quay trở về trang giao diện hiển thị kết quả dò tìm lỗi ban đầu        
    return redirect(url_for('giaodien_ket_qua_do_tim'))

# Xuất sao kê từ bảng ket_qua_do_tim_loi
@app.route('/bc48/export-sao-ke', methods=['POST'])
@login_required
@admin_or_user_sd_bc48
def export_sao_ke():
    # Ưu tiên lấy dữ liệu trực tiếp từ Form đang hiển thị trên màn hình lúc bấm nút
    if request.method == 'POST':
        filters = request.form.to_dict()
        filters = {k: v for k, v in filters.items() if v and v.strip() != ""}
        # Đồng bộ cập nhật lại session để giữ trạng thái đồng nhất
        if filters:
            session['last_filters'] = filters
    else:
        filters = session.get('last_filters', {})
    
    try:
        # Gọi hàm xử lý và truyền trực tiếp bộ lọc chuẩn vừa lấy từ bảng mới ket_qua_do_tim_loi
        file_data = bc48.export_sao_ke_duy_nhat(db, filters)
        
        # =========================================================================
        # BỔ SUNG: Xây dựng chuỗi tiêu chí lọc gắn vào tên File
        # =========================================================================
        # 1. Loại báo cáo (Mặc định ALL nếu trống)
        loai_bc_fn = filters.get('loai_bc', 'ALL').strip().upper()
        
        # 2. Mã chi nhánh (Nếu có thì lấy, không có ghi ALL_CN)
        macn_fn = filters.get('macn', '').strip()
        macn_str = f"_CN{macn_fn.zfill(8)}" if macn_fn else "_ALL_CN"
        
        # 3. Khoảng thời gian báo cáo
        tu_ngay_fn = filters.get('tu_ngay', '').replace('-', '')
        den_ngay_fn = filters.get('den_ngay', '').replace('-', '')
        
        if tu_ngay_fn and den_ngay_fn:
            time_str = f"_{tu_ngay_fn}_TO_{den_ngay_fn}"
        elif tu_ngay_fn:
            time_str = f"_FROM_{tu_ngay_fn}"
        elif den_ngay_fn:
            time_str = f"_UNTIL_{den_ngay_fn}"
        else:
            time_str = "_ALL_TIME"
            
        # Thời gian xuất file thực tế của hệ thống để tránh trùng tên file
        now_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Tạo tên file hoàn chỉnh: ví dụ: "sao_ke_duy_nhat_CTR_CN00012345_20260101_TO_20260331_20260702_174500.xlsx"
        filename = f"sao_ke_duy_nhat_{loai_bc_fn}{macn_str}{time_str}_{now_str}.xlsx"
        # =========================================================================
        
        return send_file(
            file_data,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        flash(f"Lỗi khi xuất file sao kê: {str(e)}", "danger")
        return redirect(url_for('giaodien_gui_mail_chi_nhanh'))


@app.route('/bc48/xu-ly-xuat-csv-zip', methods=['POST'])
@login_required
@admin_or_user_sd_bc48
def xu_ly_xuat_csv_zip():
    # 1. Lấy dữ liệu từ form
    filters = request.form.to_dict()
    
    tu_ngay = filters.get('tu_ngay')
    den_ngay = filters.get('den_ngay')
    loai_bc = filters.get('loai_bc', 'ALL')
    loai_mat_khau = filters.get('loai_mat_khau', 'fixed')
    
    if not tu_ngay or not den_ngay:
        flash("Vui lòng chọn Từ ngày và Đến ngày!", "warning")
        return redirect(url_for('giaodien_gui_mail_chi_nhanh'))

    # SỬA LỖI TẠI ĐÂY: Lấy định danh an toàn từ current_user (Thử ma_nhan_vien, ma_nv rồi đến username)
    nguoi_dung_hien_tai = (
        getattr(current_user, 'ma_nhan_vien', None) or 
        getattr(current_user, 'ma_nv', None) or 
        getattr(current_user, 'username', 'Hệ thống')
    )
    
    # 2. Khởi tạo engine và gọi hàm xử lý từ bc48.py
    engine = bc48.get_bc48_engine(db)
    result = bc48.chi_xuat_csv_ket_qua_do_tim_loi(
        engine=engine, 
        loai_bc=loai_bc, 
        tu_ngay=tu_ngay, 
        den_ngay=den_ngay, 
        loai_mat_khau=loai_mat_khau, 
        nguoi_thuc_hien=nguoi_dung_hien_tai,
        filters_data=filters  # <-- TRUYỀN THÊM TOÀN BỘ FORM ĐỂ LỌC THEO MACN, MA_HIEU_1, MA_HIEU_2
    )

    # 3. Phản hồi kết quả dựa trên kiểu dữ liệu trả về
    if isinstance(result, list): 
        # Nếu thành công trả về danh sách, lưu vào session phục vụ nút Gửi Mail độc lập
        session['files_to_send'] = result
        success_count = len(result)
        flash(f"Đã xuất thành công {success_count} file ZIP mã hóa cho các đơn vị! Dữ liệu đã sẵn sàng để gửi Mail.", "success")
    else:
        # Nếu trả về thông báo chuỗi (Lỗi hệ thống hoặc Không tìm thấy dữ liệu)
        # Xóa dữ liệu cũ trong session cũ (nếu có) để tránh gửi nhầm
        session.pop('files_to_send', None)
        flash(result, "danger" if "Lỗi" in result else "warning")
        
    return redirect(url_for('giaodien_gui_mail_chi_nhanh'))
####################################################################################
# Nhật ký mật khẩu nén file csv sau khi bóc tách mota_loi của CSV
# tránh việc gửi kèm mật khẩu vào email làm lộ dữ liệu => giải pháp đóng gói mã hóa và gửi mail theo đúng các thông tin cấu hình từ bảng danh_sach_mail_chi_nhanh
####################################################################################
@app.route('/bc48/nhat-ky-mat-khau', methods=['GET'])
@login_required
@admin_or_user_sd_bc48
def nhat_ky_mat_khau():
    page = request.args.get('page', 1, type=int)
    per_page = 20  # Số lượng dòng trên mỗi trang
    offset = (page - 1) * per_page

    # 1. KHỞI TẠO ENGINE ĐỂ TRỎ ĐÚNG SANG DATABASE BC48
    engine = bc48.get_bc48_engine(db)

    # 2. Kiểm tra quyền kiểm soát tối cao (Admin)
    is_admin = (getattr(current_user, 'role', '') == 'admin') or getattr(current_user, 'is_admin', False)
    
    user_manh8so_moi = None

    # 3. Nếu KHÔNG PHẢI ADMIN, tiến hành truy vết lấy mã Chi nhánh (manh8so_moi) của User
    if not is_admin:
        try:
            # Lưu ý: Thông tin tài khoản users và thông tin nhân sự thong_tin_nguoi_lao_dong
            # nằm ở DB gốc (db.session). Lấy trường ma_hieu_1 từ bảng don_vi
            query_user_branch = text("""
                SELECT dv.ma_hieu_1 
                FROM users u
                INNER JOIN thong_tin_nguoi_lao_dong nv ON u.ma_nhan_vien = nv.ma_nhan_vien
                INNER JOIN don_vi dv ON nv.ma_hieu_2 = dv.ma_hieu_2
                WHERE u.ma_nhan_vien = :ma_nv
                LIMIT 1
            """)
            user_manh8so_moi = db.session.execute(query_user_branch, {'ma_nv': current_user.ma_nhan_vien}).scalar()
            
            if not user_manh8so_moi:
                flash("Tài khoản của bạn chưa được liên kết với mã đơn vị quản lý trên hệ thống. Vui lòng liên hệ Admin!", "danger")
                return redirect(url_for('index'))
                
        except Exception as e:
            print(f"Lỗi truy vết đơn vị cán bộ: {str(e)}")
            flash("Có lỗi xảy ra khi xác thực quyền đơn vị!", "danger")
            return redirect(url_for('index'))

    # 4. Thiết lập câu lệnh SQL lấy danh sách mật khẩu dựa trên phân quyền (Đồng bộ theo manh8so_moi)
    if is_admin:
        count_query = text("SELECT COUNT(*) FROM log_mat_khau_xuat_file")
        data_query = text("""
            SELECT ngay_xuat, manh8so_moi, loai_bao_cao, ten_file, mat_khau, nguoi_xuat 
            FROM log_mat_khau_xuat_file 
            ORDER BY ngay_xuat DESC 
            LIMIT :limit OFFSET :offset
        """)
        params = {'limit': per_page, 'offset': offset}
    else:
        count_query = text("SELECT COUNT(*) FROM log_mat_khau_xuat_file WHERE manh8so_moi = :manh8so_moi")
        data_query = text("""
            SELECT ngay_xuat, manh8so_moi, loai_bao_cao, ten_file, mat_khau, nguoi_xuat 
            FROM log_mat_khau_xuat_file 
            WHERE manh8so_moi = :manh8so_moi
            ORDER BY ngay_xuat DESC 
            LIMIT :limit OFFSET :offset
        """)
        params = {'manh8so_moi': user_manh8so_moi, 'limit': per_page, 'offset': offset}

    # 5. THỰC THI TRUY VẤN TRÊN ENGINE CỦA BC48
    try:
        with engine.connect() as conn:
            # Thực thi đếm tổng số dòng trên DB BC48
            total_rows = conn.execute(count_query, {'manh8so_moi': user_manh8so_moi} if not is_admin else {}).scalar() or 0
            total_pages = math.ceil(total_rows / per_page) if total_rows > 0 else 1
            
            if page > total_pages:
                page = total_pages
                if not is_admin:
                    params['offset'] = (page - 1) * per_page

            # Thực thi lấy dữ liệu log mật khẩu trên DB BC48
            result = conn.execute(data_query, params).mappings().fetchall()
            
            # Đóng gói dữ liệu sang mảng Dict để hiển thị lên giao diện Jinja2
            pass_logs = []
            for row in result:
                pass_logs.append({
                    'ngay_xuat': row['ngay_xuat'],
                    # Giữ nguyên key 'ma_hieu_1' để tương thích với cấu trúc của file template HTML hiện tại
                    'ma_hieu_1': row['manh8so_moi'], 
                    'loai_bao_cao': row['loai_bao_cao'],
                    'ten_file': row['ten_file'],
                    'mat_khau': row['mat_khau'],
                    'nguoi_xuat': row['nguoi_xuat']
                })
            
    except Exception as e:
        print(f"Lỗi thực thi dữ liệu mật khẩu trên DB BC48: {str(e)}")
        pass_logs = []
        total_pages = 1

    return render_template(
        'bc48_nhat_ky_mat_khau.html',
        title='Nhật ký cấp mật khẩu',
        pass_logs=pass_logs,
        page=page,
        total_pages=total_pages
    )

# Sau khi xuất CSV, nén đặt mật khẩu thì gửi mail
@app.route('/bc48/gui-mail-sau-xuat', methods=['POST'])
@login_required
@admin_or_user_sd_bc48
def gui_mail_sau_xuat():
    # 1. Kiểm tra dữ liệu tệp tin đệm trong session
    files_to_send = session.get('files_to_send')
    if not files_to_send or not isinstance(files_to_send, list):
        flash("Không tìm thấy danh sách tệp tin sẵn sàng để gửi. Vui lòng bấm nút 'Xuất CSV Nén (ZIP)' trước!", "warning")
        return redirect(url_for('giaodien_gui_mail_chi_nhanh'))

    # 2. Lấy cấu hình SMTP từ app.config đã khai báo
    smtp_setting = app.config.get('SMTP_CONFIG', {})
    mail_config = {
        'SMTP_SERVER': smtp_setting.get('server'),
        'SMTP_PORT': int(smtp_setting.get('port', 25)) if smtp_setting.get('port') else 25,
        'SMTP_USERNAME': smtp_setting.get('email'),
        'SMTP_PASSWORD': smtp_setting.get('password'),
        'SENDER_EMAIL': smtp_setting.get('email'),
        'USE_TLS': smtp_setting.get('use_tls', False)
    }

    success_mail_count = 0
    fail_mail_count = 0
    errors_list = []
    
    # Lấy thông tin cán bộ thực hiện thao tác gửi mail
    user_action = getattr(current_user, 'username', 'Hệ thống') 
    
    # Khởi tạo engine kết nối chính xác tới cơ sở dữ liệu phân tách db_bc48
    engine_bc48 = bc48.get_bc48_engine(db)

    # 3. Vòng lặp duyệt danh sách gửi mail cho từng đơn vị
    for file_info in files_to_send:
        ma_cn = file_info.get('ma_cn')
        email_nhan = file_info.get('email')
        zip_path = file_info.get('zip_path')
        zip_name = file_info.get('zip_name')

        # Bóc tách khoảng thời gian và loại BC từ file_info để định vị dòng dữ liệu master
        f_tu_ngay = file_info.get('tu_ngay')
        f_den_ngay = file_info.get('den_ngay')
        loai_bc = zip_name.split('_')[0] if zip_name else "BC48"

        # Lấy thông tin email CC từ danh bạ hệ thống (sử dụng engine_bc48 luôn cho đồng bộ)
        email_cc = ""
        try:
            with engine_bc48.connect() as conn:
                df_mail_sys = bc48.lay_danh_sach_mail_toan_he_thong(conn)
                if not df_mail_sys.empty:
                    df_current = df_mail_sys[df_mail_sys['macn'] == str(ma_cn).strip().upper()]
                    if not df_current.empty and 'email_cc' in df_current.columns:
                        email_cc = str(df_current['email_cc'].iloc[0]).strip()
        except Exception as cc_ex:
            print(f"[CẢNH BÁO] Không lấy được Email CC cho đơn vị {ma_cn}: {str(cc_ex)}")
            email_cc = ""

        # Gọi hàm core thực hiện gửi mail đính kèm tệp tin ZIP
        status = bc48.gui_mail_dinh_kem_zip(
            mail_config=mail_config,
            target_email=email_nhan,
            cc_email=email_cc,
            file_path=zip_path,
            file_name=zip_name,
            loai_bc=loai_bc,
            ma_cn=ma_cn
        )

        thoi_gian_gui = datetime.now()

        if status == "SUCCESS":
            success_mail_count += 1
            trang_thai_log = "Thành công"
            
            # CẬP NHẬT TRẠNG THÁI TIẾP NHẬN: Chuyển sang Chờ các đơn vị nhận phản hồi xử lý
            try:
                with engine_bc48.begin() as update_conn:
                    # Bước A: Khởi tạo đợt gửi mới vào bảng lich_su_gui_file để sinh ID tự động
                    sql_insert_history = text("""
                        INSERT INTO lich_su_gui_file (
                            ten_file, manh8so_moi, loai_bao_cao, ngay_gui, 
                            trang_thai_tiep_nhan, nguoi_gui, email_nhan
                        ) VALUES (
                            :ten_file, :ma_cn, :loai_bc, :ngay_gui, 
                            'DA_GUI_MAIL', :nguoi_gui, :email_nhan
                        )
                    """)

                    res_insert = update_conn.execute(sql_insert_history, {
                        "ten_file": zip_name,
                        "ma_cn": str(ma_cn).strip().upper(),
                        "loai_bc": loai_bc,
                        "ngay_gui": thoi_gian_gui,
                        "nguoi_gui": user_action,
                        "email_nhan": email_nhan
                    })
                    
                    # Lấy ra id_dot vừa được sinh tự động (Last Inserted ID)
                    new_id_dot = res_insert.lastrowid

                    # Bước B: Xây dựng các điều kiện quét để tìm chính xác các giao dịch thuộc file này
                    where_clauses = ["k.macn = :ma_cn"]
                    params = {"ma_cn": str(ma_cn).strip().upper()}

                    if loai_bc != 'ALL':
                        where_clauses.append("k.loai_bc = :loai_bc")
                        params["loai_bc"] = loai_bc
                    if f_tu_ngay:
                        where_clauses.append("k.ngay_baocao >= :tu_ngay")
                        params["tu_ngay"] = f_tu_ngay
                    if f_den_ngay:
                        where_clauses.append("k.ngay_baocao <= :den_ngay")
                        params["den_ngay"] = f_den_ngay

                    where_sql = " AND ".join(where_clauses)

                    # Bước C: Chèn hoặc Cập nhật bảng theo_doi_khac_phuc_loi gán kèm mã `id_dot` và ghi nhận tiến trình gửi mail
                    sql_update_khac_phuc = f"""
                        INSERT INTO theo_doi_khac_phuc_loi (
                            ma_giao_dich, ngay_baocao, bang_goc_tim_thay, 
                            trang_thai_khac_phuc, id_dot, ngay_cap_nhat,
                            ngay_gui_mail_dau, ngay_gui_gan_nhat, so_lan_gui_mail
                        )
                        SELECT 
                            k.ma_giao_dich, k.ngay_baocao, CONCAT(LOWER(k.loai_bc), '_error'), 
                            'DANG_SUA', :id_dot, NOW(),
                            NOW(), NOW(), 1
                        FROM ket_qua_do_tim_loi k
                        WHERE {where_sql}
                        ON DUPLICATE KEY UPDATE 
                            trang_thai_khac_phuc = 'DANG_SUA',
                            id_dot = :id_dot,
                            ngay_cap_nhat = NOW(),
                            -- Nếu chưa từng gửi (NULL), gán bằng thời gian hiện tại. Nếu có rồi thì giữ nguyên.
                            ngay_gui_mail_dau = COALESCE(ngay_gui_mail_dau, NOW()), 
                            -- Luôn luôn cập nhật thời điểm gửi mới nhất
                            ngay_gui_gan_nhat = NOW(), 
                            -- Tự động tăng lũy tiến số lần gửi mail lên +1
                            so_lan_gui_mail = COALESCE(so_lan_gui_mail, 0) + 1;
                    """

                    # Thêm id_dot vào params để truyền vào câu lệnh SQL
                    params["id_dot"] = new_id_dot
                    update_conn.execute(text(sql_update_khac_phuc), params)
                    
                    # Bước D: Cập nhật trạng thái xử lý tổng quát (trang_thai_xuly = 1) bên bảng gốc
                    sql_update_loi_master = f"""
                        UPDATE ket_qua_do_tim_loi k
                        SET k.trang_thai_xuly = 1
                        WHERE {where_sql};
                    """
                    # Xóa id_dot khỏi params trước khi chạy lệnh update bảng lỗi gốc nếu cần thiết (ở đây dùng chung params có dư id_dot vẫn không sao)
                    update_conn.execute(text(sql_update_loi_master), params)

                    print(f"[AML SYSTEM] Đã ghi nhận lịch sử đợt {new_id_dot} vào lich_su_gui_file cho chi nhánh {ma_cn}")
            except Exception as ex:
                print(f"[LỖI TIẾN TRÌNH DB] Không thể cập nhật trạng thái khắc phục cho đơn vị {ma_cn}: {str(ex)}")
        else:
            fail_mail_count += 1
            trang_thai_log = f"Thất bại: {status}"
            errors_list.append(f"Đơn vị {ma_cn}: {status}")

        # GHI NHẬT KÝ: Insert thông tin vết gửi vào bảng lich_su_gui_file; log_gui_mail_bc48
        try:
            with engine_bc48.begin() as log_conn:
                log_conn.execute(text("""
                    INSERT INTO log_gui_mail_bc48 (
                        manh8so_moi, ngay_gui, email_nhan, ten_file_dinh_kem, duong_dan_file, trang_thai, nguoi_thuc_hien
                    ) VALUES (:ma_cn, :ngay_gui, :email_nhan, :zip_name, :zip_path, :trang_thai, :nguoi)
                """), {
                    "ma_cn": ma_cn, "ngay_gui": thoi_gian_gui, "email_nhan": email_nhan,
                    "zip_name": zip_name, "zip_path": zip_path, "trang_thai": trang_thai_log, "nguoi": user_action
                })
        except Exception as log_ex:
            print(f"[LỖI GHI LOG DATABASE] Không thể ghi nhật ký cho đơn vị {ma_cn}: {str(log_ex)}")

    # 4. Phản hồi kết quả tổng quát ra màn hình ứng dụng
    if fail_mail_count == 0:
        flash(f"Thực thi hoàn tất! Đã gửi thành công toàn bộ {success_mail_count} Email, chuyển trạng thái dữ liệu sang 'Đã gửi mail' và lưu nhật ký hệ thống.", "success")
        session.pop('files_to_send', None)  # Xóa sạch session đệm sau khi hoàn thành quy trình
    else:
        flash(f"Tiến trình kết thúc có lỗi cục bộ: Gửi thành công {success_mail_count} thư, thất bại {fail_mail_count} thư. Chi tiết xem tại Terminal hoặc bảng log_gui_mail_bc48.", "danger")

    return redirect(url_for('giaodien_gui_mail_chi_nhanh'))

@app.route('/bc48/gui-mail-sau-xuat', methods=['POST'])
@login_required
@admin_or_user_sd_bc48
def gui_mail_sau_xuat_v01():
    # 1. Kiểm tra dữ liệu tệp tin đệm trong session
    files_to_send = session.get('files_to_send')
    if not files_to_send or not isinstance(files_to_send, list):
        flash("Không tìm thấy danh sách tệp tin sẵn sàng để gửi. Vui lòng bấm nút 'Xuất CSV Nén (ZIP)' trước!", "warning")
        return redirect(url_for('giaodien_gui_mail_chi_nhanh'))

    # 2. Lấy cấu hình SMTP từ app.config đã khai báo
    smtp_setting = app.config.get('SMTP_CONFIG', {})
    mail_config = {
        'SMTP_SERVER': smtp_setting.get('server'),
        'SMTP_PORT': int(smtp_setting.get('port', 25)) if smtp_setting.get('port') else 25,
        'SMTP_USERNAME': smtp_setting.get('email'),
        'SMTP_PASSWORD': smtp_setting.get('password'),
        'SENDER_EMAIL': smtp_setting.get('email'),
        'USE_TLS': smtp_setting.get('use_tls', False)
    }

    success_mail_count = 0
    fail_mail_count = 0
    errors_list = []
    
    # Lấy thông tin cán bộ thực hiện thao tác gửi mail
    user_action = getattr(current_user, 'username', 'Hệ thống') 
    
    # Khởi tạo engine kết nối chính xác tới cơ sở dữ liệu phân tách db_bc48
    engine_bc48 = bc48.get_bc48_engine(db)

    # 3. Vòng lặp duyệt danh sách gửi mail cho từng đơn vị
    for file_info in files_to_send:
        ma_cn = file_info.get('ma_cn')
        email_nhan = file_info.get('email')
        zip_path = file_info.get('zip_path')
        zip_name = file_info.get('zip_name')
        id_dot_theo_doi = file_info.get('id_dot') # Lấy ID đợt vừa khởi tạo ở bước xuất file
        
        f_tu_ngay = file_info.get('tu_ngay')
        f_den_ngay = file_info.get('den_ngay')
        loai_bc = zip_name.split('_')[0] if zip_name else "BC48"

        # Lấy thông tin email CC từ danh bạ hệ thống (sử dụng engine_bc48 luôn cho đồng bộ)
        email_cc = ""
        try:
            with engine_bc48.connect() as conn:
                df_mail_sys = bc48.lay_danh_sach_mail_toan_he_thong(conn)
                if not df_mail_sys.empty:
                    df_current = df_mail_sys[df_mail_sys['macn'] == str(ma_cn).strip().upper()]
                    if not df_current.empty and 'email_cc' in df_current.columns:
                        email_cc = str(df_current['email_cc'].iloc[0]).strip()
        except Exception as cc_ex:
            print(f"[CẢNH BÁO] Không lấy được Email CC cho đơn vị {ma_cn}: {str(cc_ex)}")
            email_cc = ""

        # Gọi hàm core thực hiện gửi mail đính kèm tệp tin ZIP
        status = bc48.gui_mail_dinh_kem_zip(
            mail_config=mail_config,
            target_email=email_nhan,
            cc_email=email_cc,
            file_path=zip_path,
            file_name=zip_name,
            loai_bc=loai_bc,
            ma_cn=ma_cn
        )

        thoi_gian_gui = datetime.now()

        if status == "SUCCESS":
            success_mail_count += 1
            trang_thai_log = "Thành công"
            
            # CẬP NHẬT TRẠNG THÁI TIẾP NHẬN: Chuyển sang Chờ các đơn vị nhận phản hồi xử lý
            try:
                with engine_bc48.begin() as update_conn:
                    update_conn.execute(text("""
                        UPDATE lich_su_gui_file 
                        SET trang_thai_tiep_nhan = 'DA_GUI_MAIL', ngay_gui = :ngay_gui
                        WHERE id_dot = :id_dot
                    """), {"ngay_gui": thoi_gian_gui, "id_dot": id_dot_theo_doi})
                    # Đồng thời cập nhật trạng thái chi tiết của từng mã lỗi thuộc đợt này
                    update_conn.execute(text("""
                        UPDATE theo_doi_khac_phuc_loi 
                        SET trang_thai_khac_phuc = 'DANG_SUA' 
                        WHERE id_dot = :id_dot
                    """), {"id_dot": id_dot_theo_doi})
                    print(f"[AML SYSTEM] Đã chuyển trang_thai_xuly = 1 cho đơn vị {ma_cn} [{loai_bc}]")
            except Exception as ex:
                print(f"[LỖI TIẾN TRÌNH DB] Không thể cập nhật trạng thái xử lý cho đơn vị {ma_cn}: {str(ex)}")
        else:
            fail_mail_count += 1
            trang_thai_log = f"Thất bại: {status}"
            errors_list.append(f"Đơn vị {ma_cn}: {status}")

        # GHI NHẬT KÝ: Insert thông tin vết gửi vào bảng log_gui_mail_bc48
        try:
            with engine_bc48.begin() as log_conn:
                log_conn.execute(text("""
                    INSERT INTO log_gui_mail_bc48 (
                        manh8so_moi, ngay_gui, email_nhan, ten_file_dinh_kem, duong_dan_file, trang_thai, nguoi_thuc_hien
                    ) VALUES (:ma_cn, :ngay_gui, :email_nhan, :zip_name, :zip_path, :trang_thai, :nguoi)
                """), {
                    "ma_cn": ma_cn, "ngay_gui": thoi_gian_gui, "email_nhan": email_nhan,
                    "zip_name": zip_name, "zip_path": zip_path, "trang_thai": trang_thai_log, "nguoi": user_action
                })
        except Exception as log_ex:
            print(f"[LỖI GHI LOG DATABASE] Không thể ghi nhật ký cho đơn vị {ma_cn}: {str(log_ex)}")

    # 4. Phản hồi kết quả tổng quát ra màn hình ứng dụng
    if fail_mail_count == 0:
        flash(f"Thực thi hoàn tất! Đã gửi thành công toàn bộ {success_mail_count} Email, chuyển trạng thái dữ liệu sang 'Đã gửi mail' và lưu nhật ký hệ thống.", "success")
        session.pop('files_to_send', None)  # Xóa sạch session đệm sau khi hoàn thành quy trình
    else:
        flash(f"Tiến trình kết thúc có lỗi cục bộ: Gửi thành công {success_mail_count} thư, thất bại {fail_mail_count} thư. Chi tiết xem tại Terminal hoặc bảng log_gui_mail_bc48.", "danger")

    return redirect(url_for('giaodien_gui_mail_chi_nhanh'))


####################################################################################
# Nhận file excel phản hồi từ chi nhánh báo cáo kết quả khắc phục chỉnh sửa CSV do Cục PCRT trả lỗi
####################################################################################
@app.route('/bc48/import-phan-hoi', methods=['GET'])
@login_required
@admin_or_user_sd_bc48
def giaodien_import_phan_hoi():
    """
    Hiển thị trang giao diện Import file phản hồi 
    và danh sách 20 bản ghi được chi nhánh sửa đổi cập nhật gần đây nhất trong ngày.
    """
    engine = bc48.get_bc48_engine(db)
    query_history = text("""
        SELECT ma_giao_dich, DATE_FORMAT(ngay_baocao, '%Y-%m-%d') as ngay_baocao, 
               bang_goc_tim_thay, trang_thai_khac_phuc, ghi_chu_xu_ly,
               DATE_FORMAT(ngay_cap_nhat, '%d/%m/%Y %H:%i') as ngay_cap_nhat
        FROM theo_doi_khac_phuc_loi
        ORDER BY ngay_cap_nhat DESC
        LIMIT 20;
    """)
    
    lich_su_logs = []
    try:
        with engine.connect() as conn:
            res = conn.execute(query_history)
            lich_su_logs = [dict(row._mapping) for row in res]
    except Exception as e:
        print(f"Lỗi lấy lịch sử import: {str(e)}")
        
    return render_template('bc48_import_phan_hoi.html', lich_su_logs=lich_su_logs)


@app.route('/bc48/import-phan-hoi-action', methods=['POST'])
@login_required
@admin_or_user_sd_bc48
def import_phan_hoi_chi_nhanh():
    """
    Xử lý đọc file Excel phản hồi từ các chi nhánh gửi lên,
    linh hoạt ánh xạ các tên cột thực tế của chi nhánh về tên cột chuẩn hệ thống;
    chỉ lọc lấy đúng các cột cần thiết và cập nhật tiến trình vào DB.
    """
    if 'file_excel' not in request.files:
        flash('Lỗi: Không tìm thấy tệp tin gửi lên.', 'danger')
        return redirect(url_for('giaodien_import_phan_hoi'))
        
    file = request.files['file_excel']
    if file.filename == '':
        flash('Vui lòng chọn một file Excel hợp lệ từ máy tính!', 'warning')
        return redirect(url_for('giaodien_import_phan_hoi'))
        
    try:
        # Đọc dữ liệu Excel trực tiếp từ bộ nhớ
        df = pd.read_excel(file)

        # 1. Chuẩn hóa tên cột thô ban đầu (bỏ khoảng trắng đầu cuối, giữ nguyên hoa thường để map chính xác)
        df.columns = [str(c).strip() for c in df.columns]

        # 2. Định nghĩa từ điển ánh xạ (Từ tên cột của chi nhánh -> Tên cột chuẩn của hệ thống)
        MAPPING_COLS = {
            'magiaodichtrongfile': 'ma_giao_dich',
            'ngaydulieubaocao': 'ngay_baocao',
            'Lý do chưa chỉnh sửa': 'trang_thai_ghi_nhan',
            'maipcas': 'ma_hieu_2',
            'motaloinhnn': 'mota_loi'
        }
        # Tiến hành đổi tên các cột nếu tìm thấy trong file của chi nhánh
        df.rename(columns=MAPPING_COLS, inplace=True)

        # 3. Chuẩn hóa lại toàn bộ cột về chữ thường để khớp với các logic xử lý bên dưới của bạn
        df.columns = [str(c).lower().strip() for c in df.columns]
        
        # Định danh các cột cốt lõi bắt buộc phải xuất hiện trong file (lúc này đã được đổi tên thành chuẩn)
        required_cols = ['ma_giao_dich', 'ngay_baocao', 'trang_thai_ghi_nhan']
        if not all(col in df.columns for col in required_cols):
            flash('Import thất bại: Tệp Excel sai cấu trúc chuẩn! Yêu cầu phải có các cột: ma_giao_dich, ngay_baocao, trang_thai_ghi_nhan.', 'danger')
            return redirect(url_for('giaodien_import_phan_hoi'))
            
        engine = bc48.get_bc48_engine(db)
        success_count = 0
        
        with engine.begin() as transaction_conn:
            for _, row in df.iterrows():
                # Kiểm tra tính hợp lệ của trạng thái khắc phục do chi nhánh điền
                trang_thai_raw = str(row['trang_thai_ghi_nhan']).strip().upper() if not pd.isna(row['trang_thai_ghi_nhan']) else ""
                if not trang_thai_raw or trang_thai_raw in ['NAN', 'ALL', '']:
                    continue
                
                ma_gd_clean = str(row['ma_giao_dich']).strip()
                if pd.isna(row['ma_giao_dich']) or ma_gd_clean == "":
                    continue
                
                # Làm sạch giá trị ngày báo cáo về định dạng chuỗi chuẩn YYYY-MM-DD
                try:
                    if isinstance(row['ngay_baocao'], (datetime, pd.Timestamp)):
                        ngay_bc_clean = row['ngay_baocao'].strftime('%Y-%m-%d')
                    else:
                        ngay_bc_clean = str(row['ngay_baocao']).strip().split()[0]
                except:
                    continue

                # Xác định bảng gốc tìm thấy (khóa định danh phụ cùng ma_giao_dich)
                if 'bang_goc_tim_thay' in df.columns and not pd.isna(row['bang_goc_tim_thay']):
                    bang_goc = str(row['bang_goc_tim_thay']).strip().lower()
                else:
                    loai_bc_raw = str(row.get('loai_bc', 'ALL')).strip().lower()
                    bang_goc = f"{loai_bc_raw}_error" if loai_bc_raw in ['ctr', 'dwt', 'eft', 'ptr'] else "unknown_error"

                # CHỈ LẤY CÁC CỘT CẦN THIẾT THEO YÊU CẦU
                ma_hieu_2_val = str(row['ma_hieu_2']).strip() if 'ma_hieu_2' in df.columns and not pd.isna(row['ma_hieu_2']) else None
                ma_loi_f_ao_val = str(row['ma_loi_f_ao']).strip() if 'ma_loi_f_ao' in df.columns and not pd.isna(row['ma_loi_f_ao']) else None
                
                mota_loi_val = None
                if 'mota_loi' in df.columns and not pd.isna(row['mota_loi']):
                    mota_loi_val = str(row['mota_loi']).strip()
                elif 'mota_loi_don_le' in df.columns and not pd.isna(row['mota_loi_don_le']):
                    mota_loi_val = str(row['mota_loi_don_le']).strip()

                # Cột lý do nguyên nhân sai sót
                ghi_chu_raw = str(row.get('nguyen_nhan_sai_sot', '')).strip()
                ghi_chu_clean = None if ghi_chu_raw in ['nan', 'None', ''] else ghi_chu_raw

                # Thực thi ghi đè tiến trình: Lưu thêm NOW() vào ngày_nhan_file
                sql_save = text("""
                    INSERT INTO theo_doi_khac_phuc_loi (
                        ma_giao_dich, ngay_baocao, bang_goc_tim_thay, ma_hieu_2, ma_loi_f_ao, mota_loi,
                        trang_thai_khac_phuc, ghi_chu_xu_ly, id_dot, ngay_cap_nhat, ngay_nhan_file
                    ) VALUES (
                        :ma_giao_dich, :ngay_baocao, :bang_goc_tim_thay, :ma_hieu_2, :ma_loi_f_ao, :mota_loi,
                        :trang_thai, :ghi_chu, 0, NOW(), NOW()
                    )
                    ON DUPLICATE KEY UPDATE 
                        trang_thai_khac_phuc = :trang_thai,
                        ghi_chu_xu_ly = :ghi_chu,
                        ma_hieu_2 = COALESCE(:ma_hieu_2, ma_hieu_2),
                        ma_loi_f_ao = COALESCE(:ma_loi_f_ao, ma_loi_f_ao),
                        mota_loi = COALESCE(:mota_loi, mota_loi),
                        id_dot = IF(id_dot IS NULL OR id_dot = 0, 0, id_dot),
                        ngay_cap_nhat = NOW(),
                        ngay_nhan_file = NOW();
                """)
                
                transaction_conn.execute(sql_save, {
                    "ma_giao_dich": ma_gd_clean,
                    "ngay_baocao": ngay_bc_clean,
                    "bang_goc_tim_thay": bang_goc,
                    "ma_hieu_2": ma_hieu_2_val,
                    "ma_loi_f_ao": ma_loi_f_ao_val,
                    "mota_loi": mota_loi_val,
                    "trang_thai": trang_thai_raw,
                    "ghi_chu": ghi_chu_clean
                })
                success_count += 1
                
        if success_count > 0:
            flash(f"Hệ thống xử lý thành công! Đã ghi nhận thời gian nhận file và cập nhật tiến độ khắc phục cho {success_count} dòng.", "success")
        else:
            flash("Không có bản ghi hợp lệ nào được cập nhật.", "warning")
            
    except Exception as e:
        flash(f"Lỗi quy trình xử lý dữ liệu: {str(e)}", "danger")
        
    return redirect(url_for('giaodien_import_phan_hoi'))

####################################################################################
# KHÔNG DÙNG: Nút lệnh chỉ tách mô tả lỗi của CSV process_and_save_split_errors; process_extract_errors_only
####################################################################################
# Bảng ket_qua_do_tim_loi (id Khóa chính): Lưu kết quả sau khi chạy Procedure kiểm tra. Dữ liệu ở đây thường ở dạng thô, cột ma_loi và mota_loi còn bị gộp bởi dấu phẩy
# Bảng log_chi_tiet_loi_phan_tach (id_goc liên kết với id của bảng ket_qua_do_tim_loi): Lưu dữ liệu đã "làm sạch". Mỗi dòng chỉ có 1 mã lỗi duy nhất, giúp bạn làm báo cáo thống kê mã lỗi dễ dàng
# Bảng log_gui_mail_bc48 (ma_hieu_1 và ngay_gui liên kết logic): Lưu vết việc gửi mail. Giúp bạn trả lời câu hỏi: "File có chứa lỗi đó đã được gửi đi lúc nào và gửi cho ai?"
# Nếu ma_loi trống (độ dài = 0) và mota_loi có 3 giá trị (độ dài = 3), thì max_l sẽ là 3; Hệ thống thấy cột ma_loi đang thiếu 3 giá trị so với mức tối đa, nó sẽ tự động thêm 3 giá trị mặc định là "N/A"; Sau đó nó mới tiến hành tách thành 3 dòng
# Dù không có mã lỗi (ma_loi), nhưng chi nhánh vẫn nhận được đầy đủ 3 dòng mô tả lỗi để họ biết cần phải sửa những gì; Cả 3 dòng lỗi này vẫn được gắn chặt với ma_giao_dich và ma_hieu_1 ban đầu;
#(1) Trích xuất một tập dữ liệu dựa trên bộ lọc điều kiện (ví dụ: lọc theo loai_bc hoặc lọc theo thang_nam); (2) Chỉ xóa các dòng trong bảng log_chi_tiet_loi_phan_tach có id_goc nằm trong tập dữ liệu vừa lọc đó; (3) Tiến hành tách và nạp lại
@app.route('/boctach-mota-loi', methods=['GET'])
@login_required
@admin_or_user_sd_bc48
def giaodien_boctach_mota_loi():
    engine = bc48.get_bc48_engine(db)
    
    ma_giao_dich = request.args.get('ma_giao_dich', '').strip()
    trang_thai = request.args.get('trang_thai', '')
    filter_loai_bc = request.args.get('filter_loai_bc', '').strip() # Nhận tham số lọc loại BC

    ma_don_vi = request.args.get('ma_don_vi', '').strip()
    ngay_baocao = request.args.get('ngay_baocao', '').strip()
    ma_loi_f_ao = request.args.get('ma_loi_f_ao', '').strip()
    
    page = request.args.get('page', 1, type=int)
    per_page = 20

    # Lấy danh sách kèm theo đủ 3 bộ lọc
    data, total_rows = bc48.lay_danh_sach_log_loi(
        engine, 
        ma_giao_dich=ma_giao_dich, 
        trang_thai=trang_thai, 
        filter_loai_bc=filter_loai_bc,
        ma_don_vi=ma_don_vi,
        ngay_baocao=ngay_baocao,
        ma_loi_f_ao=ma_loi_f_ao,
        page=page, 
        per_page=per_page
    )
    
    total_pages = math.ceil(total_rows / per_page)
    
    # --- LẤY ĐỘNG LOAI_BC_LIST TỪ BẢNG KET_QUA_DO_TIM_LOI ---
    loai_bc_list = []
    try:
        # Truy vấn trực tiếp lấy danh sách loại báo cáo không trùng nhau từ bảng ket_qua_do_tim_loi
        sql_query = text("""
            SELECT DISTINCT loai_bc 
            FROM ket_qua_do_tim_loi 
            WHERE loai_bc IS NOT NULL AND loai_bc != ''
            ORDER BY loai_bc ASC
        """)
        with engine.connect() as connection:
            result = connection.execute(sql_query)
            # Lấy phần tử đầu tiên của mỗi dòng và chuẩn hóa viết hoa
            loai_bc_list = [r[0].upper() for r in result]
            
    except Exception as e:
        # Cơ chế fallback phòng ngừa lỗi bảng cấu trúc hoặc dữ liệu trống
        print(f"Lưu ý: Không lấy được danh sách loai_bc từ ket_qua_do_tim_loi ({str(e)}), dùng fallback trống.")
        loai_bc_list = []

    # Nếu bảng hoàn toàn trống, ta truy vấn dự phòng thêm từ bảng đích log_chi_tiet_loi_phan_tach để tăng tính chính xác
    if not loai_bc_list:
        try:
            sql_fallback = text("SELECT DISTINCT loai_bc FROM log_chi_tiet_loi_phan_tach WHERE loai_bc IS NOT NULL")
            with engine.connect() as connection:
                loai_bc_list = [r[0].upper() for r in connection.execute(sql_fallback)]
        except:
            pass
    
    current_month = datetime.now().strftime('%Y-%m') # Định dạng YYYY-MM cho thẻ input type="month"
    
    return render_template(
        'boctach_mota_loi.html',
        danh_sach_loi=data,
        ma_giao_dich=ma_giao_dich,
        trang_thai=trang_thai,
        filter_loai_bc=filter_loai_bc,
        page=page,
        total_pages=total_pages,
        total_rows=total_rows,
        loai_bc_list=loai_bc_list, 
        current_month=current_month 
    )

# nút lệnh "Chỉ bóc tách mota_loi của CSV" trên site: /boctach-mota-loi
@app.route('/xu-ly-boctach-mota-loi', methods=['POST'])
@login_required
@admin_or_user_sd_bc48
def xu_ly_boctach_mota_loi():
    # --- CHỐNG SPAM REQUEST SONG SONG BẰNG SESSION LOCK ---
    if session.get('dang_xu_ly_boctach') == True:
        flash("Hệ thống đang thực thi một tiến trình bóc tách mota_loi. Vui lòng đợi trong giây lát!", "warning")
        return redirect(url_for('giaodien_boctach_mota_loi'))

    # Bật cờ khóa tiến trình
    session['dang_xu_ly_boctach'] = True
    
    engine = bc48.get_bc48_engine(db)
    loai_bc = request.form.get('loai_bc', 'ALL')
    
    if request.form.get('all_months') == '1':
        raw_thang_nam = 'ALL'
    else:
        raw_thang_nam = request.form.get('thang_nam', 'ALL') 
    
    try:
        # Gọi hàm bóc tách chạy nền xử lý Vectorization
        ket_qua = bc48.process_extract_errors_only(engine, loai_bc, raw_thang_nam)
        
        if isinstance(ket_qua, list):
            flash(f"Đã thực hiện bóc tách mota_loi của CSV, dọn dẹp log cũ và phân tách dữ liệu lỗi thành công ({loai_bc} - Kỳ: {raw_thang_nam})!", "success")
        else:
            if "Lỗi hệ thống" in ket_qua:
                flash(ket_qua, "danger")
            else:
                flash(ket_qua, "warning")
                
    except Exception as e:
        flash(f"Có lỗi phát sinh trong quá trình xử lý bóc tách: {str(e)}", "danger")
        
    finally:
        # --- ĐẢM BẢO LUÔN GIẢI PHÓNG KHÓA KHI KẾT THÚC (Kể cả khi thành công hay gặp ngoại lệ crash) ---
        session.pop('dang_xu_ly_boctach', None)
        
    return redirect(url_for('giaodien_boctach_mota_loi'))


@app.route('/xuat-excel-boctach', methods=['GET'])
@login_required
@admin_or_user_sd_bc48
def xuat_excel_boctach():
    engine = bc48.get_bc48_engine(db)
    
    # Lấy các tham số lọc từ URL (cùng tên với bộ lọc trên giao diện)
    params = {
        'ma_giao_dich': request.args.get('ma_giao_dich', '').strip(),
        'trang_thai': request.args.get('trang_thai', ''),
        'filter_loai_bc': request.args.get('filter_loai_bc', '').strip(),
        'ma_don_vi': request.args.get('ma_don_vi', '').strip(),
        'ngay_baocao': request.args.get('ngay_baocao', '').strip(),
        'ma_loi_f_ao': request.args.get('ma_loi_f_ao', '').strip()
    }
    
    try:
        # Gọi hàm lấy dữ liệu toàn bộ (per_page=None để bỏ LIMIT)
        data, _ = bc48.lay_danh_sach_log_loi(engine, **params, page=1, per_page=None)
        
        if not data:
            flash("Không có dữ liệu phù hợp để xuất file!", "warning")
            return redirect(url_for('giaodien_boctach_mota_loi'))
        
        # Gọi hàm export từ bc48
        output, file_name = bc48.xuat_bao_cao_chi_tiet_loi_phan_tach(data)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=file_name
        )
        
    except Exception as e:
        flash(f"Lỗi khi xuất file Excel: {str(e)}", "danger")
        return redirect(url_for('giaodien_boctach_mota_loi'))



@app.route('/capnhat-trangthai-loi/<int:id_chi_tiet>', methods=['POST'])
@login_required
@admin_or_user_sd_bc48
def xu_ly_cap_nhat_loi(id_chi_tiet):
    engine = bc48.get_bc48_engine(db)
    
    trang_thai_moi = request.form.get('trang_thai_xuly')
    file_phan_hoi = request.form.get('file_phan_hoi_tu_cn', '').strip()
    
    try:
        bc48.cap_nhat_trang_thai_loi(
            engine, 
            id_chi_tiet, 
            trang_thai_moi, 
            file_phan_hoi if file_phan_hoi else None
        )
        flash("Cập nhật trạng thái xử lý lỗi thành công!", "success")
    except Exception as e:
        flash(f"Có lỗi xảy ra: {str(e)}", "danger")
        
    return redirect(request.referrer or url_for('giaodien_boctach_mota_loi'))
####################################################################################
# Dashboard để theo dõi trạng thái xử lý lỗi
####################################################################################
@app.route('/bc48/dashboard-theo-doi')
@login_required
@admin_or_user_sd_bc48
def dashboard_theo_doi_loi():
    engine = bc48.get_bc48_engine(db)
    
    # Gọi logic từ bc48.py
    stats, summary_totals, details = bc48.get_dashboard_stats(engine)

    return render_template('bc48_dashboard.html', 
                           stats=stats, 
                           totals=summary_totals,
                           details=details,
                           now=datetime.now())


################################################################################################################
# Xem bảng cau_hinh_file_nghiep_vu trên db bc48
################################################################################################################
@app.route('/admin/view-cau-hinh-nghiep-vu')
@login_required
def view_cau_hinh_nghiep_vu():
    try:
        # Gọi hàm từ module bc48, truyền db từ app.py vào
        data = bc48.view_cau_hinh_nghiep_vu(db)
        
        return render_template(
            'admin/view_cau_hinh_nghiep_vu.html', 
            data=data
        )
    except Exception as e:
        flash(f"Lỗi hệ thống: {str(e)}", "danger")
        return redirect(url_for('index'))

################################################################################################################
# Xem bảng log_loi_logic_LoaiKH_LoaiGT, kết quả của việc chạy sp_KiemTraLogic_LoaiKH_LoaiGT
# Thong ke so luong loi logic giua loai_khach_hang va loai_giay_to duoc dinh nghia mapping, chi tiet tai bang log_loi_logic_LoaiKH_LoaiGT
# Có số lượng >0 có nghĩa là TXT chưa chuẩn nhé
################################################################################################################
@app.route('/ket-qua-logic-kh-gt')
@login_required
@admin_or_user_sd_bc48
def ket_qua_logic_kh_gt():
    # 1. Lấy số trang từ URL (mặc định là trang 1)
    page = request.args.get('page', 1, type=int)
    per_page = 50  # Số bản ghi mỗi trang
    
    # 2. Lấy bộ lọc ngày từ URL, nếu không có thì mặc định lấy từ đầu tháng hiện hành đến ngày hôm nay
    default_tu_ngay = datetime.now().strftime('%Y-%m-01')
    default_den_ngay = datetime.now().strftime('%Y-%m-%d')
    
    tu_ngay = request.args.get('tu_ngay', default_tu_ngay)
    den_ngay = request.args.get('den_ngay', default_den_ngay)
    
    try:
        # Truyền thêm tham số lọc ngày vào hàm lấy dữ liệu dưới DB
        data, total_records = bc48.get_log_loi_logic_data(db, page, per_page, tu_ngay, den_ngay)
        
        # Tính toán tổng số trang
        total_pages = math.ceil(total_records / per_page) if total_records > 0 else 1
        
        return render_template('view_logic_error_kh_gt.html', 
                               data=data, 
                               page=page, 
                               total_pages=total_pages,
                               total_records=total_records,
                               tu_ngay=tu_ngay,
                               den_ngay=den_ngay)

    except Exception as e:
        flash(f"Lỗi hệ thống: {str(e)}", "danger")
        return redirect(url_for('index'))


@app.route('/run-check-logic', methods=['POST'])
@login_required
@admin_or_user_sd_bc48
def run_check_logic_LoaiKH_LoaiGT():
    # 1. Lấy tham số ngày chọn từ AJAX gửi lên
    tu_ngay = request.form.get('tu_ngay')
    den_ngay = request.form.get('den_ngay')
    
    if not tu_ngay or not den_ngay:
        return jsonify({
            'status': 'error',
            'message': 'Tham số đầu vào Từ ngày hoặc Đến ngày không được để trống.'
        }), 400
        
    # 2. Thực thi Procedure với tham số động do người dùng lựa chọn
    success = bc48.run_sp_kiem_tra_logic_LoaiKH_LoaiGT(db, tu_ngay, den_ngay)
    
    if success:
        return jsonify({
            'status': 'success', 
            'message': f'Đã hoàn thành rà soát logic dữ liệu từ ngày {tu_ngay} đến ngày {den_ngay} thành công!'
        })
    else:
        return jsonify({
            'status': 'error', 
            'message': 'Lỗi trong quá trình thực thi Procedure sp_KiemTraLogic_LoaiKH_LoaiGT trên Hệ quản trị cơ sở dữ liệu.'
        }), 500

################################################################################################################
# KQ tự dò tìm lỗi logic: LoaiTien != VND và SoTien = QuyDoi ; bảng log_loi_logic_TyGia lưu kết quả chạy sp_KiemTraLogic_TyGia
# duyệt qua toàn bộ các bảng nghiệp vụ có đuôi dạng _yyyymm (bắt đầu bằng ctr_, dwt_, eft_, ptr_), 
# kiểm tra lỗi logic: loaitien khác 'VND' nhưng sotien lại bằng quydoi (ngoại tệ thì số tiền giao dịch và số tiền quy đổi ra VND không thể bằng nhau)
# TRUNCATE log_loi_logic_TyGia;
# CALL sp_KiemTraLogic_TyGia;
################################################################################################################
@app.route('/ket-qua-logic-ty-gia')
@login_required
@admin_or_user_sd_bc48
def ket_qua_logic_ty_gia():
    # Lấy số trang và điều kiện lọc từ các tham số GET URL
    page = request.args.get('page', 1, type=int)
    per_page = 50  # Số bản ghi mỗi trang
    
    thang_nam_raw = request.args.get('thang_nam', '') # Định dạng nhận: 'YYYY-MM' hoặc rỗng
    ngay_baocao = request.args.get('ngay_baocao', '')   # Định dạng nhận: 'YYYY-MM-DD' hoặc rỗng
    
    # Chuẩn hóa định dạng tháng_nam về 'YYYYMM' để khớp khi gửi vào tầng xử lý DB bc48 nếu cần
    thang_nam_db = thang_nam_raw.replace('-', '') if thang_nam_raw else None
    ngay_baocao_db = ngay_baocao if ngay_baocao else None

    try:
        # Gọi hàm lấy dữ liệu từ module bc48 (Bổ sung truyền tham số lọc tháng/ngày)
        # Lưu ý: Bạn cần cập nhật hàm này tại module bc48 để thêm mệnh đề WHERE tương ứng khi truy vấn bảng log.
        data, total_records = bc48.get_log_loi_logic_ty_gia_data(
            db, page, per_page, thang_nam=thang_nam_db, ngay_baocao=ngay_baocao_db
        )
        
        # Tính toán tổng số trang
        total_pages = math.ceil(total_records / per_page)
        
        return render_template('view_logic_error_ty_gia.html', 
                               data=data, 
                               page=page, 
                               total_pages=total_pages,
                               total_records=total_records,
                               selected_thang_nam=thang_nam_raw,  # Trả lại giữ nguyên hiển thị trên ô Input HTML
                               selected_ngay_baocao=ngay_baocao)
    except Exception as e:
        flash(f"Lỗi: {str(e)}", "danger")
        return redirect(url_for('index'))


@app.route('/run-check-logic-ty-gia', methods=['POST'])
@login_required
@admin_or_user_sd_bc48
def run_check_logic_ty_gia():
    try:
        # Lấy tham số cấu hình gửi từ JSON AJAX body
        req_data = request.get_json() or {}
        thang_nam = req_data.get('thang_nam')      # Thường là chuỗi '202605' hoặc None
        ngay_baocao = req_data.get('ngay_baocao')  # Thường là chuỗi '2026-05-22' hoặc None

        # Gọi hàm thực thi Stored Procedure từ module bc48, có bổ sung chuyển đổi tham số
        # Tiến hành truyền thang_nam và ngay_baocao tương ứng vào tầng DB
        success, msg_detail = bc48.run_sp_kiem_tra_logic_ty_gia(db, thang_nam, ngay_baocao)
        
        if success:
            # Tạo chuỗi phản hồi trực quan dựa theo tham số vừa chọn
            scope_desc = "toàn bộ lịch sử"
            if ngay_baocao:
                scope_desc = f"ngày {ngay_baocao}"
            elif thang_nam:
                scope_desc = f"tháng {thang_nam}"
                
            return jsonify({
                'status': 'success', 
                'message': f'Đã hoàn thành rà soát logic Tỷ giá cho phạm vi: {scope_desc}!'
            })
        else:
            error_msg = msg_detail if msg_detail else 'Lỗi khi thực thi Procedure sp_KiemTraLogic_TyGia.'
            return jsonify({'status': 'error', 'message': error_msg}), 500
            
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Lỗi hệ thống backend: {str(e)}'}), 500

################################################################################################################
# Kiểm tra dữ liệu trong các bảng _yyyymm thay thế sp_TuDongKiemTraDuLieu (event 11h00 hàng ngày) trong mysql
# 1. Kiểm tra macn -> don_vi
# 2. Kiểm tra quoctich -> quoc_gia
# 3. Kiểm tra kieukh -> loai_khach_hang
# 4. Kiểm tra loaitien -> loai_tien
# 5. Kiểm tra loaigd -> ma_loai_nghiep_vu_pcrt
# 6. Kiểm tra kenhct -> kenh_chuyen_tien
# 7. Kiểm tra loaigt (DWT, EFT, PTR dùng 'loaigt') -> loai_giay_to
# 8. Kiểm tra loaigto (Riêng CTR dùng 'loaigto') -> loai_giay_to
# 9. Kiểm tra loaihanghoa -> loai_hang_hoa
# 10. Kiểm tra loaitk -> loai_tai_khoan
# TRUNCATE danh_sach_bang_da_kiem_tra;
# TRUNCATE log_kiem_tra_du_lieu;
### 🔒 Lưu ý về logic bảo toàn:
# Khi bạn gõ đích danh bảng hoặc chọn quét `all`, hệ thống sẽ **không lưu** bản ghi vào bảng `danh_sach_bang_da_kiem_tra` sau khi chạy xong.
# Điều này giúp bảo toàn cơ chế quét tự động: các bảng đó vẫn được giữ nguyên trạng thái để lần sau bạn gõ Enter thì hệ thống vẫn nhận diện nó là bảng cần quét. 
# Lệnh `INSERT IGNORE` được sử dụng để tránh lỗi trùng lặp dữ liệu (`Duplicate entry`) nếu có xung đột chỉ mục bảng.
################################################################################################################
@app.route('/ket-qua-kiem-tra-danh-muc')
@login_required
@admin_or_user_sd_bc48
def ket_qua_kiem_tra_danh_muc():
    # 1. Lấy tham số phân trang và lọc dữ liệu từ URL
    page = request.args.get('page', 1, type=int)
    per_page = 50  # Số bản ghi hiển thị trên mỗi trang
    
    thang_nam_raw = request.args.get('thang_nam', '')  # Nhận định dạng: 'YYYY-MM' hoặc rỗng
    ngay_baocao = request.args.get('ngay_baocao', '')   # Nhận định dạng: 'YYYY-MM-DD' hoặc rỗng
    
    # Chuẩn hóa 'YYYY-MM' thành 'YYYYMM' để khớp chuỗi đuôi tên bảng (_202605)
    thang_nam_db = thang_nam_raw.replace('-', '') if thang_nam_raw else None
    
    # Chuẩn hóa 'YYYY-MM-DD' thành 'YYYYMMDD' để lọc chuẩn cột thoidiem (ví dụ: '20260520%')
    ngay_baocao_db = ngay_baocao.replace('-', '') if ngay_baocao else None

    try:
        # 2. Gọi tầng xử lý dữ liệu từ bc48.py
        data, total_records = bc48.get_log_kiem_tra_du_lieu_data(
            db, page, per_page, thang_nam=thang_nam_db, ngay_baocao=ngay_baocao_db
        )

        # 3. LẤY DỮ LIỆU THỐNG KÊ DASHBOARD (Mới bổ sung)
        dashboard_summary = bc48.get_dashboard_summary_danh_muc(
            db, thang_nam=thang_nam_db, ngay_baocao=ngay_baocao_db
        )
        
        # 4. Tính toán tổng số trang
        total_pages = math.ceil(total_records / per_page)
        
        return render_template(
            'view_loi_khong_thuoc_danh_muc.html', 
            data=data, 
            page=page, 
            total_pages=total_pages,
            total_records=total_records,
            dashboard_summary=dashboard_summary,
            selected_thang_nam=thang_nam_raw,  # Giữ trạng thái hiển thị trên Form lọc HTML
            selected_ngay_baocao=ngay_baocao
        )
    except Exception as e:
        flash(f"Lỗi hệ thống: {str(e)}", "danger")
        return redirect(url_for('index'))


@app.route('/run-check-danh-muc', methods=['POST'])
@login_required
@admin_or_user_sd_bc48
def run_check_danh_muc():
    global IS_CHECKING_DANH_MUC

    if IS_CHECKING_DANH_MUC:
        return jsonify({
            'status': 'error', 
            'message': 'Hệ thống đang có một tiến trình rà soát danh mục đang chạy ngầm. Vui lòng đợi trong giây lát!'
        }), 429  # Too Many Requests
    
    try:
        IS_CHECKING_DANH_MUC = True
        
        # Lấy tham số cấu hình gửi lên từ AJAX JSON Body
        # "Nếu chọn Ngày -> Chỉ quét ngày đó; Nếu chọn Tháng -> Chỉ quét các bảng thuộc tháng đó; Nếu trống cả hai -> Mới quét toàn bộ lịch sử"
        req_data = request.get_json() or {}
        thang_nam = req_data.get('thang_nam')      # Dạng chuỗi '202605' hoặc None
        ngay_baocao = req_data.get('ngay_baocao')  # Dạng chuỗi '2026-05-20' hoặc None/Rỗng

        # Chuẩn hóa ngày thành chuỗi 'YYYYMMDD' cho khớp cấu trúc tìm kiếm text của thoidiem
        target_date = ngay_baocao.replace('-', '') if ngay_baocao else ""

        # Gọi hàm xử lý cốt lõi từ module bc48.py
        success, msg_detail = bc48.process_run_check_danh_muc(db, target_date=target_date, target_thang=thang_nam)
        
        if success:
            scope_desc = "toàn bộ lịch sử hệ thống"
            if target_date:
                scope_desc = f"ngày {ngay_baocao}"
            elif thang_nam:
                scope_desc = f"tháng {thang_nam[:4]}/{thang_nam[4:]}"
                
            return jsonify({
                'status': 'success', 
                'message': f'Đã hoàn thành quét rà soát danh mục dữ liệu báo cáo cho phạm vi: {scope_desc}!'
            })
        else:
            error_msg = msg_detail if msg_detail else 'Gặp lỗi trong tiến trình đối chiếu danh mục dữ liệu.'
            return jsonify({'status': 'error', 'message': error_msg}), 500
            
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Lỗi hệ thống backend: {str(e)}'}), 500
        
    finally:
        # Khối này luôn chạy cuối cùng, bất kể tiến trình thành công, thất bại hay crash hệ thống
        IS_CHECKING_DANH_MUC = False

@app.route('/xuat-excel-loi-danh-muc')
@login_required
@admin_or_user_sd_bc48
def xuat_excel_loi_danh_muc():
    """Route xử lý truy xuất toàn bộ dữ liệu lỗi theo bộ lọc và kết xuất file Excel trực tiếp"""
    thang_nam_raw = request.args.get('thang_nam', '')
    ngay_baocao = request.args.get('ngay_baocao', '')
    
    thang_nam_db = thang_nam_raw.replace('-', '') if thang_nam_raw else None
    ngay_baocao_db = ngay_baocao.replace('-', '') if ngay_baocao else None

    try:
        # Gọi hàm lấy toàn bộ dữ liệu không phân trang phục vụ xuất Excel
        raw_error_data = bc48.get_all_log_errors_for_excel(
            db, thang_nam=thang_nam_db, ngay_baocao=ngay_baocao_db
        )
        
        if not raw_error_data:
            flash("Không có dữ liệu lỗi danh mục để xuất file Excel trong phạm vi bộ lọc!", "info")
            return redirect(url_for('ket_qua_kiem_tra_danh_muc', thang_nam=thang_nam_raw, ngay_baocao=ngay_baocao))
            
        # Chuyển đổi sang DataFrame của Pandas để xử lý cấu trúc cột
        df = pd.DataFrame(raw_error_data)
        
        # Định nghĩa lại tiêu đề cột bằng Tiếng Việt ký sự chuyên nghiệp
        df.columns = [
            'ID Hệ Thống', 
            'Tên Bảng Gốc Dữ Liệu', 
            'Cột Dữ Liệu Phát Hiện Lỗi', 
            'Giá Trị Lỗi Hệ Thống / Trống', 
            'Mã Giao Dịch (magd)', 
            'Thời Điểm Phát Sinh Giao Dịch', 
            'Thời Điểm Hệ Thống Quét Log'
        ]
        
        # Xử lý ghi dữ liệu trực tiếp vào bộ nhớ đệm RAM (BytesIO) để xuất file tốc độ cao
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Lỗi Danh Mục Dữ Liệu AML', index=False)
        output.seek(0)
        
        # Thiết lập tên file động dựa trên cấu hình bộ lọc người dùng chọn
        suffix_name = ngay_baocao_db if ngay_baocao_db else (thang_nam_db if thang_nam_db else "Toan_Bo_Lich_Su")
        filename = f"Bao_Cao_Loi_Danh_Muc_AML_{suffix_name}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        flash(f"Lỗi trong quá trình xuất file Excel: {str(e)}", "danger")
        return redirect(url_for('ket_qua_kiem_tra_danh_muc', thang_nam=thang_nam_raw, ngay_baocao=ngay_baocao))

################################################################################################################
# Tao Procedure thong ke hinh thuc GLD, GBS, GLA tu tat ca bang _yyyymm
# Thống kê theo hình thức file gửi (GLD, GBS, GLA) từ tất cả các bảng ctr_yyyymm; dwt_yyyymm; eft_yyyymm; ptr_yyyymm
# Trong khoảng thời gian có bao nhiêu dòng của loại báo cáo nào được GLA?
################################################################################################################
@app.route('/bc48/dashboard-hinh-thuc')
@login_required
@admin_or_user_sd_bc48
def bc48_dashboard_hinh_thuc():
    # Gọi hàm xử lý logic từ module bc48 đã import ở đầu file, truyền vào đối tượng db của app.py
    matrix_data, ratio_data, hinh_thuc_list, details_data = bc48.get_dashboard_hinh_thuc_data(db)
    
    # Trả giao diện HTML kèm dữ liệu sạch
    return render_template(
        'dashboard_hinh_thuc.html', 
        matrix=matrix_data, 
        ratio=ratio_data,
        hinh_thuc_list=hinh_thuc_list,
        details=details_data
    )

@app.route('/bc48/sao-ke-gla')
@login_required
@admin_or_user_sd_bc48
def bc48_sao_ke_gla():
    # Nhận tham số từ request
    start = int(request.args.get('start', 0))
    length = int(request.args.get('length', 25))
    search_params = {
        "hinh_thuc": request.args.get('hinh_thuc'),
        "loai_bc": request.args.get('loai_bc'),
        "tu_ngay": request.args.get('tu_ngay'),
        "den_ngay": request.args.get('den_ngay')
    }

    # Gọi hàm đã viết ở bc48.py
    data, total = bc48.get_sao_ke_gla(db, search_params, start, length)
    
    return jsonify({
        "draw": int(request.args.get('draw', 1)),
        "recordsTotal": total,
        "recordsFiltered": total,
        "data": data
    })

# Tab "SLgFile, SLgDong, HinhThuc, quydoi 4 loai_bc" lấy thông tin toàn bộ cột trong bảng dashboard_tke_slgfile_sodong_hinh_thuc
@app.route('/bc48/api-tke-slgfile')
@login_required
@admin_or_user_sd_bc48
def bc48_api_tke_slgfile():
    start = int(request.args.get('start', 0))
    length = int(request.args.get('length', 25))
    search_value = request.args.get('search[value]', '') # Lấy ô tìm kiếm mặc định của DataTable

    data, total = bc48.get_dashboard_tke_slgfile(db, start, length, search_value)
    
    return jsonify({
        "draw": int(request.args.get('draw', 1)),
        "recordsTotal": total,
        "recordsFiltered": total,
        "data": data
    })


################################################################################################################
# Thống kê từng ngày số lượng dòng KTraGDLoi, KTraFileThanhCong của từng loại báo cáo (CTR; DWT; EFT; PTR); bảng bang_thong_ke_loi_daily kết quả của sp_tong_hop_sodong_bang_error
################################################################################################################
@app.route('/bc48/thong-ke-trang-thai') # Route 1: Hiển thị trang giao diện (chỉ render HTML)
@login_required
@admin_or_user_sd_bc48
def bc48_thong_ke_trang_thai():
    return render_template('thong_ke_trang_thai.html')

@app.route('/api/bc48/thong-ke-trang-thai') # Route 2: API cung cấp dữ liệu cho DataTables
@login_required
@admin_or_user_sd_bc48
def api_thong_ke_trang_thai():
    start = int(request.args.get('start', 0))
    length = int(request.args.get('length', 20))
    search_value = request.args.get('search[value]') # Lấy nội dung ô Search
    tu_ngay = request.args.get('tu_ngay')
    den_ngay = request.args.get('den_ngay')
    
    data, total = bc48.get_thong_ke_trang_thai_data(db, start, length, search_value, tu_ngay, den_ngay)
    
    return jsonify({
        "draw": int(request.args.get('draw', 1)),
        "recordsTotal": total,
        "recordsFiltered": total, # Có thể dùng để lọc tìm kiếm nếu cần
        "data": data
    })

@app.route('/api/bc48/export-thong-ke')
@login_required
@admin_or_user_sd_bc48
def export_thong_ke_to_excel_route():
    tu_ngay = request.args.get('tu_ngay')
    den_ngay = request.args.get('den_ngay')
    
    # Lấy thời gian hiện tại để làm timestamp cho tên file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Tạo tên file: Thong_Ke_Trang_Thai_20260608_214530.xlsx
    # Hoặc bao gồm cả khoảng thời gian lọc nếu cần
    file_name = f"SoDong_LOI_tung_ngay_cac_loai_baocao_{timestamp}.xlsx"
    
    # Gọi hàm từ bc48
    excel_file = bc48.export_thong_ke_trang_thai_to_excel(db, tu_ngay, den_ngay)
    
    return send_file(excel_file, download_name=file_name, as_attachment=True)

################################################################################################################
# Truy vấn thông tin bảng thongtinkhachhang (kết quả chạy của sp_Laythongtinkhachhang; evt_Lay_thongtin_khachhang)
################################################################################################################
@app.route('/ttkh_bc48', methods=['GET'])
@login_required
@admin_or_user_sd_bc48
def ttkh_bc48():
    # Kiểm tra quyền Admin
    if not current_user.is_admin:
        abort(403)
        
    # Lấy các tham số tìm kiếm từ request
    sogt = request.args.get('sogt', '').strip()
    tenkh = request.args.get('tenkh', '').strip()
    data_month = request.args.get('data_month', '').strip()
    
    # Xử lý phân trang
    try:
        page = int(request.args.get('page', 1))
    except ValueError:
        page = 1
    per_page = 50  # Số lượng bản ghi trên một trang

    # Gọi tầng xử lý nghiệp vụ tại bc48.py
    # Truyền tham số db vào để thực thi truy vấn qua Engine bind 'db_bc48'
    result = bc48.get_thong_tin_khach_hang(
        db=db, 
        sogt=sogt, 
        tenkh=tenkh, 
        data_month=data_month, 
        page=page, 
        per_page=per_page
    )

    return render_template(
        'ttkh_bc48.html',  # Bạn cần tạo thêm template giao diện này để hiển thị bảng dữ liệu
        data=result['data'],
        total=result['total'],
        page=page,
        per_page=per_page,
        total_pages=result['total_pages'],
        sogt=sogt,
        tenkh=tenkh,
        data_month=data_month
    )

@app.route('/ttkh_bc48/export', methods=['GET'])
@login_required
@admin_or_user_sd_bc48
def ttkh_bc48_export():
    # 1. Kiểm tra quyền bảo mật
    if not current_user.is_admin:
        abort(403)
        
    # 2. Tiếp nhận tham số lọc từ URL do Ajax/Giao diện gửi lên
    sogt = request.args.get('sogt', '').strip()
    tenkh = request.args.get('tenkh', '').strip()
    data_month = request.args.get('data_month', '').strip()

    # 3. Tạo tên file động theo thời gian thực kết xuất
    filename = f"XUAT_DULIEU_TTKH_BC48_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    # 4. Trả về luồng Response đồng thời gọi trực tiếp Generator từ tầng nghiệp vụ bc48.py
    response = Response(
        stream_with_context(bc48.generate_ttkh_csv_stream(db, sogt, tenkh, data_month)),
        mimetype="text/csv; charset=utf-8"
    )
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


# ----------------------------------------------------------------------
# CHẠY ỨNG DỤNG FLASK
# ----------------------------------------------------------------------
if __name__ == '__main__':
    ##app.run(host='0.0.0.0', port=5001, debug=False) # Chạy ứng dụng trên cổng 5001
    # SỬ DỤNG socketio.run thay vì app.run
    # Lưu ý: use_reloader=False cực kỳ quan trọng khi dùng Watchdog 
    # để tránh việc tạo 2 tiến trình giám sát cùng lúc gây lỗi khóa file.
    socketio.run(app, host='0.0.0.0', port=5001, debug=False, use_reloader=False)
